import os, csv
from datetime import date
from collections import defaultdict
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

dest = r'C:\Users\DRKairport\OneDrive - Deutsches Rotes Kreuz - Kreisverband KÃ¶ln e.V\Desktop\bei\csv'
os.makedirs(dest, exist_ok=True)

einsaetze = [
    {
        "datum": "09.03.2026", "stichwort": "Einsatz: Chirurgisch 1",
        "entnehmer": "Blei", "artikel": [
            ("Aluderm VerbandpÃ¤ckchen mittel 4m x 8cm", 2),
            ("aluderm Kompresse 10 x 10 cm", 3),
            ("Einmalhandschuhe L", 4),
        ]
    },
    {
        "datum": "09.03.2026", "stichwort": "Einsatz: Intern 1",
        "entnehmer": "Granrath", "artikel": [
            ("NaCl 0,9% 500ml", 1),
            ("VenenverweilkanÃ¼le 18G", 2),
        ]
    },
    {
        "datum": "15.03.2026", "stichwort": "Aus- / Fortbildung: EH-Ãœbung",
        "entnehmer": "Idic", "artikel": [
            ("Dreiecktuch", 5),
            ("VerbandpÃ¤ckchen groÃŸ", 3),
            ("Rettungsdecke silber/gold", 2),
        ]
    },
    {
        "datum": "20.03.2026", "stichwort": "Ablauf / MHD Ã¼berschritten",
        "entnehmer": "", "artikel": [
            ("NaCl 0,9% 500ml", 4),
            ("Einmalhandschuhe M", 10),
        ]
    },
    {
        "datum": "02.04.2026", "stichwort": "Einsatz: Chirurgisch 1",
        "entnehmer": "GroÃŸ", "artikel": [
            ("Aluderm VerbandpÃ¤ckchen mittel 4m x 8cm", 1),
            ("Wundschnellverband 6cm x 10m", 2),
            ("Pflaster-Strips sortiert", 6),
        ]
    },
]

# â”€â”€â”€ Stil-Konstanten â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
CLR_DRK_ROT   = "00CC0000"   # DRK-Rot
CLR_HEADER_BG = "001565A8"   # Blau (wie Nesk3-UI)
CLR_HEADER_FG = "00FFFFFF"
CLR_EINSATZ   = "00D6E4F7"   # Hellblau â€“ Einsatz-Gruppe
CLR_EINSATZ2  = "00EAF4FF"
CLR_ITEM_ODD  = "00F7FBFF"
CLR_ITEM_EVEN = "00FFFFFF"
CLR_SUMME_BG  = "00E8F5E9"
CLR_STAT_BG   = "00FFF8E1"
CLR_TITLE_BG  = "001A237E"   # Dunkelblau Titel

def thin_border():
    s = Side(style="thin", color="00CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(color=CLR_HEADER_BG): return PatternFill("solid", fgColor=color)
def bg_fill(color):                    return PatternFill("solid", fgColor=color)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A3"):
    ws.freeze_panes = cell

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  Variante A â€“ Klassisch strukturiert
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Verbrauch"

# Titelzeile
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "ðŸ§°  Sanitätsmaterial-Verbrauchsprotokoll â€“ DRK Erste-Hilfe-Station FKB"
c.font = Font(bold=True, size=14, color=CLR_HEADER_FG)
c.fill = bg_fill(CLR_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Datum-Zeile
ws.merge_cells("A2:E2")
c = ws["A2"]
c.value = f"Exportiert am: {date.today().strftime('%d.%m.%Y')}   |   Zeitraum: 01.03.2026 â€“ 02.04.2026"
c.font = Font(italic=True, size=10, color="00555555")
c.fill = bg_fill("00E8EAF6")
c.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[2].height = 18

# Spalten-Header
headers = ["Datum", "Artikel", "Menge", "Entnehmer", "Einsatz / Grund"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color=CLR_HEADER_FG, size=11)
    c.fill = header_fill()
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws.row_dimensions[3].height = 22

r = 4
group_colors = [CLR_EINSATZ, CLR_EINSATZ2]
for gi, e in enumerate(einsaetze):
    grp_bg = group_colors[gi % 2]
    # Gruppen-Header
    for col in range(1, 6):
        c = ws.cell(row=r, column=col)
        c.fill = bg_fill("001976D2" if "Einsatz" in e["stichwort"] else "00388E3C")
        c.border = thin_border()
    ws.cell(row=r, column=1, value=e["datum"]).font = Font(bold=True, color=CLR_HEADER_FG)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    c2 = ws.cell(row=r, column=2, value=f"  {e['stichwort']}")
    c2.font = Font(bold=True, color=CLR_HEADER_FG, size=11)
    ws.merge_cells(f"B{r}:E{r}")
    r += 1
    for i, (art, menge) in enumerate(e["artikel"]):
        ibg = CLR_ITEM_ODD if i % 2 == 0 else CLR_ITEM_EVEN
        ws.cell(row=r, column=1, value="").fill = bg_fill(ibg)
        c_art = ws.cell(row=r, column=2, value=f"    â†³  {art}")
        c_art.fill = bg_fill(ibg)
        c_menge = ws.cell(row=r, column=3, value=menge)
        c_menge.font = Font(bold=True, color="00C0392B")
        c_menge.alignment = Alignment(horizontal="center")
        c_menge.fill = bg_fill(ibg)
        ws.cell(row=r, column=4, value=e["entnehmer"] if i == 0 else "").fill = bg_fill(ibg)
        ws.cell(row=r, column=5, value="").fill = bg_fill(ibg)
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border()
        r += 1
    # Summenzeile
    gesamt = sum(m for _, m in e["artikel"])
    for col in range(1, 6): ws.cell(row=r, column=col).fill = bg_fill(CLR_SUMME_BG)
    ws.cell(row=r, column=2, value=f"  Î£ {len(e['artikel'])} Artikel").font = Font(italic=True, color="00555555")
    c_sum = ws.cell(row=r, column=3, value=gesamt)
    c_sum.font = Font(bold=True, color="00388E3C")
    c_sum.alignment = Alignment(horizontal="center")
    for col in range(1, 6): ws.cell(row=r, column=col).border = thin_border()
    r += 1
    ws.row_dimensions[r].height = 6  # LÃ¼cke
    r += 1

set_col_widths(ws, [12, 42, 8, 15, 28])
freeze(ws, "A4")
ws.sheet_view.showGridLines = False

wb.save(os.path.join(dest, "vA_klassisch_strukturiert.xlsx"))
print("vA gespeichert")

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  Variante B â€“ Nach Datum + Gesamtstatistik
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Verbrauch"

ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "Sanitätsmaterial-Verbrauchsprotokoll â€“ DRK Erste-Hilfe-Station FKB"
c.font = Font(bold=True, size=14, color=CLR_HEADER_FG)
c.fill = bg_fill("00334155")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:F2")
c = ws2["A2"]
c.value = f"Exportiert am: {date.today().strftime('%d.%m.%Y')}"
c.font = Font(italic=True, size=10, color="00666666")
c.fill = bg_fill("00F5F5F5")
c.alignment = Alignment(horizontal="right")

headers2 = ["Datum", "Einsatz / Grund", "Artikel", "Menge", "Entnehmer", "Notiz"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color=CLR_HEADER_FG)
    c.fill = header_fill()
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()
ws2.row_dimensions[3].height = 20

nach_datum = defaultdict(list)
for e in einsaetze:
    nach_datum[e["datum"]].append(e)

artikel_gesamt = defaultdict(int)
r2 = 4
date_colors = ["00F5F5F5", "00FAFAFA"]
for di, datum in enumerate(sorted(nach_datum.keys(), key=lambda d: tuple(reversed(d.split("."))))):  
    # Datums-Trennzeile
    ws2.merge_cells(f"A{r2}:F{r2}")
    c = ws2.cell(row=r2, column=1, value=f"  {datum}")
    c.font = Font(bold=True, size=11, color="00FFFFFF")
    c.fill = bg_fill("00546E7A")
    c.border = thin_border()
    ws2.row_dimensions[r2].height = 18
    r2 += 1

    item_bg = date_colors[di % len(date_colors)]
    for e in nach_datum[datum]:
        for i, (art, menge) in enumerate(e["artikel"]):
            ws2.cell(row=r2, column=1, value=datum if i == 0 else "").fill = bg_fill(item_bg)
            ws2.cell(row=r2, column=2, value=e["stichwort"] if i == 0 else "").fill = bg_fill(item_bg)
            ws2.cell(row=r2, column=3, value=art).fill = bg_fill(item_bg)
            c_m = ws2.cell(row=r2, column=4, value=menge)
            c_m.font = Font(bold=True, color="00263238")
            c_m.alignment = Alignment(horizontal="center")
            c_m.fill = bg_fill(item_bg)
            ws2.cell(row=r2, column=5, value=e["entnehmer"] if i == 0 else "").fill = bg_fill(item_bg)
            ws2.cell(row=r2, column=6, value="").fill = bg_fill(item_bg)
            for col in range(1, 7): ws2.cell(row=r2, column=col).border = thin_border()
            artikel_gesamt[art] += menge
            r2 += 1
    r2 += 1  # Leerzeile

# GesamtÃ¼bersicht alle Artikel
r2 += 1
ws2.merge_cells(f"A{r2}:F{r2}")
c = ws2.cell(row=r2, column=1, value="  Alle verbrauchten Artikel â€“ GesamtÃ¼bersicht")
c.font = Font(bold=True, size=12, color=CLR_HEADER_FG)
c.fill = bg_fill("00455A64")
c.border = thin_border()
ws2.row_dimensions[r2].height = 20
r2 += 1

for col, h in enumerate(["Artikel", "", "", "Verbraucht", "Einheit", ""], 1):
    c = ws2.cell(row=r2, column=col, value=h)
    c.font = Font(bold=True, color=CLR_HEADER_FG)
    c.fill = bg_fill("00607D8B")
    c.border = thin_border()
r2 += 1

for i, (art, menge) in enumerate(sorted(artikel_gesamt.items(), key=lambda x: x[0])):
    bg = "00F5F5F5" if i % 2 == 0 else "00FFFFFF"
    ws2.cell(row=r2, column=1, value=f"  {art}").fill = bg_fill(bg)
    ws2.cell(row=r2, column=1).border = thin_border()
    c_m = ws2.cell(row=r2, column=4, value=menge)
    c_m.font = Font(bold=True)
    c_m.alignment = Alignment(horizontal="center")
    c_m.fill = bg_fill(bg)
    c_m.border = thin_border()
    ws2.cell(row=r2, column=5, value="StÃ¼ck").fill = bg_fill(bg)
    ws2.cell(row=r2, column=5).border = thin_border()
    for col in [2, 3, 6]:
        ws2.cell(row=r2, column=col).fill = bg_fill(bg)
        ws2.cell(row=r2, column=col).border = thin_border()
    r2 += 1

set_col_widths(ws2, [12, 30, 42, 8, 15, 20])
freeze(ws2, "A4")
ws2.sheet_view.showGridLines = False

wb2.save(os.path.join(dest, "vB_nach_datum_mit_statistik.xlsx"))
print("vB gespeichert")

print("\nAlle Dateien:")
for f in sorted(os.listdir(dest)):
    print(f"  {f}  ({os.path.getsize(os.path.join(dest,f))} Bytes)")


einsaetze = [
    {
        "datum": "09.03.2026", "stichwort": "Einsatz: Chirurgisch 1",
        "entnehmer": "Blei", "artikel": [
            ("Aluderm VerbandpÃ¤ckchen mittel 4m x 8cm", 2),
            ("aluderm Kompresse 10 x 10 cm", 3),
            ("Einmalhandschuhe L", 4),
        ]
    },
    {
        "datum": "09.03.2026", "stichwort": "Einsatz: Intern 1",
        "entnehmer": "Granrath", "artikel": [
            ("NaCl 0,9% 500ml", 1),
            ("VenenverweilkanÃ¼le 18G", 2),
        ]
    },
    {
        "datum": "15.03.2026", "stichwort": "Aus- / Fortbildung: EH-Ãœbung",
        "entnehmer": "Idic", "artikel": [
            ("Dreiecktuch", 5),
            ("VerbandpÃ¤ckchen groÃŸ", 3),
            ("Rettungsdecke silber/gold", 2),
        ]
    },
    {
        "datum": "20.03.2026", "stichwort": "Ablauf / MHD Ã¼berschritten",
        "entnehmer": "", "artikel": [
            ("NaCl 0,9% 500ml", 4),
            ("Einmalhandschuhe M", 10),
        ]
    },
    {
        "datum": "02.04.2026", "stichwort": "Einsatz: Chirurgisch 1",
        "entnehmer": "GroÃŸ", "artikel": [
            ("Aluderm VerbandpÃ¤ckchen mittel 4m x 8cm", 1),
            ("Wundschnellverband 6cm x 10m", 2),
            ("Pflaster-Strips sortiert", 6),
        ]
    },
]

# â”€â”€ Variante 1: Flach (aktuell) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with open(os.path.join(dest, "v1_flach_aktuell.csv"), "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["Datum", "Artikel", "Menge", "Entnehmer", "Bemerkung"])
    for e in einsaetze:
        for art, menge in e["artikel"]:
            w.writerow([e["datum"], art, menge, e["entnehmer"], e["stichwort"]])

# â”€â”€ Variante 2: Einsatz als Trennzeile â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with open(os.path.join(dest, "v2_einsatz_als_kopfzeile.csv"), "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["Datum", "Artikel", "Menge", "Entnehmer", "Einsatz / Grund"])
    for e in einsaetze:
        w.writerow([e["datum"], f">>> {e['stichwort']} <<<", "", e["entnehmer"], ""])
        for art, menge in e["artikel"]:
            w.writerow(["", f"    {art}", menge, "", ""])
        w.writerow(["", "", "", "", ""])

# â”€â”€ Variante 3: Strukturiert mit Summe je Einsatz â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with open(os.path.join(dest, "v3_mit_einsatzsummen.csv"), "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["Sanitätsmaterial-Verbrauchsprotokoll â€“ DRK Erste-Hilfe-Station Flughafen KÃ¶ln/Bonn"])
    w.writerow([f"Exportiert am: {date.today().strftime('%d.%m.%Y')}"])
    w.writerow([])
    w.writerow(["Datum", "Artikel", "Menge", "Entnehmer", "Einsatz / Grund"])
    for e in einsaetze:
        w.writerow([e["datum"], f"=== {e['stichwort']} ===", "", e["entnehmer"], ""])
        gesamt = 0
        for art, menge in e["artikel"]:
            w.writerow(["", f"  {art}", menge, "", ""])
            gesamt += menge
        w.writerow(["", f"  Gesamt: {gesamt} Einheiten", "", "", ""])
        w.writerow([])

# â”€â”€ Variante 4: Pro Tag gruppiert + Gesamtstatistik â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with open(os.path.join(dest, "v4_nach_datum_mit_statistik.csv"), "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["Sanitätsmaterial-Verbrauchsprotokoll â€“ DRK Erste-Hilfe-Station Flughafen KÃ¶ln/Bonn"])
    w.writerow([f"Exportiert am: {date.today().strftime('%d.%m.%Y')}"])
    w.writerow([f"Zeitraum: 01.03.2026 â€“ 02.04.2026"])
    w.writerow([])
    w.writerow(["Datum", "Einsatz / Grund", "Artikel", "Menge", "Entnehmer"])

    # Nach Datum gruppieren
    from collections import defaultdict
    nach_datum = defaultdict(list)
    for e in einsaetze:
        nach_datum[e["datum"]].append(e)

    artikel_gesamt = defaultdict(int)
    for datum in sorted(nach_datum.keys(), key=lambda d: tuple(reversed(d.split(".")))):
        w.writerow([f"--- {datum} ---", "", "", "", ""])
        for e in nach_datum[datum]:
            for i, (art, menge) in enumerate(e["artikel"]):
                stichwort = e["stichwort"] if i == 0 else ""
                entnehmer = e["entnehmer"] if i == 0 else ""
                w.writerow(["", stichwort, f"  {art}", menge, entnehmer])
                artikel_gesamt[art] += menge
        w.writerow([])

    # Gesamtstatistik
    w.writerow(["GESAMTÃœBERSICHT"])
    w.writerow(["", "Artikel", "", "Gesamt verbraucht", ""])
    for art, menge in sorted(artikel_gesamt.items(), key=lambda x: -x[1]):
        w.writerow(["", art, "", menge, ""])

print("Erstellt:")
for fname in sorted(os.listdir(dest)):
    fullp = os.path.join(dest, fname)
    print(f"  {fname}  ({os.path.getsize(fullp)} Bytes)")
