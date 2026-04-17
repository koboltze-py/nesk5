"""
Schulungen-Kalender-Widget
Großer Monatskalender mit farbkodierten Ablaufterminen für Mitarbeiter-Schulungen.
Farben: Gelb (3 Monate), Orange (2 Monate), Rot (1 Monat), Dunkelrot (abgelaufen).
"""
import calendar
from datetime import date, datetime, timedelta

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QDialog, QFormLayout, QLineEdit, QComboBox,
    QDateEdit, QTextEdit, QScrollArea, QSizePolicy, QMessageBox,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QSplitter, QGroupBox, QCheckBox, QProgressDialog,
)
from PySide6.QtCore import Qt, QDate, QSize, Signal
from PySide6.QtGui import QFont, QColor, QPainter, QPen, QBrush, QCursor

import os

from config import FIORI_BLUE, BASE_DIR as _SCHULUNG_BASE_DIR

_SCHULUNG_EXPORT_DIR = os.path.join(_SCHULUNG_BASE_DIR, "Daten", "Schulungen")

# ─── Farb-Konstanten ──────────────────────────────────────────────────────────
_FARBEN = {
    "abgelaufen": ("#b71c1c", "#ffffff"),   # bg, text
    "rot":        ("#e53935", "#ffffff"),
    "orange":     ("#ef6c00", "#ffffff"),
    "gelb":       ("#f9a825", "#000000"),
    "ok":         ("#2e7d32", "#ffffff"),
    "einmalig":   ("#546e7a", "#ffffff"),   # done-once, no expiry
}
_ZELL_BG = {
    "abgelaufen": "#ffebee",
    "rot":        "#fff3e0",
    "orange":     "#fff8e1",
    "gelb":       "#fffde7",
}
_WICHTIGKEIT = ["abgelaufen", "rot", "orange", "gelb", "ok", "einmalig", ""]


def _chip_farbe(dring: str) -> tuple[str, str]:
    return _FARBEN.get(dring, ("#90a4ae", "#ffffff"))


def _btn(text: str, color: str = FIORI_BLUE, hover: str = "#0057b8",
         height: int = 30) -> QPushButton:
    b = QPushButton(text)
    b.setFixedHeight(height)
    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
    b.setStyleSheet(
        f"QPushButton{{background:{color};color:#fff;border:none;"
        f"border-radius:4px;padding:2px 12px;font-size:12px;}}"
        f"QPushButton:hover{{background:{hover};}}"
    )
    return b


def _btn_flat(text: str) -> QPushButton:
    b = QPushButton(text)
    b.setFixedHeight(28)
    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
    b.setStyleSheet(
        "QPushButton{background:#eceff1;color:#333;border:none;"
        "border-radius:4px;padding:2px 10px;font-size:12px;}"
        "QPushButton:hover{background:#cfd8dc;}"
    )
    return b


# ─── Kalender-Zelle ───────────────────────────────────────────────────────────
class _TagZelle(QFrame):
    """Eine einzelne Tageszelle im Kalender."""
    geklickt = Signal(date, list)   # emittiert (datum, eintraege)

    MAX_CHIPS = 4

    def __init__(self, parent=None):
        super().__init__(parent)
        self._datum: date | None = None
        self._eintraege: list = []
        self._ist_heute = False
        self._anderer_monat = False
        self.setFixedHeight(100)
        self.setMinimumWidth(80)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)

        self._tag_lbl = QLabel()
        self._tag_lbl.setFont(QFont("Arial", 9, QFont.Weight.Bold))
        layout.addWidget(self._tag_lbl)

        self._chips_layout = QVBoxLayout()
        self._chips_layout.setSpacing(1)
        self._chips_layout.setContentsMargins(0, 0, 0, 0)
        layout.addLayout(self._chips_layout)
        layout.addStretch()

    def setze_datum(self, d: date | None, eintraege: list, anderer_monat: bool = False):
        self._datum          = d
        self._eintraege      = eintraege
        self._ist_heute      = (d == date.today()) if d else False
        self._anderer_monat  = anderer_monat
        self._aktualisieren()

    def _aktualisieren(self):
        # Chips löschen
        while self._chips_layout.count():
            item = self._chips_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if self._datum is None:
            self.setStyleSheet("QFrame{background:#f5f5f5;border:1px solid #e0e0e0;border-radius:4px;}")
            self._tag_lbl.setText("")
            return

        # Hintergrundfarbe je nach dringlichstem Eintrag
        dringlichkeiten = [e.get("_dringlichkeit", "") for e in self._eintraege]
        dring_bg = ""
        for d_key in _WICHTIGKEIT:
            if d_key in dringlichkeiten:
                dring_bg = d_key
                break

        bg = _ZELL_BG.get(dring_bg, "#ffffff")
        if self._anderer_monat:
            bg = "#f9f9f9"
        rand = "#1565c0" if self._ist_heute else "#e0e0e0"
        rand_w = "2" if self._ist_heute else "1"
        self.setStyleSheet(
            f"QFrame{{background:{bg};border:{rand_w}px solid {rand};"
            f"border-radius:4px;}}"
        )

        # Tag-Nummer
        txt = str(self._datum.day)
        farbe = "#1565c0" if self._ist_heute else ("#999" if self._anderer_monat else "#333")
        self._tag_lbl.setText(txt)
        self._tag_lbl.setStyleSheet(f"color:{farbe};background:transparent;")

        # Chips
        zeige = self._eintraege[:self.MAX_CHIPS]
        for e in zeige:
            chip = self._chip_label(e)
            self._chips_layout.addWidget(chip)
        rest = len(self._eintraege) - self.MAX_CHIPS
        if rest > 0:
            mehr = QLabel(f"  +{rest} weitere")
            mehr.setStyleSheet("color:#666;font-size:9px;background:transparent;")
            self._chips_layout.addWidget(mehr)

    def _chip_label(self, eintrag: dict) -> QLabel:
        dring      = eintrag.get("_dringlichkeit", "")
        vorwarnung = eintrag.get("_vorwarnung", False)
        bg, fg     = _chip_farbe(dring)
        typ  = SCHULUNGSTYP_KURZ.get(eintrag.get("schulungstyp",""), eintrag.get("schulungstyp","")[:6])
        name = eintrag.get("_name", "?")
        if vorwarnung:
            # Ablaufdatum anzeigen, damit klar ist wann es wirklich abläuft
            gb_str = eintrag.get("gueltig_bis", "?")
            # nur Tag.Monat (ohne Jahr) für kompakte Anzeige
            datum_kurz = gb_str[:5] if len(gb_str) >= 5 else gb_str
            text = f"  ⚠ {name[:12]} · {typ} ({datum_kurz})"
        else:
            text = f"  {name[:14]} · {typ}"
        lbl = QLabel(text)
        # Vorwarnungen etwas transparenter (gestrichelt wäre ideal, aber QLabel-Grenzen)
        stil = (f"QLabel{{background:{bg};color:{fg};border-radius:3px;"
                f"font-size:10px;font-weight:bold;padding:1px 3px;"
                + ("border:1px dashed " + fg + ";" if vorwarnung else "")
                + "}")
        lbl.setStyleSheet(stil)
        gb_anzeige = eintrag.get("gueltig_bis", "?")
        lbl.setToolTip(
            f"{'⚠ Vorwarnung – läuft demnächst ab!' + chr(10) if vorwarnung else ''}"
            f"{eintrag.get('_name')} – {SCHULUNGSTYPEN_CFG_K.get(eintrag.get('schulungstyp',''),{}).get('anzeige','')}\n"
            f"Gültig bis: {gb_anzeige}"
        )
        return lbl

    def mousePressEvent(self, ev):
        if self._datum and self._eintraege:
            self.geklickt.emit(self._datum, self._eintraege)
        super().mousePressEvent(ev)


# ─── Monats-Kalender-Grid ─────────────────────────────────────────────────────
class _MonatsKalender(QWidget):
    tagesklick = Signal(date, list)

    WOCHENTAGE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._jahr  = date.today().year
        self._monat = date.today().month
        self._daten: dict = {}   # date → [eintrag, ...]
        self._build_ui()

    def _build_ui(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(4)

        # Wochentag-Header
        header = QHBoxLayout()
        header.setSpacing(4)
        for wt in self.WOCHENTAGE:
            lbl = QLabel(wt)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setFixedHeight(24)
            lbl.setStyleSheet(
                f"QLabel{{background:{'#1565c0' if wt in ('Sa','So') else FIORI_BLUE};"
                f"color:#fff;border-radius:3px;font-size:11px;font-weight:bold;}}"
            )
            header.addWidget(lbl)
        v.addLayout(header)

        # Tageszellen (6 Zeilen × 7 Spalten)
        self._zellen: list[list[_TagZelle]] = []
        for row in range(6):
            zeile = []
            hl = QHBoxLayout()
            hl.setSpacing(4)
            for col in range(7):
                z = _TagZelle()
                z.geklickt.connect(self.tagesklick.emit)
                hl.addWidget(z)
                zeile.append(z)
            v.addLayout(hl)
            self._zellen.append(zeile)

        self._render()

    def setze_monat(self, jahr: int, monat: int, daten: dict):
        self._jahr  = jahr
        self._monat = monat
        self._daten = daten
        self._render()

    def _render(self):
        cal = calendar.monthcalendar(self._jahr, self._monat)
        # calendar.monthcalendar kann 4-6 Zeilen haben, wir haben immer 6
        while len(cal) < 6:
            cal.append([0, 0, 0, 0, 0, 0, 0])

        for r, woche in enumerate(cal[:6]):
            for c, tag_nr in enumerate(woche):
                zelle = self._zellen[r][c]
                if tag_nr == 0:
                    zelle.setze_datum(None, [], False)
                else:
                    try:
                        d = date(self._jahr, self._monat, tag_nr)
                    except ValueError:
                        zelle.setze_datum(None, [], False)
                        continue
                    eintraege = self._daten.get(d, [])
                    zelle.setze_datum(d, eintraege, False)


# ─── Tages-Detail-Dialog ──────────────────────────────────────────────────────
class _TagesDetailDialog(QDialog):
    def __init__(self, d: date, eintraege: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"📅 {d.strftime('%d. %B %Y')} – Ablaufende Schulungen")
        self.resize(560, 380)
        v = QVBoxLayout(self)
        v.setSpacing(10)

        titel = QLabel(f"Ablaufende Schulungen am {d.strftime('%d.%m.%Y')}:")
        titel.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        v.addWidget(titel)

        tbl = QTableWidget()
        tbl.setColumnCount(4)
        tbl.setHorizontalHeaderLabels(["Mitarbeiter", "Schulungsart", "Gültig bis", "Status"])
        tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tbl.setAlternatingRowColors(True)
        hh = tbl.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        tbl.verticalHeader().setVisible(False)
        tbl.setRowCount(len(eintraege))

        for row, e in enumerate(eintraege):
            dring = e.get("_dringlichkeit", "")
            bg, _ = _chip_farbe(dring)
            cfg_an = SCHULUNGSTYPEN_CFG_K.get(e.get("schulungstyp",""), {}).get("anzeige", e.get("schulungstyp",""))
            daten = [
                e.get("_name", ""),
                cfg_an,
                e.get("gueltig_bis", ""),
                e.get("status", ""),
            ]
            pastel = _ZELL_BG.get(dring, "#ffffff")
            for col, text in enumerate(daten):
                item = QTableWidgetItem(text)
                item.setBackground(QColor(pastel))
                tbl.setItem(row, col, item)
        v.addWidget(tbl)

        btn = _btn("Schließen", "#546e7a", "#455a64")
        btn.clicked.connect(self.accept)
        v.addWidget(btn, alignment=Qt.AlignmentFlag.AlignRight)


# ─── Neuer-Mitarbeiter-Dialog ─────────────────────────────────────────────────
class NeuerMitarbeiterDialog(QDialog):
    """
    Dialog zum Anlegen eines neuen Mitarbeiters – inklusive
    Schulungsdaten und optionalem Sync in mitarbeiter.db / nesk3.db.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("👤 Neuer Mitarbeiter anlegen")
        self.setMinimumWidth(620)
        self.resize(680, 780)
        self._result_ma    = {}
        self._result_syncs = {}
        self._build_ui()

    def _build_ui(self):
        from functions.mitarbeiter_sync import (
            lade_positionen_ma_db, lade_abteilungen_ma_db
        )
        v = QVBoxLayout(self)
        v.setSpacing(10)
        v.setContentsMargins(16, 16, 16, 12)

        titel = QLabel("👤 Neuen Mitarbeiter anlegen")
        titel.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        titel.setStyleSheet(f"color:{FIORI_BLUE};")
        v.addWidget(titel)

        tabs = QTabWidget()
        v.addWidget(tabs)

        # ── Tab 1: Stammdaten ──────────────────────────────────────────────
        t1 = QWidget()
        f1 = QFormLayout(t1)
        f1.setSpacing(8)
        f1.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self._nachname     = QLineEdit(); self._nachname.setPlaceholderText("Nachname *")
        self._vorname      = QLineEdit(); self._vorname.setPlaceholderText("Vorname *")
        self._anstellung   = QComboBox()
        self._anstellung.addItems(["NA", "Hauptamt", "Nebenamt", "FSJ", "Praktikant", "Sonstiges"])
        self._qualifikation = QComboBox()
        self._qualifikation.setEditable(True)
        self._qualifikation.addItems(["PRM", "RS", "NotSan", "FSJ", "SB", "Sonstiges"])
        self._bemerkung_ma = QTextEdit()
        self._bemerkung_ma.setMaximumHeight(70)
        self._bemerkung_ma.setPlaceholderText("Interne Bemerkung …")

        f1.addRow("Nachname *:", self._nachname)
        f1.addRow("Vorname *:", self._vorname)
        f1.addRow("Anstellung:", self._anstellung)
        f1.addRow("Qualifikation:", self._qualifikation)
        f1.addRow("Bemerkung:", self._bemerkung_ma)
        tabs.addTab(t1, "👤 Stammdaten")

        # ── Tab 2: Schulungen ──────────────────────────────────────────────
        t2 = QWidget()
        f2 = QFormLayout(t2)
        f2.setSpacing(8)
        f2.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        def _date_edit(special="— (noch nicht absolviert)"):
            w = QDateEdit()
            w.setCalendarPopup(True)
            w.setDisplayFormat("dd.MM.yyyy")
            w.setSpecialValueText(special)
            w.setMinimumDate(QDate(2000, 1, 1))
            w.setDate(QDate(2000, 1, 1))
            return w

        self._zuep_bis       = _date_edit(); f2.addRow("ZÜP gültig bis:", self._zuep_bis)
        self._eh_datum       = _date_edit(); f2.addRow("EH (letztes Datum):", self._eh_datum)
        self._refresher      = _date_edit(); f2.addRow("Refresher (letztes Datum):", self._refresher)
        self._aerztl_bis     = _date_edit(); f2.addRow("Ärztl. Untersuchung gültig bis:", self._aerztl_bis)
        self._fuehr_kont     = _date_edit(); f2.addRow("Führerschein Kontrolle:", self._fuehr_kont)
        self._arbeitsschutz  = _date_edit(); f2.addRow("Arbeitsschutz:", self._arbeitsschutz)
        self._sicherheit_von = _date_edit(); f2.addRow("Sicherheitsschulung (Von):", self._sicherheit_von)
        self._vorfeld_bis    = _date_edit(); f2.addRow("Vorfeldschulung (Gültig bis):", self._vorfeld_bis)
        tabs.addTab(t2, "🎓 Schulungen")

        # ── Tab 3: Weitere Infos (für andere DBs) ─────────────────────────
        t3 = QWidget()
        f3 = QFormLayout(t3)
        f3.setSpacing(8)
        f3.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self._personalnr  = QLineEdit(); self._personalnr.setPlaceholderText("Optional")
        self._funktion    = QComboBox()
        self._funktion.addItems(["Schichtleiter", "Dispo", "Betreuer"])
        self._position    = QComboBox()
        self._position.setEditable(True)
        _entfernte_pos = {"Arzt", "Führungskraft", "Verwaltung", "Sanitätshelfer"}
        pos = [p for p in lade_positionen_ma_db() if p not in _entfernte_pos]
        if "Betreuer" not in pos:
            pos = sorted(pos + ["Betreuer"])
        self._position.addItems(pos if pos else ["Betreuer", "PRM", "Schichtleiter"])
        self._abteilung   = QComboBox()
        self._abteilung.setEditable(True)
        abt = lade_abteilungen_ma_db()
        self._abteilung.addItems(abt if abt else ["Passagierbetreuung"])
        self._email       = QLineEdit(); self._email.setPlaceholderText("max.mustermann@drk-koeln.de")
        self._telefon     = QLineEdit(); self._telefon.setPlaceholderText("0221 / …")
        self._eintrittsd  = QDateEdit(QDate.currentDate())
        self._eintrittsd.setCalendarPopup(True)
        self._eintrittsd.setDisplayFormat("dd.MM.yyyy")

        f3.addRow("Personalnummer:", self._personalnr)
        f3.addRow("Funktion:", self._funktion)
        f3.addRow("Position:", self._position)
        f3.addRow("Abteilung:", self._abteilung)
        f3.addRow("E-Mail:", self._email)
        f3.addRow("Telefon:", self._telefon)
        f3.addRow("Eintrittsdatum:", self._eintrittsd)
        tabs.addTab(t3, "📋 Mitarbeiterdaten")

        # ── Sync-Optionen ──────────────────────────────────────────────────
        sync_frame = QFrame()
        sync_frame.setStyleSheet(
            "QFrame{background:#e3f2fd;border:1px solid #90caf9;"
            "border-radius:4px;padding:4px;}"
        )
        sf = QHBoxLayout(sync_frame)
        info = QLabel("Auch anlegen in:")
        info.setStyleSheet("font-weight:bold;color:#1565c0;")
        sf.addWidget(info)
        self._sync_ma   = QCheckBox("mitarbeiter.db")
        self._sync_ma.setChecked(True)
        self._sync_nesk = QCheckBox("nesk3.db")
        self._sync_nesk.setChecked(True)
        sf.addWidget(self._sync_ma)
        sf.addWidget(self._sync_nesk)
        sf.addStretch()
        v.addWidget(sync_frame)

        # ── Buttons ────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        self._btn_speichern = _btn("💾  Speichern", "#2e7d32", "#1b5e20", 34)
        self._btn_speichern.clicked.connect(self._speichern)
        btn_abbrechen = _btn_flat("Abbrechen")
        btn_abbrechen.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(self._btn_speichern)
        btn_row.addWidget(btn_abbrechen)
        v.addLayout(btn_row)

    def _speichern(self):
        if not self._nachname.text().strip():
            QMessageBox.warning(self, "Pflichtfeld", "Bitte Nachname eingeben.")
            return
        if not self._vorname.text().strip():
            QMessageBox.warning(self, "Pflichtfeld", "Bitte Vorname eingeben.")
            return
        self.accept()

    def _datum_oder_leer(self, widget: "QDateEdit") -> str:
        d = widget.date()
        if d == QDate(2000, 1, 1):
            return ""
        return d.toString("dd.MM.yyyy")

    def _datum_obj(self, widget: "QDateEdit"):
        from datetime import date as _date
        d = widget.date()
        if d == QDate(2000, 1, 1):
            return None
        return _date(d.year(), d.month(), d.day())

    def get_stamm_daten(self) -> dict:
        return {
            "nachname":      self._nachname.text().strip(),
            "vorname":       self._vorname.text().strip(),
            "geburtsdatum":  "",
            "anstellung":    self._anstellung.currentText(),
            "qualifikation": self._qualifikation.currentText(),
            "bemerkung":     self._bemerkung_ma.toPlainText().strip(),
        }

    def get_schulungs_daten(self) -> dict:
        """Gibt Schulungseinträge-Rohdaten zurück (gueltig_bis als String)."""
        from functions.schulungen_db import _berechne_gueltig_bis, _datum_str
        from datetime import date as _date
        def _d(w):
            if w is None:
                return None
            d = w.date()
            return d if d != QDate(2000, 1, 1) else None
        def _ds(qd): return _date(qd.year(), qd.month(), qd.day()) if qd else None

        eintraege = {}

        def _add(key, datum_absolviert_qd, gueltig_bis_qd=None):
            d_abs   = _ds(_d(datum_absolviert_qd))
            d_gb    = _ds(_d(gueltig_bis_qd)) if gueltig_bis_qd else None
            gb_calc = _berechne_gueltig_bis(key, d_abs, d_gb or d_abs)
            if d_abs or d_gb:
                eintraege[key] = {
                    "schulungstyp":     key,
                    "datum_absolviert": _datum_str(d_abs),
                    "gueltig_bis":      _datum_str(gb_calc or d_gb),
                }

        _add("ZÜP",                 None, self._zuep_bis)
        _add("EH",                  self._eh_datum)
        _add("Refresher",           self._refresher)
        _add("Aerztl_Untersuchung", None, self._aerztl_bis)
        _add("Fuehrerschein_Kont",  self._fuehr_kont)
        _add("Arbeitsschutz",       self._arbeitsschutz)
        _add("Sicherheitsschulung", self._sicherheit_von)
        _add("Vorfeldschulung",     None, self._vorfeld_bis)
        return eintraege

    def get_ma_db_daten(self) -> dict:
        return {
            "personalnummer": self._personalnr.text().strip(),
            "funktion":       self._funktion.currentText(),
            "position":       self._position.currentText(),
            "abteilung":      self._abteilung.currentText(),
            "email":          self._email.text().strip(),
            "telefon":        self._telefon.text().strip(),
            "eintrittsdatum": self._eintrittsd.date().toString("dd.MM.yyyy"),
        }

    def sync_aktiviert_ma(self)   -> bool: return self._sync_ma.isChecked()
    def sync_aktiviert_nesk(self) -> bool: return self._sync_nesk.isChecked()


# ─── Schulungstyp-Kurznahmen (für Chips) ─────────────────────────────────────
# Späte Imports aus schulungen_db –  wird zur Laufzeit befüllt
SCHULUNGSTYP_KURZ: dict = {}
SCHULUNGSTYPEN_CFG_K: dict = {}


def _lade_typen():
    global SCHULUNGSTYP_KURZ, SCHULUNGSTYPEN_CFG_K
    try:
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        SCHULUNGSTYPEN_CFG_K = SCHULUNGSTYPEN_CFG
        SCHULUNGSTYP_KURZ = {
            "ZÜP":                 "ZÜP",
            "EH":                  "EH",
            "Refresher":           "Ref.",
            "Aerztl_Untersuchung": "Ärztl.",
            "Fuehrerschein_Kont":  "FS-Kont.",
            "Einw_Zertifikate":    "Zert.",
            "Fixierung":           "Fix.",
            "Einw_eMobby":         "e-Mobby",
            "Bulmor":              "Bulmor",
            "Arbeitsschutz":       "ArbSchut.",
            "Einw_QM":             "QM",
            "Fragebogen_Schulung": "Frageb.",
            "Personalausweis":     "PA/Pass",
            "Sicherheitsschulung": "Sich.Sch.",
            "Vorfeldschulung":      "Vorfeld",
            "Sonstiges":           "Sonst.",
        }
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════════════
#  Schulungen-Excel-Export
# ══════════════════════════════════════════════════════════════════════════════

def _schulungen_als_excel_speichern(eintraege: list, pfad: str) -> None:
    """Speichert ablaufende Schulungseinträge als Excel-Datei (2 Blätter)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ablaufende Schulungen"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1565A8")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    spalten = [
        ("Mitarbeiter",    28),
        ("Schulungsart",   22),
        ("Gültig bis",     14),
        ("Tage",           10),
        ("Status",         18),
    ]
    for col_idx, (titel, breite) in enumerate(spalten, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titel)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = breite
    ws.row_dimensions[1].height = 22

    from functions.schulungen_db import SCHULUNGSTYPEN_CFG
    fill_rot   = PatternFill("solid", fgColor="FFCDD2")
    fill_gelb  = PatternFill("solid", fgColor="FFF9C4")
    fill_orange = PatternFill("solid", fgColor="FFE0B2")
    fill_abgl  = PatternFill("solid", fgColor="FFEBEE")
    fill_white = PatternFill("solid", fgColor="FFFFFF")
    fill_alt   = PatternFill("solid", fgColor="F5F5F5")

    for row_idx, e in enumerate(eintraege, start=2):
        tage  = e.get("_tage_rest", 0)
        dring = e.get("_dringlichkeit", "")
        cfg   = SCHULUNGSTYPEN_CFG.get(e.get("schulungstyp", ""), {})
        anzeige = cfg.get("anzeige", e.get("schulungstyp", ""))

        if tage < 0:
            tage_txt = f"ÜBERFÄLLIG {-tage}d"
            row_fill = fill_abgl
        elif tage <= 31:
            tage_txt = f"{tage} Tage"
            row_fill = fill_rot
        elif tage <= 61:
            tage_txt = f"{tage} Tage"
            row_fill = fill_orange
        elif tage <= 92:
            tage_txt = f"{tage} Tage"
            row_fill = fill_gelb
        else:
            tage_txt = f"{tage} Tage"
            row_fill = fill_alt if row_idx % 2 == 0 else fill_white

        status_txt = "Abgelaufen" if tage < 0 else "Gültig"

        zeile = [
            e.get("_name", ""),
            anzeige,
            e.get("gueltig_bis", ""),
            tage_txt,
            status_txt,
        ]
        for col_idx, wert in enumerate(zeile, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=wert)
            cell.border    = thin_border
            cell.fill      = row_fill
            cell.alignment = left_align if col_idx in (1, 2) else center_align
        ws.row_dimensions[row_idx].height = 17

    ws.auto_filter.ref = f"A1:{get_column_letter(len(spalten))}1"

    # ── Übersichts-Blatt: Pro Mitarbeiter ─────────────────────────────────────
    from collections import defaultdict
    ws2 = wb.create_sheet("Übersicht je Mitarbeiter")

    ma_gruppen: dict = defaultdict(list)
    for e in eintraege:
        ma_gruppen[e.get("_name", "Unbekannt")].append(e)

    ü_spalten = [
        ("Mitarbeiter",   28),
        ("Schulungsart",  22),
        ("Gültig bis",    14),
        ("Tage",          10),
        ("Status",        18),
    ]
    for col_idx, (titel, breite) in enumerate(ü_spalten, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=titel)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = breite
    ws2.row_dimensions[1].height = 22

    row_idx = 2
    for ma_name in sorted(ma_gruppen.keys()):
        for e in sorted(ma_gruppen[ma_name], key=lambda x: x.get("gueltig_bis", "")):
            tage  = e.get("_tage_rest", 0)
            cfg   = SCHULUNGSTYPEN_CFG.get(e.get("schulungstyp", ""), {})
            anzeige = cfg.get("anzeige", e.get("schulungstyp", ""))
            tage_txt   = f"ÜBERFÄLLIG {-tage}d" if tage < 0 else f"{tage} Tage"
            status_txt = "Abgelaufen" if tage < 0 else "Gültig"
            zeile = [ma_name, anzeige, e.get("gueltig_bis", ""), tage_txt, status_txt]
            row_fill = fill_abgl if tage < 0 else (
                fill_rot if tage <= 31 else (
                fill_orange if tage <= 61 else (
                fill_gelb if tage <= 92 else (fill_alt if row_idx % 2 == 0 else fill_white)
            )))
            for col_idx, wert in enumerate(zeile, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=wert)
                cell.border    = thin_border
                cell.fill      = row_fill
                cell.alignment = left_align if col_idx in (1, 2) else center_align
            ws2.row_dimensions[row_idx].height = 17
            row_idx += 1

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(ü_spalten))}1"

    os.makedirs(os.path.dirname(pfad), exist_ok=True)
    wb.save(pfad)


# ── Export-Dialog ─────────────────────────────────────────────────────────────
class _SchulungExportDialog(QDialog):
    """Dialog zur Auswahl von Schulungsarten, Zeitraum und Speicherpfad."""

    def __init__(self, parent=None):
        super().__init__(parent)
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        self._typen_cfg = SCHULUNGSTYPEN_CFG
        self.setWindowTitle("📊 Excel-Export – Ablaufende Schulungen")
        self.setMinimumWidth(560)
        self._build_ui()

    def _build_ui(self):
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        v = QVBoxLayout(self)
        v.setSpacing(10)

        # ── Schulungsarten ────────────────────────────────────────────────────
        grp_art = QGroupBox("Schulungsarten")
        grp_art.setStyleSheet(
            "QGroupBox{font-weight:bold;border:1px solid #a5d6a7;"
            "border-radius:4px;margin-top:8px;padding-top:8px;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}"
        )
        art_v = QVBoxLayout(grp_art)

        # Nur Typen die ablaufen können
        self._art_checks: dict[str, QCheckBox] = {}
        ablaufende_typen = {
            k: cfg for k, cfg in SCHULUNGSTYPEN_CFG.items()
            if not cfg.get("laeuft_nicht_ab", False)
        }
        row_l = QHBoxLayout()
        for key, cfg in ablaufende_typen.items():
            cb = QCheckBox(cfg["anzeige"])
            cb.setChecked(True)
            self._art_checks[key] = cb
            row_l.addWidget(cb)
        art_v.addLayout(row_l)

        btn_alle = QPushButton("Alle")
        btn_alle.setFixedWidth(60)
        btn_alle.clicked.connect(lambda: [cb.setChecked(True) for cb in self._art_checks.values()])
        btn_keine = QPushButton("Keine")
        btn_keine.setFixedWidth(60)
        btn_keine.clicked.connect(lambda: [cb.setChecked(False) for cb in self._art_checks.values()])
        btn_row = QHBoxLayout()
        btn_row.addWidget(btn_alle)
        btn_row.addWidget(btn_keine)
        btn_row.addStretch()
        art_v.addLayout(btn_row)
        v.addWidget(grp_art)

        # ── Zeitraum ──────────────────────────────────────────────────────────
        grp_zeit = QGroupBox("Zeitraum (Gültig-bis-Datum)")
        grp_zeit.setStyleSheet(
            "QGroupBox{font-weight:bold;border:1px solid #90caf9;"
            "border-radius:4px;margin-top:8px;padding-top:8px;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}"
        )
        zeit_v = QVBoxLayout(grp_zeit)

        # Quick-Buttons
        qbtn_row = QHBoxLayout()
        qbtn_row.setSpacing(6)
        for label, monate in [
            ("Nur abgelaufen",  0),
            ("≤ 1 Monat",       1),
            ("≤ 2 Monate",      2),
            ("≤ 3 Monate",      3),
            ("≤ 6 Monate",      6),
            ("≤ 12 Monate",    12),
        ]:
            b = QPushButton(label)
            b.setFixedHeight(26)
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            b.setStyleSheet(
                "QPushButton{background:#e3f2fd;color:#1565c0;border:1px solid #90caf9;"
                "border-radius:4px;font-size:11px;padding:1px 8px;}"
                "QPushButton:hover{background:#bbdefb;}"
            )
            b.clicked.connect(lambda checked=False, m=monate: self._schnell_setzen(m))
            qbtn_row.addWidget(b)
        qbtn_row.addStretch()
        zeit_v.addLayout(qbtn_row)

        # Von/Bis
        vb_row = QHBoxLayout()
        vb_row.addWidget(QLabel("Von:"))
        self._de_von = QDateEdit()
        self._de_von.setCalendarPopup(True)
        self._de_von.setDisplayFormat("dd.MM.yyyy")
        self._de_von.setDate(QDate(2000, 1, 1))
        vb_row.addWidget(self._de_von)
        vb_row.addSpacing(12)
        vb_row.addWidget(QLabel("Bis:"))
        self._de_bis = QDateEdit()
        self._de_bis.setCalendarPopup(True)
        self._de_bis.setDisplayFormat("dd.MM.yyyy")
        today = date.today()
        self._de_bis.setDate(QDate(today.year, today.month + 3 if today.month <= 9 else today.month - 9,
                                    today.day))
        vb_row.addWidget(self._de_bis)
        vb_row.addStretch()
        zeit_v.addLayout(vb_row)
        v.addWidget(grp_zeit)

        # Standardmäßig ≤ 3 Monate
        self._schnell_setzen(3)

        # ── Speicherort ───────────────────────────────────────────────────────
        pf_row = QHBoxLayout()
        pf_row.addWidget(QLabel("Speicherpfad:"))
        self._pfad_edit = QLineEdit()
        self._pfad_edit.setText(os.path.join(_SCHULUNG_EXPORT_DIR,
                                              self._dateiname_berechnen()))
        pf_row.addWidget(self._pfad_edit, 1)
        btn_browse = QPushButton("...")
        btn_browse.setFixedWidth(28)
        btn_browse.clicked.connect(self._browse)
        pf_row.addWidget(btn_browse)
        v.addLayout(pf_row)

        # ── OK / Abbrechen ────────────────────────────────────────────────────
        dlg_btns = QHBoxLayout()
        dlg_btns.addStretch()
        btn_ok = QPushButton("📊 Exportieren")
        btn_ok.setFixedHeight(30)
        btn_ok.setStyleSheet(
            "QPushButton{background:#1565c0;color:#fff;border:none;"
            "border-radius:4px;padding:2px 16px;font-size:12px;}"
            "QPushButton:hover{background:#0d47a1;}"
        )
        # ── Optionen ──────────────────────────────────────────────────────────
        opt_row = QHBoxLayout()
        self._cb_ohne_abgelaufene = QCheckBox("Abgelaufene ausschließen")
        self._cb_ohne_abgelaufene.setChecked(False)
        self._cb_ohne_abgelaufene.setToolTip(
            "Wenn aktiviert, werden Einträge mit überschrittenem Ablaufdatum nicht exportiert."
        )
        opt_row.addWidget(self._cb_ohne_abgelaufene)
        opt_row.addStretch()
        v.addLayout(opt_row)

        btn_ok.clicked.connect(self.accept)
        btn_ab = QPushButton("Abbrechen")
        btn_ab.setFixedHeight(30)
        btn_ab.clicked.connect(self.reject)
        dlg_btns.addWidget(btn_ab)
        dlg_btns.addWidget(btn_ok)
        v.addLayout(dlg_btns)

    def _dateiname_berechnen(self) -> str:
        today = date.today()
        return f"Schulungen_Ablaufend_{today.strftime('%Y-%m-%d')}.xlsx"

    def _schnell_setzen(self, monate: int):
        today = date.today()
        q_von = QDate(2000, 1, 1)   # weit in der Vergangenheit
        if monate == 0:
            # Nur abgelaufen: bis gestern
            gd = today - timedelta(days=1)
            q_bis = QDate(gd.year, gd.month, gd.day)
        else:
            from datetime import date as _date
            try:
                bis = today.replace(month=today.month + monate)
            except ValueError:
                # Monatsüberlauf
                import calendar as _cal
                total = today.month + monate
                year  = today.year + (total - 1) // 12
                month = ((total - 1) % 12) + 1
                day   = min(today.day, _cal.monthrange(year, month)[1])
                bis   = today.replace(year=year, month=month, day=day)
            q_bis = QDate(bis.year, bis.month, bis.day)
        self._de_von.setDate(q_von)
        self._de_bis.setDate(q_bis)

    def _browse(self):
        from PySide6.QtWidgets import QFileDialog
        start = self._pfad_edit.text() or _SCHULUNG_EXPORT_DIR
        pfad, _ = QFileDialog.getSaveFileName(
            self, "Excel speichern unter", start, "Excel (*.xlsx)"
        )
        if pfad:
            if not pfad.lower().endswith(".xlsx"):
                pfad += ".xlsx"
            self._pfad_edit.setText(pfad)

    def get_werte(self):
        """Gibt (von_datum, bis_datum, speicherpfad, schulungstypen_liste, ohne_abgelaufene) zurück."""
        vq = self._de_von.date()
        bq = self._de_bis.date()
        von = date(vq.year(), vq.month(), vq.day())
        bis = date(bq.year(), bq.month(), bq.day())
        typen = [k for k, cb in self._art_checks.items() if cb.isChecked()]
        return von, bis, self._pfad_edit.text(), typen, self._cb_ohne_abgelaufene.isChecked()


# ─── Agenda-Liste ─────────────────────────────────────────────────────────────
class _AgendaWidget(QWidget):
    """Liste der nächsten ablaufenden Schulungen (nächste 3 Monate)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(4)

        hdr = QLabel("⚠️  Bald ablaufende Schulungen (nächste 3 Monate)")
        hdr.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        hdr.setStyleSheet(f"color:{FIORI_BLUE};")
        v.addWidget(hdr)

        self._tbl = QTableWidget()
        self._tbl.setColumnCount(5)
        self._tbl.setHorizontalHeaderLabels(
            ["Mitarbeiter", "Schulungsart", "Gültig bis", "Tage", "Status"]
        )
        hh = self._tbl.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setStyleSheet("font-size:12px;")
        self._tbl.verticalHeader().setVisible(False)
        v.addWidget(self._tbl)

    def aktualisieren(self):
        from functions.schulungen_db import lade_ablaufende, SCHULUNGSTYPEN_CFG
        try:
            eintraege = lade_ablaufende(3)
        except Exception as exc:
            self._tbl.setRowCount(0)
            return

        # Sortierung: abgelaufen zuerst, dann nach Tagen
        priori = {"abgelaufen": -1, "rot": 0, "orange": 1, "gelb": 2, "ok": 3}
        eintraege.sort(key=lambda e: (priori.get(e.get("_dringlichkeit",""), 9),
                                       e.get("_tage_rest", 9999)))
        self._tbl.setRowCount(len(eintraege))
        FARB_MAP = {
            "abgelaufen": "#ffcdd2",
            "rot":        "#ffe0b2",
            "orange":     "#fff9c4",
            "gelb":       "#f9fbe7",
        }
        for row, e in enumerate(eintraege):
            dring = e.get("_dringlichkeit", "")
            bg    = FARB_MAP.get(dring, "#ffffff")
            cfg   = SCHULUNGSTYPEN_CFG.get(e.get("schulungstyp",""), {})
            anzeige = cfg.get("anzeige", e.get("schulungstyp",""))
            tage    = e.get("_tage_rest", 0)
            tage_txt = f"{tage} Tage" if tage >= 0 else f"ÜBERFÄLLIG {-tage}d"
            vals = [
                e.get("_name", ""),
                anzeige,
                e.get("gueltig_bis", ""),
                tage_txt,
                e.get("status", ""),
            ]
            for col, text in enumerate(vals):
                item = QTableWidgetItem(text)
                item.setBackground(QColor(bg))
                if col == 3 and tage < 0:
                    item.setForeground(QColor("#b71c1c"))
                self._tbl.setItem(row, col, item)


# ─── Mitarbeiter-Detail-Dialog ────────────────────────────────────────────────
# ─── Einzelne-Schulung bearbeiten ─────────────────────────────────────────────
class _SchulungBearbeitenDialog(QDialog):
    """Kleiner Dialog: Datum einer einzelnen Schulung neu setzen."""

    def __init__(self, ma_id: int, schulungstyp: str, eintrag: dict | None, parent=None):
        super().__init__(parent)
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        self._ma_id  = ma_id
        self._typ    = schulungstyp
        self._eintrag = eintrag
        cfg = SCHULUNGSTYPEN_CFG.get(schulungstyp, {})
        self._cfg    = cfg
        self._ablauf = cfg.get("ablauf", "einmalig")
        anzeige = cfg.get("anzeige", schulungstyp)
        self.setWindowTitle(f"✏️  {anzeige} bearbeiten")
        self.setMinimumWidth(400)
        self._build_ui(anzeige, eintrag)

    def _build_ui(self, anzeige: str, eintrag: dict | None):
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        v = QVBoxLayout(self)
        v.setSpacing(10)
        v.setContentsMargins(18, 14, 18, 14)

        titel = QLabel(f"✏️  {anzeige}")
        titel.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        titel.setStyleSheet(f"color:{FIORI_BLUE};")
        v.addWidget(titel)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        def _date_edit(val: str = ""):
            w = QDateEdit()
            w.setCalendarPopup(True)
            w.setDisplayFormat("dd.MM.yyyy")
            w.setSpecialValueText("— (nicht gesetzt)")
            w.setMinimumDate(QDate(2000, 1, 1))
            if val:
                parts = str(val).strip().split(".")
                if len(parts) == 3:
                    try:
                        w.setDate(QDate(int(parts[2]), int(parts[1]), int(parts[0])))
                        return w
                    except Exception:
                        pass
            w.setDate(QDate(2000, 1, 1))
            return w

        # Gültig bis: bei ablauf="direkt" (ZÜP, Ärztl.) ist das das Hauptfeld
        self._datum_gb = _date_edit(eintrag.get("gueltig_bis", "") if eintrag else "")

        if self._ablauf == "direkt":
            # Gültig bis ist Pflichtfeld; Absolviert am optional
            form.addRow("Ablaufdatum (Gültig bis) *:", self._datum_gb)
            self._datum_abs = _date_edit(eintrag.get("datum_absolviert", "") if eintrag else "")
            form.addRow("Absolviert am (optional):", self._datum_abs)
        elif self._ablauf == "intervall":
            # Absolviert am ist Pflichtfeld; Gültig bis wird auto-berechnet
            self._datum_abs = _date_edit(eintrag.get("datum_absolviert", "") if eintrag else "")
            form.addRow("Absolviert am *:", self._datum_abs)
            intervall = self._cfg.get("intervall", 1)
            info = QLabel(f"ℹ️  Wird automatisch berechnet: Absolviert + {intervall} Jahr(e)")
            info.setStyleSheet("color:#555;font-size:10px;font-style:italic;")
            form.addRow("Gültig bis:", info)
            self._datum_abs.dateChanged.connect(self._auto_gueltig_bis)
        else:
            # einmalig: nur Absolviert am
            self._datum_abs = _date_edit(eintrag.get("datum_absolviert", "") if eintrag else "")
            form.addRow("Absolviert am *:", self._datum_abs)

        self._bemerkung = QLineEdit()
        self._bemerkung.setPlaceholderText("Optional …")
        if eintrag:
            self._bemerkung.setText(eintrag.get("bemerkung", "") or "")
        form.addRow("Bemerkung:", self._bemerkung)

        # Informiert-Zeile
        w_info = QWidget()
        h_info = QHBoxLayout(w_info)
        h_info.setContentsMargins(0, 0, 0, 0)
        h_info.setSpacing(6)
        self._chk_informiert = QCheckBox("informiert am")
        h_info.addWidget(self._chk_informiert)
        self._datum_informiert = _date_edit(eintrag.get("informiert_am", "") if eintrag else "")
        self._datum_informiert.setEnabled(False)
        h_info.addWidget(self._datum_informiert)
        btn_info_clr = QPushButton("🗑")
        btn_info_clr.setFixedSize(26, 24)
        btn_info_clr.setToolTip("Informiert-Eintrag löschen")
        btn_info_clr.setStyleSheet(
            "QPushButton{background:#ffebee;border:none;border-radius:3px;font-size:13px;}"
            "QPushButton:hover{background:#ef9a9a;}"
        )
        btn_info_clr.clicked.connect(self._info_loeschen)
        h_info.addWidget(btn_info_clr)
        h_info.addStretch()
        if eintrag and eintrag.get("informiert"):
            self._chk_informiert.setChecked(True)
            self._datum_informiert.setEnabled(True)
        self._chk_informiert.toggled.connect(self._datum_informiert.setEnabled)
        form.addRow("Mitarbeiter:", w_info)

        v.addLayout(form)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_speichern = _btn("💾  Speichern", "#2e7d32", "#1b5e20")
        btn_speichern.clicked.connect(self._speichern)
        btn_abbrechen = _btn("Abbrechen", "#546e7a", "#455a64")
        btn_abbrechen.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(btn_speichern)
        btn_row.addWidget(btn_abbrechen)
        v.addLayout(btn_row)

    def _info_loeschen(self):
        self._chk_informiert.setChecked(False)
        self._datum_informiert.setDate(QDate(2000, 1, 1))

    def _auto_gueltig_bis(self, d: QDate):
        if d == QDate(2000, 1, 1):
            return
        intervall = self._cfg.get("intervall", 1)
        try:
            gb = QDate(d.year() + intervall, d.month(), d.day())
            self._datum_gb.setDate(gb)
        except Exception:
            pass

    def _datum_str(self, qdate: QDate) -> str:
        if qdate == QDate(2000, 1, 1):
            return ""
        return qdate.toString("dd.MM.yyyy")

    def _speichern(self):
        from functions.schulungen_db import (
            aktualisiere_schulungseintrag, speichere_schulungseintrag,
            SCHULUNGSTYPEN_CFG, _berechne_status, _parse_datum
        )
        cfg    = SCHULUNGSTYPEN_CFG.get(self._typ, {})
        ablauf = cfg.get("ablauf", "einmalig")

        abs_str = self._datum_str(self._datum_abs.date())
        gb_str  = self._datum_str(self._datum_gb.date())

        if ablauf == "direkt":
            # Pflichtfeld: Ablaufdatum
            if not gb_str:
                QMessageBox.warning(self, "Pflichtfeld", "Bitte Ablaufdatum (Gültig bis) eingeben.")
                return
        elif ablauf == "intervall":
            # Pflichtfeld: Absolviert am; Gültig bis wird berechnet
            if not abs_str:
                QMessageBox.warning(self, "Pflichtfeld", "Bitte Datum 'Absolviert am' eingeben.")
                return
            d = self._datum_abs.date()
            intervall = cfg.get("intervall", 1)
            try:
                gb_date = QDate(d.year() + intervall, d.month(), d.day())
            except Exception:
                gb_date = QDate(2000, 1, 1)
            gb_str = gb_date.toString("dd.MM.yyyy")
        else:
            # einmalig: Absolviert am Pflicht, kein Ablaufdatum
            if not abs_str:
                QMessageBox.warning(self, "Pflichtfeld", "Bitte Datum 'Absolviert am' eingeben.")
                return
            gb_str = ""

        gb_obj = _parse_datum(gb_str)
        status = _berechne_status(gb_obj, cfg.get("laeuft_nicht_ab", False))

        daten = {
            "mitarbeiter_id":  self._ma_id,
            "schulungstyp":    self._typ,
            "datum_absolviert": abs_str,
            "gueltig_bis":     gb_str,
            "laeuft_nicht_ab": int(cfg.get("laeuft_nicht_ab", False)),
            "status":          status,
            "bemerkung":       self._bemerkung.text().strip(),
            "informiert":      int(self._chk_informiert.isChecked()),
            "informiert_am":   self._datum_str(self._datum_informiert.date()) if self._chk_informiert.isChecked() else "",
        }
        if self._eintrag and self._eintrag.get("id"):
            aktualisiere_schulungseintrag(self._eintrag["id"], daten)
        else:
            speichere_schulungseintrag(daten)
        self.accept()


# ─── Mitarbeiter-Detail-Dialog ────────────────────────────────────────────────
class _MitarbeiterDetailDialog(QDialog):
    """Zeigt alle Schulungstypen für einen einzelnen Mitarbeiter.
    Doppelklick oder Klick auf ✏️ öffnet den Bearbeiten-Dialog.
    """

    geaendert = Signal()   # feuert wenn ein Eintrag gespeichert wurde

    _FARB_MAP = {
        "abgelaufen": "#ffcdd2",
        "rot":        "#ffe0b2",
        "orange":     "#fff9c4",
        "gelb":       "#f9fbe7",
        "ok":         "#e8f5e9",
        "einmalig":   "#eceff1",
    }

    def __init__(self, ma: dict, parent=None):
        super().__init__(parent)
        self._ma = ma
        name = f"{ma.get('nachname', '')} {ma.get('vorname', '')}".strip()
        self.setWindowTitle(f"🎓 Schulungen: {name}")
        self.resize(720, 520)
        self._build_ui()

    def _build_ui(self):
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        v = QVBoxLayout(self)
        v.setSpacing(10)
        v.setContentsMargins(16, 14, 16, 14)

        ma   = self._ma
        name = f"{ma.get('nachname', '')}, {ma.get('vorname', '')}".strip(", ")
        qual = ma.get("qualifikation", "") or ""
        titel = QLabel(f"👤  {name}   {('('+qual+')') if qual else ''}")
        titel.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        titel.setStyleSheet(f"color:{FIORI_BLUE};")
        v.addWidget(titel)

        schulungen = ma.get("schulungen", {})
        if not schulungen:
            warn = QLabel("⚠️  Dieser Mitarbeiter hat noch keine Schulungseinträge in der Datenbank.")
            warn.setStyleSheet("color:#b71c1c;font-size:12px;padding:4px 0;")
            v.addWidget(warn)

        self._tbl = QTableWidget()
        self._tbl.setColumnCount(6)
        self._tbl.setHorizontalHeaderLabels(
            ["Schulungstyp", "Absolviert am", "Gültig bis", "Status", "Informiert", ""]
        )
        self._tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setStyleSheet("font-size:12px;")
        hh = self._tbl.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.doubleClicked.connect(lambda idx: self._bearbeiten(idx.row()))
        v.addWidget(self._tbl, 1)

        self._tabelle_befuellen()

        hint = QLabel("💡 Doppelklick oder ✏️-Button zum Bearbeiten eines Eintrags")
        hint.setStyleSheet("color:#777;font-size:10px;font-style:italic;")
        v.addWidget(hint)

        # Legende
        leg_lay = QHBoxLayout()
        for key, label in [
            ("abgelaufen", "Abgelaufen"), ("rot", "≤1 Mon."),
            ("orange", "≤2 Mon."), ("gelb", "≤3 Mon."), ("ok", "OK"),
        ]:
            bg, fg = _chip_farbe(key)
            lbl = QLabel(f"  {label}  ")
            lbl.setStyleSheet(
                f"background:{bg};color:{fg};border-radius:3px;"
                f"font-size:10px;padding:1px 5px;"
            )
            leg_lay.addWidget(lbl)
        no_lbl = QLabel("  Kein Eintrag  ")
        no_lbl.setStyleSheet("color:#bdbdbd;font-size:10px;padding:1px 5px;")
        leg_lay.addWidget(no_lbl)
        leg_lay.addStretch()
        v.addLayout(leg_lay)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_mail = _btn("📧  E-Mail senden", "#1565c0", "#0d47a1")
        btn_mail.setToolTip("Outlook-Entwurf mit abgelaufenen Schulungen erstellen")
        btn_mail.clicked.connect(self._sende_email)
        btn_row.addWidget(btn_mail)
        btn_row.addStretch()
        btn = _btn("Schließen", "#546e7a", "#455a64")
        btn.clicked.connect(self.accept)
        btn_row.addWidget(btn)
        v.addLayout(btn_row)

    def _sende_email(self):
        import urllib.parse, subprocess, os
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        ma = self._ma
        email = ""
        try:
            from database.connection import get_ma_connection
            import sqlite3 as _sqlite3
            conn = get_ma_connection()
            conn.row_factory = _sqlite3.Row
            row = conn.execute(
                "SELECT email FROM mitarbeiter WHERE nachname=? AND vorname=? LIMIT 1",
                (ma.get("nachname", ""), ma.get("vorname", ""))
            ).fetchone()
            conn.close()
            if row:
                email = row["email"] or ""
        except Exception:
            pass
        DRING_LABEL = {
            "abgelaufen": "abgelaufen",
            "rot":        "läuft in ≤1 Monat ab",
            "orange":     "läuft in ≤2 Monaten ab",
            "gelb":       "läuft in ≤3 Monaten ab",
        }
        probleme = []
        for typ, eintrag in ma.get("schulungen", {}).items():
            dring = eintrag.get("_dringlichkeit", "")
            if dring in DRING_LABEL:
                anzeige = SCHULUNGSTYPEN_CFG.get(typ, {}).get("anzeige", typ)
                gb      = eintrag.get("gueltig_bis", "—") or "—"
                probleme.append(f"  • {anzeige}: {gb} ({DRING_LABEL[dring]})")
        if not probleme:
            QMessageBox.information(
                self, "Keine Probleme",
                f"{ma.get('vorname', '')} {ma.get('nachname', '')} hat keine abgelaufenen "
                f"oder demnächst ablaufenden Schulungen."
            )
            return
        vorname = ma.get("vorname", "")
        liste   = "\n".join(probleme)
        betreff = "Schulungsstatus – Handlungsbedarf"
        text = (
            f"Hallo {vorname},\n\n"
            f"bei der Überprüfung deiner Schulungsunterlagen haben wir festgestellt, dass "
            f"folgende Schulungen abgelaufen sind oder demnächst ablaufen:\n\n"
            f"{liste}\n\n"
            f"Bitte kümmere dich zeitnah um die Erneuerung.\n\n"
            f"Bei Fragen stehen wir gerne zur Verfügung.\n\n"
            f"Viele Grüße\nDRK Erste-Hilfe-Station FKB"
        )
        mailto = (
            f"mailto:{urllib.parse.quote(email)}"
            f"?subject={urllib.parse.quote(betreff)}"
            f"&body={urllib.parse.quote(text)}"
        )
        try:
            subprocess.Popen(["outlook", "/c", "ipm.note", "/m", mailto])
        except FileNotFoundError:
            os.startfile(mailto)

    def _tabelle_befuellen(self):
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG, lade_mitarbeiter_mit_schulungen
        # Daten neu laden für aktuellen Stand
        alle = lade_mitarbeiter_mit_schulungen()
        ma_id = self._ma["id"]
        for m in alle:
            if m["id"] == ma_id:
                self._ma = m
                break
        schulungen = self._ma.get("schulungen", {})

        self._tbl.setRowCount(len(SCHULUNGSTYPEN_CFG))
        self._tbl.blockSignals(True)
        for row, (key, cfg) in enumerate(SCHULUNGSTYPEN_CFG.items()):
            anzeige = cfg["anzeige"]
            eintrag = schulungen.get(key)
            if eintrag:
                dring = eintrag.get("_dringlichkeit", "")
                bg    = self._FARB_MAP.get(dring, "#ffffff")
                dat   = eintrag.get("datum_absolviert", "") or "—"
                gb    = eintrag.get("gueltig_bis", "") or "—"
                st    = eintrag.get("status", "") or "—"
                # Informiert-Anzeige
                if eintrag.get("informiert"):
                    inf_am = eintrag.get("informiert_am", "") or ""
                    inf_txt = f"✓ {inf_am}" if inf_am else "✓"
                else:
                    inf_txt = ""
                items = [
                    QTableWidgetItem(anzeige),
                    QTableWidgetItem(dat),
                    QTableWidgetItem(gb),
                    QTableWidgetItem(st),
                    QTableWidgetItem(inf_txt),
                ]
                if bg != "#ffffff":
                    for it in items[:4]:
                        it.setBackground(QColor(bg))
                if eintrag.get("informiert"):
                    items[4].setForeground(QColor("#2e7d32"))
            else:
                items = [QTableWidgetItem(anzeige)] + [QTableWidgetItem("") for _ in range(4)]
                for it in items:
                    it.setForeground(QColor("#bdbdbd"))
                items[0].setForeground(QColor("#9e9e9e"))

            # Typ + Eintrag für spätere Bearbeitung im Item speichern
            items[0].setData(Qt.ItemDataRole.UserRole, (key, eintrag))
            for col, it in enumerate(items):
                self._tbl.setItem(row, col, it)

            # ✏️-Bearbeiten-Button in letzter Spalte (col 5)
            btn_edit = QPushButton("✏️")
            btn_edit.setFixedSize(28, 24)
            btn_edit.setToolTip("Eintrag bearbeiten")
            btn_edit.setStyleSheet(
                "QPushButton{background:#e3f2fd;border:none;border-radius:3px;font-size:13px;}"
                "QPushButton:hover{background:#90caf9;}"
            )
            btn_edit.clicked.connect(lambda _=False, r=row: self._bearbeiten(r))
            self._tbl.setCellWidget(row, 5, btn_edit)
        self._tbl.blockSignals(False)

    def _bearbeiten(self, row: int):
        item = self._tbl.item(row, 0)
        if not item:
            return
        key, eintrag = item.data(Qt.ItemDataRole.UserRole)
        dlg = _SchulungBearbeitenDialog(self._ma["id"], key, eintrag, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._tabelle_befuellen()   # Tabelle sofort refreshen
            self.geaendert.emit()       # Signal nach oben weiterleiten


# ─── Mitarbeiter-Liste (durchsuchbar) ─────────────────────────────────────────
class _MitarbeiterListeWidget(QWidget):
    """Durchsuchbare, filterbare Liste aller Mitarbeiter mit Schulungsstatus.
    Doppelklick auf eine Zeile öffnet den Detaildialog mit allen 14 Schulungstypen.
    """

    # Schlüsseltypen für die Matrix-Spalten
    _MATRIX = [
        ("EH",                 "EH"),
        ("Refresher",          "Ref."),
        ("ZÜP",                "ZÜP"),
        ("Aerztl_Untersuchung","Ärztl."),
        ("Sicherheitsschulung","Sich.Sch."),
        ("PRM_Schulung",       "PRM-Sch."),
    ]

    _FARB_MAP = {
        "abgelaufen": "#ffcdd2",
        "rot":        "#ffe0b2",
        "orange":     "#fff9c4",
        "gelb":       "#f9fbe7",
        "ok":         "#e8f5e9",
        "einmalig":   "#eceff1",
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self._alle_daten: list[dict] = []
        self._build_ui()

    def _build_ui(self):
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 4, 0, 0)
        v.setSpacing(6)

        # ── Filterleiste ──────────────────────────────────────────────────
        fframe = QFrame()
        fframe.setStyleSheet(
            "QFrame{background:#e3f2fd;border:1px solid #90caf9;"
            "border-radius:4px;}"
        )
        fl = QHBoxLayout(fframe)
        fl.setContentsMargins(8, 5, 8, 5)
        fl.setSpacing(8)

        fl.addWidget(QLabel("🔍"))
        self._suche = QLineEdit()
        self._suche.setPlaceholderText("Name suchen …")
        self._suche.setMinimumWidth(170)
        self._suche.setClearButtonEnabled(True)
        self._suche.textChanged.connect(self._filter_anwenden)
        fl.addWidget(self._suche)

        fl.addWidget(QLabel("Status:"))
        self._filter_status = QComboBox()
        self._filter_status.addItems([
            "Alle", "Abgelaufen", "≤ 1 Monat", "≤ 2 Monate",
            "≤ 3 Monate", "OK", "Kein Eintrag",
        ])
        self._filter_status.currentIndexChanged.connect(self._filter_anwenden)
        fl.addWidget(self._filter_status)

        fl.addWidget(QLabel("Schulung:"))
        self._filter_typ = QComboBox()
        self._filter_typ.addItem("Alle Schulungen", None)
        for key, cfg in SCHULUNGSTYPEN_CFG.items():
            self._filter_typ.addItem(cfg["anzeige"], key)
        self._filter_typ.currentIndexChanged.connect(self._filter_anwenden)
        fl.addWidget(self._filter_typ)

        btn_reset = _btn_flat("✕ Zurücksetzen")
        btn_reset.clicked.connect(self._filter_zuruecksetzen)
        fl.addWidget(btn_reset)
        fl.addStretch()

        self._anzahl_lbl = QLabel()
        self._anzahl_lbl.setStyleSheet("color:#555;font-size:11px;")
        fl.addWidget(self._anzahl_lbl)

        btn_prm = _btn("📥  PRM importieren", "#6a1b9a", "#4a148c")
        btn_prm.setToolTip("PRM-Schulungen aus zertifikate_aktuell.xlsx importieren")
        btn_prm.clicked.connect(self._prm_importieren)
        fl.addWidget(btn_prm)
        v.addWidget(fframe)

        # ── Tabelle ───────────────────────────────────────────────────────
        self._tbl = QTableWidget()
        self._tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setStyleSheet("font-size:12px;")
        self._tbl.doubleClicked.connect(lambda idx: self._zeige_detail(idx.row()))
        self._tbl.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._tbl.customContextMenuRequested.connect(self._kontext_menu)
        v.addWidget(self._tbl, 1)

        hint = QLabel("💡 Doppelklick auf einen Mitarbeiter → alle 14 Schulungstypen anzeigen")
        hint.setStyleSheet("color:#777;font-size:10px;font-style:italic;padding:2px 0;")
        v.addWidget(hint)

    # ── Daten laden ───────────────────────────────────────────────────────
    def aktualisieren(self):
        from functions.schulungen_db import lade_mitarbeiter_mit_schulungen
        try:
            self._alle_daten = lade_mitarbeiter_mit_schulungen()
        except Exception:
            self._alle_daten = []
        self._filter_anwenden()

    # ── Filter ────────────────────────────────────────────────────────────
    _STATUS_DRING = {
        1: "abgelaufen",
        2: "rot",
        3: "orange",
        4: "gelb",
        5: "ok",
        6: "__kein_eintrag__",
    }
    _DRING_REIHE = ["abgelaufen", "rot", "orange", "gelb", "ok", "einmalig", ""]

    def _schlechtester(self, schulungen: dict) -> str:
        worst = len(self._DRING_REIHE) - 1
        for e in schulungen.values():
            d = e.get("_dringlichkeit", "")
            if d in self._DRING_REIHE:
                idx = self._DRING_REIHE.index(d)
                if idx < worst:
                    worst = idx
        return self._DRING_REIHE[worst]

    def _filter_anwenden(self):
        stext      = self._suche.text().strip().lower()
        status_idx = self._filter_status.currentIndex()
        typ_key    = self._filter_typ.currentData()
        filter_dr  = self._STATUS_DRING.get(status_idx)

        gefiltert = []
        for ma in self._alle_daten:
            name = f"{ma.get('nachname','')} {ma.get('vorname','')}".lower()
            if stext and stext not in name:
                continue

            schulungen = ma.get("schulungen", {})

            if typ_key:
                eintrag = schulungen.get(typ_key)
                if filter_dr == "__kein_eintrag__":
                    if eintrag:
                        continue
                elif filter_dr:
                    if not eintrag or eintrag.get("_dringlichkeit") != filter_dr:
                        continue
            else:
                if filter_dr == "__kein_eintrag__":
                    if schulungen:
                        continue
                elif filter_dr:
                    if self._schlechtester(schulungen) != filter_dr:
                        continue

            gefiltert.append(ma)

        # ── Mitarbeiter OHNE Einträge gesondert ans Ende ──────────────────
        mit_eintraegen   = [m for m in gefiltert if m.get("schulungen")]
        ohne_eintraegen  = [m for m in gefiltert if not m.get("schulungen")]
        geordnet = mit_eintraegen + ohne_eintraegen

        self._tabelle_befuellen(geordnet, len(ohne_eintraegen))
        self._anzahl_lbl.setText(
            f"{len(gefiltert)} Mitarbeiter"
            + (f"  ({len(ohne_eintraegen)} ohne Einträge)" if ohne_eintraegen else "")
        )

    def _tabelle_befuellen(self, daten: list[dict], anzahl_ohne: int):
        typ_key = self._filter_typ.currentData()

        if typ_key:
            from functions.schulungen_db import SCHULUNGSTYPEN_CFG
            anzeige = SCHULUNGSTYPEN_CFG.get(typ_key, {}).get("anzeige", typ_key)
            cols = ["Name", "Qualifikation", anzeige, "Gültig bis", "Status", ""]
        else:
            cols = ["Name", "Qualifikation"] + [k for _, k in self._MATRIX] + [""]

        self._tbl.setColumnCount(len(cols))
        self._tbl.setHorizontalHeaderLabels(cols)
        self._tbl.setRowCount(len(daten))

        hh = self._tbl.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        for i in range(2, len(cols) - 1):
            hh.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(len(cols) - 1, QHeaderView.ResizeMode.Fixed)
        self._tbl.setColumnWidth(0, 220)
        self._tbl.setColumnWidth(len(cols) - 1, 36)

        grenze = len(daten) - anzahl_ohne  # erste Zeile der "ohne Einträge"-Gruppe

        for row, ma in enumerate(daten):
            schulungen = ma.get("schulungen", {})
            name  = f"{ma.get('nachname', '')}, {ma.get('vorname', '')}".strip(", ")
            qual  = ma.get("qualifikation", "") or ""
            kein  = not schulungen  # Mitarbeiter ohne jegliche Einträge

            name_item = QTableWidgetItem(name)
            name_item.setData(Qt.ItemDataRole.UserRole, ma)
            qual_item = QTableWidgetItem(qual)

            if kein:
                # Gesamte Zeile grau einfärben
                for it in [name_item, qual_item]:
                    it.setForeground(QColor("#9e9e9e"))

            self._tbl.setItem(row, 0, name_item)
            self._tbl.setItem(row, 1, qual_item)

            if typ_key:
                eintrag = schulungen.get(typ_key)
                if eintrag:
                    dring = eintrag.get("_dringlichkeit", "")
                    bg    = self._FARB_MAP.get(dring, "#ffffff")
                    for col, text in enumerate([
                        SCHULUNGSTYPEN_CFG_K.get(typ_key, {}).get("anzeige", typ_key),
                        eintrag.get("gueltig_bis", "—") or "—",
                        eintrag.get("status", "—") or "—",
                    ], start=2):
                        it = QTableWidgetItem(text)
                        it.setBackground(QColor(bg))
                        self._tbl.setItem(row, col, it)
                else:
                    for col, text in enumerate(["—", "—", "—"], start=2):
                        it = QTableWidgetItem(text)
                        it.setForeground(QColor("#bdbdbd"))
                        self._tbl.setItem(row, col, it)
            else:
                for col, (typ, _) in enumerate(self._MATRIX, start=2):
                    eintrag = schulungen.get(typ)
                    if eintrag:
                        dring = eintrag.get("_dringlichkeit", "")
                        bg    = self._FARB_MAP.get(dring, "#ffffff")
                        it    = QTableWidgetItem(eintrag.get("gueltig_bis", "—") or "—")
                        it.setBackground(QColor(bg))
                    else:
                        it = QTableWidgetItem("—")
                        it.setForeground(QColor("#bdbdbd"))
                    self._tbl.setItem(row, col, it)

            # E-Mail-Button
            mail_col = len(cols) - 1
            btn_mail = QPushButton("📧")
            btn_mail.setFixedSize(28, 24)
            btn_mail.setToolTip("E-Mail mit abgelaufenen Schulungen senden")
            btn_mail.setStyleSheet(
                "QPushButton{background:#e3f2fd;border:none;border-radius:3px;font-size:13px;}"
                "QPushButton:hover{background:#90caf9;}"
            )
            btn_mail.clicked.connect(lambda _=False, m=ma: self._sende_email(m))
            self._tbl.setCellWidget(row, mail_col, btn_mail)

    def _zeige_detail(self, row: int):
        item = self._tbl.item(row, 0)
        if not item:
            return
        ma = item.data(Qt.ItemDataRole.UserRole)
        if not ma:
            return
        dlg = _MitarbeiterDetailDialog(ma, self)
        dlg.geaendert.connect(self.aktualisieren)   # nach Speichern Liste neu laden
        dlg.exec()

    def _sende_email(self, ma: dict):
        import urllib.parse, subprocess
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG

        # E-Mail-Adresse aus mitarbeiter.db nachschlagen
        email = ""
        try:
            from database.connection import get_ma_connection
            import sqlite3 as _sqlite3
            conn = get_ma_connection()
            conn.row_factory = _sqlite3.Row
            row = conn.execute(
                "SELECT email FROM mitarbeiter WHERE nachname=? AND vorname=? LIMIT 1",
                (ma.get("nachname", ""), ma.get("vorname", ""))
            ).fetchone()
            conn.close()
            if row:
                email = row["email"] or ""
        except Exception:
            pass

        # Abgelaufene / kritische Schulungen ermitteln
        DRING_LABEL = {
            "abgelaufen": "abgelaufen",
            "rot":        "läuft in ≤1 Monat ab",
            "orange":     "läuft in ≤2 Monaten ab",
            "gelb":       "läuft in ≤3 Monaten ab",
        }
        probleme = []
        for typ, eintrag in ma.get("schulungen", {}).items():
            dring = eintrag.get("_dringlichkeit", "")
            if dring in DRING_LABEL:
                anzeige = SCHULUNGSTYPEN_CFG.get(typ, {}).get("anzeige", typ)
                gb      = eintrag.get("gueltig_bis", "—") or "—"
                probleme.append(f"  • {anzeige}: {gb} ({DRING_LABEL[dring]})")

        if not probleme:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(
                self, "Keine Probleme",
                f"{ma.get('vorname', '')} {ma.get('nachname', '')} hat keine abgelaufenen "
                f"oder demnächst ablaufenden Schulungen."
            )
            return

        vorname  = ma.get("vorname", "")
        liste    = "\n".join(probleme)
        betreff  = "Schulungsstatus – Handlungsbedarf"
        text = (
            f"Hallo {vorname},\n\n"
            f"bei der Überprüfung deiner Schulungsunterlagen haben wir festgestellt, "
            f"dass folgende Schulungen abgelaufen sind oder demnächst ablaufen:\n\n"
            f"{liste}\n\n"
            f"Bitte kümmere dich zeitnah um die Erneuerung.\n\n"
            f"Bei Fragen stehen wir gerne zur Verfügung.\n\n"
            f"Viele Grüße\nDRK Erste-Hilfe-Station FKB"
        )
        mailto = (
            f"mailto:{urllib.parse.quote(email)}"
            f"?subject={urllib.parse.quote(betreff)}"
            f"&body={urllib.parse.quote(text)}"
        )
        try:
            subprocess.Popen(["outlook", "/c", "ipm.note", "/m", mailto])
        except FileNotFoundError:
            import os
            os.startfile(mailto)

    def _prm_importieren(self):
        """PRM-Schulungen aus zertifikate_aktuell.xlsx importieren."""
        from functions.schulungen_db import prm_schulung_importieren, _PRM_EXCEL_PFAD
        import os

        pfad = str(_PRM_EXCEL_PFAD)
        if not os.path.isfile(pfad):
            QMessageBox.warning(
                self, "Datei nicht gefunden",
                f"Die PRM-Schulungsdatei wurde nicht gefunden:\n{pfad}\n\n"
                "Bitte Datei prüfen und erneut versuchen.",
            )
            return

        antwort = QMessageBox.question(
            self, "PRM-Schulungen importieren",
            f"Sollen die PRM-Schulungen aus\n<b>{pfad}</b>\nimportiert werden?\n\n"
            "Bereits vorhandene PRM-Einträge werden aktualisiert.\n"
            "Nicht vorhandene Mitarbeiter werden neu angelegt.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort != QMessageBox.StandardButton.Yes:
            return

        fortschritt = QProgressDialog("PRM-Schulungen werden importiert …", None, 0, 0, self)
        fortschritt.setWindowModality(Qt.WindowModality.WindowModal)
        fortschritt.show()

        try:
            ergebnis = prm_schulung_importieren(pfad)
        except Exception as exc:
            fortschritt.close()
            QMessageBox.critical(self, "Import-Fehler", f"Fehler beim Import:\n{exc}")
            return
        finally:
            fortschritt.close()

        importiert   = ergebnis["importiert"]
        neu_angelegt = ergebnis["neu_angelegt"]
        kein_match   = ergebnis["kein_match"]

        bericht = (
            f"<b>Import abgeschlossen</b><br><br>"
            f"Einträge verarbeitet:  <b>{importiert}</b><br>"
            f"Neue Mitarbeiter angelegt: <b>{neu_angelegt}</b><br>"
        )
        if kein_match:
            bericht += (
                f"<br>Folgende Namen aus der Excel-Datei konnten keinem vorhandenen "
                f"Mitarbeiter zugeordnet werden und wurden <b>neu angelegt</b>:<br>"
                + "".join(f"&nbsp;&nbsp;• {n}<br>" for n in kein_match)
            )

        dlg = QMessageBox(self)
        dlg.setWindowTitle("PRM-Import Ergebnis")
        dlg.setIcon(QMessageBox.Icon.Information)
        dlg.setTextFormat(Qt.TextFormat.RichText)
        dlg.setText(bericht)
        dlg.exec()

        self.aktualisieren()

    def _filter_zuruecksetzen(self):
        self._suche.clear()
        self._filter_status.setCurrentIndex(0)
        self._filter_typ.setCurrentIndex(0)

    def _kontext_menu(self, pos):
        row = self._tbl.rowAt(pos.y())
        if row < 0:
            return
        item = self._tbl.item(row, 0)
        if not item:
            return
        ma = item.data(Qt.ItemDataRole.UserRole)
        if not ma:
            return
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self._tbl)
        akt_detail = menu.addAction("🎓  Schulungen anzeigen")
        menu.addSeparator()
        akt_loeschen = menu.addAction("🗑️  Mitarbeiter löschen")
        akt_loeschen.setObjectName("loeschen")
        aktion = menu.exec(self._tbl.viewport().mapToGlobal(pos))
        if aktion is akt_detail:
            self._zeige_detail(row)
        elif aktion is akt_loeschen:
            self._loesche_mitarbeiter(ma)

    def _loesche_mitarbeiter(self, ma: dict):
        name = f"{ma.get('nachname', '')}, {ma.get('vorname', '')}".strip(", ")
        antwort = QMessageBox.question(
            self, "Mitarbeiter löschen",
            f"Soll <b>{name}</b> wirklich gelöscht werden?<br>"
            "Alle Schulungseinträge dieses Mitarbeiters werden ebenfalls entfernt.<br><br>"
            "Diese Aktion kann nicht rückgängig gemacht werden.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort != QMessageBox.StandardButton.Yes:
            return
        try:
            from functions.schulungen_db import loesche_mitarbeiter
            loesche_mitarbeiter(ma["id"])
            self.aktualisieren()
        except Exception as exc:
            QMessageBox.critical(self, "Fehler beim Löschen", str(exc))


# ─── Haupt-Widget ─────────────────────────────────────────────────────────────
class SchulungenKalenderWidget(QWidget):
    """
    Großes Kalender-Widget für den Schulungen-Tab.
    Zeigt monatliche Übersicht + Agenda-Liste + Import/Anlegen-Buttons.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        _lade_typen()
        self._jahr  = date.today().year
        self._monat = date.today().month
        self._build_ui()
        # Einmaliger Auto-Import beim ersten Start (wenn DB noch leer)
        try:
            from functions.schulungen_db import erstimport_wenn_leer
            erstimport_wenn_leer()
        except Exception:
            pass
        self._aktualisieren()

    def _build_ui(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)

        # ── Steuerleiste ───────────────────────────────────────────────────
        ctrl = QFrame()
        ctrl.setStyleSheet(
            "QFrame{background:#e8f5e9;border:1px solid #a5d6a7;"
            "border-radius:4px;}"
        )
        cl = QHBoxLayout(ctrl)
        cl.setContentsMargins(10, 6, 10, 6)
        cl.setSpacing(10)

        self._btn_prev = _btn("◀", "#546e7a", "#455a64", 32)
        self._btn_prev.setFixedWidth(36)
        self._btn_prev.clicked.connect(self._vorheriger_monat)

        self._monat_lbl = QLabel()
        self._monat_lbl.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        self._monat_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._monat_lbl.setMinimumWidth(180)

        self._btn_next = _btn("▶", "#546e7a", "#455a64", 32)
        self._btn_next.setFixedWidth(36)
        self._btn_next.clicked.connect(self._naechster_monat)

        self._btn_heute = _btn_flat("Heute")
        self._btn_heute.clicked.connect(self._gehe_zu_heute)

        cl.addWidget(self._btn_prev)
        cl.addWidget(self._monat_lbl)
        cl.addWidget(self._btn_next)
        cl.addWidget(self._btn_heute)
        cl.addStretch()

        self._btn_import = _btn("📥  Excel importieren", "#1565c0", "#0d47a1", 32)
        self._btn_import.clicked.connect(self._excel_importieren)
        self._btn_neuer_ma = _btn("👤  Neuer Mitarbeiter", "#2e7d32", "#1b5e20", 32)
        self._btn_neuer_ma.clicked.connect(self._neuer_mitarbeiter)
        self._btn_excel_export = _btn("📊  Excel-Export", "#6a1b9a", "#4a148c", 32)
        self._btn_excel_export.clicked.connect(self._schulungen_excel_export)
        cl.addWidget(self._btn_import)
        cl.addWidget(self._btn_neuer_ma)
        cl.addWidget(self._btn_excel_export)
        v.addWidget(ctrl)

        # ── Haupt-Tabs: Kalender | Mitarbeiter-Liste ───────────────────────
        self._haupt_tabs = QTabWidget()
        self._haupt_tabs.currentChanged.connect(self._on_tab_wechsel)

        # Tab 0: Kalender mit Legende + Splitter
        kalender_widget = QWidget()
        kv = QVBoxLayout(kalender_widget)
        kv.setContentsMargins(0, 4, 0, 0)
        kv.setSpacing(6)

        # Legende
        leg = QHBoxLayout()
        leg.setSpacing(8)
        for key, label in [
            ("abgelaufen", "Abgelaufen"),
            ("rot", "≤ 1 Monat"),
            ("orange", "≤ 2 Monate"),
            ("gelb", "≤ 3 Monate"),
            ("ok", "OK"),
        ]:
            bg, fg = _chip_farbe(key)
            lbl = QLabel(f"  {label}  ")
            lbl.setStyleSheet(
                f"QLabel{{background:{bg};color:{fg};border-radius:4px;"
                f"font-size:11px;padding:2px 6px;}}"
            )
            leg.addWidget(lbl)
        leg.addStretch()
        kv.addLayout(leg)

        splitter = QSplitter(Qt.Orientation.Vertical)
        self._kalender = _MonatsKalender()
        self._kalender.tagesklick.connect(self._tag_geklickt)
        splitter.addWidget(self._kalender)
        self._agenda = _AgendaWidget()
        splitter.addWidget(self._agenda)
        splitter.setSizes([480, 260])
        kv.addWidget(splitter, 1)

        self._haupt_tabs.addTab(kalender_widget, "📅  Kalender / Agenda")

        # Tab 1: Mitarbeiter-Liste
        self._ma_liste = _MitarbeiterListeWidget()
        self._haupt_tabs.addTab(self._ma_liste, "👥  Mitarbeiter-Liste")

        v.addWidget(self._haupt_tabs, 1)

    def _on_tab_wechsel(self, idx: int):
        """Navigations-Buttons nur im Kalender-Tab einblenden."""
        kalender = (idx == 0)
        self._btn_prev.setVisible(kalender)
        self._monat_lbl.setVisible(kalender)
        self._btn_next.setVisible(kalender)
        self._btn_heute.setVisible(kalender)
        if not kalender:
            self._ma_liste.aktualisieren()


    def _aktualisieren(self):
        from functions.schulungen_db import lade_kalender_daten
        monate_de = [
            "", "Januar", "Februar", "März", "April", "Mai", "Juni",
            "Juli", "August", "September", "Oktober", "November", "Dezember"
        ]
        self._monat_lbl.setText(f"{monate_de[self._monat]} {self._jahr}")
        try:
            daten = lade_kalender_daten(self._jahr, self._monat)
        except Exception:
            daten = {}
        self._kalender.setze_monat(self._jahr, self._monat, daten)
        self._agenda.aktualisieren()
        # MA-Liste neu laden wenn aktiver Tab
        if self._haupt_tabs.currentIndex() == 1:
            self._ma_liste.aktualisieren()

    def _vorheriger_monat(self):
        if self._monat == 1:
            self._monat = 12
            self._jahr -= 1
        else:
            self._monat -= 1
        self._aktualisieren()

    def _naechster_monat(self):
        if self._monat == 12:
            self._monat = 1
            self._jahr += 1
        else:
            self._monat += 1
        self._aktualisieren()

    def _gehe_zu_heute(self):
        heute = date.today()
        self._jahr  = heute.year
        self._monat = heute.month
        self._aktualisieren()

    def _tag_geklickt(self, d: date, eintraege: list):
        dlg = _TagesDetailDialog(d, eintraege, self)
        dlg.exec()

    def _excel_importieren(self):
        from PySide6.QtWidgets import QFileDialog
        from functions.schulungen_db import excel_importieren, _EXCEL_PFAD
        pfad, _ = QFileDialog.getOpenFileName(
            self, "Excel-Stammdatei auswählen",
            str(_EXCEL_PFAD.parent),
            "Excel (*.xlsx *.xls)"
        )
        if not pfad:
            return
        prog = QProgressDialog("Importiere …", None, 0, 0, self)
        prog.setWindowTitle("Excel-Import")
        prog.setWindowModality(Qt.WindowModality.WindowModal)
        prog.show()
        try:
            imp, skip = excel_importieren(pfad)
            prog.close()
            QMessageBox.information(
                self, "Import abgeschlossen",
                f"✅ Erfolgreich importiert: {imp} Mitarbeiter\n"
                f"⏭ Übersprungen (leer): {skip}"
            )
            self._aktualisieren()
        except ImportError as e:
            prog.close()
            QMessageBox.critical(self, "Fehler", str(e))
        except Exception as e:
            prog.close()
            QMessageBox.critical(self, "Import-Fehler", str(e))

    def _schulungen_excel_export(self):
        """Öffnet den Export-Dialog und speichert ablaufende Schulungen als Excel."""
        from functions.schulungen_db import lade_eintraege_fuer_export

        dlg = _SchulungExportDialog(parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        von, bis, speicherpfad, typen, ohne_abgelaufene = dlg.get_werte()

        if not speicherpfad:
            QMessageBox.warning(self, "Kein Pfad", "Bitte einen Speicherpfad angeben.")
            return
        if not typen:
            QMessageBox.warning(self, "Keine Schulungsart", "Bitte mindestens eine Schulungsart auswählen.")
            return

        try:
            eintraege = lade_eintraege_fuer_export(von, bis, typen)
        except Exception as exc:
            QMessageBox.critical(self, "Datenbankfehler", str(exc))
            return

        if not eintraege:
            QMessageBox.information(self, "Keine Einträge",
                                     "Für die gewählten Filter wurden keine Einträge gefunden.")
            return

        if ohne_abgelaufene:
            eintraege = [e for e in eintraege if e.get("_tage_rest", 0) >= 0]

        # Sortieren: abgelaufen zuerst, dann nach Tagen
        eintraege.sort(key=lambda e: e.get("_tage_rest", 9999))

        try:
            _schulungen_als_excel_speichern(eintraege, speicherpfad)
        except Exception as exc:
            QMessageBox.critical(self, "Excel-Export fehlgeschlagen", str(exc))
            return

        antwort = QMessageBox.question(
            self, "Excel-Export erfolgreich",
            f"✅ {len(eintraege)} Einträge wurden exportiert:\n\n📄 {speicherpfad}\n\nDatei jetzt öffnen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort == QMessageBox.StandardButton.Yes:
            try:
                os.startfile(speicherpfad)
            except Exception as exc:
                QMessageBox.warning(self, "Öffnen fehlgeschlagen", str(exc))

    def _neuer_mitarbeiter(self):
        from functions.schulungen_db import (
            speichere_mitarbeiter, speichere_schulungseintrag,
            SCHULUNGSTYPEN_CFG, _berechne_status
        )
        from functions.mitarbeiter_sync import sync_neuer_mitarbeiter
        from datetime import date as _date

        dlg = NeuerMitarbeiterDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        try:
            stamm  = dlg.get_stamm_daten()
            schul  = dlg.get_schulungs_daten()
            ma_dat = dlg.get_ma_db_daten()

            # 1. Mitarbeiter-Stamm in schulungen.db
            ma_id = speichere_mitarbeiter(stamm)

            # 2. Schulungseinträge
            now = datetime.now().isoformat(timespec="seconds")
            for key, s in schul.items():
                cfg = SCHULUNGSTYPEN_CFG.get(key, {})
                from functions.schulungen_db import _parse_datum, _berechne_status as _bs
                gb = _parse_datum(s.get("gueltig_bis"))
                speichere_schulungseintrag({
                    "mitarbeiter_id":  ma_id,
                    "schulungstyp":    key,
                    "datum_absolviert": s.get("datum_absolviert",""),
                    "gueltig_bis":     s.get("gueltig_bis",""),
                    "laeuft_nicht_ab": int(cfg.get("laeuft_nicht_ab", False)),
                    "status":          _bs(gb, cfg.get("laeuft_nicht_ab", False)),
                    "bemerkung":       "",
                })

            # 3. Sync in andere DBs
            sync_cfg = {}
            if dlg.sync_aktiviert_ma():
                sync_cfg["mitarbeiter.db"] = True
            if dlg.sync_aktiviert_nesk():
                sync_cfg["nesk3.db"] = True

            ergebnis = sync_neuer_mitarbeiter(stamm, ma_dat)
            meldungen = [f"✅ schulungen.db – Mitarbeiter + {len(schul)} Schulungseinträge"]
            for db, ok in ergebnis.items():
                if db == "mitarbeiter.db" and not dlg.sync_aktiviert_ma():
                    continue
                if db == "nesk3.db" and not dlg.sync_aktiviert_nesk():
                    continue
                meldungen.append(f"{'✅' if ok else '⏭'} {db} – {'angelegt' if ok else 'bereits vorhanden'}")

            QMessageBox.information(
                self, "Mitarbeiter angelegt",
                f"{stamm['nachname']}, {stamm['vorname']} wurde erfolgreich angelegt.\n\n"
                + "\n".join(meldungen)
            )
            self._aktualisieren()

        except Exception as exc:
            QMessageBox.critical(self, "Fehler beim Anlegen", str(exc))
