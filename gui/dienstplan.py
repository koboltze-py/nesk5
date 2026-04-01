"""
Dienstplan-Widget
Schichten anzeigen, hinzufügen, bearbeiten, löschen
Excel-Import und Word-Export (Stärkemeldung)
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QDialog,
    QFormLayout, QComboBox, QDateEdit, QTimeEdit, QTextEdit,
    QMessageBox, QFileDialog, QSpinBox, QFrame, QLineEdit,
    QTreeView, QSplitter, QFileSystemModel,
    QScrollArea, QCheckBox, QRadioButton, QButtonGroup
)
from PySide6.QtCore import Qt, QDate, QTime, Signal, QFileSystemWatcher
from PySide6.QtGui import QFont, QColor

from config import FIORI_BLUE, FIORI_TEXT, FIORI_ERROR
from database.models import Dienstplan


class DienstplanDialog(QDialog):
    """Dialog zum Erstellen/Bearbeiten einer Schicht."""
    def __init__(self, schicht: Dienstplan = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Schicht" + (" bearbeiten" if schicht else " hinzufügen"))
        self.setMinimumWidth(420)
        self.result_schicht: Dienstplan | None = None
        self._s = schicht or Dienstplan()
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(10)

        # Mitarbeiter-Auswahl
        self._mitarbeiter_combo = QComboBox()
        self._ma_map: dict[str, int] = {}
        try:
            from functions.mitarbeiter_functions import get_alle_mitarbeiter
            for ma in get_alle_mitarbeiter(nur_aktive=True):
                label = f"{ma.vollname} ({ma.personalnummer})"
                self._mitarbeiter_combo.addItem(label)
                self._ma_map[label] = ma.id
                if ma.id == self._s.mitarbeiter_id:
                    self._mitarbeiter_combo.setCurrentText(label)
        except Exception:
            self._mitarbeiter_combo.addItem("(Kein Mitarbeiter verfügbar)")

        self._datum = QDateEdit()
        self._datum.setCalendarPopup(True)
        self._datum.setDisplayFormat("dd.MM.yyyy")
        if self._s.datum:
            self._datum.setDate(QDate(self._s.datum.year, self._s.datum.month, self._s.datum.day))
        else:
            self._datum.setDate(QDate.currentDate())

        self._start = QTimeEdit()
        self._start.setDisplayFormat("HH:mm")
        self._start.setTime(QTime(6, 0) if not self._s.start_uhrzeit
                            else QTime(self._s.start_uhrzeit.hour, self._s.start_uhrzeit.minute))

        self._end = QTimeEdit()
        self._end.setDisplayFormat("HH:mm")
        self._end.setTime(QTime(14, 0) if not self._s.end_uhrzeit
                          else QTime(self._s.end_uhrzeit.hour, self._s.end_uhrzeit.minute))

        self._position = QComboBox()
        try:
            from functions.mitarbeiter_functions import get_positionen
            self._position.addItems(get_positionen())
        except Exception:
            self._position.addItems(["Notfallsanitäter", "Rettungssanitäter"])
        if self._s.position:
            idx = self._position.findText(self._s.position)
            if idx >= 0:
                self._position.setCurrentIndex(idx)

        self._typ = QComboBox()
        self._typ.addItems(["regulär", "nacht", "bereitschaft"])
        self._typ.setCurrentText(self._s.schicht_typ)

        self._notizen = QTextEdit(self._s.notizen)
        self._notizen.setMaximumHeight(80)

        form.addRow("Mitarbeiter *:", self._mitarbeiter_combo)
        form.addRow("Datum *:",       self._datum)
        form.addRow("Von:",           self._start)
        form.addRow("Bis:",           self._end)
        form.addRow("Position:",      self._position)
        form.addRow("Schicht-Typ:",   self._typ)
        form.addRow("Notizen:",       self._notizen)
        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Speichern")
        save_btn.setMinimumHeight(40)
        save_btn.setStyleSheet(f"background-color: {FIORI_BLUE}; color: white; font-size: 13px; border-radius: 4px;")
        save_btn.clicked.connect(self._save)
        cancel_btn = QPushButton("Abbrechen")
        cancel_btn.setMinimumHeight(40)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)

    def _save(self):
        selected_label = self._mitarbeiter_combo.currentText()
        ma_id = self._ma_map.get(selected_label)
        if not ma_id:
            QMessageBox.warning(self, "Fehler", "Bitte einen Mitarbeiter auswählen.")
            return

        from datetime import date, time
        qd = self._datum.date()
        qs = self._start.time()
        qe = self._end.time()

        self._s.mitarbeiter_id  = ma_id
        self._s.datum           = date(qd.year(), qd.month(), qd.day())
        self._s.start_uhrzeit   = time(qs.hour(), qs.minute())
        self._s.end_uhrzeit     = time(qe.hour(), qe.minute())
        self._s.position        = self._position.currentText()
        self._s.schicht_typ     = self._typ.currentText()
        self._s.notizen         = self._notizen.toPlainText().strip()

        self.result_schicht = self._s
        self.accept()


# Dienste die zum Standard gehören  -  Sonderdienste werden im Dialog angezeigt
_STANDARD_DIENSTE = frozenset({'N', 'N10', 'T', 'T10', 'DT', 'DT3', 'DN', 'DN3'})

_TAG_DIENSTE   = frozenset({'T', 'T10', 'T8', 'DT', 'DT3'})
_NACHT_DIENSTE = frozenset({'N', 'N10', 'NF', 'DN', 'DN3'})


class ExportDialog(QDialog):
    """Dialog für Word-Export: Zeitraum, PAX-Zahl, Ausgabepfad, Sonderdienst-Filter."""

    _STAERKEMELDUNG_DIR = (
        r"C:\Users\DRKairport\OneDrive - Deutsches Rotes Kreuz - Kreisverband Köln e.V"
        r"\Dateien von Erste-Hilfe-Station-Flughafen - DRK Köln e.V_ - !Gemeinsam.26"
        r"\06_Stärkemeldung"
    )

    def __init__(self, parsed_data: dict | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Word-Export  -  Stärkemeldung")
        self.setMinimumWidth(500)
        self.setMinimumHeight(460)
        self._parsed_data = parsed_data or {}
        self.result: dict | None = None
        self._checkboxen: list[tuple] = []   # (QCheckBox, vollname_lower)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form   = QFormLayout()
        form.setSpacing(10)

        today = QDate.currentDate()

        self._von = QDateEdit(today)
        self._von.setCalendarPopup(True)
        self._von.setDisplayFormat("dd.MM.yyyy")

        self._bis = QDateEdit(today)
        self._bis.setCalendarPopup(True)
        self._bis.setDisplayFormat("dd.MM.yyyy")

        self._pax = QSpinBox()
        self._pax.setRange(0, 99999)
        self._pax.setValue(0)

        self._bulmor = QSpinBox()
        self._bulmor.setRange(0, 5)
        self._bulmor.setValue(5)
        self._bulmor.setToolTip("Anzahl fahrbereiter Bulmor-Fahrzeuge (wird im Dashboard-Format angezeigt)")

        self._einsaetze = QSpinBox()
        self._einsaetze.setRange(0, 9999)
        self._einsaetze.setValue(0)
        self._einsaetze.setToolTip("Anzahl der Schichtleiter-Einsaetze / Behandlungen dieser Schicht")

        self._sl_tag = QLineEdit()
        self._sl_tag.setPlaceholderText("Name Schichtleiter Tag")
        self._sl_nacht = QLineEdit()
        self._sl_nacht.setPlaceholderText("Name Schichtleiter Nacht")

        form.addRow("Von:",              self._von)
        form.addRow("Bis:",              self._bis)
        form.addRow("PAX-Zahl:",        self._pax)
        form.addRow("Aktive Bulmor:",   self._bulmor)
        form.addRow("SL-Einsaetze:",   self._einsaetze)
        form.addRow("SL Tag (Name):",   self._sl_tag)
        form.addRow("SL Nacht (Name):", self._sl_nacht)
        layout.addLayout(form)

        # -- Format-Auswahl -------------------------------------------
        sep_fmt = QFrame()
        sep_fmt.setFrameShape(QFrame.Shape.HLine)
        sep_fmt.setStyleSheet("color: #ddd;")
        layout.addWidget(sep_fmt)

        fmt_lbl = QLabel("Exportformat:")
        fmt_lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        fmt_lbl.setStyleSheet("color: #333; padding: 2px 0px;")
        layout.addWidget(fmt_lbl)

        self._fmt_gruppe = QButtonGroup(self)
        self._rb_klassisch = QRadioButton("Klassische Stärkemeldung  (bisheriges Format)")
        self._rb_dashboard = QRadioButton("Mit Dashboard-Panel links  (neues Format)")
        self._rb_klassisch.setChecked(True)
        self._fmt_gruppe.addButton(self._rb_klassisch, 0)
        self._fmt_gruppe.addButton(self._rb_dashboard, 1)
        layout.addWidget(self._rb_klassisch)
        layout.addWidget(self._rb_dashboard)

        # -- Sonderdienst-Abschnitt -------------------------------------
        sonder_personen = []
        for listen_key in ('betreuer', 'dispo'):
            for p in self._parsed_data.get(listen_key, []):
                dk = (p.get('dienst_kategorie') or '').upper()
                if dk not in _STANDARD_DIENSTE:
                    sonder_personen.append(p)

        if sonder_personen:
            sep1 = QFrame()
            sep1.setFrameShape(QFrame.Shape.HLine)
            sep1.setStyleSheet("color: #ddd;")
            layout.addWidget(sep1)

            sonder_lbl = QLabel(
                " Mitarbeiter mit Sonderdiensten:\n"
                "   Haken setzen = vom Export ausschließen"
            )
            sonder_lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
            sonder_lbl.setStyleSheet("color: #555; padding: 2px 0px;")
            layout.addWidget(sonder_lbl)

            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setMaximumHeight(160)
            scroll.setStyleSheet(
                "QScrollArea { border: 1px solid #dce8f5; border-radius: 4px; }"
            )
            inner        = QWidget()
            inner_layout = QVBoxLayout(inner)
            inner_layout.setSpacing(3)
            inner_layout.setContentsMargins(8, 6, 8, 6)

            try:
                from functions.settings_functions import get_ausgeschlossene_namen
                settings_ausgeschlossen = set(get_ausgeschlossene_namen())
            except Exception:
                settings_ausgeschlossen = set()

            for p in sonder_personen:
                vollname_lower = p.get('vollname', '').lower()
                dienst         = p.get('dienst_kategorie', '') or '?'
                anzeige        = p.get('anzeigename', p.get('vollname', ''))
                cb = QCheckBox(f"  {anzeige}  [{dienst}]")
                cb.setFont(QFont("Arial", 10))
                # vorbelegen: ausschließen wenn bereits in Einstellungen
                cb.setChecked(vollname_lower in settings_ausgeschlossen)
                inner_layout.addWidget(cb)
                self._checkboxen.append((cb, vollname_lower))

            inner_layout.addStretch()
            scroll.setWidget(inner)
            layout.addWidget(scroll)

        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet("color: #ddd;")
        layout.addWidget(sep2)

        btn_row = QHBoxLayout()
        export_btn = QPushButton("Exportieren")
        export_btn.setMinimumHeight(40)
        export_btn.setStyleSheet(
            f"background-color: {FIORI_BLUE}; color: white; font-size: 13px; border-radius: 4px;"
        )
        export_btn.clicked.connect(self._export)
        cancel_btn = QPushButton("Abbrechen")
        cancel_btn.setMinimumHeight(40)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(export_btn)
        layout.addLayout(btn_row)

    def _export(self):
        # Speicherort wählen
        ziel_dir = self._STAERKEMELDUNG_DIR if os.path.isdir(self._STAERKEMELDUNG_DIR) else os.path.expanduser("~")
        qv = self._von.date()
        qb = self._bis.date()
        von_str = f"{qv.day():02d}.{qv.month():02d}.{qv.year()}"
        bis_str = f"{qb.day():02d}.{qb.month():02d}.{qb.year()}"
        if qv == qb:
            default_name = f"Stärkemeldung {von_str}.docx"
        else:
            default_name = f"Stärkemeldung {von_str} - {bis_str}.docx"
        ausgabe_pfad, _ = QFileDialog.getSaveFileName(
            self, "Speicherort wählen", os.path.join(ziel_dir, default_name),
            "Word-Dokument (*.docx)"
        )
        if not ausgabe_pfad:
            return

        # PAX-Zahl = 0 → Nutzer darauf aufmerksam machen
        if self._pax.value() == 0:
            ret = QMessageBox.warning(
                self, "PAX-Zahl ist 0",
                "Die PAX-Zahl ist aktuell 0.\n\n"
                "Bitte tragen Sie die Anzahl der Passagiere ein,\n"
                "oder klicken Sie auf 'Trotzdem exportieren'.",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Ignore,
                QMessageBox.StandardButton.Ok,
            )
            if ret != QMessageBox.StandardButton.Ignore:
                self._pax.setFocus()
                self._pax.selectAll()
                return

        ausgeschlossene = {vn for cb, vn in self._checkboxen if cb.isChecked()}
        self.result = {
            'von_datum':               datetime(qv.year(), qv.month(), qv.day()),
            'bis_datum':               datetime(qb.year(), qb.month(), qb.day()),
            'pax_zahl':                self._pax.value(),
            'bulmor_aktiv':            self._bulmor.value(),
            'einsaetze_zahl':          self._einsaetze.value(),
            'sl_tag_name':             self._sl_tag.text().strip(),
            'sl_nacht_name':           self._sl_nacht.text().strip(),
            'ausgabe_pfad':            ausgabe_pfad,
            'ausgeschlossene_vollnamen': ausgeschlossene,
            'format':                  'dashboard' if self._rb_dashboard.isChecked() else 'klassisch',
        }
        self.accept()


_ALLE_DIENST_TYPEN = [
    'T', 'T10', 'N', 'N10', 'NF',
    'DT', 'DT3', 'DN', 'DN3', 'D',
    'FB', 'FB1', 'FB2',
    'R', 'B1', 'B2',
    'KRANK',
]
_ALLE_ZEITEN = ['%02d:%02d' % (h, m) for h in range(24) for m in (0, 15, 30, 45)]


class EditDienstDialog(QDialog):
    """Dialog zum Bearbeiten von Dienst, Von- und Bis-Zeit einer Zeile."""

    def __init__(self, person: dict, parent=None, hinweis: str = ''):
        super().__init__(parent)
        self.setWindowTitle('Dienst bearbeiten')
        self.setMinimumWidth(340)
        self.result_data: dict | None = None
        self._person = person
        self._hinweis = hinweis
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        name_lbl = QLabel(self._person.get('vollname', ''))
        name_lbl.setFont(QFont('Arial', 12, QFont.Weight.Bold))
        name_lbl.setStyleSheet(f'color: {FIORI_TEXT};')
        layout.addWidget(name_lbl)

        if self._hinweis:
            hint_lbl = QLabel(f'ℹ  {self._hinweis}')
            hint_lbl.setWordWrap(True)
            hint_lbl.setFont(QFont('Arial', 9))
            hint_lbl.setStyleSheet(
                'background:#fff8e1; border:1px solid #f0c040; border-radius:4px;'
                'color:#7a5000; padding:5px 8px;'
            )
            layout.addWidget(hint_lbl)

        form = QFormLayout()
        form.setSpacing(10)

        self._dienst_cb = QComboBox()
        self._dienst_cb.setEditable(True)
        self._dienst_cb.addItems(_ALLE_DIENST_TYPEN)
        current = (self._person.get('dienst_kategorie') or '').upper()
        idx = self._dienst_cb.findText(current)
        if idx >= 0:
            self._dienst_cb.setCurrentIndex(idx)
        else:
            self._dienst_cb.setCurrentText(current)

        self._von_cb = QComboBox()
        self._von_cb.setEditable(True)
        self._von_cb.addItems(_ALLE_ZEITEN)
        self._von_cb.setCurrentText(self._person.get('start_zeit', '') or '')

        self._bis_cb = QComboBox()
        self._bis_cb.setEditable(True)
        self._bis_cb.addItems(_ALLE_ZEITEN)
        self._bis_cb.setCurrentText(self._person.get('end_zeit', '') or '')

        form.addRow('Dienst:', self._dienst_cb)
        form.addRow('Von:', self._von_cb)
        form.addRow('Bis:', self._bis_cb)
        layout.addLayout(form)

        btn_row = QHBoxLayout()
        save_btn = QPushButton('\U0001f4be Speichern')
        save_btn.setMinimumHeight(38)
        save_btn.setStyleSheet(
            f'background-color: {FIORI_BLUE}; color: white; '
            f'font-size: 13px; border-radius: 4px;'
        )
        save_btn.clicked.connect(self._save)
        cancel_btn = QPushButton('Abbrechen')
        cancel_btn.setMinimumHeight(38)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(save_btn)
        layout.addLayout(btn_row)

    def _save(self):
        dienst = self._dienst_cb.currentText().strip().upper()
        von    = self._von_cb.currentText().strip()
        bis    = self._bis_cb.currentText().strip()
        self.result_data = {'dienst': dienst, 'von': von, 'bis': bis}
        self.accept()


class _DispoZeitenVorschauDialog(QDialog):
    """
    Zeigt alle Dispo- und Betreuer-Eintraege des Dienstplans vor dem Export zur
    Kontrolle an. Zeiten koennen per Doppelklick oder 'Bearbeiten'-Knopf
    angepasst werden; die Aenderungen fliessen in den Word-Export ein.
    """

    def __init__(self, parsed_data: dict, parent=None, raw_data: dict | None = None):
        super().__init__(parent)
        import copy
        self.setWindowTitle("Dispo-Zeiten – Vorschau & Bearbeitung")
        self.setMinimumWidth(720)
        self.setMinimumHeight(440)
        self._data = copy.deepcopy(parsed_data)
        # Originale Excel-Zeiten vor dem Runden sichern (für Vergleichsspalten)
        # Bevorzugt aus raw_data (Parser ohne Rundung), sonst Fallback auf parsed_data
        src = raw_data if raw_data else parsed_data
        self._orig_zeiten: dict[tuple, tuple] = {}
        for i, p in enumerate(src.get('dispo', [])):
            self._orig_zeiten[('dispo', i)] = (
                p.get('start_zeit') or '–', p.get('end_zeit') or '–'
            )
        for i, p in enumerate(src.get('betreuer', [])):
            self._orig_zeiten[('betreuer', i)] = (
                p.get('start_zeit') or '–', p.get('end_zeit') or '–'
            )
        # Dispo-Zeiten vorab auf volle Stunden runden (so wie der Word-Exporter es tut)
        for p in self._data.get('dispo', []):
            for key in ('start_zeit', 'end_zeit'):
                t = p.get(key) or ''
                if t and ':' in t:
                    p[key] = f"{int(t.split(':')[0]):02d}:00"
        self._row_to_person: list[tuple[str, int]] = []
        self._build_ui()

    # ── UI ──────────────────────────────────────────────────────────────────

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        title = QLabel("✏  Dispo-Zeiten prüfen und anpassen")
        title.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {FIORI_TEXT};")
        layout.addWidget(title)

        hint = QLabel(
            "Doppelklick auf eine Zeile oder «Bearbeiten»-Knopf, um Dienst / Von / Bis "
            "zu ändern. Geänderte Werte werden im Word-Export übernommen."
        )
        hint.setWordWrap(True)
        hint.setFont(QFont("Arial", 9))
        hint.setStyleSheet("color: #666;")
        layout.addWidget(hint)

        self._table = QTableWidget()
        self._table.setColumnCount(6)
        self._table.setHorizontalHeaderLabels([
            "Name (Kategorie)", "Dienst",
            "Von (Excel)", "Bis (Excel)",
            "Von (Export)", "Bis (Export)"
        ])
        self._table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        for col in (1, 2, 3, 4, 5):
            self._table.horizontalHeader().setSectionResizeMode(
                col, QHeaderView.ResizeMode.ResizeToContents
            )
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.setStyleSheet(
            "QTableWidget{border:1px solid #ddd; font-size:12px;}"
        )
        self._table.verticalHeader().setVisible(False)
        self._table.itemDoubleClicked.connect(lambda item: self._edit_row(item.row()))
        layout.addWidget(self._table, 1)

        self._rebuild_table()

        # ── Buttons ──────────────────────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#ddd; margin:2px 0px;")
        layout.addWidget(sep)

        btn_row = QHBoxLayout()

        btn_edit = QPushButton("✏  Bearbeiten")
        btn_edit.setFixedHeight(36)
        btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_edit.setToolTip("Markierte Zeile bearbeiten (Dienst, Von, Bis)")
        btn_edit.setStyleSheet(
            f"QPushButton{{background:{FIORI_BLUE};color:white;border:none;"
            f"border-radius:4px;padding:4px 14px;font-size:12px;}}"
            f"QPushButton:hover{{background:#0855a9;}}"
        )
        btn_edit.clicked.connect(self._edit_selected)
        btn_row.addWidget(btn_edit)
        btn_row.addStretch()

        btn_cancel = QPushButton("Abbrechen")
        btn_cancel.setFixedHeight(36)
        btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancel.clicked.connect(self.reject)

        btn_weiter = QPushButton("Weiter  →")
        btn_weiter.setFixedHeight(36)
        btn_weiter.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_weiter.setToolTip("Zeiten übernehmen und zum Export-Dialog weitergehend")
        btn_weiter.setStyleSheet(
            "QPushButton{background:#107e3e;color:white;border:none;"
            "border-radius:4px;padding:4px 18px;font-size:12px;}"
            "QPushButton:hover{background:#0d6131;}"
        )
        btn_weiter.clicked.connect(self.accept)

        btn_row.addWidget(btn_cancel)
        btn_row.addSpacing(6)
        btn_row.addWidget(btn_weiter)
        layout.addLayout(btn_row)

    # ── Tabellenlogik ────────────────────────────────────────────────────────

    def _rebuild_table(self):
        self._row_to_person.clear()
        rows: list[tuple[str, int, dict]] = []
        for kat_key in ('dispo', 'betreuer'):
            for idx, p in enumerate(self._data.get(kat_key, [])):
                rows.append((kat_key, idx, p))

        self._table.setRowCount(len(rows))
        for row, (kat_key, idx, p) in enumerate(rows):
            self._row_to_person.append((kat_key, idx))
            kat_lbl = "Dispo" if kat_key == "dispo" else "Betreuer"
            anzeige = p.get("anzeigename") or p.get("vollname") or "–"
            name_item = QTableWidgetItem(f"[{kat_lbl}]  {anzeige}")
            name_item.setForeground(
                QColor("#0a5ba4") if kat_key == "dispo" else QColor(FIORI_TEXT)
            )
            self._table.setItem(row, 0, name_item)
            self._table.setItem(row, 1, QTableWidgetItem(p.get("dienst_kategorie", "") or "–"))

            orig_von, orig_bis = self._orig_zeiten.get((kat_key, idx), ('–', '–'))
            exp_von = p.get("start_zeit", "") or "–"
            exp_bis = p.get("end_zeit", "") or "–"

            self._table.setItem(row, 2, QTableWidgetItem(orig_von))
            self._table.setItem(row, 3, QTableWidgetItem(orig_bis))

            it_von = QTableWidgetItem(exp_von)
            it_bis = QTableWidgetItem(exp_bis)
            # Exportzeit blau hervorheben wenn sie von der Excel-Zeit abweicht
            if exp_von != orig_von:
                it_von.setForeground(QColor("#0a6ed1"))
                it_von.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            if exp_bis != orig_bis:
                it_bis.setForeground(QColor("#0a6ed1"))
                it_bis.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self._table.setItem(row, 4, it_von)
            self._table.setItem(row, 5, it_bis)

    def _edit_selected(self):
        self._edit_row(self._table.currentRow())

    def _edit_row(self, row: int):
        if row < 0 or row >= len(self._row_to_person):
            return
        kat_key, idx = self._row_to_person[row]
        person = self._data[kat_key][idx]
        hinweis = (
            'Dispo-Zeiten werden im Word-Export auf volle Stunden gerundet '
            '(z. B. 07:30 → 07:00).'
            if kat_key == 'dispo' else ''
        )
        dlg = EditDienstDialog(person, parent=self, hinweis=hinweis)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.result_data:
            d = dlg.result_data
            person["dienst_kategorie"] = d["dienst"]
            person["start_zeit"] = d["von"]
            person["end_zeit"] = d["bis"]
            person["manuell_geaendert"] = True
            self._rebuild_table()
            self._table.selectRow(row)

    # ── Ergebnis ─────────────────────────────────────────────────────────────

    @property
    def modified_data(self) -> dict:
        """Gibt die (ggf. bearbeiteten) Dienstplan-Daten zurück."""
        return self._data


class _DienstplanPane(QWidget):
    """Einzelne Tabellen-Ansicht fuer einen geoeffneten Dienstplan."""

    # Signal: dieser Pane wurde als Export-Quelle aktiviert (sendet eigenen Index)
    export_selected = Signal(int)

    def __init__(self, pane_index: int = 0, parent=None):
        super().__init__(parent)
        self._pane_index: int           = pane_index
        self._parsed_data: dict | None  = None
        self._display_data: dict | None = None
        self._table_row_data: list      = []
        self._excel_path: str           = ''
        self._is_export_active: bool    = False
        self._build_ui()

    # ---------- Eigenschaften ----------

    @property
    def excel_path(self) -> str:
        return self._excel_path

    @property
    def parsed_data(self) -> dict | None:
        return self._parsed_data

    @property
    def is_empty(self) -> bool:
        return not bool(self._excel_path)

    def set_export_active(self, active: bool):
        """Markiert diese Pane visuell als Export-Quelle."""
        self._is_export_active = active
        self._update_header_style()
        self._export_btn.setEnabled(not active)
        self._export_btn.setText(
            '✓  Für Wordexport gewählt' if active
            else 'Hier klicken um Datei als Wordexport auszuwählen'
        )

    def _update_header_style(self):
        if self._is_export_active:
            self._header_bar.setStyleSheet(
                'background: #0a5ba4; border-radius: 4px 4px 0 0; padding: 2px 4px;'
            )
            self._title_lbl.setStyleSheet('color: white; font-weight: bold; font-size: 11px;')
        else:
            self._header_bar.setStyleSheet(
                'background: #f0f4f8; border: 1px solid #dce8f5; '
                'border-radius: 4px 4px 0 0; padding: 2px 4px;'
            )
            self._title_lbl.setStyleSheet(f'color: {FIORI_TEXT}; font-size: 11px;')

    def clear(self):
        self._parsed_data    = None
        self._display_data   = None
        self._table_row_data = []
        self._excel_path     = ''
        self._table.clearContents()
        self._table.setRowCount(0)
        self._datum_lbl.setVisible(False)
        self._status_lbl.setText('Doppelklick auf eine Datei im Baum, um sie zu laden.')
        self._status_lbl.setStyleSheet('color: #888; padding: 2px 0px;')
        self._row_count_lbl.setText('0 Eintraege')
        self._title_lbl.setText(f'Dienstplan {self._pane_index + 1}')

    # ---------- UI ----------

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        # Header-Zeile: Titel + Export-Button + Schliessen
        self._header_bar = QWidget()
        header_layout = QHBoxLayout(self._header_bar)
        header_layout.setContentsMargins(6, 3, 4, 3)
        header_layout.setSpacing(4)

        self._title_lbl = QLabel(f'Dienstplan {self._pane_index + 1}')
        self._title_lbl.setFont(QFont('Arial', 11))
        header_layout.addWidget(self._title_lbl)
        header_layout.addStretch()

        self._export_btn = QPushButton('Fuer Export auswaehlen')
        self._export_btn.setFixedHeight(22)
        self._export_btn.setToolTip('Diesen Dienstplan-Tab für den Stundenlisten-Export auswählen')
        self._export_btn.setStyleSheet(
            'font-size: 10px; padding: 0 6px; border-radius: 3px; '
            f'background: {FIORI_BLUE}; color: white; border: none;'
        )
        self._export_btn.clicked.connect(lambda: self.export_selected.emit(self._pane_index))
        header_layout.addWidget(self._export_btn)

        self._excel_open_btn = QPushButton('📊 In Excel öffnen')
        self._excel_open_btn.setFixedHeight(22)
        self._excel_open_btn.setEnabled(False)
        self._excel_open_btn.setToolTip('Dienstplan-Excel-Datei in Excel öffnen')
        self._excel_open_btn.setStyleSheet(
            'font-size: 10px; padding: 0 6px; border-radius: 3px; '
            'background: #107e3e; color: white; border: none;'
        )
        self._excel_open_btn.clicked.connect(self._open_in_excel)
        header_layout.addWidget(self._excel_open_btn)

        self._close_btn = QPushButton('X')
        self._close_btn.setFixedSize(22, 22)
        self._close_btn.setStyleSheet(
            'font-size: 10px; font-weight: bold; padding: 0; border-radius: 3px; '
            'background: #c0392b; color: white; border: none;'
        )
        self._close_btn.setToolTip('Dienstplan schliessen')
        self._close_btn.clicked.connect(self._on_close_clicked)
        header_layout.addWidget(self._close_btn)

        self._update_header_style()
        layout.addWidget(self._header_bar)

        # Hauptinhalt
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(4, 2, 4, 4)
        content_layout.setSpacing(3)
        layout.addWidget(content, 1)

        # Ersetzt layout durch content_layout fuer den Rest der UI
        layout = content_layout

        self._datum_lbl = QLabel('')
        self._datum_lbl.setFont(QFont('Arial', 12, QFont.Weight.Bold))
        self._datum_lbl.setStyleSheet(f'color: {FIORI_TEXT}; padding: 2px 0px;')
        self._datum_lbl.setVisible(False)
        layout.addWidget(self._datum_lbl)

        self._status_lbl = QLabel('Doppelklick auf eine Datei im Baum, um sie zu laden.')
        self._status_lbl.setFont(QFont('Arial', 9))
        self._status_lbl.setWordWrap(True)
        self._status_lbl.setStyleSheet('color: #888; padding: 1px 0px;')
        layout.addWidget(self._status_lbl)

        self._table = QTableWidget()
        self._table.setColumnCount(5)
        self._table.setHorizontalHeaderLabels([
            'Kategorie', 'Name', 'Dienst', 'Von', 'Bis'
        ])
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(False)
        self._table.horizontalHeader().setMinimumSectionSize(40)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(18)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.itemDoubleClicked.connect(self._on_table_double_click)
        self._table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                gridline-color: #e8ecf0;
            }
            QTableWidget::item { padding: 0px 4px; }
            QHeaderView::section {
                background-color: #f0f4f8;
                border: none;
                border-bottom: 1px solid #c8d4e0;
                padding: 2px 4px;
                font-size: 13px;
                font-weight: bold;
            }
        """)
        layout.addWidget(self._table, 1)

        self._row_count_lbl = QLabel('0 Eintraege')
        self._row_count_lbl.setFont(QFont('Arial', 9))
        self._row_count_lbl.setWordWrap(True)
        self._row_count_lbl.setStyleSheet('color: #888;')
        layout.addWidget(self._row_count_lbl)

    def _on_close_clicked(self):
        """Pane leeren und ausblenden (ausser Pane 0 - wird nur geleert)."""
        self.clear()
        if self._pane_index > 0:
            self.setVisible(False)

    # ---------- Laden ----------

    def _open_in_excel(self):
        """Geladene Dienstplan-Excel-Datei direkt in Excel öffnen."""
        if not self._excel_path or not os.path.isfile(self._excel_path):
            QMessageBox.warning(self, 'Nicht verfügbar', 'Datei nicht gefunden.')
            return
        try:
            os.startfile(self._excel_path)
        except Exception as exc:
            QMessageBox.critical(self, 'Fehler', f'Datei konnte nicht geöffnet werden:\n{exc}')

    def load(self, path: str) -> bool:
        """Excel-Datei einlesen und Tabelle befüllen. Gibt True bei Erfolg zurück."""
        self._status_lbl.setText(' Datei wird eingelesen ...')
        self._status_lbl.setStyleSheet('color: #555; padding: 2px 0px;')
        self._status_lbl.repaint()

        try:
            from functions.dienstplan_parser import DienstplanParser

            display_result = DienstplanParser(path, alle_anzeigen=True).parse()
            export_result  = DienstplanParser(path, alle_anzeigen=False).parse()

            if not display_result['success']:
                QMessageBox.critical(
                    self, 'Fehler beim Einlesen',
                    f"Die Datei konnte nicht geparst werden:\n\n{display_result['error']}"
                )
                self._status_lbl.setText('Fehler: Fehler beim Einlesen.')
                self._status_lbl.setStyleSheet('color: #bb0000; padding: 2px 0px;')
                return False

            self._excel_path   = path
            self._display_data = display_result
            self._parsed_data  = export_result
            self._excel_open_btn.setEnabled(True)

            # Dateiname im Header anzeigen
            dateiname = os.path.basename(path)
            self._title_lbl.setText(dateiname)

            datum = display_result.get('datum')
            if datum:
                self._datum_lbl.setText(f'Datum: {datum}')
                self._datum_lbl.setVisible(True)
            else:
                self._datum_lbl.setVisible(False)

            self._render_table_parsed(display_result)

            unbekannte = display_result.get('unbekannte_dienste', [])
            if unbekannte:
                QMessageBox.warning(
                    self, 'Unbekannte Dienst-Kürzel',
                    'Folgende Dienst-Kürzel wurden nicht erkannt:\n'
                    + '\n'.join(f'  - {d}' for d in sorted(unbekannte))
                    + '\n\nSie werden trotzdem angezeigt.'
                )

            self._status_lbl.setText(f'Geladen: {dateiname}')
            self._status_lbl.setStyleSheet('color: #107e3e; padding: 2px 0px;')
            return True

        except Exception as e:
            QMessageBox.critical(self, 'Fehler', f'Unerwarteter Fehler:\n{e}')
            self._status_lbl.setText(f'Fehler: {e}')
            self._status_lbl.setStyleSheet('color: #bb0000; padding: 2px 0px;')
            return False

    # ---------- Tabelle rendern ----------

    def _render_table_parsed(self, data: dict):
        """Tabelleninhalt aus geparsten Excel-Daten aufbauen."""
        tag_personen   = []
        nacht_personen = []
        sonst_personen = []

        # Namen die als Stationsleitung angezeigt werden (Kleinbuchstaben)
        STATIONSLEITUNG = {'lars peters'}

        for kat, liste in (('Dispo', data.get('dispo', [])),
                           ('Betreuer', data.get('betreuer', []))):
            for p in liste:
                name_lower = p.get('vollname', '').strip().lower()
                effekt_kat = 'Stationsleitung' if name_lower in STATIONSLEITUNG else kat
                dk = (p.get('dienst_kategorie') or '').upper()
                if dk in _TAG_DIENSTE:
                    tag_personen.append((effekt_kat, p))
                elif dk in _NACHT_DIENSTE:
                    nacht_personen.append((effekt_kat, p))
                else:
                    sonst_personen.append((effekt_kat, p))

        # ── Kranke Mitarbeiter nach Schichttyp + Dispo/Betreuer gruppieren ──────────
        krank_tag_dispo    = []
        krank_tag_betr     = []
        krank_nacht_dispo  = []
        krank_nacht_betr   = []
        krank_sonder       = []

        for p in data.get('kranke', []):
            stype = p.get('krank_schicht_typ') or 'sonderdienst'
            is_d  = p.get('krank_ist_dispo', False)
            if stype == 'tagdienst':
                if is_d:
                    krank_tag_dispo.append(p)
                else:
                    krank_tag_betr.append(p)
            elif stype == 'nachtdienst':
                if is_d:
                    krank_nacht_dispo.append(p)
                else:
                    krank_nacht_betr.append(p)
            else:
                krank_sonder.append(p)

        # Krank-Abschnitte: Dispo zuerst, dann Betreuer
        krank_tag_personen   = [('KrankDispo', p) for p in krank_tag_dispo]   + \
                               [('Krank',      p) for p in krank_tag_betr]
        krank_nacht_personen = [('KrankDispo', p) for p in krank_nacht_dispo] + \
                               [('Krank',      p) for p in krank_nacht_betr]
        krank_sonder_personen = [('Krank',     p) for p in krank_sonder]

        abschnitte = []
        if tag_personen:
            abschnitte.append(('Tagdienst',         '#1565a8', '#ffffff', tag_personen))
        if nacht_personen:
            abschnitte.append(('Nachtdienst',       '#0d2b4a', '#e8eeff', nacht_personen))
        if sonst_personen:
            abschnitte.append(('Sonstige',          '#555555', '#ffffff', sonst_personen))
        if krank_tag_personen:
            abschnitte.append(('Krank – Tagdienst',  '#8b0000', '#ffe8e8', krank_tag_personen))
        if krank_nacht_personen:
            abschnitte.append(('Krank – Nachtdienst','#5a0000', '#f5d0d0', krank_nacht_personen))
        if krank_sonder_personen:
            abschnitte.append(('Krank – Sonderdienst','#6b3300', '#fdebd0', krank_sonder_personen))

        total_rows = sum(1 + len(personen) for _, _, _, personen in abschnitte)
        self._table.clearSpans()
        self._table.setRowCount(total_rows)

        farben = {
            'Dispo':           QColor('#dce8f5'),
            'Betreuer':        QColor('#ffffff'),
            'Stationsleitung': QColor('#fff8e1'),
            'Krank':           QColor('#fce8e8'),
            'KrankDispo':      QColor('#f0d0d0'),   # Dispo-Krank: etwas dunkler
        }
        text_farben = {
            'Dispo':           QColor('#0a5ba4'),
            'Betreuer':        QColor('#1a1a1a'),
            'Stationsleitung': QColor('#7a5000'),
            'Krank':           QColor('#bb0000'),
            'KrankDispo':      QColor('#7a0000'),   # dunkler Rotton für Dispo-Krank
        }

        self._table_row_data = []
        row = 0
        sep_font = QFont('Arial', 11, QFont.Weight.Bold)
        for label_text, hdr_bg, hdr_fg, personen in abschnitte:
            self._table_row_data.append(None)
            self._table.setSpan(row, 0, 1, 5)
            sep_item = QTableWidgetItem(label_text)
            sep_item.setBackground(QColor(hdr_bg))
            sep_item.setForeground(QColor(hdr_fg))
            sep_item.setFont(sep_font)
            sep_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self._table.setItem(row, 0, sep_item)
            self._table.verticalHeader().resizeSection(row, 24)
            row += 1

            for kategorie, p in personen:
                self._table_row_data.append(p)
                bg = farben.get(kategorie, QColor('#ffffff'))
                fg = text_farben.get(kategorie, QColor('#1a1a1a'))

                # Spalte 0: Kategorie-Anzeige (für Krank: Dispo/Betreuer)
                if kategorie == 'KrankDispo':
                    kat_anzeige = 'Dispo'
                elif kategorie == 'Krank':
                    kat_anzeige = 'Betreuer'
                else:
                    kat_anzeige = kategorie

                # Spalte 2: Dienst-Kürzel – bei Kranken den abgeleiteten Wert zeigen
                if p.get('ist_krank'):
                    dienst_anzeige = p.get('krank_abgeleiteter_dienst') or ''
                else:
                    dienst_anzeige = p.get('dienst_kategorie') or ''

                vals = [
                    kat_anzeige,
                    p.get('anzeigename', ''),
                    dienst_anzeige,
                    p.get('start_zeit', '') or '',
                    p.get('end_zeit',   '') or '',
                ]
                for col, val in enumerate(vals):
                    item = QTableWidgetItem(val)
                    item.setBackground(bg)
                    item.setForeground(fg)
                    self._table.setItem(row, col, item)

                if p.get('ist_bulmorfahrer'):
                    for col in range(5):
                        self._table.item(row, col).setBackground(QColor('#fff3b0'))
                row += 1

        # ── Statuszeile ───────────────────────────────────────────────────
        tag_n   = len(tag_personen)
        nacht_n = len(nacht_personen)
        sonst_n = len(sonst_personen)

        # Tagdienst: Dispo / Betreuer trennen
        tag_dispo_n  = sum(1 for kat, _ in tag_personen   if kat == 'Dispo')
        tag_betr_n   = tag_n - tag_dispo_n

        # Nachtdienst: Dispo / Betreuer trennen
        nacht_dispo_n = sum(1 for kat, _ in nacht_personen if kat == 'Dispo')
        nacht_betr_n  = nacht_n - nacht_dispo_n

        # Krank-Zählung getrennt nach Dispo / Betreuer / Sonder
        n_kd_tag   = len(krank_tag_dispo)
        n_kb_tag   = len(krank_tag_betr)
        n_kd_nacht = len(krank_nacht_dispo)
        n_kb_nacht = len(krank_nacht_betr)
        n_k_sonder = len(krank_sonder)
        krank_gesamt = n_kd_tag + n_kb_tag + n_kd_nacht + n_kb_nacht + n_k_sonder

        teile = []
        if tag_n:
            tag_sub = []
            if tag_betr_n:  tag_sub.append(f'Betreuer {tag_betr_n}')
            if tag_dispo_n: tag_sub.append(f'Dispo {tag_dispo_n}')
            teile.append(f'{tag_n} Tagdienst ({", ".join(tag_sub)})' if tag_sub else f'{tag_n} Tagdienst')
        if nacht_n:
            nacht_sub = []
            if nacht_betr_n:  nacht_sub.append(f'Betreuer {nacht_betr_n}')
            if nacht_dispo_n: nacht_sub.append(f'Dispo {nacht_dispo_n}')
            teile.append(f'{nacht_n} Nachtdienst ({", ".join(nacht_sub)})' if nacht_sub else f'{nacht_n} Nachtdienst')
        if sonst_n: teile.append(f'{sonst_n} Sonstige')

        if krank_gesamt:
            # Betreuer-Block
            betr_teile = []
            if n_kb_tag:   betr_teile.append(f'{n_kb_tag} Tag')
            if n_kb_nacht: betr_teile.append(f'{n_kb_nacht} Nacht')
            if n_k_sonder: betr_teile.append(f'{n_k_sonder} Sonder')
            betr_gesamt = n_kb_tag + n_kb_nacht + n_k_sonder

            # Dispo-Block
            dispo_teile = []
            if n_kd_tag:   dispo_teile.append(f'{n_kd_tag} Tag')
            if n_kd_nacht: dispo_teile.append(f'{n_kd_nacht} Nacht')
            dispo_gesamt = n_kd_tag + n_kd_nacht

            krank_blocks = []
            if betr_gesamt:
                krank_blocks.append(
                    f'Betreuer {betr_gesamt} ({" / ".join(betr_teile)})'
                )
            if dispo_gesamt:
                krank_blocks.append(
                    f'Dispo {dispo_gesamt} ({" / ".join(dispo_teile)})'
                )
            teile.append(f'{krank_gesamt} Krank  –  {" | ".join(krank_blocks)}')

        self._row_count_lbl.setText('  |  '.join(teile) if teile else '0 Eintraege')
        self._row_count_lbl.setStyleSheet('color: #555; font-weight: bold; padding: 2px 0px;')

    # ---------- Doppelklick Tabellenzeile ----------

    def _on_table_double_click(self, item):
        """Ã–ffnet Edit-Dialog, speichert Ã„nderungen in Excel und lädt neu."""
        row = item.row()
        if row < 0 or row >= len(self._table_row_data):
            return
        person = self._table_row_data[row]
        if person is None:
            return
        if not self._display_data:
            return

        dlg = EditDienstDialog(person, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted or not dlg.result_data:
            return

        res = dlg.result_data
        try:
            self._save_to_excel(person, res['dienst'], res['von'], res['bis'])
        except Exception as e:
            QMessageBox.critical(
                self, 'Fehler beim Speichern',
                f'Die Excel-Datei konnte nicht gespeichert werden:\n{e}'
            )
            return

        excel_path = self._display_data.get('excel_path', '')
        if not excel_path:
            return
        try:
            from functions.dienstplan_parser import DienstplanParser
            display_result = DienstplanParser(excel_path, alle_anzeigen=True).parse()
            export_result  = DienstplanParser(excel_path, alle_anzeigen=False).parse()
            if display_result['success']:
                self._display_data = display_result
                self._parsed_data  = export_result
                datum = display_result.get('datum')
                if datum:
                    self._datum_lbl.setText(f'\U0001f4c5 Datum: {datum}')
                    self._datum_lbl.setVisible(True)
                self._render_table_parsed(display_result)
                self._status_lbl.setText(
                    f'\u2705 Gespeichert: {person.get("anzeigename", "")} -> {res["dienst"]}'
                    f'  \u26a0\ufe0f Andere Nutzer m\u00fcssen die Datei neu \u00f6ffnen, '
                    f'sonst werden die \u00c4nderungen \u00fcberschrieben!'
                )
                self._status_lbl.setStyleSheet('color: #b85c00; font-weight: bold; padding: 2px 0px;')
        except Exception as e:
            QMessageBox.critical(self, 'Fehler', f'Fehler beim Neu-Laden:\n{e}')

    # ---------- Excel-Schutz ----------

    @staticmethod
    def _check_excel_locked(excel_path: str):
        """Prüft ob die Excel-Datei von einem anderen Programm geöffnet ist."""
        import os
        ordner     = os.path.dirname(excel_path)
        dateiname  = os.path.basename(excel_path)
        lock_datei = os.path.join(ordner, '~$' + dateiname)
        if os.path.exists(lock_datei):
            raise IOError(
                f'Die Excel-Datei ist zurzeit auf einem anderen PC oder in Excel '
                f'geöffnet:\n\n{dateiname}\n\n'
                f'Bitte die Datei dort schließen und dann erneut speichern.'
            )
        try:
            fh = open(excel_path, 'r+b')
            fh.close()
        except PermissionError:
            raise IOError(
                f'Die Excel-Datei ist gesperrt (kein Schreibzugriff):\n\n{dateiname}\n\n'
                f'Bitte sicherstellen, dass die Datei nicht anderweitig geöffnet ist.'
            )

    def _save_to_excel(self, person: dict, dienst: str, von: str, bis: str):
        """Schreibt Dienst/Von/Bis sicher in die Excel-Datei zurück (3-Ebenen-Schutz)."""
        import os
        import openpyxl
        from openpyxl.styles import PatternFill

        excel_path = (self._display_data or {}).get('excel_path', '')
        column_map = (self._display_data or {}).get('column_map', {})
        excel_row  = person.get('excel_row')
        if not excel_path or not column_map or not excel_row:
            raise ValueError('Excel-Pfad oder Zeilennummer nicht gefunden.')

        ordner    = os.path.dirname(excel_path)
        dateiname = os.path.basename(excel_path)
        temp_path = os.path.join(ordner, dateiname + '.nesk3tmp')
        nesk_lock = os.path.join(ordner, dateiname + '.nesk3lock')

        self._check_excel_locked(excel_path)

        if os.path.exists(nesk_lock):
            import time
            alter = time.time() - os.path.getmtime(nesk_lock)
            if alter < 30:
                raise IOError(
                    'Eine andere Nesk3-Instanz speichert gerade dieselbe Datei.\n'
                    'Bitte kurz warten und erneut versuchen.'
                )
            os.remove(nesk_lock)

        try:
            with open(nesk_lock, 'w') as lf:
                lf.write('nesk3')

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            dienst_cell = ws.cell(row=excel_row, column=column_map['dienst'] + 1)
            dienst_cell.value = dienst or None

            if dienst.upper() in ('KRANK', 'K'):
                dienst_cell.fill = PatternFill(patternType='solid', fgColor='FFFF0000')
            else:
                dienst_cell.fill = PatternFill(patternType='none')

            if column_map.get('beginn') is not None:
                ws.cell(row=excel_row, column=column_map['beginn'] + 1).value = \
                    self._parse_time_str(von)
            if column_map.get('ende') is not None:
                ws.cell(row=excel_row, column=column_map['ende'] + 1).value = \
                    self._parse_time_str(bis)

            wb.save(temp_path)
            wb.close()
            self._check_excel_locked(excel_path)
            os.replace(temp_path, excel_path)
            self._backup_excel_save(excel_path)

        finally:
            for f in (nesk_lock, temp_path):
                try:
                    if os.path.exists(f):
                        os.remove(f)
                except OSError:
                    pass

    @staticmethod
    def _backup_excel_save(excel_path: str):
        """Erstellt nach jedem erfolgreichen Excel-Save eine Sicherungskopie."""
        import os, glob, shutil
        from datetime import datetime
        from config import BASE_DIR
        try:
            backup_dir = os.path.join(BASE_DIR, 'Backup Data', 'excel_saves')
            os.makedirs(backup_dir, exist_ok=True)
            name  = os.path.splitext(os.path.basename(excel_path))[0]
            datum = datetime.now().strftime('%Y%m%d_%H%M%S')
            dst   = os.path.join(backup_dir, f'{name}_{datum}.xlsx')
            shutil.copy2(excel_path, dst)
            alle = sorted(glob.glob(os.path.join(backup_dir, f'{name}_????????_??????.xlsx')))
            for alt in alle[:-20]:
                try:
                    os.remove(alt)
                except OSError:
                    pass
        except Exception:
            pass

    @staticmethod
    def _parse_time_str(s: str):
        """Wandelt 'HH:MM'-String in datetime.time um, oder gibt None zurück."""
        from datetime import time as dtime
        if not s or not s.strip():
            return None
        try:
            h, m = s.strip().split(':')
            return dtime(int(h), int(m))
        except Exception:
            return None


class PaxMonatsübersichtDialog(QDialog):
    """
    Zeigt vor dem Word-Export die Monatsübersicht der PAX- und Einsatz-Einträge.
    Jahr und Monat sind frei wählbar. Durch Anklicken einer Zeile kann die
    PAX-Zahl des gewählten Tages für den Export übernommen werden.
    """

    _MONATE = ["Januar", "Februar", "März", "April", "Mai", "Juni",
               "Juli", "August", "September", "Oktober", "November", "Dezember"]

    def __init__(self, von_datum, pax_heute: int, einsaetze_heute: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Monatsübersicht vor dem Export")
        self.setMinimumSize(720, 580)
        self._von       = von_datum
        self._pax_heute = pax_heute
        self._ein_heute = einsaetze_heute
        # Rückgabewerte für den Export; None = Original-Wert behalten
        self.selected_pax  : int | None = None
        self.selected_ein  : int | None = None
        self.selected_datum: str | None = None
        self._build_ui()

    # ------------------------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── Warnung ──────────────────────────────────────────────────────────
        warn = QLabel(
            "⚠️  Bitte prüfen: Für eine korrekte Jahressumme sollte jeder Betriebstag "
            "eine PAX-Zahl und eine SL-Einsatzzahl haben.\n"
            "Fehlende Einträge (Wert = 0) sind rot markiert. "
            "Du kannst den Export trotzdem fortsetzen."
        )
        warn.setWordWrap(True)
        warn.setStyleSheet(
            "background:#fff3cd; border:1px solid #ffc107; border-radius:5px; "
            "padding:8px 12px; color:#856404; font-size:12px;"
        )
        layout.addWidget(warn)

        # ── Jahr + Monat Navigation ───────────────────────────────────────────
        nav_row = QHBoxLayout()
        nav_row.addWidget(QLabel("<b>Jahr:</b>"))
        self._jahr_spin = QSpinBox()
        self._jahr_spin.setRange(2020, 2099)
        self._jahr_spin.setValue(self._von.year)
        self._jahr_spin.setFixedWidth(75)
        nav_row.addWidget(self._jahr_spin)

        nav_row.addSpacing(14)
        nav_row.addWidget(QLabel("<b>Monat:</b>"))
        self._monat_combo = QComboBox()
        for _m in self._MONATE:
            self._monat_combo.addItem(_m)
        self._monat_combo.setCurrentIndex(self._von.month - 1)
        self._monat_combo.setFixedWidth(130)
        nav_row.addWidget(self._monat_combo)
        nav_row.addStretch()

        self._monat_lbl = QLabel("")
        self._monat_lbl.setStyleSheet("font-weight:bold; font-size:13px; color:#103060;")
        nav_row.addWidget(self._monat_lbl)
        layout.addLayout(nav_row)

        # ── Tabelle ───────────────────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setColumnCount(4)
        self._table.setHorizontalHeaderLabels(["Datum", "Wochentag", "PAX-Zahl", "SL-Einsätze"])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setStyleSheet(
            "QTableWidget { border:1px solid #dce8f5; border-radius:4px; }"
            "QHeaderView::section { background:#e8f0fb; font-weight:bold; padding:4px; }"
            "QTableWidget::item:selected { background:#1565a8; color:white; font-weight:bold; }"
        )
        layout.addWidget(self._table)

        # ── Auswahl-Info ─────────────────────────────────────────────────────
        self._auswahl_lbl = QLabel(
            "🔵  Zeile anklicken um die PAX-Zahl dieses Tages für den Export zu übernehmen."
        )
        self._auswahl_lbl.setWordWrap(True)
        self._auswahl_lbl.setStyleSheet(
            "background:#e8f0fb; border:1px solid #b0c8f0; border-radius:4px; "
            "padding:7px 10px; color:#1a4a8a; font-size:12px;"
        )
        layout.addWidget(self._auswahl_lbl)

        # ── Fehlende-Einträge-Info ────────────────────────────────────────────
        self._fehler_lbl = QLabel("")
        self._fehler_lbl.setWordWrap(True)
        self._fehler_lbl.setStyleSheet("color:#b00000; font-size:11px; padding:2px 0px;")
        layout.addWidget(self._fehler_lbl)

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()

        verwaltung_btn = QPushButton("PAX && Einsätze jetzt nachtragen ...")
        verwaltung_btn.setMinimumHeight(32)
        verwaltung_btn.setStyleSheet(
            "background:#5a3e8c;color:white;border:none;border-radius:4px;padding:0 12px;"
        )
        verwaltung_btn.clicked.connect(self._verwaltung_oeffnen)
        btn_row.addWidget(verwaltung_btn)

        btn_row.addStretch()

        self._export_btn = QPushButton("Trotzdem exportieren")
        self._export_btn.setMinimumHeight(32)
        self._export_btn.setStyleSheet(
            f"background:{FIORI_BLUE};color:white;border:none;border-radius:4px;padding:0 16px;"
        )
        self._export_btn.clicked.connect(self._do_export)
        btn_row.addWidget(self._export_btn)

        abbruch_btn = QPushButton("Abbrechen")
        abbruch_btn.setMinimumHeight(32)
        abbruch_btn.clicked.connect(self.reject)
        btn_row.addWidget(abbruch_btn)

        layout.addLayout(btn_row)

        # Signale verbinden und Tabelle einmalig befüllen
        self._jahr_spin.valueChanged.connect(self._aktualisiere_tabelle)
        self._monat_combo.currentIndexChanged.connect(self._aktualisiere_tabelle)
        self._table.itemSelectionChanged.connect(self._on_auswahl)
        self._aktualisiere_tabelle()

    # ------------------------------------------------------------------
    def _aktualisiere_tabelle(self):
        import calendar
        from datetime import date
        from database.pax_db import lade_alle_eintraege

        jahr      = self._jahr_spin.value()
        monat     = self._monat_combo.currentIndex() + 1
        heute_str = self._von.strftime("%Y-%m-%d")

        self._monat_lbl.setText(
            f"Monat: {self._MONATE[monat - 1]} {jahr}  "
            f"(Heute eingetragen:  PAX = {self._pax_heute:,}   |   SL-Einsätze = {self._ein_heute})"
        )

        # Alle DB-Einträge des gewählten Jahres laden
        try:
            alle = {e["datum"]: e for e in lade_alle_eintraege(jahr)}
        except Exception:
            alle = {}

        heute_date = date.today()

        # Zukunftsdaten erkennen
        zukunfts_eintraege = sorted(
            d_str for d_str in alle
            if date.fromisoformat(d_str) > heute_date
        )

        # Heutigen Eintrag mit frisch eingegebenen Werten überschreiben
        if self._pax_heute > 0 or self._ein_heute > 0:
            alle[heute_str] = {
                "datum": heute_str,
                "pax_zahl": self._pax_heute,
                "einsaetze_zahl": self._ein_heute,
            }

        wochentage   = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
        tage_im_monat = calendar.monthrange(jahr, monat)[1]
        fehlende     = 0
        scroll_to_row = -1

        self._table.blockSignals(True)
        self._table.setRowCount(0)

        for tag in range(1, tage_im_monat + 1):
            d = date(jahr, monat, tag)
            if d > heute_date:
                continue  # Zukunftstage nicht anzeigen
            d_str   = d.strftime("%Y-%m-%d")
            eintrag = alle.get(d_str)
            pax     = eintrag["pax_zahl"]       if eintrag else 0
            ein     = eintrag["einsaetze_zahl"] if eintrag else 0
            fehlt   = (pax == 0) or (ein == 0)
            if fehlt:
                fehlende += 1

            row = self._table.rowCount()
            self._table.insertRow(row)

            datum_item = QTableWidgetItem(d_str)
            wt_item    = QTableWidgetItem(wochentage[d.weekday()])
            pax_item   = QTableWidgetItem(str(pax) if pax else "—")
            ein_item   = QTableWidgetItem(str(ein) if ein else "—")

            for item in (datum_item, wt_item, pax_item, ein_item):
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if fehlt:
                    item.setBackground(QColor("#ffe0e0"))
                    item.setForeground(QColor("#b00000"))
                if d_str == heute_str:
                    item.setBackground(QColor("#d4edda"))
                    item.setForeground(QColor("#155724"))
                    scroll_to_row = row

            self._table.setItem(row, 0, datum_item)
            self._table.setItem(row, 1, wt_item)
            self._table.setItem(row, 2, pax_item)
            self._table.setItem(row, 3, ein_item)

        self._table.blockSignals(False)

        # Zur heute-Zeile scrollen und vorauswählen
        if scroll_to_row >= 0:
            self._table.scrollToItem(self._table.item(scroll_to_row, 0))
            self._table.selectRow(scroll_to_row)

        # Fehlende-Info aktualisieren
        if fehlende:
            self._fehler_lbl.setText(
                f"❌  {fehlende} Tag(e) haben fehlende oder unvollständige Einträge (rot markiert). "
                "Bitte im Verwaltungs-Dialog nachtragen oder jetzt trotzdem exportieren."
            )
        else:
            self._fehler_lbl.setText("")

        if zukunfts_eintraege:
            QMessageBox.warning(
                self,
                "Zukunftsdaten gefunden",
                f"⚠️  ACHTUNG: Es gibt {len(zukunfts_eintraege)} Eintrag/Einträge "
                f"mit Datum in der Zukunft:\n  "
                + ", ".join(zukunfts_eintraege)
                + "\n\nBitte im Verwaltungs-Dialog korrigieren oder löschen!"
            )

    # ------------------------------------------------------------------
    def _on_auswahl(self):
        """Aktualisiert Auswahl-Info wenn eine Zeile angeklickt wird."""
        row = self._table.currentRow()
        if row < 0 or not self._table.selectedItems():
            self._auswahl_lbl.setText(
                "🔵  Zeile anklicken um die PAX-Zahl dieses Tages für den Export zu übernehmen."
            )
            self._auswahl_lbl.setStyleSheet(
                "background:#e8f0fb; border:1px solid #b0c8f0; border-radius:4px; "
                "padding:7px 10px; color:#1a4a8a; font-size:12px;"
            )
            self._export_btn.setText("Trotzdem exportieren")
            return

        datum_item = self._table.item(row, 0)
        pax_item   = self._table.item(row, 2)
        ein_item   = self._table.item(row, 3)

        datum   = datum_item.text() if datum_item else "?"
        pax_txt = pax_item.text()   if pax_item   else "—"
        ein_txt = ein_item.text()   if ein_item   else "—"

        pax_val = int(pax_txt) if pax_txt not in ("—", "-", "") else 0
        ein_val = int(ein_txt) if ein_txt not in ("—", "-", "") else 0

        self._auswahl_lbl.setText(
            f"✅  Ausgewählt: {datum}  │  PAX: {pax_val:,}  │  SL-Einsätze: {ein_val}"
            "  —  Diese Werte werden für den Export verwendet."
        )
        self._auswahl_lbl.setStyleSheet(
            "background:#d4edda; border:2px solid #28a745; border-radius:4px; "
            "padding:7px 10px; color:#155724; font-size:12px; font-weight:bold;"
        )
        self._export_btn.setText(f"Exportieren  (PAX: {pax_val:,} | Einsätze: {ein_val})")

    # ------------------------------------------------------------------
    def _do_export(self):
        """Übernimmt die ausgewählten Werte und bestätigt den Dialog."""
        row = self._table.currentRow()
        if row >= 0 and self._table.selectedItems():
            datum_item = self._table.item(row, 0)
            pax_item   = self._table.item(row, 2)
            ein_item   = self._table.item(row, 3)
            self.selected_datum = datum_item.text() if datum_item else None
            pax_txt = pax_item.text() if pax_item else "—"
            ein_txt = ein_item.text() if ein_item else "—"
            self.selected_pax = int(pax_txt) if pax_txt not in ("—", "-", "") else 0
            self.selected_ein = int(ein_txt) if ein_txt not in ("—", "-", "") else 0
        self.accept()

    def _verwaltung_oeffnen(self):
        dlg = PaxEinsatzVerwaltungDialog(parent=self)
        dlg.exec()
        self.reject()


class PaxEinsatzVerwaltungDialog(QDialog):
    """Dialog zum Anzeigen und Bearbeiten der täglichen PAX- und Einsatzzahlen."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("PAX & SL-Einsätze verwalten")
        self.setMinimumSize(680, 500)
        self._build_ui()
        self._lade_daten()

    # ------------------------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── Jahr + Monat Auswahl ────────────────────────────────────────
        top = QHBoxLayout()
        top.addWidget(QLabel("Jahr:"))
        self._jahr_spin = QSpinBox()
        self._jahr_spin.setRange(2020, 2099)
        self._jahr_spin.setValue(datetime.now().year)
        self._jahr_spin.setFixedWidth(80)
        self._jahr_spin.valueChanged.connect(self._lade_daten)
        top.addWidget(self._jahr_spin)

        top.addSpacing(14)
        top.addWidget(QLabel("Monat:"))
        self._monat_combo = QComboBox()
        for _m in ["Januar", "Februar", "März", "April", "Mai", "Juni",
                   "Juli", "August", "September", "Oktober", "November", "Dezember"]:
            self._monat_combo.addItem(_m)
        self._monat_combo.setCurrentIndex(datetime.now().month - 1)
        self._monat_combo.setFixedWidth(120)
        self._monat_combo.currentIndexChanged.connect(self._lade_daten)
        top.addWidget(self._monat_combo)
        top.addStretch()

        self._summen_lbl = QLabel("")
        self._summen_lbl.setStyleSheet("font-weight: bold; color: #0a5ba4;")
        top.addWidget(self._summen_lbl)
        layout.addLayout(top)

        # ── Hinweis ─────────────────────────────────────────────────────
        hint = QLabel(
            "Alle Tage des gewählten Monats werden angezeigt. "
            "Rot = kein Eintrag vorhanden. "
            "Doppelklick zum Bearbeiten."
        )
        hint.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(hint)

        # ── Tabelle ─────────────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setColumnCount(3)
        self._table.setHorizontalHeaderLabels(["Datum", "PAX-Zahl", "SL-Einsätze"])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked |
                                    QTableWidget.EditTrigger.SelectedClicked)
        self._table.itemChanged.connect(self._on_item_changed)
        self._table.setStyleSheet(
            "QTableWidget { border: 1px solid #dce8f5; border-radius: 4px; }"
            "QHeaderView::section { background: #e8f0fb; font-weight: bold; padding: 4px; }"
        )
        layout.addWidget(self._table)

        # ── Buttons ─────────────────────────────────────────────────────
        btn_row = QHBoxLayout()

        add_btn = QPushButton("+ Eintrag hinzufügen")
        add_btn.setMinimumHeight(32)
        add_btn.setStyleSheet(
            f"background:{FIORI_BLUE};color:white;border:none;border-radius:4px;padding:0 12px;"
        )
        add_btn.clicked.connect(self._eintrag_hinzufuegen)
        btn_row.addWidget(add_btn)

        del_btn = QPushButton("Zeile löschen")
        del_btn.setMinimumHeight(32)
        del_btn.setStyleSheet(
            f"background:{FIORI_ERROR};color:white;border:none;border-radius:4px;padding:0 12px;"
        )
        del_btn.clicked.connect(self._zeile_loeschen)
        btn_row.addWidget(del_btn)

        btn_row.addStretch()

        save_btn = QPushButton("Speichern")
        save_btn.setMinimumHeight(32)
        save_btn.setStyleSheet(
            "background:#107e3e;color:white;border:none;border-radius:4px;padding:0 16px;"
        )
        save_btn.clicked.connect(self._speichern)
        btn_row.addWidget(save_btn)

        close_btn = QPushButton("Schließen")
        close_btn.setMinimumHeight(32)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)

        layout.addLayout(btn_row)

        self._aenderungen: set[int] = set()  # geänderte Zeilen-Indices

    # ------------------------------------------------------------------
    def _lade_daten(self):
        import calendar
        from database.pax_db import lade_alle_eintraege
        from datetime import date as _date
        self._table.blockSignals(True)
        self._table.setRowCount(0)
        self._aenderungen.clear()

        jahr  = self._jahr_spin.value()
        monat = self._monat_combo.currentIndex() + 1
        heute = _date.today()

        # Alle DB-Einträge des Jahres als dict laden
        try:
            alle_eintraege = {e["datum"]: e for e in lade_alle_eintraege(jahr)}
        except Exception:
            alle_eintraege = {}

        zukunft_gefunden = []
        tage_im_monat = calendar.monthrange(jahr, monat)[1]

        for tag in range(1, tage_im_monat + 1):
            d     = _date(jahr, monat, tag)
            d_str = d.strftime("%Y-%m-%d")
            eintrag = alle_eintraege.get(d_str)

            row = self._table.rowCount()
            self._table.insertRow(row)

            if eintrag:
                item_datum = QTableWidgetItem(d_str)
                item_pax   = QTableWidgetItem(str(eintrag["pax_zahl"]))
                item_ein   = QTableWidgetItem(str(eintrag["einsaetze_zahl"]))
                # Zukunftsdaten orange markieren
                if d > heute:
                    zukunft_gefunden.append(d_str)
                    for item in (item_datum, item_pax, item_ein):
                        item.setBackground(QColor("#fff3cd"))
                        item.setForeground(QColor("#856404"))
                        item.setToolTip("⚠ Zukunftsdatum – bitte löschen!")
            else:
                # Fehlender Eintrag: rot markieren, aber editierbar
                item_datum = QTableWidgetItem(d_str)
                item_pax   = QTableWidgetItem("0")
                item_ein   = QTableWidgetItem("0")
                if d <= heute:
                    for item in (item_datum, item_pax, item_ein):
                        item.setBackground(QColor("#ffe0e0"))
                        item.setForeground(QColor("#b00000"))
                        item.setToolTip("Kein Eintrag – Doppelklick zum Nachtragen")
                else:
                    # Zukunftstag ohne Eintrag: grau anzeigen, nicht editierbar
                    for item in (item_datum, item_pax, item_ein):
                        item.setBackground(QColor("#f5f5f5"))
                        item.setForeground(QColor("#aaaaaa"))
                        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            self._table.setItem(row, 0, item_datum)
            self._table.setItem(row, 1, item_pax)
            self._table.setItem(row, 2, item_ein)

        self._table.blockSignals(False)
        self._update_summen()

        # Zur heutigen Zeile scrollen (wenn im aktuellen Monat)
        heute_str = heute.strftime("%Y-%m-%d")
        for r in range(self._table.rowCount()):
            if self._table.item(r, 0) and self._table.item(r, 0).text() == heute_str:
                self._table.scrollToItem(self._table.item(r, 0))
                break

        if zukunft_gefunden:
            QMessageBox.warning(
                self,
                "Zukunftsdaten gefunden",
                f"⚠️  Achtung: Es gibt {len(zukunft_gefunden)} Eintrag/Einträge "
                f"mit Datum in der Zukunft:\n\n  "
                + "\n  ".join(zukunft_gefunden)
                + "\n\nBitte diese Zeilen löschen! Orange markierte Zeilen anwählen "
                "und die Schaltfläche 'Zeile löschen' nutzen."
            )

    def _update_summen(self):
        total_pax = 0
        total_ein = 0
        for r in range(self._table.rowCount()):
            try:
                total_pax += int(self._table.item(r, 1).text())
            except Exception:
                pass
            try:
                total_ein += int(self._table.item(r, 2).text())
            except Exception:
                pass
        self._summen_lbl.setText(
            f"Gesamt PAX: {total_pax:,}   |   Gesamt SL-Einsätze: {total_ein}"
        )

    def _on_item_changed(self, item: QTableWidgetItem):
        self._aenderungen.add(item.row())
        self._update_summen()

    # ------------------------------------------------------------------
    def _eintrag_hinzufuegen(self):
        row = self._table.rowCount()
        self._table.insertRow(row)
        heute = datetime.now().strftime("%Y-%m-%d")
        self._table.blockSignals(True)
        self._table.setItem(row, 0, QTableWidgetItem(heute))
        self._table.setItem(row, 1, QTableWidgetItem("0"))
        self._table.setItem(row, 2, QTableWidgetItem("0"))
        self._table.blockSignals(False)
        self._aenderungen.add(row)
        self._table.scrollToBottom()
        self._table.setCurrentCell(row, 0)
        self._table.editItem(self._table.item(row, 0))

    def _zeile_loeschen(self):
        selected = self._table.selectedItems()
        if not selected:
            QMessageBox.information(self, "Hinweis", "Bitte erst eine Zeile auswählen.")
            return
        row = self._table.currentRow()
        datum_item = self._table.item(row, 0)
        datum = datum_item.text().strip() if datum_item else ""
        antwort = QMessageBox.question(
            self, "Eintrag löschen",
            f"Eintrag für {datum or 'diese Zeile'} wirklich löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort != QMessageBox.StandardButton.Yes:
            return
        if datum:
            try:
                from database.pax_db import loesche_eintrag
                loesche_eintrag(datum)
            except Exception as exc:
                QMessageBox.warning(self, "Fehler", f"Löschen fehlgeschlagen:\n{exc}")
                return
        self._table.removeRow(row)
        self._aenderungen.discard(row)
        self._update_summen()

    # ------------------------------------------------------------------
    def _speichern(self):
        from database.pax_db import speichere_tages_pax, speichere_tages_einsaetze
        fehler = []
        gespeichert = 0

        rows_to_save = sorted(self._aenderungen)
        if not rows_to_save:
            QMessageBox.information(self, "Hinweis", "Keine Änderungen vorhanden.")
            return

        for row in rows_to_save:
            if row >= self._table.rowCount():
                continue
            datum_item = self._table.item(row, 0)
            pax_item   = self._table.item(row, 1)
            ein_item   = self._table.item(row, 2)

            datum = datum_item.text().strip() if datum_item else ""
            if not datum:
                fehler.append(f"Zeile {row+1}: Kein Datum eingetragen.")
                continue

            # Datumsformat prüfen
            try:
                datetime.strptime(datum, "%Y-%m-%d")
            except ValueError:
                fehler.append(f"Zeile {row+1}: Datum '{datum}' nicht im Format JJJJ-MM-TT.")
                continue

            try:
                pax = int(pax_item.text()) if pax_item else 0
            except ValueError:
                fehler.append(f"Zeile {row+1}: Ungültige PAX-Zahl.")
                continue
            try:
                ein = int(ein_item.text()) if ein_item else 0
            except ValueError:
                fehler.append(f"Zeile {row+1}: Ungültige Einsatz-Zahl.")
                continue

            try:
                if pax > 0:
                    speichere_tages_pax(datum, pax)
                if ein > 0:
                    speichere_tages_einsaetze(datum, ein)
                gespeichert += 1
            except Exception as exc:
                fehler.append(f"Zeile {row+1}: DB-Fehler: {exc}")

        if fehler:
            QMessageBox.warning(self, "Fehler beim Speichern",
                                "Folgende Fehler sind aufgetreten:\n\n" + "\n".join(fehler))
        else:
            self._aenderungen.clear()
            QMessageBox.information(self, "Gespeichert",
                                    f"{gespeichert} Eintrag/Einträge erfolgreich gespeichert.")
        self._lade_daten()


class DienstplanWidget(QWidget):
    MAX_PANES = 4   # bis zu 4 Dienstplaene nebeneinander

    def __init__(self, parent=None):
        super().__init__(parent)
        self._alle: list[Dienstplan]            = []
        self._fs_model: QFileSystemModel | None = None
        self._export_pane_idx: int              = 0   # Index der fuer Export aktiven Pane
        self._html_generiert:  bool             = False  # True nach erstem HTML-Export
        self._html_watcher = QFileSystemWatcher(parent=self)
        self._html_watcher.fileChanged.connect(self._on_excel_geaendert)
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(24, 24, 24, 24)
        outer.setSpacing(8)

        # Titelzeile
        top = QHBoxLayout()
        title = QLabel("Dienstplan")
        title.setFont(QFont("Arial", 22, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {FIORI_TEXT};")
        top.addWidget(title)
        top.addStretch()

        self._export_lbl = QLabel("")
        self._export_lbl.setFont(QFont("Arial", 10))
        self._export_lbl.setStyleSheet("color: #888;")
        top.addWidget(self._export_lbl)
        top.addSpacing(8)

        word_btn = QPushButton("Word exportieren")
        word_btn.setMinimumHeight(36)
        word_btn.setToolTip("Stärkemeldung als Word-Dokument exportieren und öffnen")
        word_btn.setStyleSheet(
            f"background-color: {FIORI_BLUE}; color: white; "
            f"border-radius: 4px; padding: 0 12px;"
        )
        word_btn.clicked.connect(self._word_exportieren)
        top.addWidget(word_btn)

        pax_btn = QPushButton("PAX & Einsätze verwalten")
        pax_btn.setMinimumHeight(36)
        pax_btn.setToolTip("PAX-Zahlen und SL-Einsätze pro Tag einsehen und korrigieren")
        pax_btn.setStyleSheet(
            "background-color: #5a3e8c; color: white; "
            "border-radius: 4px; padding: 0 12px;"
        )
        pax_btn.clicked.connect(self._pax_verwalten)
        top.addWidget(pax_btn)

        html_btn = QPushButton("🌐  Als Webseite anzeigen")
        html_btn.setMinimumHeight(36)
        html_btn.setToolTip(
            "Dienstplan als HTML-Seite generieren und im Browser öffnen.\n"
            "Die Seite wird automatisch neu erzeugt, wenn sich die Excel-Datei ändert."
        )
        html_btn.setStyleSheet(
            "background-color: #1e7e34; color: white; "
            "border-radius: 4px; padding: 0 12px;"
        )
        html_btn.clicked.connect(self._html_exportieren)
        top.addWidget(html_btn)

        reload_btn = QPushButton("Neu laden")
        reload_btn.setToolTip("Ordner-Ansicht neu laden")
        reload_btn.setMinimumHeight(36)
        reload_btn.clicked.connect(self.reload_tree)
        top.addWidget(reload_btn)

        outer.addLayout(top)

        # Info-Banner: Mehrere Dienstpläne gleichzeitig
        info_banner = QLabel(
            "ℹ️   Bis zu 4 Dienstpläne gleichzeitig öffnen: "
            "Klicken Sie im Dateibaum links auf eine Excel-Datei – "
            "jede Datei erscheint als eigene Spalte nebeneinander. "
            "Um eine Datei für den Word-Export auszuwählen, klicken Sie in der "
            "jeweiligen Spalte auf die Schaltfläche \"Hier klicken um Datei als Wordexport auszuwählen\"."
        )
        info_banner.setWordWrap(True)
        info_banner.setStyleSheet(
            "background-color: #e8f0fb; border: 1px solid #b0c8f0; "
            "border-radius: 6px; padding: 8px 14px; color: #1a4a8a; font-size: 12px;"
        )
        outer.addWidget(info_banner)

        # Splitter: Dateibaum links | Panes rechts
        outer_splitter = QSplitter(Qt.Orientation.Horizontal)
        outer_splitter.setHandleWidth(4)

        # Linke Seite: Dateibaum
        tree_panel = QWidget()
        tree_panel.setMinimumWidth(180)
        tree_panel.setMaximumWidth(340)
        tree_layout = QVBoxLayout(tree_panel)
        tree_layout.setContentsMargins(0, 0, 6, 0)
        tree_layout.setSpacing(4)

        tree_header = QLabel("Dienstplaene")
        tree_header.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        tree_header.setStyleSheet(f"color: {FIORI_TEXT}; padding: 4px 0px;")
        tree_layout.addWidget(tree_header)

        self._manuell_btn = QPushButton("📂  Datei öffnen ...")
        self._manuell_btn.setFixedHeight(28)
        self._manuell_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._manuell_btn.setToolTip(
            "Excel-Dienstplan manuell auswählen (ohne Ordner-Konfiguration)"
        )
        self._manuell_btn.setStyleSheet(
            f"QPushButton {{font-size:10px; padding: 2px 8px; border-radius:4px;"
            f"background:{FIORI_BLUE}; color:white; border:none;}}"
            f"QPushButton:hover {{background:#0855a9;}}"
        )
        self._manuell_btn.clicked.connect(self._manuell_datei_oeffnen)
        tree_layout.addWidget(self._manuell_btn)

        self._tree = QTreeView()
        self._tree.setStyleSheet("""
            QTreeView {
                background-color: white;
                border: 1px solid #dce8f5;
                border-radius: 6px;
                font-size: 12px;
            }
            QTreeView::item { padding: 3px 2px; }
            QTreeView::item:selected { background-color: #dce8f5; color: #0a5ba4; }
            QTreeView::item:hover    { background-color: #f0f4f8; }
        """)
        self._tree.setAnimated(True)
        self._tree.setSortingEnabled(True)
        self._tree.activated.connect(self._on_tree_activated)
        tree_layout.addWidget(self._tree, 1)

        self._ordner_lbl = QLabel("")
        self._ordner_lbl.setWordWrap(True)
        self._ordner_lbl.setStyleSheet("color: #aaa; font-size: 9px; padding: 2px;")
        tree_layout.addWidget(self._ordner_lbl)

        outer_splitter.addWidget(tree_panel)

        # Rechte Seite: horizontaler Splitter mit bis zu MAX_PANES Panes
        self._pane_splitter = QSplitter(Qt.Orientation.Horizontal)
        self._pane_splitter.setHandleWidth(6)

        self._panes: list[_DienstplanPane] = []
        for i in range(self.MAX_PANES):
            pane = _DienstplanPane(pane_index=i)
            pane.export_selected.connect(self._set_export_pane)
            self._panes.append(pane)
            self._pane_splitter.addWidget(pane)

        # Beim Start nur 1 Pane sichtbar
        for i in range(1, self.MAX_PANES):
            self._panes[i].setVisible(False)

        outer_splitter.addWidget(self._pane_splitter)
        outer_splitter.setSizes([240, 900])
        outer.addWidget(outer_splitter, 1)

        # Export-Pane initial markieren
        self._set_export_pane(0)
        self._setup_tree()

    # ------------------------------------------------------------------
    # Dateibaum
    # ------------------------------------------------------------------

    def _setup_tree(self):
        """Dateibaum fuer den konfigurierten Ordner aufbauen."""
        try:
            from functions.settings_functions import get_setting
            ordner = get_setting('dienstplan_ordner')
        except Exception:
            ordner = ''

        if not ordner or not os.path.isdir(ordner):
            self._ordner_lbl.setText(
                "Kein gueltiger Ordner konfiguriert.\n"
                "Bitte unter Einstellungen einen Ordner festlegen."
            )
            self._ordner_lbl.setStyleSheet("color: #bb6600; font-size: 10px; padding: 4px;")
            return

        self._fs_model = QFileSystemModel(self)
        self._fs_model.setNameFilters(['*.xlsx', '*.xls'])
        self._fs_model.setNameFilterDisables(False)
        root_idx = self._fs_model.setRootPath(ordner)

        self._tree.setModel(self._fs_model)
        self._tree.setRootIndex(root_idx)

        for col in range(1, 4):
            self._tree.hideColumn(col)
        self._tree.header().setVisible(False)

        self._ordner_lbl.setText(ordner)
        self._ordner_lbl.setStyleSheet("color: #aaa; font-size: 9px; padding: 2px;")

    def reload_tree(self):
        """Ordner-Konfiguration neu lesen und Baum neu aufbauen."""
        if self._fs_model is not None:
            self._tree.setModel(None)
            self._fs_model.deleteLater()
            self._fs_model = None
        self._ordner_lbl.setText("")
        self._setup_tree()

    def _on_tree_activated(self, index):
        """Eintrag im Baum aktiviert (Enter / Doppelklick)."""
        if self._fs_model is None:
            return
        path = self._fs_model.filePath(index)
        if os.path.isfile(path) and path.lower().endswith(('.xlsx', '.xls')):
            self._open_in_next_pane(path)

    def _manuell_datei_oeffnen(self):
        """Öffnet einen Datei-Dialog und lädt die gewählte Datei in die nächste freie Pane."""
        try:
            from functions.settings_functions import get_setting
            start_dir = get_setting('dienstplan_ordner') or os.path.expanduser('~')
        except Exception:
            start_dir = os.path.expanduser('~')
        path, _ = QFileDialog.getOpenFileName(
            self, "Dienstplan-Excel öffnen", start_dir,
            "Excel-Dateien (*.xlsx *.xls)"
        )
        if path:
            self._open_in_next_pane(path)

    # ------------------------------------------------------------------
    # Pane-Verwaltung
    # ------------------------------------------------------------------

    def _set_export_pane(self, idx: int):
        """Markiert Pane idx als aktiv fuer den Word-Export."""
        self._export_pane_idx = idx
        for i, pane in enumerate(self._panes):
            pane.set_export_active(i == idx)
        # Titelzeile aktualisieren
        pane = self._panes[idx]
        if not pane.is_empty:
            name = os.path.basename(pane.excel_path)
            self._export_lbl.setText(f"Export: {name}")
            self._export_lbl.setStyleSheet(
                "color: #0a5ba4; font-weight: bold; "
                "background: #dce8f5; border-radius: 4px; padding: 2px 6px;"
            )
        else:
            self._export_lbl.setText("")

    def _export_pane(self) -> '_DienstplanPane':
        return self._panes[self._export_pane_idx]

    def _open_in_next_pane(self, path: str):
        """
        Datei in der naechsten freien Pane oeffnen.
        Bereits geoeffnet -> neu laden.
        Alle 4 voll -> erste Pane ersetzen.
        """
        # Bereits geoeffnet?
        for i, pane in enumerate(self._panes):
            if pane.excel_path == path:
                pane.load(path)
                self._set_export_pane(i)
                return

        # Naechste freie sichtbare Pane
        for i, pane in enumerate(self._panes):
            if pane.isVisible() and pane.is_empty:
                pane.load(path)
                self._set_export_pane(i)
                # Naechste Pane sichtbar machen (fuer naechsten Ladevorgang)
                next_idx = i + 1
                if next_idx < self.MAX_PANES:
                    self._panes[next_idx].setVisible(True)
                    self._pane_splitter.setSizes(
                        [1] * (next_idx + 1) + [0] * (self.MAX_PANES - next_idx - 1)
                    )
                return

        # Alle sichtbaren Panes belegt → neue Pane oeffnen wenn moeglich
        visible_count = sum(1 for p in self._panes if p.isVisible())
        if visible_count < self.MAX_PANES:
            pane = self._panes[visible_count]
            pane.setVisible(True)
            pane.load(path)
            self._set_export_pane(visible_count)
            self._pane_splitter.setSizes([1] * (visible_count + 1) + [0] * (self.MAX_PANES - visible_count - 1))
            return

        # Alle 4 Panes belegt → Export-Pane ersetzen
        pane = self._panes[self._export_pane_idx]
        pane.load(path)
        self._set_export_pane(self._export_pane_idx)

    def _watch_excel(self, path: str):
        """Fügt eine Excel-Datei dem Datei-Watcher hinzu (einmalig)."""
        if path and path not in self._html_watcher.files():
            self._html_watcher.addPath(path)

    def _on_excel_geaendert(self, path: str):
        """
        Wird aufgerufen wenn eine geladene Excel-Datei auf der Festplatte
        geändert wird. Erzeugt die HTML-Seite neu – aber nur wenn der
        Benutzer sie mindestens einmal manuell generiert hat.
        """
        if not self._html_generiert:
            return
        # Kleines Delay: Excel schreibt manchmal in mehreren Schritten
        from PySide6.QtCore import QTimer
        QTimer.singleShot(1500, lambda: self._html_auto_update(path))

    def _html_auto_update(self, path: str):
        """Stille automatische HTML-Aktualisierung nach Dateiänderung."""
        # Passende Pane suchen
        for pane in self._panes:
            if pane.excel_path == path and pane._display_data:
                try:
                    from functions.dienstplan_parser import DienstplanParser
                    result = DienstplanParser(path, alle_anzeigen=True).parse()
                    if not result.get('success'):
                        return
                    from functions.dienstplan_html_export import generiere_html
                    generiere_html(result)
                    # Status-Zeile aktualisieren
                    from datetime import datetime as _dt
                    pane._status_lbl.setText(
                        f'🌐 Webseite automatisch aktualisiert – {_dt.now().strftime("%H:%M:%S")}'
                    )
                    pane._status_lbl.setStyleSheet('color: #1e7e34; font-weight: bold; padding: 2px 0px;')
                except Exception:
                    pass
                return

    # ------------------------------------------------------------------
    # Datenladen
    # ------------------------------------------------------------------

    def refresh(self):
        self._alle = []
        for pane in self._panes:
            pane.clear()
        self._set_export_pane(0)
        self._html_generiert = False

    # ------------------------------------------------------------------
    # HTML-Export (Webseite)
    # ------------------------------------------------------------------

    def _html_exportieren(self):
        """Generatert die HTML-Ansicht des aktiven Export-Dienstplans und öffnet sie im Browser."""
        pane = self._export_pane()
        if pane._display_data is None:
            QMessageBox.information(
                self, "Kein Dienstplan",
                "Bitte zuerst eine Excel-Datei im Dateibaum laden (Doppelklick)."
            )
            return
        try:
            from functions.dienstplan_html_export import generiere_html, html_pfad
            pfad = generiere_html(pane._display_data)
            self._html_generiert = True

            # Excel in Watcher aufnehmen für automatische Aktualisierung
            self._watch_excel(pane.excel_path)

            # Im Standard-Browser öffnen
            import webbrowser
            url = "file:///" + pfad.replace("\\", "/")
            webbrowser.open(url)

            pane._status_lbl.setText(
                f'🌐 Webseite generiert – {os.path.basename(pfad)}'
            )
            pane._status_lbl.setStyleSheet('color: #1e7e34; font-weight: bold; padding: 2px 0px;')

        except Exception as e:
            QMessageBox.critical(self, "Fehler beim HTML-Export", f"Fehler:\n{e}")

    # ------------------------------------------------------------------
    # Word-Export
    # ------------------------------------------------------------------

    def _pax_verwalten(self):
        """Öffnet den Dialog zur Verwaltung von PAX- und Einsatzzahlen."""
        dlg = PaxEinsatzVerwaltungDialog(parent=self)
        dlg.exec()

    def _word_exportieren(self):
        pane = self._export_pane()
        if pane.parsed_data is None:
            QMessageBox.information(
                self, "Kein Dienstplan",
                "Bitte zuerst eine Datei im Dateibaum auswaehlen (Doppelklick)."
            )
            return

        # ── Schritt 1: Dispo-Zeiten Vorschau / Bearbeitung ──────────────────
        # Roh-Zeiten aus Excel ohne jede Rundung für die Vergleichsspalten
        try:
            from functions.dienstplan_parser import DienstplanParser
            raw_result = DienstplanParser(
                pane.excel_path, alle_anzeigen=True, round_dispo=False
            ).parse()
            raw_data = raw_result if raw_result.get('success') else None
        except Exception:
            raw_data = None

        vorschau = _DispoZeitenVorschauDialog(pane.parsed_data, raw_data=raw_data, parent=self)
        if vorschau.exec() != QDialog.DialogCode.Accepted:
            return
        export_data = vorschau.modified_data

        # ── Schritt 2: Export-Einstellungen (Zeitraum, PAX, Pfad ...) ────────
        dlg = ExportDialog(parsed_data=export_data, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted or not dlg.result:
            return

        params = dlg.result

        # ── Schritt 3: Monatsübersicht anzeigen (nur Dashboard-Format) ────────
        if params.get('format') == 'dashboard':
            monats_dlg = PaxMonatsübersichtDialog(
                von_datum      = params['von_datum'],
                pax_heute      = params.get('pax_zahl', 0),
                einsaetze_heute= params.get('einsaetze_zahl', 0),
                parent         = self,
            )
            if monats_dlg.exec() != QDialog.DialogCode.Accepted:
                return
            # Vom Nutzer ausgewählten Tag/PAX übernehmen (falls eine Zeile gewählt)
            if monats_dlg.selected_pax is not None:
                params['pax_zahl'] = monats_dlg.selected_pax
            if monats_dlg.selected_ein is not None:
                params['einsaetze_zahl'] = monats_dlg.selected_ein

        try:
            # PAX-Zahl + Einsätze VOR Export in DB speichern, damit die
            # Jahressummen im Dashboard-Word bereits den heutigen Wert enthalten.
            try:
                from database.pax_db import speichere_tages_pax, speichere_tages_einsaetze
                datum_str = params['von_datum'].strftime('%Y-%m-%d')
                if params.get('pax_zahl', 0) > 0:
                    speichere_tages_pax(datum_str, params['pax_zahl'])
                if params.get('einsaetze_zahl', 0) > 0:
                    speichere_tages_einsaetze(datum_str, params['einsaetze_zahl'])
            except Exception:
                pass  # nicht-kritisch

            if params.get('format') == 'dashboard':
                from functions.staerkemeldung_dashboard_export import StaerkemeldungDashboardExport
                exporter = StaerkemeldungDashboardExport(
                    dienstplan_data           = export_data,
                    ausgabe_pfad              = params['ausgabe_pfad'],
                    von_datum                 = params['von_datum'],
                    bis_datum                 = params['bis_datum'],
                    pax_zahl                  = params['pax_zahl'],
                    bulmor_aktiv              = params.get('bulmor_aktiv', 5),
                    einsaetze_zahl            = params.get('einsaetze_zahl', 0),
                    sl_tag_name               = params.get('sl_tag_name', ''),
                    sl_nacht_name             = params.get('sl_nacht_name', ''),
                    ausgeschlossene_vollnamen  = params.get('ausgeschlossene_vollnamen', set()),
                )
            else:
                from functions.staerkemeldung_export import StaerkemeldungExport
                exporter = StaerkemeldungExport(
                    dienstplan_data           = export_data,
                    ausgabe_pfad              = params['ausgabe_pfad'],
                    von_datum                 = params['von_datum'],
                    bis_datum                 = params['bis_datum'],
                    pax_zahl                  = params['pax_zahl'],
                    ausgeschlossene_vollnamen  = params.get('ausgeschlossene_vollnamen', set()),
                )
            pfad, warnungen = exporter.export()

            if warnungen:
                QMessageBox.warning(
                    self, "Hinweise",
                    "Export abgeschlossen mit Hinweisen:\n\n" + "\n".join(warnungen)
                )

            QMessageBox.information(
                self, "Export erfolgreich",
                f"Stärkemeldung gespeichert unter:\n{pfad}"
            )

            pane._status_lbl.setText(f"Word-Export: {os.path.basename(pfad)}")
            pane._status_lbl.setStyleSheet("color: #107e3e; padding: 2px 0px;")

        except Exception as e:
            QMessageBox.critical(self, "Fehler beim Export", f"Fehler:\n{e}")
            return

    # ------------------------------------------------------------------
    # Stubs (DB-Anbindung folgt)
    # ------------------------------------------------------------------

    def _render_table(self, schichten: list[Dienstplan]):
        pass

    def _get_selected_id(self) -> int | None:
        return None

    def _add_schicht(self):
        pass

    def _edit_schicht(self):
        pass

    def _delete_schicht(self):
        pass
