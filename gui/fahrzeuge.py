"""
Fahrzeug-Widget
Verwaltung von Fahrzeugen inkl. Status-Historie, Schäden und Terminen
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import date

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QSplitter, QTextEdit, QLineEdit,
    QComboBox, QFormLayout, QMessageBox, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QDateEdit,
    QDialog, QDialogButtonBox, QCheckBox, QSizePolicy, QInputDialog,
    QFileDialog
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QColor

from config import FIORI_BLUE, FIORI_TEXT, FIORI_WHITE, FIORI_BORDER, FIORI_SUCCESS, FIORI_ERROR, FIORI_SIDEBAR_BG

from functions.fahrzeug_functions import (
    erstelle_fahrzeug, aktualisiere_fahrzeug, loesche_fahrzeug,
    lade_alle_fahrzeuge, lade_fahrzeug,
    setze_fahrzeug_status, lade_status_historie, aktueller_status, loesche_status_eintrag,
    aktualisiere_status_eintrag,
    erstelle_schaden, aktualisiere_schaden, lade_schaeden,
    markiere_schaden_behoben, loesche_schaden,
    erstelle_termin, aktualisiere_termin, lade_termine,
    markiere_termin_erledigt, loesche_termin,
    lade_komplette_historie,
)

# ── Hilfsdaten ────────────────────────────────────────────────────────────────
STATUS_META = {
    "fahrbereit":   {"label": "✓ Fahrbereit",    "color": "#107e3e", "bg": "#e8f5e9"},
    "defekt":       {"label": "⚠ Defekt",         "color": "#bb0000", "bg": "#ffebee"},
    "werkstatt":    {"label": "🔧 Werkstatt",      "color": "#e67e22", "bg": "#fff3e0"},
    "ausser_dienst":{"label": "⊘ Außer Dienst",   "color": "#7b1fa2", "bg": "#f3e5f5"},
    "sonstiges":    {"label": "· Sonstiges",       "color": "#5c6bc0", "bg": "#ede7f6"},
}
STATUS_KEYS  = list(STATUS_META.keys())
STATUS_LABELS = [v["label"] for v in STATUS_META.values()]

TERMIN_TYP_META = {
    "tuev":              "🔍 TÜV",
    "hauptuntersuchung": "📋 Hauptuntersuchung",
    "inspektion":        "🛢 Inspektion",
    "reparatur":         "🔧 Reparatur",
    "sonstiges":         "📌 Sonstiges",
}
SCHWERE_META = {
    "gering": {"label": "● Gering",  "color": "#43a047"},
    "mittel": {"label": "● Mittel",  "color": "#fb8c00"},
    "schwer": {"label": "● Schwer",  "color": "#e53935"},
}

def _field_style() -> str:
    return "border:1px solid #ccc; border-radius:3px; padding:4px 8px; background:white;"

def _btn_style(bg: str, hover: str, fg: str = "white") -> str:
    return (f"QPushButton{{background:{bg};color:{fg};border:none;border-radius:4px;"
            f"padding:5px 14px;font-weight:bold;}}"
            f"QPushButton:hover{{background:{hover};}}"
            f"QPushButton:disabled{{background:#ccc;color:#999;}}")

def _fmt_date(d: str) -> str:
    """YYYY-MM-DD → DD.MM.YYYY"""
    try:
        from datetime import datetime
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d or "–"

def _status_badge(status: str) -> QLabel:
    m = STATUS_META.get(status, STATUS_META["sonstiges"])
    lbl = QLabel(m["label"])
    lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
    lbl.setStyleSheet(
        f"color:{m['color']};background:{m['bg']};border:1px solid {m['color']};"
        f"border-radius:4px;padding:2px 8px;"
    )
    return lbl


# ══════════════════════════════════════════════════════════════════════════════
#  Dialog: Fahrzeug anlegen / bearbeiten
# ══════════════════════════════════════════════════════════════════════════════

class _FahrzeugDialog(QDialog):
    def __init__(self, fahrzeug: dict | None = None, parent=None):
        super().__init__(parent)
        self._fid = fahrzeug["id"] if fahrzeug else None
        self.setWindowTitle("Fahrzeug anlegen" if fahrzeug is None else "Fahrzeug bearbeiten")
        self.setMinimumWidth(420)
        self._build(fahrzeug)

    def _build(self, f):
        layout = QVBoxLayout(self)
        fl = QFormLayout()
        fl.setSpacing(8)

        self._kz    = QLineEdit(f.get("kennzeichen","") if f else "")
        self._typ   = QLineEdit(f.get("typ","") if f else "")
        self._tuev  = QDateEdit()
        self._tuev.setCalendarPopup(True)
        self._tuev.setDisplayFormat("dd.MM.yyyy")
        self._tuev.setSpecialValueText("–  kein Datum")
        self._kein_tuev = QCheckBox("Kein TÜV erforderlich")
        if f and f.get("tuev_datum"):
            qd = QDate.fromString(f["tuev_datum"], "yyyy-MM-dd")
            self._tuev.setDate(qd if qd.isValid() else QDate.currentDate())
        else:
            self._tuev.setDate(QDate.currentDate())
            if f is not None:
                self._kein_tuev.setChecked(True)
                self._tuev.setEnabled(False)
        self._kein_tuev.toggled.connect(lambda on: self._tuev.setEnabled(not on))
        self._notiz = QTextEdit(f.get("notizen","") if f else "")
        self._notiz.setFixedHeight(70)

        for w in [self._kz, self._typ]:
            w.setStyleSheet(_field_style())

        self._kz.setPlaceholderText("z.B. K-DRK 123")
        self._typ.setPlaceholderText("z.B. RTW, KTW, PKW")

        fl.addRow("Kennzeichen *:", self._kz)
        fl.addRow("Typ:",           self._typ)
        fl.addRow("TÜV bis:",       self._tuev)
        fl.addRow("",               self._kein_tuev)
        fl.addRow("Notizen:",       self._notiz)

        layout.addLayout(fl)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _validate(self):
        if not self._kz.text().strip():
            QMessageBox.warning(self, "Pflichtfeld", "Bitte Kennzeichen eingeben.")
            return
        self.accept()

    def get_data(self) -> dict:
        return dict(
            kennzeichen   = self._kz.text().strip().upper(),
            typ           = self._typ.text().strip(),
            tuev_datum    = "" if self._kein_tuev.isChecked() else self._tuev.date().toString("yyyy-MM-dd"),
            notizen       = self._notiz.toPlainText().strip(),
        )


# ══════════════════════════════════════════════════════════════════════════════
#  Dialog: Status setzen
# ══════════════════════════════════════════════════════════════════════════════

class _StatusDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Status ändern")
        self.setMinimumWidth(380)
        layout = QVBoxLayout(self)
        fl = QFormLayout()
        fl.setSpacing(8)

        self._status = QComboBox()
        for key, meta in STATUS_META.items():
            self._status.addItem(meta["label"], key)

        self._von = QDateEdit()
        self._von.setCalendarPopup(True)
        self._von.setDate(QDate.currentDate())
        self._von.setDisplayFormat("dd.MM.yyyy")

        self._bis = QDateEdit()
        self._bis.setCalendarPopup(True)
        self._bis.setDate(QDate.currentDate())
        self._bis.setDisplayFormat("dd.MM.yyyy")
        self._bis.setEnabled(False)

        self._unbestimmt = QCheckBox("Unbestimmt (kein Enddatum)")
        self._unbestimmt.setChecked(True)
        self._unbestimmt.toggled.connect(lambda c: self._bis.setEnabled(not c))

        self._grund = QTextEdit()
        self._grund.setPlaceholderText("Grund / Bemerkung (optional)")
        self._grund.setFixedHeight(70)
        self._grund.setStyleSheet(_field_style())

        fl.addRow("Status:",      self._status)
        fl.addRow("Gültig von:", self._von)
        fl.addRow("",             self._unbestimmt)
        fl.addRow("Gültig bis:", self._bis)
        fl.addRow("Grund:",      self._grund)
        layout.addLayout(fl)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_data(self) -> dict:
        return dict(
            status = self._status.currentData(),
            von    = self._von.date().toString("yyyy-MM-dd"),
            bis    = "" if self._unbestimmt.isChecked() else self._bis.date().toString("yyyy-MM-dd"),
            grund  = self._grund.toPlainText().strip(),
        )


# ══════════════════════════════════════════════════════════════════════════════
#  Dialog: Status-Eintrag bearbeiten
# ══════════════════════════════════════════════════════════════════════════════

class _StatusBearbeitenDialog(QDialog):
    """Bestehendes Status-Historieneintrag bearbeiten."""
    def __init__(self, eintrag: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Status-Eintrag bearbeiten")
        self.setMinimumWidth(380)
        layout = QVBoxLayout(self)
        fl = QFormLayout()
        fl.setSpacing(8)

        self._status = QComboBox()
        for key, meta in STATUS_META.items():
            self._status.addItem(meta["label"], key)
        # Vorauswahl des gespeicherten Status
        for i in range(self._status.count()):
            if self._status.itemData(i) == eintrag.get("status", ""):
                self._status.setCurrentIndex(i)
                break

        self._von = QDateEdit()
        self._von.setCalendarPopup(True)
        self._von.setDisplayFormat("dd.MM.yyyy")
        von_qdate = QDate.fromString(eintrag.get("von", ""), "yyyy-MM-dd")
        self._von.setDate(von_qdate if von_qdate.isValid() else QDate.currentDate())

        self._bis = QDateEdit()
        self._bis.setCalendarPopup(True)
        self._bis.setDisplayFormat("dd.MM.yyyy")
        bis_raw = eintrag.get("bis", "") or ""
        bis_qdate = QDate.fromString(bis_raw, "yyyy-MM-dd")
        has_bis = bis_qdate.isValid()
        self._bis.setDate(bis_qdate if has_bis else QDate.currentDate())
        self._bis.setEnabled(has_bis)

        self._unbestimmt = QCheckBox("Unbestimmt (kein Enddatum)")
        self._unbestimmt.setChecked(not has_bis)
        self._unbestimmt.toggled.connect(lambda c: self._bis.setEnabled(not c))

        self._grund = QTextEdit()
        self._grund.setPlaceholderText("Grund / Bemerkung (optional)")
        self._grund.setFixedHeight(70)
        self._grund.setStyleSheet(_field_style())
        self._grund.setPlainText(eintrag.get("grund", "") or "")

        fl.addRow("Status:",      self._status)
        fl.addRow("Gültig von:", self._von)
        fl.addRow("",             self._unbestimmt)
        fl.addRow("Gültig bis:", self._bis)
        fl.addRow("Grund:",      self._grund)
        layout.addLayout(fl)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_data(self) -> dict:
        return dict(
            status = self._status.currentData(),
            von    = self._von.date().toString("yyyy-MM-dd"),
            bis    = "" if self._unbestimmt.isChecked() else self._bis.date().toString("yyyy-MM-dd"),
            grund  = self._grund.toPlainText().strip(),
        )


# ══════════════════════════════════════════════════════════════════════════════
#  Dialog: Schaden erfassen
# ══════════════════════════════════════════════════════════════════════════════

class _SchadenDialog(QDialog):
    def __init__(self, schaden: dict | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Schaden erfassen" if schaden is None else "Schaden bearbeiten")
        self.setMinimumWidth(400)
        layout = QVBoxLayout(self)
        fl = QFormLayout(); fl.setSpacing(8)

        self._datum = QDateEdit()
        self._datum.setCalendarPopup(True)
        self._datum.setDisplayFormat("dd.MM.yyyy")
        self._datum.setDate(
            QDate.fromString(schaden["datum"], "yyyy-MM-dd")
            if schaden and schaden.get("datum") else QDate.currentDate()
        )

        self._beschreibung = QTextEdit(schaden.get("beschreibung","") if schaden else "")
        self._beschreibung.setPlaceholderText("Schadensbeschreibung ...")
        self._beschreibung.setFixedHeight(80)
        self._beschreibung.setStyleSheet(_field_style())

        self._schwere = QComboBox()
        for key, meta in SCHWERE_META.items():
            self._schwere.addItem(meta["label"], key)
        if schaden:
            idx = list(SCHWERE_META.keys()).index(schaden.get("schwere","gering"))
            self._schwere.setCurrentIndex(idx)

        self._kommentar = QTextEdit(schaden.get("kommentar","") if schaden else "")
        self._kommentar.setPlaceholderText("Kommentar / weitere Infos ...")
        self._kommentar.setFixedHeight(70)
        self._kommentar.setStyleSheet(_field_style())

        fl.addRow("Datum:",        self._datum)
        fl.addRow("Beschreibung:", self._beschreibung)
        fl.addRow("Schwere:",      self._schwere)
        fl.addRow("Kommentar:",    self._kommentar)
        layout.addLayout(fl)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _validate(self):
        if not self._beschreibung.toPlainText().strip():
            QMessageBox.warning(self, "Pflichtfeld", "Beschreibung ist erforderlich.")
            return
        self.accept()

    def get_data(self) -> dict:
        return dict(
            datum        = self._datum.date().toString("yyyy-MM-dd"),
            beschreibung = self._beschreibung.toPlainText().strip(),
            schwere      = self._schwere.currentData(),
            kommentar    = self._kommentar.toPlainText().strip(),
        )


# ══════════════════════════════════════════════════════════════════════════════
#  Dialog: Termin erfassen
# ══════════════════════════════════════════════════════════════════════════════

class _TerminDialog(QDialog):
    def __init__(self, termin: dict | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Termin erfassen" if termin is None else "Termin bearbeiten")
        self.setMinimumWidth(420)
        layout = QVBoxLayout(self)
        fl = QFormLayout(); fl.setSpacing(8)

        self._titel = QLineEdit(termin.get("titel","") if termin else "")
        self._titel.setPlaceholderText("z.B. TÜV-Hauptuntersuchung")
        self._titel.setStyleSheet(_field_style())

        self._typ = QComboBox()
        for key, lbl in TERMIN_TYP_META.items():
            self._typ.addItem(lbl, key)
        if termin:
            idx = list(TERMIN_TYP_META.keys()).index(termin.get("typ","sonstiges"))
            self._typ.setCurrentIndex(idx)

        self._datum = QDateEdit()
        self._datum.setCalendarPopup(True)
        self._datum.setDisplayFormat("dd.MM.yyyy")
        self._datum.setDate(
            QDate.fromString(termin["datum"],"yyyy-MM-dd")
            if termin and termin.get("datum") else QDate.currentDate()
        )

        self._uhrzeit = QLineEdit(termin.get("uhrzeit","") if termin else "")
        self._uhrzeit.setPlaceholderText("HH:MM (optional)")
        self._uhrzeit.setStyleSheet(_field_style())

        self._beschreibung = QTextEdit(termin.get("beschreibung","") if termin else "")
        self._beschreibung.setPlaceholderText("Beschreibung / Details ...")
        self._beschreibung.setFixedHeight(70)
        self._beschreibung.setStyleSheet(_field_style())

        self._kommentar = QTextEdit(termin.get("kommentar","") if termin else "")
        self._kommentar.setPlaceholderText("Kommentar ...")
        self._kommentar.setFixedHeight(70)
        self._kommentar.setStyleSheet(_field_style())

        fl.addRow("Titel *:",      self._titel)
        fl.addRow("Typ:",          self._typ)
        fl.addRow("Datum:",        self._datum)
        fl.addRow("Uhrzeit:",      self._uhrzeit)
        fl.addRow("Beschreibung:", self._beschreibung)
        fl.addRow("Kommentar:",    self._kommentar)
        layout.addLayout(fl)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _validate(self):
        if not self._titel.text().strip():
            QMessageBox.warning(self, "Pflichtfeld", "Titel ist erforderlich.")
            return
        self.accept()

    def get_data(self) -> dict:
        return dict(
            titel        = self._titel.text().strip(),
            typ          = self._typ.currentData(),
            datum        = self._datum.date().toString("yyyy-MM-dd"),
            uhrzeit      = self._uhrzeit.text().strip(),
            beschreibung = self._beschreibung.toPlainText().strip(),
            kommentar    = self._kommentar.toPlainText().strip(),
        )


# ══════════════════════════════════════════════════════════════════════════════
#  Haupt-Widget: FahrzeugeWidget
# ══════════════════════════════════════════════════════════════════════════════

class FahrzeugeWidget(QWidget):
    """Haupt-Widget für die Fahrzeugverwaltung."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._aktives_fid: int | None = None
        self._liste_items: dict = {}  # fid → (frame, status, aktiv)
        self._build_ui()
        self.refresh()

    # ── UI-Aufbau ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._build_header())

        # ── Top-Level-Tabs: Fahrzeuge | Ausfälle ─────────────────────────────
        self._main_tabs = QTabWidget()
        self._main_tabs.setDocumentMode(False)
        self._main_tabs.setStyleSheet("""
            QTabWidget::pane { border: none; background: #f8f9fa; }
            QTabBar::tab {
                padding: 8px 22px; font-size: 13px; font-family: 'Segoe UI';
                color: #666; background: #e8ecf0;
                border-bottom: 2px solid transparent;
                border-radius: 4px 4px 0 0; margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: #f8f9fa; color: #1565a8;
                font-weight: bold; border-bottom: 2px solid #1565a8;
            }
            QTabBar::tab:hover:!selected { background: #dde4ec; color: #1565a8; }
        """)

        # Tab 1: Fahrzeuge (bestehender Inhalt)
        fz_widget = QWidget()
        fz_layout = QVBoxLayout(fz_widget)
        fz_layout.setContentsMargins(0, 0, 0, 0)
        fz_layout.setSpacing(0)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(2)
        splitter.addWidget(self._build_liste())
        splitter.addWidget(self._build_detail())
        splitter.setSizes([260, 740])
        fz_layout.addWidget(splitter, 1)

        # Tab 2: Ausfälle
        self._ausfaelle_w = self._build_ausfaelle_tab()

        self._main_tabs.addTab(fz_widget,          "🚗 Fahrzeuge")
        self._main_tabs.addTab(self._ausfaelle_w,  "📊 Ausfälle")
        self._main_tabs.currentChanged.connect(self._on_main_tab_changed)

        root.addWidget(self._main_tabs, 1)

    def _build_header(self) -> QWidget:
        h = QFrame()
        h.setFixedHeight(64)
        h.setStyleSheet(f"background-color:{FIORI_SIDEBAR_BG};")
        layout = QHBoxLayout(h)
        layout.setContentsMargins(20, 8, 20, 8)

        title = QLabel("🚗 Fahrzeugverwaltung")
        title.setFont(QFont("Arial", 17, QFont.Weight.Bold))
        title.setStyleSheet("color:white;")
        layout.addWidget(title)
        layout.addStretch()

        btn = QPushButton("＋  Fahrzeug anlegen")
        btn.setFont(QFont("Arial", 11))
        btn.setFixedHeight(40)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setStyleSheet(_btn_style(FIORI_BLUE, "#0855a9"))
        btn.setToolTip("Neues Fahrzeug in der Datenbank anlegen")
        btn.clicked.connect(self._fahrzeug_anlegen)
        layout.addWidget(btn)
        return h

    # ── Fahrzeug-Liste ─────────────────────────────────────────────────────────

    def _build_liste(self) -> QWidget:
        container = QWidget()
        container.setMinimumWidth(230)
        container.setMaximumWidth(320)
        vlayout = QVBoxLayout(container)
        vlayout.setContentsMargins(0, 0, 0, 0)
        vlayout.setSpacing(0)

        # ── Suchleiste ────────────────────────────────────────────────────────
        search_bar = QFrame()
        search_bar.setStyleSheet("background:#f0f2f4; border-bottom:1px solid #ddd;")
        search_bar.setFixedHeight(86)
        sl = QVBoxLayout(search_bar)
        sl.setContentsMargins(8, 6, 8, 6)
        sl.setSpacing(4)

        self._fz_search = QLineEdit()
        self._fz_search.setPlaceholderText("🔍 Kennzeichen, Status, Schäden ...")
        self._fz_search.setStyleSheet(
            "background:white; border:1px solid #ccc; border-radius:3px;"
            "padding:4px 8px; font-size:11px;"
        )
        self._fz_search.setToolTip(
            "Fahrzeuge filtern \u2013 Suchbegriff eingeben und Filterkategorie wählen.\n"
            "Alle: Kennzeichen + Status + Schäden + Termine\n"
            "Status: nur nach aktuellem Fahrzeugstatus\n"
            "Schäden: in Schadenbeschreibungen suchen\n"
            "Termine: in Wartungs-/TÜV-Terminen suchen\n"
            "Historie: in der vollständigen Statushistorie suchen"
        )
        self._fz_search.textChanged.connect(self._apply_fahrzeug_filter)

        filter_row = QHBoxLayout()
        filter_lbl = QLabel("Filter:")
        filter_lbl.setStyleSheet("border:none; font-size:10px; color:#555;")
        self._fz_filter_combo = QComboBox()
        self._fz_filter_combo.addItems(["Alle", "Status", "Schäden", "Termine", "Historie"])
        self._fz_filter_combo.setStyleSheet(
            "background:white; border:1px solid #ccc; border-radius:3px;"
            "padding:2px 4px; font-size:10px;"
        )
        self._fz_filter_combo.setToolTip("Suchbereich einschränken: Status / Schäden / Termine / Historie")
        self._fz_filter_combo.currentIndexChanged.connect(self._apply_fahrzeug_filter)
        filter_row.addWidget(filter_lbl)
        filter_row.addWidget(self._fz_filter_combo, 1)

        sl.addWidget(self._fz_search)
        sl.addLayout(filter_row)
        vlayout.addWidget(search_bar)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea{border:none;}")

        self._liste_inner = QWidget()
        self._liste_layout = QVBoxLayout(self._liste_inner)
        self._liste_layout.setContentsMargins(8, 8, 8, 8)
        self._liste_layout.setSpacing(4)
        self._liste_layout.addStretch()

        scroll.setWidget(self._liste_inner)
        vlayout.addWidget(scroll, 1)
        return container

    def _refresh_liste(self):
        self._liste_items.clear()
        while self._liste_layout.count() > 1:
            item = self._liste_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        fahrzeuge = lade_alle_fahrzeuge()
        for f in fahrzeuge:
            frame = self._make_liste_item(f)
            self._liste_layout.insertWidget(self._liste_layout.count() - 1, frame)
            self._liste_items[f["id"]] = (frame, f.get("aktueller_status") or "fahrbereit", bool(f.get("aktiv", 1)))
        self._update_liste_selection()
        self._apply_fahrzeug_filter()

    def _apply_fahrzeug_filter(self):
        """Zeigt/versteckt Fahrzeuge anhand des Suchtextes und des gewählten Filters."""
        text = self._fz_search.text().strip().lower()
        filter_mode = self._fz_filter_combo.currentText()

        for fid, (frame, status, aktiv) in self._liste_items.items():
            if not text:
                frame.setVisible(True)
                continue

            fz = lade_fahrzeug(fid) or {}
            kz  = (fz.get("kennzeichen") or "").lower()
            typ = (fz.get("typ") or "").lower()
            match_basis = text in kz or text in typ

            if filter_mode == "Status":
                sm = STATUS_META.get(status, {})
                match = match_basis or text in (sm.get("label") or "").lower() or text in (status or "").lower()

            elif filter_mode == "Schäden":
                schaeden = lade_schaeden(fid) or []
                match = match_basis or any(
                    text in (s.get("beschreibung") or "").lower() or
                    text in (s.get("kommentar") or "").lower() or
                    text in (s.get("schwere") or "").lower()
                    for s in schaeden
                )

            elif filter_mode == "Termine":
                termine = lade_termine(fid) or []
                match = match_basis or any(
                    text in (t.get("typ") or "").lower() or
                    text in (t.get("notiz") or "").lower()
                    for t in termine
                )

            elif filter_mode == "Historie":
                historie = lade_komplette_historie(fid) or []
                match = match_basis or any(
                    text in (h.get("status") or "").lower() or
                    text in (h.get("notiz") or "").lower() or
                    text in (h.get("datum") or "").lower()
                    for h in historie
                )

            else:  # Alle
                schaeden = lade_schaeden(fid) or []
                termine  = lade_termine(fid) or []
                match = (
                    match_basis
                    or text in (STATUS_META.get(status, {}).get("label") or "").lower()
                    or any(text in (s.get("beschreibung") or "").lower() for s in schaeden)
                    or any(text in (t.get("typ") or "").lower() for t in termine)
                )

            frame.setVisible(match)

    def _make_liste_item(self, f: dict) -> QFrame:
        fid     = f["id"]
        status  = f.get("aktueller_status") or "fahrbereit"
        meta    = STATUS_META.get(status, STATUS_META["sonstiges"])
        aktiv   = bool(f.get("aktiv", 1))
        bg      = meta["bg"] if aktiv else "#f0f0f0"
        border  = meta["color"] if aktiv else "#aaa"

        frame = QFrame()
        frame.setProperty("normalBg", bg)
        frame.setProperty("normalBorder", border)
        frame.setStyleSheet(
            f"QFrame{{background:{bg};border:1px solid #ddd;"
            f"border-left:4px solid {border};border-radius:4px;}}"
        )
        frame.setMinimumHeight(58)
        frame.setCursor(Qt.CursorShape.PointingHandCursor)

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(10, 6, 10, 6)
        layout.setSpacing(2)

        top = QHBoxLayout()
        kz_lbl = QLabel(f.get("kennzeichen","–"))
        kz_lbl.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        kz_lbl.setStyleSheet(f"color:{FIORI_TEXT};border:none;")
        top.addWidget(kz_lbl)
        top.addStretch()
        s_lbl = QLabel(meta["label"])
        s_lbl.setFont(QFont("Arial", 9))
        s_lbl.setStyleSheet(f"color:{meta['color']};border:none;font-weight:bold;")
        top.addWidget(s_lbl)

        typ_str = f.get("typ","") or "–"
        sub = QLabel(typ_str)
        sub.setFont(QFont("Arial", 9))
        sub.setStyleSheet("color:#666;border:none;")

        layout.addLayout(top)
        layout.addWidget(sub)

        frame.mousePressEvent = lambda e, i=fid: self._zeige_fahrzeug(i)
        return frame

    def _update_liste_selection(self):
        for fid, (frame, status, aktiv) in self._liste_items.items():
            if fid == self._aktives_fid:
                frame.setStyleSheet(
                    f"QFrame{{background:#e3f0fc;border:2px solid {FIORI_BLUE};"
                    f"border-left:5px solid {FIORI_BLUE};border-radius:4px;}}"
                )
            else:
                meta   = STATUS_META.get(status, STATUS_META["sonstiges"])
                bg     = meta["bg"] if aktiv else "#f0f0f0"
                border = meta["color"] if aktiv else "#aaa"
                frame.setStyleSheet(
                    f"QFrame{{background:{bg};border:1px solid #ddd;"
                    f"border-left:4px solid {border};border-radius:4px;}}"
                )

    # ── Detail-Bereich ─────────────────────────────────────────────────────────

    def _build_detail(self) -> QWidget:
        self._detail_stack = QWidget()
        self._detail_stack.setStyleSheet("background:white;")
        layout = QVBoxLayout(self._detail_stack)
        layout.setContentsMargins(0, 0, 0, 0)

        # Leer-Zustand
        self._placeholder = QLabel("Fahrzeug in der Liste auswählen\noder neues Fahrzeug anlegen")
        self._placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setFont(QFont("Arial", 14))
        self._placeholder.setStyleSheet("color:#bbb;")
        layout.addWidget(self._placeholder)

        return self._detail_stack

    def _zeige_fahrzeug(self, fid: int):
        self._aktives_fid = fid
        self._update_liste_selection()
        f = lade_fahrzeug(fid)
        if not f:
            return

        # Detail-Bereich neu bauen
        layout = self._detail_stack.layout()
        # Alten Inhalt entfernen
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Fahrzeug-Header
        status = aktueller_status(fid)
        stat_key = status["status"] if status else "fahrbereit"
        meta = STATUS_META.get(stat_key, STATUS_META["sonstiges"])

        fh = QFrame()
        fh.setFixedHeight(64)
        fh.setStyleSheet(f"background:{meta['bg']};border-bottom:2px solid {meta['color']};")
        fhl = QHBoxLayout(fh)
        fhl.setContentsMargins(20, 8, 20, 8)

        title = QLabel(f.get("kennzeichen","–"))
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setStyleSheet(f"color:{FIORI_TEXT};")
        typ_lbl = QLabel(f.get('typ','') or "–")
        typ_lbl.setFont(QFont("Arial", 11))
        typ_lbl.setStyleSheet("color:#555;")

        fhl.addWidget(title)
        fhl.addWidget(typ_lbl)
        fhl.addStretch()

        stat_badge = _status_badge(stat_key)
        fhl.addWidget(stat_badge)

        # Bearbeiten/Löschen-Buttons
        btn_edit = QPushButton("✏  Bearbeiten")
        btn_edit.setStyleSheet(_btn_style(FIORI_BLUE, "#0855a9"))
        btn_edit.setFixedHeight(34)
        btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_edit.setToolTip("Fahrzeugdaten bearbeiten (Kennzeichen, Typ, TÜV-Datum ...)") 
        btn_edit.clicked.connect(lambda: self._fahrzeug_bearbeiten(fid))

        btn_del = QPushButton("🗑")
        btn_del.setFixedSize(34, 34)
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.setToolTip("Dieses Fahrzeug dauerhaft aus der Datenbank löschen")
        btn_del.setStyleSheet("QPushButton{background:#eee;border:none;border-radius:4px;}"
                              "QPushButton:hover{background:#ffcccc;color:#a00;}")
        btn_del.clicked.connect(lambda: self._fahrzeug_loeschen(fid))

        fhl.addWidget(btn_edit)
        fhl.addSpacing(4)
        fhl.addWidget(btn_del)

        layout.addWidget(fh)

        # Tabs
        tabs = QTabWidget()
        tabs.setDocumentMode(False)
        tabs.setStyleSheet("""
            QTabWidget::pane { border: none; background: #f8f9fa; }
            QTabBar::tab {
                padding: 8px 18px;
                font-size: 12px;
                font-family: 'Segoe UI';
                color: #666;
                background: #e8ecf0;
                border-bottom: 2px solid transparent;
                border-radius: 4px 4px 0 0;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: #f8f9fa;
                color: #1565a8;
                font-weight: bold;
                border-bottom: 2px solid #1565a8;
            }
            QTabBar::tab:hover:!selected {
                background: #dde4ec;
                color: #1565a8;
            }
        """)
        tabs.addTab(self._tab_stammdaten(f),      "📋 Stammdaten")
        tabs.addTab(self._tab_status(fid, stat_key), "🚦 Status")
        tabs.addTab(self._tab_schaeden(fid),       "⚠ Schäden")
        tabs.addTab(self._tab_termine(fid),        "📅 Termine")
        tabs.addTab(self._tab_historie(fid),       "📜 Historie")

        layout.addWidget(tabs, 1)

    # ── Zeitfilter-Hilfe ──────────────────────────────────────────────────────────

    @staticmethod
    def _add_zeitfilter(
        layout: "QVBoxLayout",
        data_list: list,
        date_key: str,
        table: "QTableWidget",
    ):
        """Fügt eine Jahr/Monat-Filterleiste direkt über die Tabelle in *layout* ein."""
        MONATE = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
                  "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]
        years = sorted(
            {(e.get(date_key) or "")[:4]
             for e in data_list
             if len(e.get(date_key) or "") >= 4},
            reverse=True,
        )

        bar = QFrame()
        bar.setStyleSheet(
            "QFrame{background:#f5f6fa; border:1px solid #e0e4ec; border-radius:4px;}"
        )
        bl = QHBoxLayout(bar)
        bl.setContentsMargins(8, 4, 8, 4)
        bl.setSpacing(6)

        _lbl_style = "border:none; font-size:11px; color:#444;"

        lbl_j = QLabel("Jahr:")
        lbl_j.setStyleSheet(_lbl_style)
        bl.addWidget(lbl_j)
        jahr_cb = QComboBox()
        jahr_cb.setToolTip("Einträge nach Jahr filtern")
        jahr_cb.addItem("Alle")
        for y in years:
            if y:
                jahr_cb.addItem(y)
        jahr_cb.setStyleSheet(
            "QComboBox{background:white;border:1px solid #ccc;border-radius:3px;"
            "padding:2px 6px;font-size:11px;}"
        )
        bl.addWidget(jahr_cb)

        bl.addSpacing(12)
        lbl_m = QLabel("Monat:")
        lbl_m.setStyleSheet(_lbl_style)
        bl.addWidget(lbl_m)
        monat_cb = QComboBox()
        monat_cb.setToolTip("Einträge nach Monat filtern")
        monat_cb.addItem("Alle", None)
        for i, m in enumerate(MONATE, 1):
            monat_cb.addItem(f"{i:02d} – {m}", i)
        monat_cb.setStyleSheet(
            "QComboBox{background:white;border:1px solid #ccc;border-radius:3px;"
            "padding:2px 6px;font-size:11px;}"
        )
        bl.addWidget(monat_cb)
        bl.addStretch()

        def _apply():
            j = jahr_cb.currentText()
            m = monat_cb.currentData()
            for r, entry in enumerate(data_list):
                d = (entry.get(date_key) or "")
                if not d:
                    table.setRowHidden(r, False)
                    continue
                match_j = (j == "Alle") or d.startswith(j)
                match_m = (m is None) or (len(d) >= 7 and int(d[5:7]) == m)
                table.setRowHidden(r, not (match_j and match_m))

        jahr_cb.currentIndexChanged.connect(lambda _: _apply())
        monat_cb.currentIndexChanged.connect(lambda _: _apply())
        layout.addWidget(bar)

    # ── Tab: Stammdaten ────────────────────────────────────────────────────────

    def _tab_stammdaten(self, f: dict) -> QWidget:
        w = QWidget()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;}")
        inner = QWidget()
        fl = QFormLayout(inner)
        fl.setContentsMargins(24, 20, 24, 20)
        fl.setSpacing(10)
        fl.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        felder = [
            ("Kennzeichen:",      f.get("kennzeichen","–")),
            ("Typ:",              f.get("typ","–") or "–"),
            ("TÜV bis:",          _fmt_date(f.get("tuev_datum","")) or "–"),
            ("Angelegt am:",      _fmt_date((f.get("erstellt_am","") or "")[:10])),
            ("Geändert am:",      _fmt_date((f.get("geaendert_am","") or "")[:10])),
            ("Status:",           "aktiv" if f.get("aktiv",1) else "ausgemustert"),
        ]
        for label, wert in felder:
            lbl_w = QLabel(wert)
            lbl_w.setFont(QFont("Arial", 12))
            lbl_w.setStyleSheet("color:#222;padding:2px 0;")
            fl.addRow(label, lbl_w)

        if f.get("notizen"):
            fl.addRow("Notizen:", QLabel(f["notizen"]))

        scroll.setWidget(inner)
        layout = QVBoxLayout(w)
        layout.setContentsMargins(0,0,0,0)
        layout.addWidget(scroll)
        return w

    # ── Tab: Status ────────────────────────────────────────────────────────────

    def _tab_status(self, fid: int, aktuell: str) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(12)

        # Aktuelle Status-Karte
        meta = STATUS_META.get(aktuell, STATUS_META["sonstiges"])
        karte = QFrame()
        karte.setStyleSheet(
            f"background:{meta['bg']};border:1px solid {meta['color']};"
            f"border-radius:6px;"
        )
        kl = QHBoxLayout(karte)
        kl.setContentsMargins(20, 14, 20, 14)
        s_lbl = QLabel(meta["label"])
        s_lbl.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        s_lbl.setStyleSheet(f"color:{meta['color']};border:none;")
        kl.addWidget(s_lbl)
        kl.addStretch()

        btn_change = QPushButton("Status ändern")
        btn_change.setFixedHeight(36)
        btn_change.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_change.setToolTip("Neuen Status setzen (fahrbereit, defekt, Werkstatt, außer Dienst ...)")
        btn_change.setStyleSheet(_btn_style(meta["color"], meta["color"]))
        btn_change.clicked.connect(lambda: self._status_aendern(fid))
        kl.addWidget(btn_change)
        layout.addWidget(karte)

        # Status-Historie Tabelle
        layout.addWidget(QLabel("Status-Verlauf:"))
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Status", "Von", "Bis", "Grund"])
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("QTableWidget{border:1px solid #ddd;font-size:12px;}")
        table.verticalHeader().setVisible(False)

        verlauf = lade_status_historie(fid)
        table.setRowCount(len(verlauf))
        for row, e in enumerate(verlauf):
            sk = e.get("status","")
            em = STATUS_META.get(sk, STATUS_META["sonstiges"])
            item_s = QTableWidgetItem(em["label"])
            item_s.setForeground(QColor(em["color"]))
            table.setItem(row, 0, item_s)
            table.setItem(row, 1, QTableWidgetItem(_fmt_date(e.get("von",""))))
            bis_str = _fmt_date(e.get("bis","")) if e.get("bis") else "unbestimmt"
            table.setItem(row, 2, QTableWidgetItem(bis_str))
            table.setItem(row, 3, QTableWidgetItem(e.get("grund","") or "–"))

        self._add_zeitfilter(layout, verlauf, "von", table)
        layout.addWidget(table, 1)

        # Eintrag bearbeiten / löschen
        def _edit_status():
            row = table.currentRow()
            if row < 0 or row >= len(verlauf):
                return
            eintrag = verlauf[row]
            dlg = _StatusBearbeitenDialog(eintrag, self)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                d = dlg.get_data()
                aktualisiere_status_eintrag(
                    eintrag["id"], d["status"], d["von"], d["bis"], d["grund"]
                )
                self._zeige_fahrzeug(fid)

        def _del_status():
            row = table.currentRow()
            if row < 0 or row >= len(verlauf):
                return
            eid = verlauf[row]["id"]
            if QMessageBox.question(self, "Löschen", "Diesen Status-Eintrag löschen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) == QMessageBox.StandardButton.Yes:
                loesche_status_eintrag(eid)
                self._zeige_fahrzeug(fid)

        table.itemDoubleClicked.connect(lambda _item: _edit_status())

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        btn_edit = QPushButton("✏  Eintrag bearbeiten")
        btn_edit.setFixedHeight(32)
        btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_edit.setToolTip("Ausgewählten Status-Historieneintrag bearbeiten (auch Doppelklick)")
        btn_edit.setStyleSheet("QPushButton{background:#eee;border:none;border-radius:4px;padding:4px 12px;}"
                               "QPushButton:hover{background:#cce5ff;color:#0057b8;}")
        btn_edit.clicked.connect(_edit_status)
        btn_row.addWidget(btn_edit)

        btn_del = QPushButton("🗑  Eintrag löschen")
        btn_del.setFixedHeight(32)
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.setToolTip("Ausgewählten Status-Historieneintrag aus der Datenbank entfernen")
        btn_del.setStyleSheet("QPushButton{background:#eee;border:none;border-radius:4px;padding:4px 12px;}"
                              "QPushButton:hover{background:#ffcccc;color:#a00;}")
        btn_del.clicked.connect(_del_status)
        btn_row.addWidget(btn_del)

        btn_row.addStretch()
        layout.addLayout(btn_row)
        return w

    # ── Tab: Schäden ───────────────────────────────────────────────────────────

    def _tab_schaeden(self, fid: int) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        btn_add = QPushButton("＋  Schaden erfassen")
        btn_add.setFixedHeight(36)
        btn_add.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_add.setStyleSheet(_btn_style("#bb0000", "#990000"))
        btn_add.clicked.connect(lambda: self._schaden_erfassen(fid))
        btn_row.addWidget(btn_add)

        btn_unfall = QPushButton("🗋  Unfallbogen")
        btn_unfall.setFixedHeight(36)
        btn_unfall.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_unfall.setStyleSheet(_btn_style("#b05000", "#8a3d00"))
        btn_unfall.clicked.connect(self._unfallbogen_dialog)
        btn_row.addWidget(btn_unfall)

        btn_rep = QPushButton("🔧  Reparaturauftrag")
        btn_rep.setFixedHeight(36)
        btn_rep.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_rep.setStyleSheet(_btn_style("#1565a8", "#0f4d8a"))
        btn_rep.clicked.connect(self._reparaturauftrag_oeffnen)
        btn_row.addWidget(btn_rep)

        btn_row.addStretch()
        layout.addLayout(btn_row)

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Datum", "Schwere", "Beschreibung", "Status", "Behoben am", "Kommentar"])
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("QTableWidget{border:1px solid #ddd;font-size:12px;}")
        table.verticalHeader().setVisible(False)

        schaeden = lade_schaeden(fid)
        table.setRowCount(len(schaeden))
        for row, s in enumerate(schaeden):
            table.setItem(row, 0, QTableWidgetItem(_fmt_date(s.get("datum",""))))
            sm = SCHWERE_META.get(s.get("schwere","gering"), SCHWERE_META["gering"])
            it_s = QTableWidgetItem(sm["label"])
            it_s.setForeground(QColor(sm["color"]))
            table.setItem(row, 1, it_s)
            table.setItem(row, 2, QTableWidgetItem(s.get("beschreibung","")))
            behoben = bool(s.get("behoben"))
            it_b = QTableWidgetItem("✓ Behoben" if behoben else "⚠ Offen")
            it_b.setForeground(QColor(FIORI_SUCCESS if behoben else "#bb0000"))
            table.setItem(row, 3, it_b)
            table.setItem(row, 4, QTableWidgetItem(
                _fmt_date(s.get("behoben_am","")) if s.get("behoben_am") else "–"
            ))
            table.setItem(row, 5, QTableWidgetItem(s.get("kommentar","") or "–"))

        self._add_zeitfilter(layout, schaeden, "datum", table)
        layout.addWidget(table, 1)

        btn_row = QHBoxLayout()
        btn_mark = QPushButton("✓  Als behoben markieren")
        btn_mark.setFixedHeight(32)
        btn_mark.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_mark.setToolTip("Markierten Schaden als behoben kennzeichnen und Datum setzen")
        btn_mark.setStyleSheet(_btn_style(FIORI_SUCCESS, "#0d6831"))
        def _mark_behoben():
            row = table.currentRow()
            if row < 0 or row >= len(schaeden): return
            sid = schaeden[row]["id"]
            markiere_schaden_behoben(sid, date.today().isoformat())
            self._zeige_fahrzeug(fid)
        btn_mark.clicked.connect(_mark_behoben)

        btn_edit = QPushButton("✏  Bearbeiten")
        btn_edit.setFixedHeight(32)
        btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_edit.setToolTip("Schadensmeldung bearbeiten (Beschreibung, Schwere, Kommentar)")
        btn_edit.setStyleSheet(_btn_style(FIORI_BLUE, "#0855a9"))
        def _edit_schaden():
            row = table.currentRow()
            if row < 0 or row >= len(schaeden): return
            dlg = _SchadenDialog(schaeden[row], self)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                d = dlg.get_data()
                aktualisiere_schaden(schaeden[row]["id"], d["beschreibung"],
                    d["schwere"], d["kommentar"],
                    schaeden[row].get("behoben",0), schaeden[row].get("behoben_am",""))
                self._zeige_fahrzeug(fid)
        btn_edit.clicked.connect(_edit_schaden)

        btn_del = QPushButton("🗑  Löschen")
        btn_del.setFixedHeight(32)
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.setToolTip("Schadensmeldung dauerhaft aus der Datenbank löschen")
        btn_del.setStyleSheet("QPushButton{background:#eee;border:none;border-radius:4px;padding:4px 10px;}"
                              "QPushButton:hover{background:#ffcccc;color:#a00;}")
        def _del_schaden():
            row = table.currentRow()
            if row < 0 or row >= len(schaeden): return
            if QMessageBox.question(self,"Löschen","Schaden-Eintrag löschen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) == QMessageBox.StandardButton.Yes:
                loesche_schaden(schaeden[row]["id"])
                self._zeige_fahrzeug(fid)
        btn_del.clicked.connect(_del_schaden)

        btn_row.addWidget(btn_mark); btn_row.addWidget(btn_edit)
        btn_row.addStretch(); btn_row.addWidget(btn_del)
        layout.addLayout(btn_row)
        return w

    # ── Tab: Termine ───────────────────────────────────────────────────────────

    def _tab_termine(self, fid: int) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        btn_add = QPushButton("＋  Termin erfassen")
        btn_add.setFixedHeight(36)
        btn_add.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_add.setStyleSheet(_btn_style("#e67e22", "#c0651a"))
        btn_add.clicked.connect(lambda: self._termin_erfassen(fid))
        layout.addWidget(btn_add, 0, Qt.AlignmentFlag.AlignLeft)

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Datum", "Uhrzeit", "Typ", "Titel", "Status", "Kommentar"])
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("QTableWidget{border:1px solid #ddd;font-size:12px;}")
        table.verticalHeader().setVisible(False)

        termine = lade_termine(fid)
        table.setRowCount(len(termine))
        for row, t in enumerate(termine):
            table.setItem(row, 0, QTableWidgetItem(_fmt_date(t.get("datum",""))))
            table.setItem(row, 1, QTableWidgetItem(t.get("uhrzeit","") or "–"))
            table.setItem(row, 2, QTableWidgetItem(TERMIN_TYP_META.get(t.get("typ","sonstiges"),"–")))
            table.setItem(row, 3, QTableWidgetItem(t.get("titel","")))
            erledigt = bool(t.get("erledigt"))
            it_e = QTableWidgetItem("✓ Erledigt" if erledigt else "· Offen")
            it_e.setForeground(QColor(FIORI_SUCCESS if erledigt else "#555"))
            table.setItem(row, 4, it_e)
            table.setItem(row, 5, QTableWidgetItem(t.get("kommentar","") or "–"))

        self._add_zeitfilter(layout, termine, "datum", table)
        layout.addWidget(table, 1)

        btn_row = QHBoxLayout()
        btn_mark = QPushButton("✓  Als erledigt markieren")
        btn_mark.setFixedHeight(32)
        btn_mark.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_mark.setToolTip("Markierten Termin als erledigt kennzeichnen")
        btn_mark.setStyleSheet(_btn_style(FIORI_SUCCESS, "#0d6831"))
        def _mark_erledigt():
            row = table.currentRow()
            if row < 0 or row >= len(termine): return
            markiere_termin_erledigt(termine[row]["id"])
            self._zeige_fahrzeug(fid)
        btn_mark.clicked.connect(_mark_erledigt)

        btn_edit = QPushButton("✏  Bearbeiten")
        btn_edit.setFixedHeight(32)
        btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_edit.setToolTip("Termin bearbeiten (Datum, Typ, Titel, Uhrzeit, Kommentar)")
        btn_edit.setStyleSheet(_btn_style(FIORI_BLUE, "#0855a9"))
        def _edit_termin():
            row = table.currentRow()
            if row < 0 or row >= len(termine): return
            dlg = _TerminDialog(termine[row], self)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                d = dlg.get_data()
                aktualisiere_termin(termine[row]["id"],
                    d["datum"], d["titel"], d["typ"],
                    d["uhrzeit"], d["beschreibung"], d["kommentar"],
                    termine[row].get("erledigt",0))
                self._zeige_fahrzeug(fid)
        btn_edit.clicked.connect(_edit_termin)

        btn_del = QPushButton("🗑  Löschen")
        btn_del.setFixedHeight(32)
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.setToolTip("Termin dauerhaft aus der Datenbank löschen")
        btn_del.setStyleSheet("QPushButton{background:#eee;border:none;border-radius:4px;padding:4px 10px;}"
                              "QPushButton:hover{background:#ffcccc;color:#a00;}")
        def _del_termin():
            row = table.currentRow()
            if row < 0 or row >= len(termine): return
            if QMessageBox.question(self,"Löschen","Termin-Eintrag löschen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) == QMessageBox.StandardButton.Yes:
                loesche_termin(termine[row]["id"])
                self._zeige_fahrzeug(fid)
        btn_del.clicked.connect(_del_termin)

        btn_row.addWidget(btn_mark); btn_row.addWidget(btn_edit)
        btn_row.addStretch(); btn_row.addWidget(btn_del)
        layout.addLayout(btn_row)
        return w

    # ── Tab: Historie ──────────────────────────────────────────────────────────

    def _tab_historie(self, fid: int) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(16, 12, 16, 12)

        INFO_COLORS = {
            "status": "#1565a8",
            "schaden": "#bb0000",
            "termin": "#e67e22",
        }
        ART_LABELS = {
            "status": "🚦 Status",
            "schaden": "⚠ Schaden",
            "termin": "📅 Termin",
        }

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Datum", "Art", "Titel/Status", "Bis", "Details", "Kommentar"])
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("QTableWidget{border:1px solid #ddd;font-size:12px;}")
        table.verticalHeader().setVisible(False)

        eintraege = lade_komplette_historie(fid)
        table.setRowCount(len(eintraege))
        for row, e in enumerate(eintraege):
            art = e.get("art","")
            farbe = INFO_COLORS.get(art, FIORI_TEXT)
            art_lbl = ART_LABELS.get(art, art)

            titel = e.get("titel","") or ""
            if art == "status":
                sm = STATUS_META.get(titel, STATUS_META["sonstiges"])
                titel = sm["label"]
                bis_raw = e.get("bis","")
                bis_str = _fmt_date(bis_raw) if bis_raw else "unbestimmt"
            else:
                bis_str = "–"

            table.setItem(row, 0, QTableWidgetItem(_fmt_date(e.get("datum",""))))
            it_art = QTableWidgetItem(art_lbl)
            it_art.setForeground(QColor(farbe))
            table.setItem(row, 1, it_art)
            table.setItem(row, 2, QTableWidgetItem(titel))
            table.setItem(row, 3, QTableWidgetItem(bis_str))
            table.setItem(row, 4, QTableWidgetItem(e.get("beschreibung","") or "–"))
            table.setItem(row, 5, QTableWidgetItem(e.get("kommentar","") or "–"))

        self._add_zeitfilter(layout, eintraege, "datum", table)
        layout.addWidget(table, 1)
        return w

    # ── Aktionen ───────────────────────────────────────────────────────────────

    def _fahrzeug_anlegen(self):
        dlg = _FahrzeugDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            try:
                fid = erstelle_fahrzeug(**d)
                self._refresh_liste()
                self._zeige_fahrzeug(fid)
            except Exception as e:
                QMessageBox.critical(self, "Fehler", f"Fahrzeug konnte nicht angelegt werden:\n{e}")

    def _fahrzeug_bearbeiten(self, fid: int):
        f = lade_fahrzeug(fid)
        if not f:
            return
        dlg = _FahrzeugDialog(f, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            try:
                aktualisiere_fahrzeug(fid, **d)
                self._refresh_liste()
                self._zeige_fahrzeug(fid)
            except Exception as e:
                QMessageBox.critical(self, "Fehler", f"Fehler beim Speichern:\n{e}")

    def _fahrzeug_loeschen(self, fid: int):
        f = lade_fahrzeug(fid)
        kz = f.get("kennzeichen","?") if f else "?"
        pw, ok = QInputDialog.getText(
            self, "Löschen bestätigen",
            "Passwort eingeben:",
            QLineEdit.EchoMode.Password
        )
        if not ok:
            return
        if pw != "mettwurst":
            QMessageBox.warning(self, "Falsches Passwort", "Das eingegebene Passwort ist falsch – Fahrzeug nicht gelöscht.")
            return
        if QMessageBox.question(
            self, "Fahrzeug löschen",
            f"Fahrzeug {kz} und alle zugehörigen Daten (Schäden, Termine, Status) dauerhaft löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            loesche_fahrzeug(fid)
            self._aktives_fid = None
            self._refresh_liste()
            # Detail-Bereich zurücksetzen
            layout = self._detail_stack.layout()
            while layout.count():
                item = layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            self._placeholder = QLabel("Fahrzeug in der Liste auswählen\noder neues Fahrzeug anlegen")
            self._placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._placeholder.setFont(QFont("Arial", 14))
            self._placeholder.setStyleSheet("color:#bbb;")
            layout.addWidget(self._placeholder)

    def _status_aendern(self, fid: int):
        dlg = _StatusDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            setze_fahrzeug_status(fid, d["status"], d["von"], d["grund"], d["bis"])
            self._refresh_liste()
            self._zeige_fahrzeug(fid)

    def _schaden_erfassen(self, fid: int):
        dlg = _SchadenDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            erstelle_schaden(fid, d["datum"], d["beschreibung"], d["schwere"], d["kommentar"])
            self._zeige_fahrzeug(fid)

    def _unfallbogen_dialog(self):
        import os as _os
        _BASE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
        folder = _os.path.join(_BASE, "Daten", "Unfallmeldebogen Vorlage")

        dlg = QDialog(self)
        dlg.setWindowTitle("🗋 Unfallbogen – Vorlagen drucken")
        dlg.setMinimumWidth(420)
        dlg_vl = QVBoxLayout(dlg)
        dlg_vl.setContentsMargins(18, 14, 18, 14)
        dlg_vl.setSpacing(10)

        title = QLabel("Unfallmeldebogen – Vorlagen")
        title.setStyleSheet("font-size:14px;font-weight:bold;")
        dlg_vl.addWidget(title)

        hint = QLabel("Datei öffnen → im PDF-Viewer über Datei › Drucken ausgeben")
        hint.setStyleSheet("color:#666;font-size:11px;")
        dlg_vl.addWidget(hint)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#ddd;")
        dlg_vl.addWidget(sep)

        try:
            files = sorted([
                f for f in _os.listdir(folder)
                if f.lower().endswith((".pdf", ".doc", ".docx", ".xls", ".xlsx"))
            ])
        except Exception:
            files = []

        if not files:
            dlg_vl.addWidget(QLabel("⚠ Keine Dateien im Ordner gefunden."))
        else:
            for fname in files:
                row = QHBoxLayout()
                lbl = QLabel(fname)
                lbl.setStyleSheet("font-size:11px;")
                lbl.setWordWrap(True)
                row.addWidget(lbl, 1)

                fpath = _os.path.join(folder, fname)

                btn_open = QPushButton("📄 Öffnen / Drucken")
                btn_open.setFixedHeight(28)
                btn_open.setStyleSheet(
                    "QPushButton{background:#0078a8;color:white;border:none;"
                    "border-radius:4px;padding:2px 10px;font-size:11px;}"
                    "QPushButton:hover{background:#005f8a;}"
                )
                btn_open.clicked.connect(lambda _checked=False, p=fpath: (
                    _os.startfile(p) if hasattr(_os, "startfile") else
                    __import__("subprocess").run(["xdg-open", p])
                ))
                row.addWidget(btn_open)
                dlg_vl.addLayout(row)

        dlg_vl.addStretch()

        btn_folder = QPushButton("📁 Ordner öffnen")
        btn_folder.setFixedHeight(28)
        btn_folder.setStyleSheet(
            "QPushButton{background:#eee;border:1px solid #ccc;"
            "border-radius:4px;padding:2px 10px;font-size:11px;}"
            "QPushButton:hover{background:#ddd;}"
        )
        btn_folder.clicked.connect(lambda: (
            __import__("os").startfile(folder)
            if hasattr(__import__("os"), "startfile") else None
        ))

        close_btn = QPushButton("Schließen")
        close_btn.setFixedHeight(28)
        close_btn.clicked.connect(dlg.accept)

        btn_row_bottom = QHBoxLayout()
        btn_row_bottom.addWidget(btn_folder)
        btn_row_bottom.addStretch()
        btn_row_bottom.addWidget(close_btn)
        dlg_vl.addLayout(btn_row_bottom)

        dlg.exec()

    def _reparaturauftrag_oeffnen(self):
        import os as _os
        _BASE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
        pdf = _os.path.join(_BASE, "Daten", "Bulmor schaden", "Reparaturauftrag Bulmor Vorlage.pdf")
        if not _os.path.isfile(pdf):
            QMessageBox.warning(self, "Datei nicht gefunden",
                                f"Reparaturauftrag-Vorlage nicht gefunden:\n{pdf}")
            return
        try:
            _os.startfile(pdf)
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Datei konnte nicht geöffnet werden:\n{e}")

    def _termin_erfassen(self, fid: int):
        dlg = _TerminDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            erstelle_termin(fid, d["datum"], d["titel"], d["typ"],
                            d["uhrzeit"], d["beschreibung"], d["kommentar"])
            self._zeige_fahrzeug(fid)

    # ── Refresh ────────────────────────────────────────────────────────────────

    def _on_main_tab_changed(self, idx: int):
        if idx == 1:  # Ausfälle-Tab
            self._refresh_ausfaelle()

    # ── Ausfälle-Tab ───────────────────────────────────────────────────────────

    def _build_ausfaelle_tab(self) -> QWidget:
        MONATE = ["Jan","Feb","Mär","Apr","Mai","Jun",
                  "Jul","Aug","Sep","Okt","Nov","Dez"]
        w = QWidget()
        w.setStyleSheet("background:white;")
        vl = QVBoxLayout(w)
        vl.setContentsMargins(16, 12, 16, 12)
        vl.setSpacing(8)

        # Titelzeile
        title_row = QHBoxLayout()
        title = QLabel("📊 Ausfallstatistik – Bulmor-Fahrzeuge")
        title.setFont(QFont("Arial", 15, QFont.Weight.Bold))
        title.setStyleSheet(f"color:{FIORI_TEXT};")
        title_row.addWidget(title)
        title_row.addStretch()

        btn_reload = QPushButton("🔄 Aktualisieren")
        btn_reload.setFixedHeight(34)
        btn_reload.setStyleSheet(_btn_style("#607d8b", "#546e7a"))
        btn_reload.clicked.connect(self._refresh_ausfaelle)
        title_row.addWidget(btn_reload)

        btn_excel = QPushButton("📤 Excel-Export")
        btn_excel.setFixedHeight(34)
        btn_excel.setStyleSheet(_btn_style("#107e3e", "#0a6630"))
        btn_excel.clicked.connect(self._export_ausfaelle_excel)
        title_row.addWidget(btn_excel)
        vl.addLayout(title_row)

        sub = QLabel("Bulmor-Fahrzeuge (5 gesamt)  •  Fahrbereit = 5 − Ausfälle  •  "
                     "basiert auf eingetragenen Statusänderungen; heute immer angezeigt")
        sub.setStyleSheet("color:#888; font-size:11px;")
        vl.addWidget(sub)

        # Filter-Bar
        fb = QFrame()
        fb.setStyleSheet("QFrame{background:#f5f6fa; border:1px solid #e0e4ec; border-radius:4px;}")
        fl = QHBoxLayout(fb)
        fl.setContentsMargins(10, 6, 10, 6)
        fl.setSpacing(8)
        _ls = "border:none; font-size:11px; color:#444;"

        fl.addWidget(self._lbl("Jahr:", _ls))
        self._af_jahr = QComboBox()
        self._af_jahr.setMinimumWidth(80)
        self._af_jahr.setStyleSheet(
            "QComboBox{background:white;border:1px solid #ccc;border-radius:3px;"
            "padding:2px 6px;font-size:11px;}")
        fl.addWidget(self._af_jahr)

        fl.addSpacing(8)
        fl.addWidget(self._lbl("Monat:", _ls))
        self._af_monat = QComboBox()
        self._af_monat.setMinimumWidth(110)
        self._af_monat.addItem("Alle", None)
        for i, mn in enumerate(MONATE, 1):
            self._af_monat.addItem(f"{i:02d} – {mn}", i)
        self._af_monat.setStyleSheet(
            "QComboBox{background:white;border:1px solid #ccc;border-radius:3px;"
            "padding:2px 6px;font-size:11px;}")
        fl.addWidget(self._af_monat)

        fl.addSpacing(8)
        fl.addWidget(self._lbl("Tag:", _ls))
        self._af_tag = QComboBox()
        self._af_tag.setMinimumWidth(70)
        self._af_tag.addItem("Alle", None)
        for d in range(1, 32):
            self._af_tag.addItem(f"{d:02d}", d)
        self._af_tag.setStyleSheet(
            "QComboBox{background:white;border:1px solid #ccc;border-radius:3px;"
            "padding:2px 6px;font-size:11px;}")
        fl.addWidget(self._af_tag)

        fl.addStretch()
        self._af_count_lbl = QLabel("")
        self._af_count_lbl.setStyleSheet("border:none; font-size:11px; color:#888;")
        fl.addWidget(self._af_count_lbl)
        vl.addWidget(fb)

        # Tabelle
        STATUS_COLS = ["fahrbereit", "defekt", "werkstatt", "ausser_dienst", "sonstiges"]
        self._af_status_cols = STATUS_COLS
        headers = ["Datum"] + [STATUS_META[s]["label"] for s in STATUS_COLS] + ["Notizen / Ausfallgrund"]

        self._af_table = QTableWidget()
        self._af_table.setColumnCount(len(headers))
        self._af_table.setHorizontalHeaderLabels(headers)
        self._af_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._af_table.setAlternatingRowColors(True)
        self._af_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._af_table.verticalHeader().setVisible(False)
        self._af_table.setStyleSheet("""
            QTableWidget { border: 1px solid #dce8f5; border-radius: 4px; }
            QTableWidget::item { padding: 4px 8px; }
            QHeaderView::section {
                background: #1565a8; color: white; font-weight: bold;
                font-size: 11px; padding: 6px 8px; border: none;
            }
            QTableWidget::item:alternate { background: #f0f5fc; }
            QTableWidget::item:selected  { background: #cfe3f8; color: #000; }
        """)
        hh = self._af_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self._af_table.setColumnWidth(0, 100)
        for i in range(1, len(headers) - 1):
            hh.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
            self._af_table.setColumnWidth(i, 115)
        # Notizen-Spalte dehnt sich
        hh.setSectionResizeMode(len(headers) - 1, QHeaderView.ResizeMode.Stretch)
        vl.addWidget(self._af_table, 1)

        # Legende
        leg_row = QHBoxLayout()
        for sk in STATUS_COLS:
            m = STATUS_META[sk]
            dot = QLabel("●")
            dot.setStyleSheet(f"color:{m['color']}; border:none; font-size:14px;")
            nm  = m["label"].split(" ", 1)[-1] if " " in m["label"] else m["label"]
            lbl = QLabel(nm)
            lbl.setStyleSheet(f"color:{m['color']}; border:none; font-size:11px; font-weight:bold;")
            leg_row.addWidget(dot)
            leg_row.addWidget(lbl)
            leg_row.addSpacing(14)
        leg_row.addStretch()
        vl.addLayout(leg_row)

        # Daten + Signale
        self._af_daten: list[dict] = []
        self._af_jahr.currentIndexChanged.connect(self._af_filter_apply)
        self._af_monat.currentIndexChanged.connect(self._af_filter_apply)
        self._af_tag.currentIndexChanged.connect(self._af_filter_apply)

        return w

    @staticmethod
    def _lbl(text: str, style: str) -> QLabel:
        l = QLabel(text)
        l.setStyleSheet(style)
        return l

    def _lade_statistik_daten(self) -> list[dict]:
        """Nur Bulmor-Fahrzeuge: Fahrbereit = BULMOR_GESAMT - Ausfälle.
        Ohne eingetragenen Status → fahrbereit. Heute immer angezeigt.
        Notizen = Kennzeichen + Grund aller nicht-fahrbereiten Fahrzeuge.
        """
        BULMOR_GESAMT = 5
        alle_fz = lade_alle_fahrzeuge()
        bulmors = [f for f in alle_fz
                   if "bulmor" in (f.get("typ") or "").lower()]
        # Fallback: alle nehmen wenn Typ nicht gesetzt
        if not bulmors:
            bulmors = alle_fz

        fids = [f["id"] for f in bulmors]
        kz_map = {f["id"]: (f.get("kennzeichen") or f.get("typ") or str(f["id"]))
                  for f in bulmors}

        alle_eintraege: dict[int, list[dict]] = {}
        for fid in fids:
            alle_eintraege[fid] = sorted(
                lade_status_historie(fid),
                key=lambda e: (e.get("von") or "", e.get("erstellt_am") or "")
            )

        # Frühestes Datum aller Einträge → lückenloser Bereich bis heute
        from datetime import date as _date, timedelta as _td
        fruehestes = None
        for eintraege in alle_eintraege.values():
            for e in eintraege:
                d = (e.get("von") or "")[:10]
                if len(d) == 10:
                    if fruehestes is None or d < fruehestes:
                        fruehestes = d

        heute = _date.today()
        start = _date.fromisoformat(fruehestes) if fruehestes else heute

        # Jeden Tag von start bis heute erzeugen
        alle_daten = []
        cur = start
        while cur <= heute:
            alle_daten.append(cur.isoformat())
            cur += _td(days=1)

        result = []
        for datum in sorted(alle_daten, reverse=True):
            counts = {s: 0 for s in STATUS_KEYS}
            notizen_liste = []
            for fid in fids:
                aktuell = None
                for e in sorted(alle_eintraege[fid],
                                key=lambda x: (x.get("von") or "",
                                               x.get("erstellt_am") or ""),
                                reverse=True):
                    if (e.get("von") or "")[:10] <= datum:
                        aktuell = e
                        break
                if aktuell:
                    s = aktuell.get("status") or "fahrbereit"
                    if s not in STATUS_KEYS:
                        s = "sonstiges"
                    counts[s] += 1
                    if s != "fahrbereit":
                        grund = (aktuell.get("grund") or "").strip()
                        kz = kz_map.get(fid, str(fid))
                        notizen_liste.append(f"{kz}: {grund}" if grund else kz)
                else:
                    # Kein Status eingetragen → fahrbereit
                    counts["fahrbereit"] += 1

            # Fehlende Bulmors (z.B. nicht in DB) → fahrbereit auffüllen
            eingetragen = sum(counts.values())
            if eingetragen < BULMOR_GESAMT:
                counts["fahrbereit"] += BULMOR_GESAMT - eingetragen

            result.append({
                "datum":   datum,
                **counts,
                "notizen": "  |  ".join(notizen_liste),
            })
        return result

    def _refresh_ausfaelle(self):
        self._af_daten = self._lade_statistik_daten()
        from datetime import date as _date
        aktuelles_jahr = str(_date.today().year)
        jahre = sorted({d["datum"][:4] for d in self._af_daten
                        if len(d.get("datum", "")) >= 4}, reverse=True)
        self._af_jahr.blockSignals(True)
        self._af_jahr.clear()
        self._af_jahr.addItem("Alle")
        for j in jahre:
            self._af_jahr.addItem(j)
        # Aktuelles Jahr vorauswählen
        idx = self._af_jahr.findText(aktuelles_jahr)
        if idx >= 0:
            self._af_jahr.setCurrentIndex(idx)
        self._af_jahr.blockSignals(False)
        self._af_filter_apply()

    def _af_filter_apply(self):
        j = self._af_jahr.currentText()
        m = self._af_monat.currentData()
        t = self._af_tag.currentData()
        gefiltert = []
        for row in self._af_daten:
            d = row.get("datum", "")
            if not d:
                continue
            if j != "Alle" and not d.startswith(j):
                continue
            if m is not None and (len(d) < 7 or int(d[5:7]) != m):
                continue
            if t is not None and (len(d) < 10 or int(d[8:10]) != t):
                continue
            gefiltert.append(row)
        self._af_fill_table(gefiltert)
        self._af_count_lbl.setText(
            f"{len(gefiltert)} / {len(self._af_daten)} Einträge")

    def _af_fill_table(self, daten: list[dict]):
        STATUS_COLS = self._af_status_cols
        self._af_table.setRowCount(len(daten))
        from datetime import date as _date
        heute = _date.today().isoformat()
        for r, row in enumerate(daten):
            datum = row["datum"]
            di = QTableWidgetItem(_fmt_date(datum))
            di.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if datum == heute:
                di.setFont(QFont("Arial", 10, QFont.Weight.Bold))
                di.setForeground(QColor(FIORI_BLUE))
            self._af_table.setItem(r, 0, di)

            for c, sk in enumerate(STATUS_COLS, 1):
                val = row.get(sk, 0)
                item = QTableWidgetItem(str(val) if val > 0 else "–")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if val > 0 and sk == "fahrbereit":
                    item.setForeground(QColor(STATUS_META["fahrbereit"]["color"]))
                    item.setFont(QFont("Arial", 10, QFont.Weight.Bold))
                elif val > 0:
                    item.setForeground(QColor(STATUS_META[sk]["color"]))
                    item.setFont(QFont("Arial", 10, QFont.Weight.Bold))
                self._af_table.setItem(r, c, item)

            notiz = row.get("notizen", "")
            n_item = QTableWidgetItem(notiz if notiz else "–")
            n_item.setTextAlignment(
                Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            if notiz:
                n_item.setForeground(QColor("#555555"))
            else:
                n_item.setForeground(QColor("#aaaaaa"))
            self._af_table.setItem(r, len(STATUS_COLS) + 1, n_item)

    def _export_ausfaelle_excel(self):
        # Daten ggf. nachladen
        if not self._af_daten:
            self._refresh_ausfaelle()
        if not self._af_daten:
            QMessageBox.information(
                self, "Keine Daten",
                "Es sind noch keine Statuseinträge vorhanden."
            )
            return

        # ── Zeitraum-Dialog ──────────────────────────────────────────────────
        from datetime import date as _date
        dlg = QDialog(self)
        dlg.setWindowTitle("Zeitraum wählen – Excel-Export")
        dlg.setMinimumWidth(320)
        dlg_vl = QVBoxLayout(dlg)
        dlg_vl.setContentsMargins(18, 14, 18, 14)
        dlg_vl.setSpacing(10)

        dlg_vl.addWidget(QLabel("Zeitraum für den Excel-Export:"))

        form = QFormLayout()
        form.setSpacing(8)

        von_edit = QDateEdit()
        von_edit.setCalendarPopup(True)
        von_edit.setDisplayFormat("dd.MM.yyyy")
        von_edit.setDate(QDate(_date.today().year, 1, 1))

        bis_edit = QDateEdit()
        bis_edit.setCalendarPopup(True)
        bis_edit.setDisplayFormat("dd.MM.yyyy")
        bis_edit.setDate(QDate.currentDate())

        form.addRow("Von:", von_edit)
        form.addRow("Bis:", bis_edit)
        dlg_vl.addLayout(form)

        btn_row = QHBoxLayout()
        ok_btn = QPushButton("Exportieren")
        ok_btn.setStyleSheet(_btn_style(FIORI_BLUE, "#0855a9"))
        ok_btn.setFixedHeight(36)
        ok_btn.clicked.connect(dlg.accept)
        ab_btn = QPushButton("Abbrechen")
        ab_btn.setFixedHeight(36)
        ab_btn.clicked.connect(dlg.reject)
        btn_row.addStretch()
        btn_row.addWidget(ab_btn)
        btn_row.addWidget(ok_btn)
        dlg_vl.addLayout(btn_row)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        qv = von_edit.date()
        qb = bis_edit.date()
        von_iso = f"{qv.year()}-{qv.month():02d}-{qv.day():02d}"
        bis_iso = f"{qb.year()}-{qb.month():02d}-{qb.day():02d}"

        gefiltert = [
            row for row in self._af_daten
            if von_iso <= row.get("datum", "") <= bis_iso
        ]

        if not gefiltert:
            QMessageBox.information(self, "Keine Daten",
                                    "Im gewählten Zeitraum sind keine Einträge vorhanden.")
            return

        # Speicherort wählen
        default_name = (f"Bulmor_Ausfallstatistik_"
                        f"{qv.day():02d}.{qv.month():02d}.{qv.year()}_"
                        f"bis_{qb.day():02d}.{qb.month():02d}.{qb.year()}.xlsx")
        pfad, _ = QFileDialog.getSaveFileName(
            self,
            "Speicherort wählen – Ausfallstatistik",
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel-Datei (*.xlsx)"
        )
        if not pfad:
            return

        try:
            import openpyxl
            from openpyxl.styles import (Font as XFont, PatternFill,
                                         Alignment, Border, Side)

            STATUS_COLS = self._af_status_cols
            col_headers = (["Datum"]
                           + [STATUS_META[s]["label"] for s in STATUS_COLS]
                           + ["Notizen / Ausfallgrund"])
            num_cols = len(col_headers)

            # ARGB-Farben (openpyxl erwartet 8-stellig)
            _HC_ARGB = {
                "fahrbereit":    "FF107E3E",
                "defekt":        "FFBB0000",
                "werkstatt":     "FFE67E22",
                "ausser_dienst": "FF7B1FA2",
                "sonstiges":     "FF5C6BC0",
            }
            _HDR_ARGB = {
                "fahrbereit":    "FF107E3E",
                "defekt":        "FFBB0000",
                "werkstatt":     "FFE67E22",
                "ausser_dienst": "FF7B1FA2",
                "sonstiges":     "FF5C6BC0",
                "datum":         "FF1565A8",
                "notizen":       "FF37474F",
            }

            def _col_letter(n: int) -> str:
                """1-basierte Spaltennummer → Buchstabe (A, B, …, Z, AA, …)"""
                result = ""
                while n > 0:
                    n, rem = divmod(n - 1, 26)
                    result = chr(65 + rem) + result
                return result

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ausfallstatistik"

            # Titelzeile
            last_col = _col_letter(num_cols)
            ws.merge_cells(f"A1:{last_col}1")
            tc = ws["A1"]
            tc.value = "Bulmor-Fahrzeuge – Ausfallstatistik"
            tc.font = XFont(bold=True, size=14, color="FF1565A8")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 26

            # Filterinfo-Zeile
            ws.merge_cells(f"A2:{last_col}2")
            filter_text = (f"Zeitraum: {qv.day():02d}.{qv.month():02d}.{qv.year()} "
                           f"– {qb.day():02d}.{qb.month():02d}.{qb.year()}"
                           f"  |  {len(gefiltert)} Einträge")
            fc = ws["A2"]
            fc.value = filter_text
            fc.font = XFont(italic=True, size=10, color="FF666666")
            fc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 16

            # Spaltenköpfe
            thin_w = Side(style="thin", color="FFDDDDDD")
            bd = Border(left=thin_w, right=thin_w, top=thin_w, bottom=thin_w)
            for ci, ch in enumerate(col_headers, 1):
                cell = ws.cell(row=3, column=ci, value=ch)
                cell.font = XFont(bold=True, color="FFFFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = bd
                if ci == 1:
                    argb = "FF1565A8"
                elif ci <= len(STATUS_COLS) + 1:
                    argb = _HDR_ARGB.get(STATUS_COLS[ci - 2], "FF607D8B")
                else:
                    argb = "FF37474F"
                cell.fill = PatternFill(fill_type="solid", fgColor=argb)
            ws.row_dimensions[3].height = 22

            # Datenzeilen
            alt_fill = PatternFill(fill_type="solid", fgColor="FFEEF4FB")
            # Maximale Inhaltsbreite je Spalte verfolgen (Startwert = Kopfbreite)
            col_max_len = {ci: len(ch) for ci, ch in enumerate(col_headers, 1)}
            for ri_offset, row in enumerate(gefiltert):
                ri = ri_offset + 4   # beginnt ab Zeile 4
                alt = (ri_offset % 2 == 1)

                # Datum
                datum_txt = _fmt_date(row["datum"])
                dc = ws.cell(row=ri, column=1, value=datum_txt)
                dc.alignment = Alignment(horizontal="center", vertical="center")
                dc.border = bd
                if alt:
                    dc.fill = alt_fill
                col_max_len[1] = max(col_max_len[1], len(datum_txt))

                # Status-Spalten
                for ci, sk in enumerate(STATUS_COLS, 2):
                    val = row.get(sk, 0)
                    sc = ws.cell(row=ri, column=ci,
                                 value=val if val > 0 else None)
                    sc.alignment = Alignment(horizontal="center",
                                             vertical="center")
                    sc.border = bd
                    if alt:
                        sc.fill = alt_fill
                    if val and val > 0:
                        sc.font = XFont(bold=True,
                                        color=_HC_ARGB.get(sk, "FF000000"))
                    col_max_len[ci] = max(col_max_len[ci], len(str(val)) if val else 0)

                # Notizen
                notiz_val = row.get("notizen") or ""
                nc = ws.cell(row=ri, column=num_cols,
                             value=notiz_val if notiz_val else "")
                nc.alignment = Alignment(horizontal="left", vertical="top",
                                         wrap_text=True)
                nc.border = bd
                if alt:
                    nc.fill = alt_fill
                col_max_len[num_cols] = max(col_max_len[num_cols], len(notiz_val))
                # Zeilenhoehe an Notiz-Laenge anpassen (Zeilenumbrueche im Text zaehlen)
                zeilen_anzahl = max(1, notiz_val.count("  |  ") + 1,
                                    notiz_val.count("\n") + 1)
                ws.row_dimensions[ri].height = max(16, 16 * zeilen_anzahl)

            # Spaltenbreiten – automatisch aus Inhaltslaenge
            NOTIZEN_MAX = 60   # Notizen-Spalte maximal 60 Zeichen breit
            for ci in range(1, num_cols + 1):
                ltr = _col_letter(ci)
                raw = col_max_len.get(ci, 8)
                if ci == num_cols:
                    # Notizen: auf sinnvollen Max-Wert begrenzen
                    w = min(raw + 4, NOTIZEN_MAX)
                else:
                    w = max(raw + 4, 10)  # mind. 10, etwas Puffer
                ws.column_dimensions[ltr].width = w

            # Autofilter
            ws.auto_filter.ref = f"A3:{last_col}3"

            wb.save(pfad)
            QMessageBox.information(
                self, "Export erfolgreich",
                f"✓ Ausfallstatistik gespeichert:\n{pfad}\n\n{len(gefiltert)} Zeilen exportiert."
            )
            os.startfile(pfad)
        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Fehler beim Excel-Export",
                f"Fehler:\n{e}\n\n{traceback.format_exc()}"
            )

    # ── Refresh ────────────────────────────────────────────────────────────────

    def refresh(self):
        self._refresh_liste()
        if self._aktives_fid:
            self._zeige_fahrzeug(self._aktives_fid)
