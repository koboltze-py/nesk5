"""
Bericht-Widget – kombinierter Excel-Export und E-Mail-Versand
Erstellt eine Excel-Datei mit bis zu 4 Abschnitten:
  • Verspätungen
  • Schulungen
  • Einsätze
  • Pat. auf Station
"""
from __future__ import annotations

import os
from datetime import date, datetime

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QGroupBox, QCheckBox, QDateEdit, QFormLayout, QFrame,
    QScrollArea, QSizePolicy, QFileDialog, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit,
    QTextEdit, QDialogButtonBox, QDialog,
)
from PySide6.QtCore import Qt, QDate, Signal
from PySide6.QtGui import QFont, QColor

from config import BASE_DIR, FIORI_BLUE, FIORI_TEXT

_BERICHT_DIR = os.path.join(BASE_DIR, "Daten", "Berichte")
_ABRECHNUNG_KOPIE_DIR = os.path.join(
    BASE_DIR, "Daten", "Berichte", "Abrechnung Sanmat Pat Einsätze"
)

# ─── Farben ───────────────────────────────────────────────────────────────────
_FARBEN = {
    "verspaetungen": ("#1565a8", "#e3f0ff"),
    "schulungen":    ("#6a1b9a", "#f3e5f5"),
    "einsaetze":     ("#2e7d32", "#e8f5e9"),
    "patienten":     ("#b71c1c", "#ffebee"),
}

# ─── Hilfs-Buttons ────────────────────────────────────────────────────────────
def _btn(text: str, color: str = FIORI_BLUE, hover: str = "#0057b8") -> QPushButton:
    b = QPushButton(text)
    b.setFixedHeight(34)
    b.setCursor(Qt.CursorShape.PointingHandCursor)
    b.setStyleSheet(
        f"QPushButton{{background:{color};color:#fff;border:none;"
        f"border-radius:4px;padding:4px 14px;font-size:12px;}}"
        f"QPushButton:hover{{background:{hover};}}"
        f"QPushButton:disabled{{background:#bbb;color:#888;}}"
    )
    return b


def _btn_light(text: str) -> QPushButton:
    b = QPushButton(text)
    b.setFixedHeight(34)
    b.setCursor(Qt.CursorShape.PointingHandCursor)
    b.setStyleSheet(
        "QPushButton{background:#eee;color:#333;border:1px solid #ccc;"
        "border-radius:4px;padding:4px 14px;font-size:12px;}"
        "QPushButton:hover{background:#ddd;}"
        "QPushButton:disabled{background:#f5f5f5;color:#aaa;}"
    )
    return b


# ─── Datum-Filter ─────────────────────────────────────────────────────────────
def _filter_nach_datum(eintraege: list[dict], von: date, bis: date) -> list[dict]:
    """Filtert Einträge mit datum='dd.MM.yyyy' auf den Bereich [von, bis]."""
    result = []
    for e in eintraege:
        try:
            d, m, y = e.get("datum", "").split(".")
            entry_date = date(int(y), int(m), int(d))
            if von <= entry_date <= bis:
                result.append(e)
        except Exception:
            pass
    return result


# ─── Abschnitts-Kontrollpanel ─────────────────────────────────────────────────
class _AbschnittPanel(QGroupBox):
    """Ein aufklappbares Panel für einen Bericht-Abschnitt."""

    def __init__(self, titel: str, farbe_ak: str, farbe_bg: str, parent=None):
        super().__init__(parent)
        self._farbe_ak = farbe_ak
        self._farbe_bg = farbe_bg
        self._build_ui(titel)

    def _build_ui(self, titel: str):
        self.setCheckable(True)
        self.setChecked(True)
        self.setTitle(f"  {titel}")
        self.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        self.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {self._farbe_ak};
                border-radius: 6px;
                margin-top: 10px;
                background: {self._farbe_bg};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 12px;
                padding: 2px 8px;
                color: {self._farbe_ak};
                font-weight: bold;
            }}
            QGroupBox:disabled {{ background: #f0f0f0; }}
        """)

        layout = QFormLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(12, 14, 12, 10)

        _field_style = (
            "QDateEdit{border:1px solid #ccc;border-radius:4px;"
            "padding:3px;font-size:12px;background:white;}"
        )

        heute = date.today()
        von_default = QDate(heute.year, heute.month, 1)
        bis_default = QDate.currentDate()

        self._von = QDateEdit()
        self._von.setCalendarPopup(True)
        self._von.setDisplayFormat("dd.MM.yyyy")
        self._von.setDate(von_default)
        self._von.setStyleSheet(_field_style)
        layout.addRow("Von:", self._von)

        self._bis = QDateEdit()
        self._bis.setCalendarPopup(True)
        self._bis.setDisplayFormat("dd.MM.yyyy")
        self._bis.setDate(bis_default)
        self._bis.setStyleSheet(_field_style)
        layout.addRow("Bis:", self._bis)

    def aktiv(self) -> bool:
        return self.isChecked()

    def von_datum(self) -> date:
        q = self._von.date()
        return date(q.year(), q.month(), q.day())

    def bis_datum(self) -> date:
        q = self._bis.date()
        return date(q.year(), q.month(), q.day())

    def zeitraum_label(self) -> str:
        return (f"{self._von.date().toString('dd.MM.yyyy')} – "
                f"{self._bis.date().toString('dd.MM.yyyy')}")


# ─── Schulungen-Panel (erweitert) ────────────────────────────────────────────
class _SchulungenPanel(_AbschnittPanel):
    """Abschnitts-Panel für Schulungen mit Typ-Auswahl und Zusatzfiltern."""

    filterGeaendert = Signal()

    def __init__(self, titel: str, farbe_ak: str, farbe_bg: str, parent=None):
        super().__init__(titel, farbe_ak, farbe_bg, parent)
        self._build_schulungen_filter()

    def _build_schulungen_filter(self):
        from functions.schulungen_db import SCHULUNGSTYPEN_CFG
        layout: QFormLayout = self.layout()
        _cb_style = "QCheckBox{font-size:11px;} QCheckBox::indicator{width:14px;height:14px;}"

        # ── Separator ────────────────────────────────────────────────────────
        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.HLine)
        sep1.setStyleSheet("color:#ccc;margin:2px 0;")
        layout.addRow(sep1)

        # ── Filter-Optionen ──────────────────────────────────────────────────
        self._cb_abgelaufene = QCheckBox("Abgelaufene ausschließen")
        self._cb_abgelaufene.setStyleSheet(_cb_style)
        self._cb_abgelaufene.setChecked(False)
        self._cb_abgelaufene.stateChanged.connect(self.filterGeaendert)
        layout.addRow(self._cb_abgelaufene)

        self._cb_bald_faellig = QCheckBox("Nur fällig in < 6 Monaten")
        self._cb_bald_faellig.setStyleSheet(_cb_style)
        self._cb_bald_faellig.setChecked(False)
        self._cb_bald_faellig.stateChanged.connect(self.filterGeaendert)
        layout.addRow(self._cb_bald_faellig)

        # ── Separator ────────────────────────────────────────────────────────
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet("color:#ccc;margin:2px 0;")
        layout.addRow(sep2)

        lbl_typen = QLabel("Schulungsarten:")
        lbl_typen.setStyleSheet("font-size:10px;font-weight:bold;color:#555;")
        layout.addRow(lbl_typen)

        self._typ_cbs: dict[str, QCheckBox] = {}
        for key, cfg in SCHULUNGSTYPEN_CFG.items():
            if cfg.get("laeuft_nicht_ab", False):
                continue  # nur Typen mit Ablaufdatum anzeigen
            cb = QCheckBox(cfg["anzeige"])
            cb.setStyleSheet(_cb_style)
            cb.setChecked(True)
            cb.stateChanged.connect(self.filterGeaendert)
            self._typ_cbs[key] = cb
            layout.addRow(cb)

    def get_aktive_typen(self) -> list[str] | None:
        """Gibt die Liste aktiver Typ-Keys zurück; None bedeutet alle Typen."""
        aktive = [k for k, cb in self._typ_cbs.items() if cb.isChecked()]
        if len(aktive) == len(self._typ_cbs):
            return None  # alle aktiv → kein Filter notwendig
        return aktive or None

    def abgelaufene_ausschliessen(self) -> bool:
        return self._cb_abgelaufene.isChecked()

    def nur_bald_faellig(self) -> bool:
        return self._cb_bald_faellig.isChecked()


# ─── Vorschau-Widget ──────────────────────────────────────────────────────────
class _Vorschau(QWidget):
    """Rechte Seite: Zeigt Zusammenfassung der geplanten Abschnitte."""

    def __init__(self, parent=None):
        super().__init__(parent)
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(8)

        titel = QLabel("📋  Vorschau")
        titel.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        titel.setStyleSheet(f"color:{FIORI_TEXT};")
        v.addWidget(titel)

        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("color:#ddd;")
        v.addWidget(line)

        self._tbl = QTableWidget()
        self._tbl.setColumnCount(4)
        self._tbl.setHorizontalHeaderLabels(["Abschnitt", "Zeitraum", "Einträge", "Status"])
        self._tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        hh = self._tbl.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setStyleSheet("font-size:12px;")
        v.addWidget(self._tbl)

        self._info_lbl = QLabel()
        self._info_lbl.setWordWrap(True)
        self._info_lbl.setStyleSheet(
            "background:#e3f2fd;color:#154360;border-radius:4px;"
            "padding:8px 12px;font-size:11px;"
        )
        v.addWidget(self._info_lbl)

    def aktualisieren(self, abschnitte: list[dict]):
        """abschnitte: [{"name", "zeitraum", "count", "aktiv", "farbe"}]"""
        self._tbl.setRowCount(len(abschnitte))
        gesamt = 0
        aktiv_count = 0
        for row, a in enumerate(abschnitte):
            items = [
                QTableWidgetItem(a["name"]),
                QTableWidgetItem(a["zeitraum"] if a["aktiv"] else "–"),
                QTableWidgetItem(str(a["count"]) if a["aktiv"] else "–"),
                QTableWidgetItem("✅ Aktiv" if a["aktiv"] else "⬜ Deaktiviert"),
            ]
            farbe = QColor(a.get("farbe_bg", "#ffffff"))
            for col, item in enumerate(items):
                item.setBackground(farbe if a["aktiv"] else QColor("#f5f5f5"))
                self._tbl.setItem(row, col, item)
            if a["aktiv"]:
                gesamt += a["count"]
                aktiv_count += 1

        if aktiv_count == 0:
            self._info_lbl.setText("⚠  Kein Abschnitt aktiviert. Bitte mindestens einen Abschnitt auswählen.")
            self._info_lbl.setStyleSheet(
                "background:#fff3e0;color:#6d3b00;border-radius:4px;padding:8px 12px;font-size:11px;"
            )
        else:
            self._info_lbl.setText(
                f"📊  {aktiv_count} Abschnitt(e) aktiv  ·  {gesamt} Einträge gesamt  "
                f"·  Excel wird mit {aktiv_count} Tabellenblatt(-blättern) erstellt."
            )
            self._info_lbl.setStyleSheet(
                "background:#e3f2fd;color:#154360;border-radius:4px;padding:8px 12px;font-size:11px;"
            )


# ─── Mail-Dialog ──────────────────────────────────────────────────────────────
class _BerichtMailDialog(QDialog):
    """Einfacher Mail-Dialog für den Bericht-Export."""

    _FIELD = (
        "QLineEdit,QTextEdit{border:1px solid #ccc;border-radius:4px;"
        "padding:4px;font-size:12px;background:white;}"
    )

    def __init__(self, excel_pfad: str, abschnitte_namen: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("📧  Bericht per E-Mail senden")
        self.setMinimumWidth(500)
        self.resize(550, 420)
        layout = QVBoxLayout(self)
        fl = QFormLayout()
        fl.setSpacing(8)

        self._empfaenger = QLineEdit("erste-hilfe-station-flughafen@drk-koeln.de")
        self._empfaenger.setStyleSheet(self._FIELD)
        fl.addRow("Empfänger:", self._empfaenger)

        heute = datetime.now().strftime("%d.%m.%Y")
        abschnitte_txt = ", ".join(abschnitte_namen)
        self._betreff = QLineEdit(f"DRK FKB – Bericht {heute}  ({abschnitte_txt})")
        self._betreff.setStyleSheet(self._FIELD)
        fl.addRow("Betreff:", self._betreff)

        layout.addLayout(fl)
        layout.addWidget(QLabel("Nachrichtentext:"))

        self._body = QTextEdit()
        self._body.setStyleSheet(self._FIELD)
        self._body.setPlainText(
            f"Hallo,\n\n"
            f"anbei der aktuelle Bericht der DRK-Station FKB ({abschnitte_txt}).\n\n"
            f"Der Bericht ist als Excel-Datei angehängt.\n\n"
            f"Mit freundlichen Grüßen\n"
            f"DRK-Kreisverband Köln e.V."
        )
        self._body.setMinimumHeight(120)
        layout.addWidget(self._body, 1)

        anhang_lbl = QLabel(f"📎  Anhang: {os.path.basename(excel_pfad)}")
        anhang_lbl.setStyleSheet(
            "background:#e3f2fd;color:#154360;border-radius:4px;padding:6px 10px;font-size:11px;"
        )
        layout.addWidget(anhang_lbl)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("📧  Outlook-Entwurf öffnen")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_daten(self) -> tuple[str, str, str]:
        return (
            self._empfaenger.text().strip(),
            self._betreff.text().strip(),
            self._body.toPlainText().strip(),
        )


# ─── Excel-Erstellung ─────────────────────────────────────────────────────────
def _erstelle_bericht_excel(abschnitte: list[dict], ziel_pfad: str) -> None:
    """
    Schreibt bis zu 4 Abschnitte als Tabellenblätter in eine Excel-Datei.
    abschnitte: [{"typ": str, "titel": str, "eintraege": list, "zeitraum": str}]
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
    thin      = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = openpyxl.Workbook()
    # Erstes leeres Blatt entfernen
    ws_dummy = wb.active

    # ── Datum-Parser: DD.MM.YYYY → datetime.date (echte Excel-Datumswerte) ──
    import re as _re_b; from datetime import date as _excel_d_b
    def _d(s):
        if not s: return s
        m = _re_b.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", str(s).strip())
        if m:
            try: return _excel_d_b(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError: pass
        return s

    def _hdr_fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    def _data_fill(hex_color: str, alpha: str = "22"):
        """Helles Tabellenblatt-Hintergrundfarbe aus der Abschnittsfarbe."""
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    def _schreibe_sheet(ws, spalten: list[tuple[str, int]], zeilen: list[list],
                        hdr_color: str, zebra1: str, zebra2: str,
                        titel_zeile: str | None = None, zeitraum: str = ""):
        """Schreibt Kopfzeile + Daten in ws."""
        row_start = 1
        if titel_zeile:
            ws.merge_cells(f"A1:{get_column_letter(len(spalten))}1")
            c = ws.cell(row=1, column=1, value=titel_zeile)
            c.font = Font(bold=True, size=13, color="FFFFFF")
            c.fill = _hdr_fill(hdr_color)
            c.alignment = center
            ws.row_dimensions[1].height = 24
            row_start = 2
            if zeitraum:
                ws.merge_cells(f"A2:{get_column_letter(len(spalten))}2")
                s = ws.cell(row=2, column=1, value=f"Zeitraum: {zeitraum}    |    Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
                s.font = Font(italic=True, size=9, color="555555")
                s.alignment = center
                ws.row_dimensions[2].height = 14
                row_start = 3

        hdr_row = row_start
        for col_idx, (name, breite) in enumerate(spalten, 1):
            c = ws.cell(row=hdr_row, column=col_idx, value=name)
            c.font = hdr_font
            c.fill = _hdr_fill(hdr_color)
            c.alignment = center
            c.border = thin
            ws.column_dimensions[get_column_letter(col_idx)].width = breite
        ws.row_dimensions[hdr_row].height = 22
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(spalten))}{hdr_row}"

        f1 = PatternFill("solid", fgColor=zebra1)
        f2 = PatternFill("solid", fgColor=zebra2)
        for r_offset, zeile in enumerate(zeilen):
            row_num = hdr_row + 1 + r_offset
            fill = f1 if r_offset % 2 == 0 else f2
            for col_idx, wert in enumerate(zeile, 1):
                c = ws.cell(row=row_num, column=col_idx, value=wert)
                if isinstance(wert, _excel_d_b):
                    c.number_format = "DD.MM.YYYY"
                c.fill = fill
                c.border = thin
                c.alignment = center if col_idx <= 3 else left
            ws.row_dimensions[row_num].height = 17

        # Frieren ab Datenzeile
        ws.freeze_panes = f"A{hdr_row + 1}"

    for ab in abschnitte:
        typ      = ab["typ"]
        eintr    = ab["eintraege"]
        zeitraum = ab.get("zeitraum", "")
        n        = len(eintr)

        # ── Verspätungen ────────────────────────────────────────────────────
        if typ == "verspaetungen":
            ws = wb.create_sheet("Verspätungen")
            spalten = [
                ("Datum", 13), ("Mitarbeiter", 26), ("Dienst", 10),
                ("Dienstbeginn", 13), ("Dienstantritt", 14), ("Verspätung", 14),
                ("Aufgen. von", 18), ("Begründung", 38),
            ]
            zeilen = []
            for e in eintr:
                vmin = e.get("verspaetung_min") or 0
                versp = f"{vmin} Min. zu spät" if vmin > 0 else (f"{abs(vmin)} Min. früh" if vmin < 0 else "Pünktlich")
                zeilen.append([
                    _d(e.get("datum", "")), e.get("mitarbeiter", ""), e.get("dienst", ""),
                    e.get("dienstbeginn", ""), e.get("dienstantritt", ""), versp,
                    e.get("aufgenommen_von", ""), e.get("begruendung", ""),
                ])
            _schreibe_sheet(ws, spalten, zeilen, "1565A8", "FFFFFF", "F5F5F5",
                            f"⏰  Verspätungen ({n} Einträge)", zeitraum)

        # ── Schulungen (je Tab: ZÜP / EH / Refresher / Sonstiges) ──────────
        elif typ == "schulungen":
            from functions.schulungen_db import SCHULUNGSTYPEN_CFG
            sub_typen = ab.get("sub_typen", {})
            heute      = date.today()
            _fill_gelb = PatternFill("solid", fgColor="FFFF00")
            _fill_rot  = PatternFill("solid", fgColor="FFCDD2")
            _fill_warn = PatternFill("solid", fgColor="FFCC80")
            _DS = 4  # Datenzeilen starten ab Zeile 4 (1=Titel,2=Zeitraum,3=Header)

            def _schreibe_schulungs_sheet(
                ws_, eintr_, sheet_titel, sheet_zeitraum, gelb_tage, hdr_color, zeige_3mon=False
            ):
                if zeige_3mon:
                    sp_ = [
                        ("Mitarbeiter", 28), ("Gültig bis", 13),
                        ("Tage Rest", 11), ("< 3 Monate?", 13), ("Status", 12),
                    ]
                else:
                    sp_ = [
                        ("Mitarbeiter", 26), ("Schulungsart", 22), ("Gültig bis", 13),
                        ("Tage Rest", 11), ("Status", 12),
                    ]
                ze_ = []
                tv_ = []
                for e_ in eintr_:
                    cfg_ = SCHULUNGSTYPEN_CFG.get(e_.get("schulungstyp", ""), {})
                    az_  = cfg_.get("anzeige", e_.get("schulungstyp", ""))
                    nm_  = f"{e_.get('nachname', '')} {e_.get('vorname', '')}".strip()
                    gb_  = e_.get("gueltig_bis", "")
                    tv_val = None
                    try:
                        _d, _m, _y = gb_.split(".")
                        gb_d = date(int(_y), int(_m), int(_d))
                        tv_val = (gb_d - heute).days
                        tt_ = f"{tv_val}" if tv_val >= 0 else f"ÜBERFÄLLIG {-tv_val}d"
                    except Exception:
                        tt_ = ""
                    tv_.append(tv_val)
                    if zeige_3mon:
                        warn_ = "⚠  Ja" if tv_val is not None and tv_val <= 90 else "✓  Nein"
                        ze_.append([nm_, gb_, tt_, warn_, e_.get("status", "gültig")])
                    else:
                        ze_.append([nm_, az_, gb_, tt_, e_.get("status", "gültig")])
                _schreibe_sheet(ws_, sp_, ze_, hdr_color, "FFFFFF", "F5F5F5",
                                sheet_titel, sheet_zeitraum)
                tage_col_ = 3 if zeige_3mon else 4
                for ro_, t_ in enumerate(tv_):
                    if t_ is not None and t_ < gelb_tage:
                        ws_.cell(row=_DS + ro_, column=tage_col_).fill = (
                            _fill_rot if t_ < 0 else _fill_gelb
                        )
                if zeige_3mon:
                    for ro_, t_ in enumerate(tv_):
                        if t_ is not None and 0 <= t_ <= 90:
                            ws_.cell(row=_DS + ro_, column=4).fill = _fill_warn

            if "ZÜP" in sub_typen:
                ws_z = wb.create_sheet("ZÜP")
                _schreibe_schulungs_sheet(
                    ws_z, sub_typen["ZÜP"],
                    f"🔐  ZÜP ({len(sub_typen['ZÜP'])} Einträge)",
                    "Gesamter Erfassungszeitraum",
                    gelb_tage=100, hdr_color="4A148C", zeige_3mon=True,
                )
            if "EH" in sub_typen:
                ws_e = wb.create_sheet("EH")
                _schreibe_schulungs_sheet(
                    ws_e, sub_typen["EH"],
                    f"🏥  Erste Hilfe ({len(sub_typen['EH'])} Einträge)",
                    zeitraum, gelb_tage=48, hdr_color="6A1B9A",
                )
            if "Refresher" in sub_typen:
                ws_r = wb.create_sheet("Refresher")
                _schreibe_schulungs_sheet(
                    ws_r, sub_typen["Refresher"],
                    f"🔄  Refresher ({len(sub_typen['Refresher'])} Einträge)",
                    zeitraum, gelb_tage=48, hdr_color="7B1FA2",
                )
            if "Sonstiges" in sub_typen:
                ws_s = wb.create_sheet("Sonstiges")
                _schreibe_schulungs_sheet(
                    ws_s, sub_typen["Sonstiges"],
                    f"📊  Sonstiges ({len(sub_typen['Sonstiges'])} Einträge)",
                    zeitraum, gelb_tage=48, hdr_color="9C27B0",
                )
            if not sub_typen:  # Fallback (Altdaten ohne sub_typen)
                ws = wb.create_sheet("Schulungen")
                _schreibe_schulungs_sheet(
                    ws, eintr, f"🎓  Schulungen ({n} Einträge)",
                    zeitraum, gelb_tage=48, hdr_color="6A1B9A",
                )

        # ── Einsätze ────────────────────────────────────────────────────────
        elif typ == "einsaetze":
            ws = wb.create_sheet("Einsätze")
            spalten = [
                ("Nr.", 5), ("Datum", 11), ("Uhrzeit", 9), ("Dauer\n(Min.)", 10),
                ("Stichwort", 20), ("Ort", 20), ("DRK-Nr.", 12),
                ("MA 1", 18), ("MA 2", 18), ("Angenom.", 10), ("Bemerkung", 30),
            ]
            zeilen = []
            for idx, e in enumerate(eintr, 1):
                zeilen.append([
                    idx, _d(e.get("datum", "")), e.get("uhrzeit", ""),
                    e.get("einsatzdauer", "") or "",
                    e.get("einsatzstichwort", ""), e.get("einsatzort", ""),
                    e.get("einsatznr_drk", ""), e.get("drk_ma1", ""), e.get("drk_ma2", ""),
                    "Ja" if e.get("angenommen", 1) else "Nein",
                    e.get("bemerkung", ""),
                ])
            _schreibe_sheet(ws, spalten, zeilen, "2E7D32", "FFFFFF", "F5F5F5",
                            f"🚑  Einsätze ({n} Einträge)", zeitraum)

        # ── Patienten ───────────────────────────────────────────────────────
        elif typ == "patienten":
            ws = wb.create_sheet("Pat. Station")
            spalten = [
                ("Nr.", 5), ("Datum", 11), ("Uhrzeit", 9), ("Dauer\n(Min.)", 10),
                ("Typ", 14), ("Alter", 7), ("Beschwerde", 22), ("Ort", 18),
                ("Maßnahmen", 22), ("DRK MA 1", 16), ("Weitergeleitet", 14), ("BG-Fall", 8),
            ]
            zeilen = []
            for idx, e in enumerate(eintr, 1):
                zeilen.append([
                    idx, _d(e.get("datum", "")), e.get("uhrzeit", ""),
                    e.get("behandlungsdauer", "") or "",
                    e.get("patient_typ", "") or "",
                    e.get("alter", "") or "",
                    e.get("beschwerde_art", "") or "",
                    e.get("unfall_ort", "") or "",
                    e.get("massnahmen", "") or "",
                    e.get("drk_ma1", "") or "",
                    e.get("weitergeleitet", "") or "",
                    "Ja" if e.get("arbeitsunfall") else "Nein",
                ])
            _schreibe_sheet(ws, spalten, zeilen, "B71C1C", "FFFFFF", "F5F5F5",
                            f"🏥  Pat. auf Station ({n} Einträge)", zeitraum)

    # Erstes Dummy-Blatt entfernen (nur wenn mindestens ein echter Sheet erstellt)
    if len(wb.worksheets) > 1:
        wb.remove(ws_dummy)
    else:
        # Fallback: kein Abschnitt hatte Daten – trotzdem speichern
        ws_dummy.title = "Bericht"
        ws_dummy.cell(row=1, column=1, value="Keine Daten im gewählten Zeitraum.")

    os.makedirs(os.path.dirname(ziel_pfad), exist_ok=True)
    wb.save(ziel_pfad)


# ─── Haupt-Widget ─────────────────────────────────────────────────────────────
class BerichtWidget(QWidget):
    """Bericht-Seite: kombinierter Export aus mehreren Modulen."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self._aktualisiere_vorschau()

    # ── UI aufbauen ───────────────────────────────────────────────────────────
    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Titelleiste
        titelbar = QFrame()
        titelbar.setStyleSheet("background:#1565a8;")
        titelbar.setFixedHeight(52)
        tbl = QHBoxLayout(titelbar)
        tbl.setContentsMargins(20, 0, 20, 0)
        lbl = QLabel("📊  Bericht erstellen")
        lbl.setFont(QFont("Arial", 15, QFont.Weight.Bold))
        lbl.setStyleSheet("color:white;")
        tbl.addWidget(lbl)
        tbl.addStretch()
        outer.addWidget(titelbar)

        # Haupt-Bereich: Links Kontrollen, Rechts Vorschau
        haupt = QHBoxLayout()
        haupt.setContentsMargins(16, 12, 16, 12)
        haupt.setSpacing(16)

        # ── Linke Spalte: Abschnitt-Panels ──────────────────────────────────
        linke_scroll = QScrollArea()
        linke_scroll.setWidgetResizable(True)
        linke_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        linke_scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        linke_scroll.setFixedWidth(310)

        linke_seite = QWidget()
        linke_layout = QVBoxLayout(linke_seite)
        linke_layout.setContentsMargins(0, 0, 4, 0)
        linke_layout.setSpacing(10)

        linke_lbl = QLabel("Abschnitte auswählen:")
        linke_lbl.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        linke_lbl.setStyleSheet(f"color:{FIORI_TEXT};")
        linke_layout.addWidget(linke_lbl)

        c = _FARBEN
        self._panel_versp = _AbschnittPanel("⏰  Verspätungen",    c["verspaetungen"][0], c["verspaetungen"][1])
        self._panel_schu  = _SchulungenPanel("🎓  Schulungen",       c["schulungen"][0],    c["schulungen"][1])
        self._panel_eins  = _AbschnittPanel("🚑  Einsätze",         c["einsaetze"][0],     c["einsaetze"][1])
        self._panel_pat   = _AbschnittPanel("🏥  Pat. auf Station", c["patienten"][0],     c["patienten"][1])

        for panel in (self._panel_versp, self._panel_schu, self._panel_eins, self._panel_pat):
            panel.toggled.connect(self._aktualisiere_vorschau)
            panel._von.dateChanged.connect(self._aktualisiere_vorschau)
            panel._bis.dateChanged.connect(self._aktualisiere_vorschau)
            linke_layout.addWidget(panel)
        self._panel_schu.filterGeaendert.connect(self._aktualisiere_vorschau)

        linke_layout.addStretch()
        linke_scroll.setWidget(linke_seite)
        haupt.addWidget(linke_scroll)

        # ── Rechte Spalte: Vorschau ──────────────────────────────────────────
        rechte_seite = QWidget()
        rechte_layout = QVBoxLayout(rechte_seite)
        rechte_layout.setContentsMargins(0, 0, 0, 0)
        rechte_layout.setSpacing(10)

        self._vorschau = _Vorschau()
        rechte_layout.addWidget(self._vorschau, 1)

        # Aktions-Buttons
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        self._btn_vorschau = _btn_light("🔄  Vorschau aktualisieren")
        self._btn_vorschau.clicked.connect(self._aktualisiere_vorschau)
        btn_row.addWidget(self._btn_vorschau)

        btn_row.addStretch()

        self._btn_export = _btn("📥  Excel exportieren", "#1565a8", "#0d47a1")
        self._btn_export.setToolTip("Excel-Datei mit allen aktiven Abschnitten speichern")
        self._btn_export.clicked.connect(self._excel_exportieren)
        btn_row.addWidget(self._btn_export)

        self._btn_mail = _btn("📧  Per E-Mail senden", "#6a1b9a", "#4a148c")
        self._btn_mail.setToolTip("Excel erstellen und als Outlook-Entwurf versenden")
        self._btn_mail.clicked.connect(self._per_mail_senden)
        btn_row.addWidget(self._btn_mail)

        rechte_layout.addLayout(btn_row)

        # ── Trennlinie + Abrechnung-Button ───────────────────────────────────
        abrech_sep = QFrame()
        abrech_sep.setFrameShape(QFrame.Shape.HLine)
        abrech_sep.setStyleSheet("color:#ddd;margin:6px 0;")
        rechte_layout.addWidget(abrech_sep)

        abrech_lbl = QLabel("Abrechnung (Pat. auf Station · Einsätze · Sanmaterial):")
        abrech_lbl.setStyleSheet(f"color:{FIORI_TEXT};font-size:11px;font-weight:bold;")
        rechte_layout.addWidget(abrech_lbl)

        abrech_row = QHBoxLayout()
        abrech_row.setSpacing(8)
        abrech_row.addStretch()

        self._btn_abrechnung = _btn(
            "📋  Abrechnung erstellen", "#E65100", "#BF360C"
        )
        self._btn_abrechnung.setToolTip(
            "Excel mit 3 Tabs (Pat., Einsätze, Sanmat) erstellen und optional per Mail versenden"
        )
        self._btn_abrechnung.clicked.connect(self._abrechnung_erstellen)
        abrech_row.addWidget(self._btn_abrechnung)

        rechte_layout.addLayout(abrech_row)

        haupt.addWidget(rechte_seite, 1)
        outer.addLayout(haupt, 1)

    # ── Daten laden ───────────────────────────────────────────────────────────
    def _lade_abschnitte(self) -> list[dict]:
        """Lädt Daten für alle aktiven Abschnitte und gibt sie zurück."""
        abschnitte = []

        if self._panel_versp.aktiv():
            try:
                from functions.verspaetung_db import lade_verspaetungen
                alle = lade_verspaetungen()
                gefiltert = _filter_nach_datum(alle, self._panel_versp.von_datum(), self._panel_versp.bis_datum())
            except Exception:
                gefiltert = []
            abschnitte.append({
                "typ": "verspaetungen", "titel": "Verspätungen",
                "eintraege": gefiltert, "zeitraum": self._panel_versp.zeitraum_label(),
                "farbe_ak": _FARBEN["verspaetungen"][0], "farbe_bg": _FARBEN["verspaetungen"][1],
            })

        if self._panel_schu.aktiv():
            try:
                from functions.schulungen_db import lade_eintraege_fuer_export
                aktive_typen = self._panel_schu.get_aktive_typen()
                _VON  = self._panel_schu.von_datum()
                _BIS  = self._panel_schu.bis_datum()
                _ALLE = aktive_typen is None

                def _lade(typen: list[str], von, bis) -> list[dict]:
                    try:
                        res = lade_eintraege_fuer_export(von, bis, schulungstypen=typen)
                        if self._panel_schu.abgelaufene_ausschliessen():
                            res = [e for e in res if e.get("_tage_rest", -1) >= 0]
                        if self._panel_schu.nur_bald_faellig():
                            res = [e for e in res if 0 <= e.get("_tage_rest", 999) <= 180]
                        return res
                    except Exception:
                        return []

                # ZÜP: immer vollständiger Zeitraum (von Anfang der Erfassung)
                sub_typen: dict[str, list[dict]] = {}
                _FRUEH = date(2000, 1, 1)
                _SPAET = date(2099, 12, 31)
                if _ALLE or "ZÜP" in aktive_typen:
                    sub_typen["ZÜP"] = _lade(["ZÜP"], _FRUEH, _SPAET)
                for key in ("EH", "Refresher", "Sonstiges"):
                    if _ALLE or key in aktive_typen:
                        sub_typen[key] = _lade([key], _VON, _BIS)

                gefiltert = [e for lst in sub_typen.values() for e in lst]
            except Exception:
                gefiltert = []
                sub_typen = {}
            abschnitte.append({
                "typ": "schulungen", "titel": "Schulungen",
                "eintraege": gefiltert, "zeitraum": self._panel_schu.zeitraum_label(),
                "farbe_ak": _FARBEN["schulungen"][0], "farbe_bg": _FARBEN["schulungen"][1],
                "sub_typen": sub_typen,
            })

        if self._panel_eins.aktiv():
            try:
                from gui.dienstliches import lade_einsaetze
                alle = lade_einsaetze()
                gefiltert = _filter_nach_datum(alle, self._panel_eins.von_datum(), self._panel_eins.bis_datum())
            except Exception:
                gefiltert = []
            abschnitte.append({
                "typ": "einsaetze", "titel": "Einsätze",
                "eintraege": gefiltert, "zeitraum": self._panel_eins.zeitraum_label(),
                "farbe_ak": _FARBEN["einsaetze"][0], "farbe_bg": _FARBEN["einsaetze"][1],
            })

        if self._panel_pat.aktiv():
            try:
                from gui.dienstliches import lade_patienten
                alle = lade_patienten()
                gefiltert = _filter_nach_datum(alle, self._panel_pat.von_datum(), self._panel_pat.bis_datum())
            except Exception:
                gefiltert = []
            abschnitte.append({
                "typ": "patienten", "titel": "Pat. auf Station",
                "eintraege": gefiltert, "zeitraum": self._panel_pat.zeitraum_label(),
                "farbe_ak": _FARBEN["patienten"][0], "farbe_bg": _FARBEN["patienten"][1],
            })

        return abschnitte

    # ── Vorschau aktualisieren ────────────────────────────────────────────────
    def _aktualisiere_vorschau(self):
        from gui.splash_screen import _mit_ladeanimation
        abschnitte, exc = _mit_ladeanimation(
            self, "Berichtdaten werden geladen …", self._lade_abschnitte
        )
        if exc is not None or abschnitte is None:
            abschnitte = []
        vorschau_daten = [
            {
                "name": a["titel"],
                "zeitraum": a["zeitraum"],
                "count": len(a["eintraege"]),
                "aktiv": True,
                "farbe_bg": a["farbe_bg"],
            }
            for a in abschnitte
        ]
        # Deaktivierte Abschnitte ebenfalls anzeigen
        alle_typen = [
            ("verspaetungen", "Verspätungen",    self._panel_versp),
            ("schulungen",    "Schulungen",       self._panel_schu),
            ("einsaetze",     "Einsätze",         self._panel_eins),
            ("patienten",     "Pat. auf Station", self._panel_pat),
        ]
        for typ, titel, panel in alle_typen:
            if not panel.aktiv():
                vorschau_daten.append({
                    "name": titel, "zeitraum": "–", "count": 0,
                    "aktiv": False, "farbe_bg": "#f5f5f5",
                })
        # Sortieren nach ursprünglicher Reihenfolge
        reihenfolge = ["verspaetungen", "schulungen", "einsaetze", "patienten"]
        typ_zu_titel = {
            "verspaetungen": "Verspätungen", "schulungen": "Schulungen",
            "einsaetze": "Einsätze", "patienten": "Pat. auf Station"
        }
        vorschau_daten.sort(key=lambda x: reihenfolge.index(
            next((k for k, v in typ_zu_titel.items() if v == x["name"]), "verspaetungen")
        ))
        self._vorschau.aktualisieren(vorschau_daten)

    # ── Excel erstellen ───────────────────────────────────────────────────────
    def _baue_ziel_pfad(self) -> str:
        heute = datetime.now().strftime("%Y-%m-%d")
        os.makedirs(_BERICHT_DIR, exist_ok=True)
        return os.path.join(_BERICHT_DIR, f"Bericht_{heute}.xlsx")

    def _excel_exportieren(self, ask_path: bool = True) -> str | None:
        from gui.splash_screen import _mit_ladeanimation
        abschnitte, exc = _mit_ladeanimation(
            self, "Berichtdaten werden geladen …", self._lade_abschnitte
        )
        if exc is not None or abschnitte is None:
            abschnitte = []
        if not abschnitte:
            QMessageBox.warning(self, "Keine Daten", "Bitte mindestens einen Abschnitt aktivieren.")
            return None

        ziel = self._baue_ziel_pfad()
        if ask_path:
            ziel, _ = QFileDialog.getSaveFileName(
                self, "Bericht speichern", ziel, "Excel-Dateien (*.xlsx)"
            )
            if not ziel:
                return None

        try:
            _erstelle_bericht_excel(abschnitte, ziel)
        except ImportError as e:
            QMessageBox.critical(self, "Modul fehlt", str(e))
            return None
        except Exception as exc:
            QMessageBox.critical(self, "Export-Fehler", f"Fehler beim Exportieren:\n{exc}")
            return None

        antwort = QMessageBox.question(
            self, "Bericht gespeichert",
            f"Bericht wurde gespeichert:\n\n{ziel}\n\nDatei jetzt öffnen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort == QMessageBox.StandardButton.Yes:
            os.startfile(ziel)

        return ziel

    # ── Per E-Mail senden ─────────────────────────────────────────────────────
    def _per_mail_senden(self):
        from gui.splash_screen import _mit_ladeanimation
        abschnitte, exc = _mit_ladeanimation(
            self, "Berichtdaten werden geladen …", self._lade_abschnitte
        )
        if exc is not None or abschnitte is None:
            abschnitte = []
        if not abschnitte:
            QMessageBox.warning(self, "Keine Daten", "Bitte mindestens einen Abschnitt aktivieren.")
            return

        # Excel ohne Datei-Dialog speichern
        ziel = self._baue_ziel_pfad()
        try:
            _erstelle_bericht_excel(abschnitte, ziel)
        except Exception as exc:
            QMessageBox.critical(self, "Export-Fehler", f"Fehler beim Erstellen der Excel:\n{exc}")
            return

        namen = [a["titel"] for a in abschnitte]
        dlg = _BerichtMailDialog(ziel, namen, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        empfaenger, betreff, body = dlg.get_daten()

        try:
            from functions.mail_functions import create_outlook_draft
            logo = os.path.join(BASE_DIR, "Daten", "Email", "Logo.jpg")
            create_outlook_draft(
                to=empfaenger,
                subject=betreff,
                body_text=body,
                attachment_path=ziel,
                logo_path=logo if os.path.isfile(logo) else None,
            )
            QMessageBox.information(
                self, "Outlook geöffnet",
                "Der Outlook-Entwurf wurde geöffnet.\n"
                "Bitte prüfen und manuell absenden."
            )
        except Exception as exc:
            QMessageBox.critical(
                self, "E-Mail-Fehler",
                f"Outlook-Entwurf konnte nicht erstellt werden:\n{exc}\n\n"
                f"Die Excel-Datei wurde trotzdem gespeichert:\n{ziel}"
            )

    # ── Abrechnung erstellen ──────────────────────────────────────────────────
    def _abrechnung_erstellen(self):
        """Öffnet den Abrechnung-Dialog und exportiert eine Excel-Datei mit 3 Tabs."""
        dlg = _AbrechnungZeitraumDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        zr = dlg.get_zeitraeume()

        # ── Daten laden ───────────────────────────────────────────────────────
        from gui.splash_screen import _mit_ladeanimation

        def _lade_alle():
            from gui.dienstliches import lade_patienten, lade_einsaetze
            from database.sanmat_db import SanmatDB

            pat_alle  = lade_patienten()
            pat = _filter_nach_datum(pat_alle, zr["pat_von"], zr["pat_bis"])

            eins_alle = lade_einsaetze()
            eins = _filter_nach_datum(eins_alle, zr["eins_von"], zr["eins_bis"])

            db = SanmatDB()
            sanmat = db.get_buchungen(
                limit=100_000,
                typ="verbrauch",
                datum_von=zr["san_von"],
                datum_bis=zr["san_bis"],
            )
            return pat, eins, sanmat

        ergebnis, exc = _mit_ladeanimation(self, "Daten werden geladen …", _lade_alle)
        if exc is not None or ergebnis is None:
            QMessageBox.critical(self, "Ladefehler", f"Daten konnten nicht geladen werden:\n{exc}")
            return
        pat_eintr, eins_eintr, sanmat_eintr = ergebnis

        # ── Speicherort wählen ────────────────────────────────────────────────
        heute = datetime.now().strftime("%Y-%m-%d")
        vorschlag = os.path.join(
            _ABRECHNUNG_KOPIE_DIR, f"Abrechnung_{heute}.xlsx"
        )
        ziel, _ = QFileDialog.getSaveFileName(
            self, "Abrechnung speichern", vorschlag, "Excel-Dateien (*.xlsx)"
        )
        if not ziel:
            return

        # ── Excel erstellen ───────────────────────────────────────────────────
        try:
            _erstelle_abrechnung_excel(
                pat_eintr, eins_eintr, sanmat_eintr,
                zr["pat_label"], zr["eins_label"], zr["san_label"],
                ziel,
            )
        except Exception as e:
            QMessageBox.critical(self, "Export-Fehler", f"Fehler beim Erstellen der Excel:\n{e}")
            return

        # ── Kopie in Abrechnung-Ordner speichern (falls abweichend) ──────────
        kopie_pfad = os.path.join(
            _ABRECHNUNG_KOPIE_DIR, os.path.basename(ziel)
        )
        if os.path.normpath(ziel) != os.path.normpath(kopie_pfad):
            try:
                import shutil
                os.makedirs(_ABRECHNUNG_KOPIE_DIR, exist_ok=True)
                shutil.copy2(ziel, kopie_pfad)
            except Exception as e:
                QMessageBox.warning(
                    self, "Kopie nicht möglich",
                    f"Hauptdatei wurde gespeichert, aber Kopie schlug fehl:\n{e}"
                )

        # ── Ergebnis-Dialog ───────────────────────────────────────────────────
        msg = (
            f"Abrechnung gespeichert:\n{ziel}\n\n"
            f"Kopie: {kopie_pfad}\n\n"
            f"Pat.: {len(pat_eintr)}  |  Einsätze: {len(eins_eintr)}  |  Sanmat: {len(sanmat_eintr)}\n\n"
            "Datei jetzt öffnen?"
        )
        antwort = QMessageBox.question(
            self, "Abrechnung gespeichert", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort == QMessageBox.StandardButton.Yes:
            os.startfile(ziel)

        # ── Per E-Mail versenden? ─────────────────────────────────────────────
        antwort_mail = QMessageBox.question(
            self, "E-Mail senden?",
            "Soll die Abrechnung auch per E-Mail versendet werden?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort_mail != QMessageBox.StandardButton.Yes:
            return

        namen = ["Pat. auf Station", "Einsätze", "Sanmaterial"]
        dlg_mail = _BerichtMailDialog(ziel, namen, self)
        dlg_mail.setWindowTitle("📧  Abrechnung per E-Mail senden")
        if dlg_mail.exec() != QDialog.DialogCode.Accepted:
            return
        empfaenger, betreff, body = dlg_mail.get_daten()
        try:
            from functions.mail_functions import create_outlook_draft
            logo = os.path.join(BASE_DIR, "Daten", "Email", "Logo.jpg")
            create_outlook_draft(
                to=empfaenger,
                subject=betreff,
                body_text=body,
                attachment_path=ziel,
                logo_path=logo if os.path.isfile(logo) else None,
            )
            QMessageBox.information(
                self, "Outlook geöffnet",
                "Der Outlook-Entwurf wurde geöffnet.\n"
                "Bitte prüfen und manuell absenden."
            )
        except Exception as exc:
            QMessageBox.critical(
                self, "E-Mail-Fehler",
                f"Outlook-Entwurf konnte nicht erstellt werden:\n{exc}\n\n"
                f"Die Abrechnung wurde trotzdem gespeichert:\n{ziel}"
            )


# ─── Abrechnung: Zeitraum-Dialog ──────────────────────────────────────────────
class _AbrechnungZeitraumDialog(QDialog):
    """Dialog: Zeiträume für Abrechnung (Pat., Einsätze, Sanmaterial) einstellen."""

    _GRP = (
        "QGroupBox{{border:2px solid {ak};border-radius:6px;margin-top:10px;"
        "background:{bg};}}"
        "QGroupBox::title{{subcontrol-origin:margin;left:12px;"
        "padding:2px 8px;color:{ak};font-weight:bold;}}"
    )
    _EDIT = (
        "QDateEdit{border:1px solid #ccc;border-radius:4px;"
        "padding:3px;font-size:12px;background:white;}"
    )

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋  Abrechnung – Zeiträume einstellen")
        self.setMinimumWidth(420)
        self.resize(460, 460)
        self._build_ui()

    def _grp_style(self, ak: str, bg: str) -> str:
        return self._GRP.format(ak=ak, bg=bg)

    def _make_date_section(
        self, titel: str, ak: str, bg: str, layout: QVBoxLayout
    ) -> tuple["QDateEdit", "QDateEdit"]:
        heute = date.today()
        von_default = QDate(heute.year, heute.month, 1)
        bis_default = QDate.currentDate()

        grp = QGroupBox(f"  {titel}")
        grp.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        grp.setStyleSheet(self._grp_style(ak, bg))
        fl = QFormLayout(grp)
        fl.setSpacing(6)
        fl.setContentsMargins(10, 14, 10, 8)

        von = QDateEdit()
        von.setCalendarPopup(True)
        von.setDisplayFormat("dd.MM.yyyy")
        von.setDate(von_default)
        von.setStyleSheet(self._EDIT)
        fl.addRow("Von:", von)

        bis = QDateEdit()
        bis.setCalendarPopup(True)
        bis.setDisplayFormat("dd.MM.yyyy")
        bis.setDate(bis_default)
        bis.setStyleSheet(self._EDIT)
        fl.addRow("Bis:", bis)

        layout.addWidget(grp)
        return von, bis

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        info = QLabel("Bitte den Zeitraum für jede Übersicht festlegen:")
        info.setStyleSheet("font-size:12px;color:#555;")
        layout.addWidget(info)

        self._pat_von,  self._pat_bis  = self._make_date_section(
            "🏥  Pat. auf Station", "#B71C1C", "#FFEBEE", layout
        )
        self._eins_von, self._eins_bis = self._make_date_section(
            "🚑  Einsätze",         "#2E7D32", "#E8F5E9", layout
        )
        self._san_von,  self._san_bis  = self._make_date_section(
            "🧰  Sanmaterial",      "#E65100", "#FFF3E0", layout
        )

        layout.addStretch()

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("✓  Weiter")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    @staticmethod
    def _qd_to_date(qd: QDate) -> date:
        return date(qd.year(), qd.month(), qd.day())

    @staticmethod
    def _qd_to_iso(qd: QDate) -> str:
        return qd.toString("yyyy-MM-dd")

    @staticmethod
    def _qd_label(von: QDate, bis: QDate) -> str:
        return (
            f"{von.toString('dd.MM.yyyy')} – {bis.toString('dd.MM.yyyy')}"
        )

    def get_zeitraeume(self) -> dict:
        return {
            "pat_von":    self._qd_to_date(self._pat_von.date()),
            "pat_bis":    self._qd_to_date(self._pat_bis.date()),
            "eins_von":   self._qd_to_date(self._eins_von.date()),
            "eins_bis":   self._qd_to_date(self._eins_bis.date()),
            "san_von":    self._qd_to_iso(self._san_von.date()),
            "san_bis":    self._qd_to_iso(self._san_bis.date()),
            "pat_label":  self._qd_label(self._pat_von.date(),  self._pat_bis.date()),
            "eins_label": self._qd_label(self._eins_von.date(), self._eins_bis.date()),
            "san_label":  self._qd_label(self._san_von.date(),  self._san_bis.date()),
        }


# ─── Abrechnung: Excel-Erstellung ─────────────────────────────────────────────
def _erstelle_abrechnung_excel(
    pat_eintr: list[dict],
    eins_eintr: list[dict],
    sanmat_eintr: list[dict],
    zeitraum_pat: str,
    zeitraum_eins: str,
    zeitraum_sanmat: str,
    ziel_pfad: str,
) -> None:
    """
    Erstellt eine Excel-Datei mit 3 Tabs:
      - Pat. auf Station, Einsätze, Sanmaterial
    Laufende Nummer: ältestes Datum = Nr. 1, neuestes = letzte Nr.
    Spaltenbreite: automatisch an Inhalt angepasst.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import re as _re
    from datetime import date as _dt

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

    def _hdr_fill(hex6: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex6.lstrip("#"))

    def _d_dmy(s: str):
        """DD.MM.YYYY → date-Objekt für echten Excel-Datumswert."""
        m = _re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", str(s or "").strip())
        if m:
            try:
                return _dt(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                pass
        return s

    def _d_iso(s: str):
        """YYYY-MM-DD → date-Objekt für echten Excel-Datumswert."""
        m = _re.match(r"^(\d{4})-(\d{2})-(\d{2})$", str(s or "").strip())
        if m:
            try:
                return _dt(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
        return s

    def _auto_breite(ws):
        """Passt alle Spaltenbreiten automatisch an den längsten Zelleninhalt an."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    for part in str(cell.value or "").split("\n"):
                        max_len = max(max_len, len(part))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 4, 65))

    def _schreibe_tab(
        ws, titel: str, zeitraum: str,
        spalten: list[str], zeilen: list[list],
        hdr_color: str, zebra1: str, zebra2: str,
    ):
        n_cols   = len(spalten)
        last_col = get_column_letter(n_cols)

        # Zeile 1 – Titel
        ws.merge_cells(f"A1:{last_col}1")
        c = ws.cell(row=1, column=1, value=titel)
        c.font      = Font(bold=True, size=13, color="FFFFFF")
        c.fill      = _hdr_fill(hdr_color)
        c.alignment = center
        ws.row_dimensions[1].height = 26

        # Zeile 2 – Zeitraum
        ws.merge_cells(f"A2:{last_col}2")
        s = ws.cell(
            row=2, column=1,
            value=(
                f"Zeitraum: {zeitraum}    |    "
                f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ),
        )
        s.font      = Font(italic=True, size=9, color="555555")
        s.alignment = center
        ws.row_dimensions[2].height = 14

        # Zeile 3 – Spaltenköpfe
        for col_idx, name in enumerate(spalten, 1):
            c = ws.cell(row=3, column=col_idx, value=name)
            c.font      = hdr_font
            c.fill      = _hdr_fill(hdr_color)
            c.alignment = center
            c.border    = thin
        ws.row_dimensions[3].height = 22
        ws.auto_filter.ref = f"A3:{last_col}3"

        f1 = PatternFill("solid", fgColor=zebra1)
        f2 = PatternFill("solid", fgColor=zebra2)
        for r_off, zeile in enumerate(zeilen):
            row_num = 4 + r_off
            fill    = f1 if r_off % 2 == 0 else f2
            for col_idx, wert in enumerate(zeile, 1):
                c = ws.cell(row=row_num, column=col_idx, value=wert)
                if isinstance(wert, _dt):
                    c.number_format = "DD.MM.YYYY"
                c.fill      = fill
                c.border    = thin
                c.alignment = center if col_idx <= 3 else left
            ws.row_dimensions[row_num].height = 18

        ws.freeze_panes = "A4"
        _auto_breite(ws)

    # Sortier-Schlüssel: aufsteigend nach Datum (ältestes = Nr. 1)
    def _key_dmy(e: dict) -> tuple:
        m = _re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", str(e.get("datum", "") or ""))
        if m:
            return (int(m.group(3)), int(m.group(2)), int(m.group(1)))
        return (9999, 99, 99)

    def _key_iso(e: dict) -> str:
        return str(e.get("datum", "") or "")

    wb      = openpyxl.Workbook()
    ws_dummy = wb.active

    # ── Tab 1: Pat. auf Station ───────────────────────────────────────────────
    pat_sorted = sorted(pat_eintr, key=_key_dmy)
    ws_pat = wb.create_sheet("Pat. auf Station")
    _schreibe_tab(
        ws_pat,
        f"🏥  Pat. auf Station  ({len(pat_sorted)} Einträge)",
        zeitraum_pat,
        [
            "Nr.", "Datum", "Uhrzeit", "Dauer (Min.)", "Typ", "Alter",
            "Beschwerde", "Ort", "Maßnahmen", "DRK MA 1", "Weitergeleitet", "BG-Fall",
        ],
        [
            [
                idx,
                _d_dmy(e.get("datum", "")),
                e.get("uhrzeit", "") or "",
                e.get("behandlungsdauer", "") or "",
                e.get("patient_typ", "") or "",
                e.get("alter", "") or "",
                e.get("beschwerde_art", "") or "",
                e.get("unfall_ort", "") or "",
                e.get("massnahmen", "") or "",
                e.get("drk_ma1", "") or "",
                e.get("weitergeleitet", "") or "",
                "Ja" if e.get("arbeitsunfall") else "Nein",
            ]
            for idx, e in enumerate(pat_sorted, 1)
        ],
        "B71C1C", "FFFFFF", "FFEBEE",
    )

    # ── Tab 2: Einsätze ───────────────────────────────────────────────────────
    eins_sorted = sorted(eins_eintr, key=_key_dmy)
    ws_eins = wb.create_sheet("Einsätze")
    _schreibe_tab(
        ws_eins,
        f"🚑  Einsätze  ({len(eins_sorted)} Einträge)",
        zeitraum_eins,
        [
            "Nr.", "Datum", "Uhrzeit", "Dauer (Min.)", "Stichwort",
            "Ort", "DRK-Nr.", "MA 1", "MA 2", "Angenommen", "Bemerkung",
        ],
        [
            [
                idx,
                _d_dmy(e.get("datum", "")),
                e.get("uhrzeit", "") or "",
                e.get("einsatzdauer", "") or "",
                e.get("einsatzstichwort", "") or "",
                e.get("einsatzort", "") or "",
                e.get("einsatznr_drk", "") or "",
                e.get("drk_ma1", "") or "",
                e.get("drk_ma2", "") or "",
                "Ja" if e.get("angenommen", 1) else "Nein",
                e.get("bemerkung", "") or "",
            ]
            for idx, e in enumerate(eins_sorted, 1)
        ],
        "2E7D32", "FFFFFF", "E8F5E9",
    )

    # ── Tab 3: Sanmaterial ────────────────────────────────────────────────────
    sanmat_sorted = sorted(sanmat_eintr, key=_key_iso)
    ws_san = wb.create_sheet("Sanmaterial")
    _schreibe_tab(
        ws_san,
        f"🧰  Sanmaterial  ({len(sanmat_sorted)} Einträge)",
        zeitraum_sanmat,
        ["Nr.", "Datum", "Artikel", "Menge", "Bemerkung"],
        [
            [
                idx,
                _d_iso(e.get("datum", "")),
                e.get("artikel_name", "") or "",
                e.get("menge", "") or "",
                e.get("bemerkung", "") or "",
            ]
            for idx, e in enumerate(sanmat_sorted, 1)
        ],
        "E65100", "FFFFFF", "FFF3E0",
    )

    if len(wb.worksheets) > 1:
        wb.remove(ws_dummy)

    os.makedirs(os.path.dirname(ziel_pfad), exist_ok=True)
    wb.save(ziel_pfad)
