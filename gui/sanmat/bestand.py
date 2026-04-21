"""
Sanitätsmaterial – Bestandsverwaltung
Bestandsübersicht, Einlagerung (Wareneingang), Mindestbestand, Korrektur.
"""

import csv
import os
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QComboBox, QLineEdit,
    QSpinBox, QDialog, QDialogButtonBox, QFormLayout, QFrame,
    QMessageBox, QHeaderView, QAbstractItemView, QDateEdit,
    QScrollArea, QFileDialog, QRadioButton, QButtonGroup as QRadioGroup,
    QGroupBox,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor


def _format_datum(iso: str) -> str:
    try:
        return datetime.strptime(iso[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return iso


class EinlagerungDialog(QDialog):
    def __init__(self, db, artikel_id: int = None, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Wareneingang buchen")
        self.setMinimumWidth(680)
        self.setModal(True)
        self._rows: list[dict] = []
        self._setup_ui(artikel_id)

    def _setup_ui(self, artikel_id):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)
        layout.addWidget(QLabel("Mehrere Positionen können gleichzeitig eingebucht werden."))

        meta = QFormLayout()
        meta.setSpacing(8)
        self.de_datum = QDateEdit(QDate.currentDate())
        self.de_datum.setDisplayFormat("dd.MM.yyyy")
        self.de_datum.setCalendarPopup(True)
        self.de_datum.setMaximumWidth(150)
        meta.addRow("Einlagerdatum:", self.de_datum)
        self.le_von = QLineEdit()
        self.le_von.setPlaceholderText("Lieferant / Kürzel / meetB ...")
        meta.addRow("Von / Lieferant:", self.le_von)
        self.le_bem = QLineEdit()
        self.le_bem.setPlaceholderText("Lieferschein-Nr., Anmerkung ...")
        meta.addRow("Bemerkung:", self.le_bem)
        layout.addLayout(meta)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        layout.addWidget(sep)

        hdr_w = QWidget()
        hdr = QHBoxLayout(hdr_w)
        hdr.setContentsMargins(0, 0, 0, 0)
        hdr.setSpacing(8)
        for txt, w in [("Artikel", 280), ("Menge", 80)]:
            lb = QLabel(txt); lb.setFixedWidth(w)
            lb.setStyleSheet("font-weight:bold; color:#444;")
            hdr.addWidget(lb)
        hdr.addStretch()
        layout.addWidget(hdr_w)

        self._rows_layout = QVBoxLayout()
        self._rows_layout.setSpacing(4)
        container_w = QWidget()
        container_w.setLayout(self._rows_layout)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setMaximumHeight(280)
        scroll.setWidget(container_w)
        layout.addWidget(scroll)

        btn_add = QPushButton("+ Zeile hinzufügen")
        btn_add.setObjectName("btn_secondary")
        btn_add.clicked.connect(lambda: self._add_zeile())
        layout.addWidget(btn_add)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Einlagern")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
        self._add_zeile(artikel_id)

    def _add_zeile(self, artikel_id=None):
        row_w = QWidget()
        hl = QHBoxLayout(row_w)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(8)
        cb_art = QComboBox()
        cb_art.setFixedWidth(280)
        for a in self.db.get_artikel():
            label = a["bezeichnung"] + (f" [{a['artikelnr']}]" if a.get("artikelnr") else "")
            cb_art.addItem(label, a["id"])
        if artikel_id:
            idx = cb_art.findData(artikel_id)
            if idx >= 0:
                cb_art.setCurrentIndex(idx)
        sb = QSpinBox()
        sb.setRange(1, 99999)
        sb.setValue(1)
        sb.setFixedWidth(80)
        btn_x = QPushButton("×")
        btn_x.setFixedSize(26, 26)
        btn_x.setStyleSheet("QPushButton{color:#B20000;font-weight:bold;font-size:15px;"
                            "border:none;background:transparent;}"
                            "QPushButton:hover{background:#fce4e4;border-radius:4px;}")
        hl.addWidget(cb_art); hl.addWidget(sb); hl.addWidget(btn_x); hl.addStretch()
        ri = {"widget": row_w, "cb_art": cb_art, "sb": sb}
        btn_x.clicked.connect(lambda _, rw=row_w, r=ri: self._remove_zeile(rw, r))
        self._rows.append(ri)
        self._rows_layout.addWidget(row_w)

    def _remove_zeile(self, row_w, ri):
        if len(self._rows) <= 1:
            return
        self._rows.remove(ri)
        self._rows_layout.removeWidget(row_w)
        row_w.deleteLater()

    def get_data(self) -> list[dict]:
        datum = self.de_datum.date().toString("yyyy-MM-dd")
        von = self.le_von.text().strip()
        bem = self.le_bem.text().strip()
        return [
            {"artikel_id": ri["cb_art"].currentData(), "artikel_name": ri["cb_art"].currentText(),
             "menge": ri["sb"].value(), "datum": datum, "von": von, "bemerkung": bem}
            for ri in self._rows
        ]


class BestandEditDialog(QDialog):
    def __init__(self, item: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Bestand bearbeiten")
        self.setMinimumWidth(380)
        self.setModal(True)
        form = QFormLayout(self)
        form.setSpacing(12)
        form.setContentsMargins(20, 20, 20, 16)
        lbl = QLabel(item.get("bezeichnung", ""))
        lbl.setStyleSheet("font-weight:bold;")
        form.addRow("Artikel:", lbl)
        form.addRow("Aktueller Bestand:", QLabel(str(item.get("menge", 0))))
        self.sb_min = QSpinBox()
        self.sb_min.setRange(0, 99999)
        self.sb_min.setValue(int(item.get("min_menge", 0)))
        form.addRow("Mindestbestand:", self.sb_min)
        self.le_lagerort = QLineEdit(item.get("lagerort", ""))
        self.le_lagerort.setPlaceholderText("z.B. Regal A3, Schrank 2 ...")
        form.addRow("Lagerort:", self.le_lagerort)
        self.le_bem = QLineEdit(item.get("bemerkung", ""))
        form.addRow("Bemerkung:", self.le_bem)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Speichern")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        form.addRow(btns)

    def get_data(self) -> dict:
        return {"min_menge": self.sb_min.value(), "lagerort": self.le_lagerort.text().strip(),
                "bemerkung": self.le_bem.text().strip()}


class KorrekturDialog(QDialog):
    def __init__(self, item: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Bestandskorrektur")
        self.setMinimumWidth(380)
        self.setModal(True)
        form = QFormLayout(self)
        form.setSpacing(12)
        form.setContentsMargins(20, 20, 20, 16)
        lbl = QLabel(item.get("bezeichnung", ""))
        lbl.setStyleSheet("font-weight:bold;")
        form.addRow("Artikel:", lbl)
        lbl_a = QLabel(str(item.get("menge", 0)))
        lbl_a.setStyleSheet("color:#8B0000; font-weight:bold;")
        form.addRow("Aktueller Bestand:", lbl_a)
        self.sb_neu = QSpinBox()
        self.sb_neu.setRange(0, 99999)
        self.sb_neu.setValue(int(item.get("menge", 0)))
        form.addRow("Neuer Bestand:", self.sb_neu)
        self.de_datum = QDateEdit(QDate.currentDate())
        self.de_datum.setDisplayFormat("dd.MM.yyyy")
        self.de_datum.setCalendarPopup(True)
        form.addRow("Datum:", self.de_datum)
        self.le_von = QLineEdit()
        self.le_von.setPlaceholderText("Kürzel/Name ...")
        form.addRow("Korrigiert von:", self.le_von)
        self.le_bem = QLineEdit()
        self.le_bem.setPlaceholderText("Inventur, Schwund ... (Optional)")
        form.addRow("Bemerkung:", self.le_bem)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Korrektur buchen")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        form.addRow(btns)

    def get_data(self) -> dict:
        return {"neue_menge": self.sb_neu.value(),
                "datum": self.de_datum.date().toString("yyyy-MM-dd"),
                "von": self.le_von.text().strip(),
                "bemerkung": self.le_bem.text().strip()}


# ────────────────────────────────────────────────────────────────────────────
# Excel-Inventur-Export
# ────────────────────────────────────────────────────────────────────────────

class InventurExportDialog(QDialog):
    """Kleiner Dialog: Filterauswahl + Speicherpfad für Excel-Inventur."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Inventur als Excel exportieren")
        self.setMinimumWidth(480)
        self.setModal(True)
        self._pfad = ""
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(14)

        # Filtergruppe
        grp = QGroupBox("Welche Artikel exportieren?")
        grp_layout = QVBoxLayout(grp)
        self._rb_alle = QRadioButton("Alle Artikel (auch ohne Bestand)")
        self._rb_bestand = QRadioButton("Nur Artikel mit vorhandenem Bestand (> 0)")
        self._rb_alle.setChecked(True)
        grp_layout.addWidget(self._rb_alle)
        grp_layout.addWidget(self._rb_bestand)
        layout.addWidget(grp)

        # Speicherpfad
        pfad_label = QLabel("Speicherort:")
        pfad_label.setStyleSheet("font-weight:bold;")
        layout.addWidget(pfad_label)

        pfad_row = QHBoxLayout()
        self._le_pfad = QLineEdit()
        vorschlag = os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"Inventur_SanMat_{date.today().strftime('%Y-%m-%d')}.xlsx"
        )
        self._le_pfad.setText(vorschlag)
        self._le_pfad.setPlaceholderText("Pfad zur Excel-Datei ...")
        pfad_row.addWidget(self._le_pfad, 1)
        btn_browse = QPushButton("Durchsuchen …")
        btn_browse.clicked.connect(self._browse)
        pfad_row.addWidget(btn_browse)
        layout.addLayout(pfad_row)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Exportieren")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self._validate_and_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _browse(self):
        vorschlag = self._le_pfad.text() or os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"Inventur_SanMat_{date.today().strftime('%Y-%m-%d')}.xlsx"
        )
        pfad, _ = QFileDialog.getSaveFileName(
            self, "Speicherort wählen", vorschlag,
            "Excel-Dateien (*.xlsx)"
        )
        if pfad:
            if not pfad.lower().endswith(".xlsx"):
                pfad += ".xlsx"
            self._le_pfad.setText(pfad)

    def _validate_and_accept(self):
        pfad = self._le_pfad.text().strip()
        if not pfad:
            QMessageBox.warning(self, "Hinweis", "Bitte einen Speicherpfad angeben.")
            return
        self._pfad = pfad
        self.accept()

    def nur_mit_bestand(self) -> bool:
        return self._rb_bestand.isChecked()

    def pfad(self) -> str:
        return self._pfad


def _thin():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def _medium():
    s = Side(style="medium", color="FF666666")
    return Border(left=s, right=s, top=s, bottom=s)


_KAT_FARBEN = {
    "Wundversorgung":       ("FFD6E4BC", "FF4A7C1F"),
    "Desinfektion":         ("FFDCE6F1", "FF174069"),
    "Schutzausrüstung":     ("FFFFF2CC", "FF7F6000"),
    "Diagnostik":           ("FFFCE4D6", "FF843C0C"),
    "Notfallausrüstung":    ("FFFFE2CC", "FF833C04"),
    "Patientenversorgung":  ("FFE2EFDA", "FF375623"),
    "Verbrauchsmaterial":   ("FFEDEDED", "FF595959"),
    "Entsorgung":           ("FFFFF0CC", "FF7F5700"),
}
_KAT_REIHENFOLGE = [
    "Wundversorgung", "Desinfektion", "Schutzausrüstung", "Diagnostik",
    "Notfallausrüstung", "Patientenversorgung", "Verbrauchsmaterial", "Entsorgung",
]
_SPALTEN = [
    "Artikelnr.", "Bezeichnung", "Kategorie", "Einheit",
    "Packungsinhalt", "PZN", "Ist-Bestand", "Min.-Bestand",
    "Lagerort", "Bemerkung",
]
_BREITEN = [14, 44, 18, 8, 16, 14, 10, 10, 16, 24]


def _erstelle_inventur_excel(bestand: list[dict], pfad: str) -> None:
    """Erzeugt die Excel-Inventurdatei aus echten DB-Daten."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventur"
    ws.sheet_view.showGridLines = False

    for col, b in enumerate(_BREITEN, start=1):
        ws.column_dimensions[get_column_letter(col)].width = b

    # Hauptheader
    ws.merge_cells(f"A1:{get_column_letter(len(_SPALTEN))}1")
    c = ws["A1"]
    c.value = (
        f"Inventurliste Sanitätsmaterial  –  DRK Flughafen Köln/Bonn"
        f"  |  Stand: {date.today().strftime('%d.%m.%Y')}"
    )
    c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor="CC0000")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Spaltenköpfe (Zeile 2)
    for col, titel in enumerate(_SPALTEN, start=1):
        c = ws.cell(row=2, column=col, value=titel)
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _medium()
    ws.row_dimensions[2].height = 24

    # Alle Artikel alphabetisch nach Bezeichnung
    artikel_liste = sorted(bestand, key=lambda x: x.get("bezeichnung", "").lower())

    row = 3
    alt = False
    for a in artikel_liste:
        bg = "F2F2F2" if not alt else "FFFFFF"
        alt = not alt
        werte = [
            a.get("artikelnr", ""),
            a.get("bezeichnung", ""),
            a.get("kategorie", ""),
            a.get("einheit", ""),
            a.get("packungsinhalt", ""),
            a.get("pzn", ""),
            a.get("menge", 0),
            a.get("min_menge", 0) or "",
            a.get("lagerort", ""),
            a.get("bemerkung", ""),
        ]
        for col, val in enumerate(werte, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=9)
            c.border = _thin()
            c.alignment = Alignment(
                vertical="center",
                wrap_text=(col == 2),
                horizontal="center" if col in (7, 8) else "left",
            )
        # Bestand rot markieren wenn 0
        menge_cell = ws.cell(row=row, column=7)
        if int(a.get("menge", 0)) == 0:
            menge_cell.font = Font(name="Calibri", size=9, bold=True, color="FF8B0000")
        elif int(a.get("min_menge") or 0) > 0 and int(a.get("menge", 0)) <= int(a.get("min_menge") or 0):
            menge_cell.font = Font(name="Calibri", size=9, bold=True, color="FF7F5700")
        ws.row_dimensions[row].height = 15
        row += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(len(_SPALTEN))}{row - 1}"

    # Seiteneinrichtung: DIN A4 Querformat, Breite = 1 Seite
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2
    )
    ws.print_options.horizontalCentered = True
    ws.freeze_panes = "A3"

    os.makedirs(os.path.dirname(os.path.abspath(pfad)), exist_ok=True)
    wb.save(pfad)


class BestandView(QWidget):
    COLS = ["Bezeichnung", "Artikelnr.", "Kategorie", "Auf Lager",
            "Mindestbest.", "Lagerort", "Einheit", "Packungsinhalt"]

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(12)

        lbl_title = QLabel("Bestandsübersicht")
        lbl_title.setObjectName("page_title")
        lbl_sub = QLabel("Aktueller Bestand aller Sanitätsmaterial-Artikel")
        lbl_sub.setObjectName("page_subtitle")
        layout.addWidget(lbl_title)
        layout.addWidget(lbl_sub)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(10)
        self.le_suche = QLineEdit()
        self.le_suche.setPlaceholderText("Bezeichnung, Artikelnr. ...")
        self.le_suche.setClearButtonEnabled(True)
        self.le_suche.textChanged.connect(self._load)
        filter_row.addWidget(self.le_suche, stretch=2)

        self.cb_kat = QComboBox()
        self.cb_kat.addItem("Alle Kategorien")
        for k in self.db.get_kategorien():
            self.cb_kat.addItem(k)
        self.cb_kat.currentIndexChanged.connect(self._load)
        filter_row.addWidget(self.cb_kat)

        self.cb_filter = QComboBox()
        self.cb_filter.addItems(["Alle", "Niedrig / Leer", "Ausreichend"])
        self.cb_filter.currentIndexChanged.connect(self._load)
        filter_row.addWidget(self.cb_filter)

        filter_row.addStretch()
        btn_einlagern = QPushButton("📦  Wareneingang buchen")
        btn_einlagern.setObjectName("btn_primary")
        btn_einlagern.clicked.connect(self._einlagern)
        filter_row.addWidget(btn_einlagern)

        btn_export = QPushButton("📄  CSV-Export")
        btn_export.setObjectName("btn_secondary")
        btn_export.clicked.connect(self._export)
        filter_row.addWidget(btn_export)

        btn_excel = QPushButton("📊  Excel-Inventur")
        btn_excel.setObjectName("btn_secondary")
        btn_excel.clicked.connect(self._export_excel)
        filter_row.addWidget(btn_excel)
        layout.addLayout(filter_row)

        self._tbl = QTableWidget(0, len(self.COLS))
        self._tbl.setHorizontalHeaderLabels(self.COLS)
        self._tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in range(1, len(self.COLS)):
            self._tbl.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.doubleClicked.connect(lambda _: self._korrektur())
        layout.addWidget(self._tbl)

        self._lbl_count = QLabel("")
        self._lbl_count.setStyleSheet("color:#888; font-size:11px;")
        layout.addWidget(self._lbl_count)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_einl_sel = QPushButton("📦  Einlagern (ausgewählt)")
        btn_einl_sel.setObjectName("btn_secondary")
        btn_einl_sel.clicked.connect(self._einlagern_from_selection)
        btn_row.addWidget(btn_einl_sel)
        btn_edit = QPushButton("✏  Mindestbest./Lagerort")
        btn_edit.setObjectName("btn_secondary")
        btn_edit.clicked.connect(self._edit_bestand)
        btn_row.addWidget(btn_edit)
        btn_korr = QPushButton("⚖  Bestandskorrektur")
        btn_korr.setObjectName("btn_secondary")
        btn_korr.clicked.connect(self._korrektur)
        btn_row.addWidget(btn_korr)
        layout.addLayout(btn_row)

    def _load(self):
        suche = self.le_suche.text().lower()
        kat = self.cb_kat.currentText()
        bestandsfilter = self.cb_filter.currentText()
        bestand = self.db.get_bestand()

        def _match(a):
            if kat != "Alle Kategorien" and a.get("kategorie", "") != kat:
                return False
            if suche:
                h = (a.get("bezeichnung", "") + " " + a.get("artikelnr", "")).lower()
                if suche not in h:
                    return False
            m = int(a.get("menge", 0))
            mn = int(a.get("min_menge", 0))
            if bestandsfilter == "Niedrig / Leer":
                return m == 0 or (mn > 0 and m <= mn)
            if bestandsfilter == "Ausreichend":
                return not (m == 0 or (mn > 0 and m <= mn))
            return True

        filtered = [a for a in bestand if _match(a)]
        warn_color = QColor("#FFF3CD")
        crit_color = QColor("#FFCCBC")
        self._tbl.setRowCount(len(filtered))
        for r, a in enumerate(filtered):
            menge = int(a.get("menge", 0))
            min_m = int(a.get("min_menge", 0))
            vals = [
                a.get("bezeichnung", ""), a.get("artikelnr", ""), a.get("kategorie", ""),
                str(menge), str(min_m) if min_m else "–", a.get("lagerort", ""),
                a.get("einheit", ""), a.get("packungsinhalt", ""),
            ]
            is_warn = min_m > 0 and menge <= min_m
            is_crit = menge == 0
            for c, v in enumerate(vals):
                it = QTableWidgetItem(str(v))
                if is_crit:
                    it.setBackground(crit_color)
                elif is_warn:
                    it.setBackground(warn_color)
                it.setData(Qt.ItemDataRole.UserRole, a.get("id"))
                self._tbl.setItem(r, c, it)
        self._lbl_count.setText(f"{len(filtered)} Artikel")

    def _get_selected(self) -> dict | None:
        row = self._tbl.currentRow()
        if row < 0:
            return None
        item = self._tbl.item(row, 0)
        if not item:
            return None
        aid = item.data(Qt.ItemDataRole.UserRole)
        return self.db.get_artikel_by_id(aid) if aid else None

    def _einlagern(self):
        dlg = EinlagerungDialog(self.db, parent=self)
        if dlg.exec():
            ok_list, err_list = [], []
            for pos in dlg.get_data():
                ok, msg = self.db.einlagern(**pos)
                (ok_list if ok else err_list).append(msg)
            msgs = []
            if ok_list:
                msgs.append("✓ " + "\n✓ ".join(ok_list))
            if err_list:
                msgs.append("✗ " + "\n✗ ".join(err_list))
            QMessageBox.information(self, "Wareneingang", "\n".join(msgs))
            self._load()

    def _einlagern_from_selection(self):
        a = self._get_selected()
        if not a:
            QMessageBox.information(self, "Hinweis", "Bitte erst einen Artikel auswählen.")
            return
        dlg = EinlagerungDialog(self.db, artikel_id=a["id"], parent=self)
        if dlg.exec():
            for pos in dlg.get_data():
                ok, msg = self.db.einlagern(**pos)
                if not ok:
                    QMessageBox.warning(self, "Fehler", msg)
            self._load()

    def _edit_bestand(self):
        a = self._get_selected()
        if not a:
            QMessageBox.information(self, "Hinweis", "Bitte erst einen Artikel auswählen.")
            return
        dlg = BestandEditDialog(a, parent=self)
        if dlg.exec():
            d = dlg.get_data()
            ok, msg = self.db.update_bestand(
                a["id"], int(a.get("menge", 0)), d["min_menge"], d["lagerort"], d["bemerkung"]
            )
            if not ok:
                QMessageBox.warning(self, "Fehler", msg)
            self._load()

    def _korrektur(self):
        a = self._get_selected()
        if not a:
            QMessageBox.information(self, "Hinweis", "Bitte erst einen Artikel auswählen.")
            return
        dlg = KorrekturDialog(a, parent=self)
        if dlg.exec():
            d = dlg.get_data()
            ok, msg = self.db.korrektur(
                a["id"], a["bezeichnung"], d["neue_menge"], d["datum"], d["von"], d["bemerkung"]
            )
            if ok:
                QMessageBox.information(self, "Korrektur", msg)
            else:
                QMessageBox.warning(self, "Fehler", msg)
            self._load()

    def _export(self):
        bestand = self.db.get_bestand()
        pfad, _ = QFileDialog.getSaveFileName(
            self, "Bestand exportieren", f"Bestand_{datetime.now().strftime('%Y%m%d')}.csv",
            "CSV-Dateien (*.csv)"
        )
        if not pfad:
            return
        try:
            with open(pfad, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["Artikelnr", "Bezeichnung", "Kategorie", "Auf_Lager",
                             "Mindestbestand", "Lagerort", "Einheit", "Packungsinhalt"])
                for b in bestand:
                    w.writerow([b.get("artikelnr", ""), b.get("bezeichnung", ""),
                                 b.get("kategorie", ""), b.get("menge", 0),
                                 b.get("min_menge", 0), b.get("lagerort", ""),
                                 b.get("einheit", ""), b.get("packungsinhalt", "")])
            QMessageBox.information(self, "Exportiert", f"Gespeichert: {pfad}")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", str(e))

    def _export_excel(self):
        if not _OPENPYXL_OK:
            QMessageBox.warning(
                self, "Fehler",
                "openpyxl ist nicht installiert.\n"
                "Bitte 'pip install openpyxl' ausführen."
            )
            return
        dlg = InventurExportDialog(parent=self)
        if not dlg.exec():
            return
        pfad = dlg.pfad()
        bestand = self.db.get_bestand()
        if dlg.nur_mit_bestand():
            bestand = [a for a in bestand if int(a.get("menge", 0)) > 0]
        if not bestand:
            QMessageBox.information(self, "Kein Bestand", "Keine Artikel für den Export vorhanden.")
            return
        try:
            _erstelle_inventur_excel(bestand, pfad)
            antwort = QMessageBox.question(
                self, "Exportiert",
                f"Datei gespeichert:\n{pfad}\n\nJetzt öffnen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if antwort == QMessageBox.StandardButton.Yes:
                import subprocess
                subprocess.Popen(["explorer", pfad])
        except Exception as e:
            QMessageBox.warning(self, "Fehler beim Export", str(e))

    def showEvent(self, event):
        super().showEvent(event)
        # Kategorien-ComboBox aktualisieren
        current_kat = self.cb_kat.currentText()
        self.cb_kat.blockSignals(True)
        self.cb_kat.clear()
        self.cb_kat.addItem("Alle Kategorien")
        for k in self.db.get_kategorien():
            self.cb_kat.addItem(k)
        idx = self.cb_kat.findText(current_kat)
        self.cb_kat.setCurrentIndex(idx if idx >= 0 else 0)
        self.cb_kat.blockSignals(False)
        self._load()
