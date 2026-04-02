"""
VerbrauchView – Zeigt alle Verbrauch-Buchungen aus Einsätzen.
Gefilterte Ansicht des Buchungsverlaufs (typ='verbrauch').
"""

from __future__ import annotations
import csv
import os
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import re

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QLineEdit, QDateEdit, QFileDialog,
    QMessageBox, QFrame, QDialog, QFormLayout,
    QSpinBox, QComboBox, QAbstractItemView, QMenu,
)
from PySide6.QtCore import Qt, QDate, QPoint
from PySide6.QtGui import QColor, QFont

from database.sanmat_db import SanmatDB

PAGE_SIZE = 100


class VerbrauchView(QWidget):
    """Zeigt alle Verbrauch-Buchungen (Einsatz-Verbrauch) mit Filter und CSV-Export."""

    _HDR = ["Datum", "Artikel", "Menge", "Entnehmer", "Bemerkung"]

    # Gründe für manuellen Verbrauch
    GRUENDE = [
        "Aus- / Fortbildung",
        "Ablauf / MHD überschritten",
        "Übung / Simulation",
        "Eigenbedarf",
        "Beschaedigt / Defekt",
        "Sonstiges",
    ]

    def __init__(self, db: SanmatDB, parent=None):
        super().__init__(parent)
        self._db = db
        self._page = 0
        self._total = 0
        # Ausschluss-Sets für CSV-Export
        self._excluded_ids:  set[int] = set()  # einzelne Buchungs-IDs
        self._excluded_eids: set[int] = set()  # ganze Einsätze per Einsatz-ID
        self._display_meta:  list[tuple] = []  # (kind, eid, buchung_id) je Tabellenzeile
        self._undo_stack:  list[dict]  = []
        self._chk_lock = False
        self._setup_ui()
        self._laden()

    # ── UI ──────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        # Titel
        title = QLabel("🧰  Einsatz-Verbrauch")
        title.setStyleSheet("font-size:16px; font-weight:bold; color:#1565a8;")
        root.addWidget(title)

        sub = QLabel("Verbrauchsmaterial aus Einsätzen – aus Sanmat-Bestand abgebucht.")
        sub.setStyleSheet("color:#666; font-size:11px;")
        root.addWidget(sub)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#ddd;")
        root.addWidget(sep)

        # Filter-Zeile
        f_lay = QHBoxLayout()
        f_lay.setSpacing(8)

        f_lay.addWidget(QLabel("Von:"))
        self._von = QDateEdit()
        self._von.setCalendarPopup(True)
        self._von.setDisplayFormat("dd.MM.yyyy")
        self._von.setDate(QDate(QDate.currentDate().year(), 1, 1))
        self._von.setFixedWidth(120)
        f_lay.addWidget(self._von)

        f_lay.addWidget(QLabel("Bis:"))
        self._bis = QDateEdit()
        self._bis.setCalendarPopup(True)
        self._bis.setDisplayFormat("dd.MM.yyyy")
        self._bis.setDate(QDate.currentDate())
        self._bis.setFixedWidth(120)
        f_lay.addWidget(self._bis)

        self._suche = QLineEdit()
        self._suche.setPlaceholderText("Entnehmer oder Artikel suchen …")
        self._suche.setFixedWidth(200)
        self._suche.returnPressed.connect(self._on_filter)
        f_lay.addWidget(self._suche)

        btn_filter = QPushButton("🔍 Filtern")
        btn_filter.clicked.connect(self._on_filter)
        f_lay.addWidget(btn_filter)

        btn_reset = QPushButton("↺ Zurücksetzen")
        btn_reset.clicked.connect(self._reset_filter)
        f_lay.addWidget(btn_reset)

        self._btn_undo = QPushButton("↩ Rükgängig")
        self._btn_undo.setEnabled(False)
        self._btn_undo.setToolTip("Letzte Löschung / Änderung rükgängig machen")
        self._btn_undo.clicked.connect(self._undo)
        f_lay.addWidget(self._btn_undo)

        f_lay.addStretch()

        btn_export = QPushButton("📄 Excel Export")
        btn_export.clicked.connect(self._export_csv)
        f_lay.addWidget(btn_export)

        btn_alle = QPushButton("☑ Alle")
        btn_alle.setFixedWidth(60)
        btn_alle.setToolTip("Alle Zeilen für Export auswählen")
        btn_alle.clicked.connect(self._auswahl_alle)
        f_lay.addWidget(btn_alle)

        btn_keine = QPushButton("☐ Keine")
        btn_keine.setFixedWidth(60)
        btn_keine.setToolTip("Alle Zeilen vom Export ausschließen")
        btn_keine.clicked.connect(self._auswahl_keine)
        f_lay.addWidget(btn_keine)

        btn_neu = QPushButton("➕ Verbrauch anlegen")
        btn_neu.setStyleSheet(
            "QPushButton{background:#1565a8;color:white;font-weight:bold;"
            "padding:5px 12px;border-radius:4px;}"
            "QPushButton:hover{background:#1976d2;}"
        )
        btn_neu.clicked.connect(self._verbrauch_anlegen)
        f_lay.addWidget(btn_neu)

        root.addLayout(f_lay)

        # Tabelle
        self._table = QTableWidget()
        self._table.setColumnCount(len(self._HDR))
        self._table.setHorizontalHeaderLabels(self._HDR)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(False)
        self._table.verticalHeader().setVisible(False)
        self._table.itemChanged.connect(self._on_checkbox_changed)
        # Kontextmenu direkt am Table-Widget (zuverlässiger als viewport-Ansatz)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._kontextmenu)
        root.addWidget(self._table, 1)

        # Paginierung
        pg_lay = QHBoxLayout()
        self._lbl_info = QLabel("")
        self._lbl_info.setStyleSheet("color:#555; font-size:11px;")
        pg_lay.addWidget(self._lbl_info)
        pg_lay.addStretch()
        self._btn_prev = QPushButton("◀ Zurück")
        self._btn_prev.clicked.connect(self._prev_page)
        self._btn_next = QPushButton("Weiter ▶")
        self._btn_next.clicked.connect(self._next_page)
        pg_lay.addWidget(self._btn_prev)
        pg_lay.addWidget(self._btn_next)
        root.addLayout(pg_lay)

    # ── Laden ────────────────────────────────────────────────────────────────

    _COL_HEADER_BG = QColor("#1565a8")
    _COL_HEADER_FG = QColor("#ffffff")
    _COL_ITEM_BG   = QColor("#DDEEFF")
    _COL_FLAT_BG   = QColor("#ffffff")

    def _laden(self):
        von_str = self._von.date().toString("yyyy-MM-dd")
        bis_str = self._bis.date().toString("yyyy-MM-dd")
        suche   = self._suche.text().strip() or None

        self._total = self._db.count_buchungen(typ="verbrauch", datum_von=von_str, datum_bis=bis_str, suche=suche)
        rows = self._db.get_buchungen(
            limit=PAGE_SIZE, offset=self._page * PAGE_SIZE,
            typ="verbrauch", datum_von=von_str, datum_bis=bis_str, suche=suche
        )

        # Gruppen-Schlüssel je Zeile: "ID_43" (Einsatz) oder "GID_..." (manuell) oder None
        eid_of: list[str | None] = []
        eid_count: dict[str, int] = {}
        for b in rows:
            m = re.search(r"\((G?ID)\s*(\d+)\)", b.get("bemerkung", "") or "")
            eid: str | None = f"{m.group(1)}_{m.group(2)}" if m else None
            eid_of.append(eid)
            if eid is not None:
                eid_count[eid] = eid_count.get(eid, 0) + 1

        # Anzeigeliste aufbauen: (kind, eid, buchung)
        # kind = 'header' | 'item' | 'flat'
        display: list[tuple] = []
        seen_eids: set[str] = set()
        for b, eid in zip(rows, eid_of):
            multi = eid is not None and eid_count.get(eid, 0) > 1
            if multi:
                if eid not in seen_eids:
                    seen_eids.add(eid)
                    display.append(("header", eid, b))
                display.append(("item", eid, b))
            else:
                display.append(("flat", eid, b))

        self._table.clearSpans()
        self._table.blockSignals(True)
        self._table.setRowCount(0)
        self._table.setRowCount(len(display))

        # Meta–Liste für Checkbox-Logik
        self._display_meta = [(kind, eid, b.get("id")) for (kind, eid, b) in display]

        for r, (kind, eid, b) in enumerate(display):
            datum_raw = b.get("datum", "")[:10]
            try:
                y, mo, d = datum_raw.split("-")
                datum_fmt = f"{d}.{mo}.{y}"
            except Exception:
                datum_fmt = datum_raw

            bemerkung_raw = b.get("bemerkung", "") or ""
            einsatz_name  = re.sub(r"\s*\(G?ID\s*\d+\).*", "", bemerkung_raw).strip()
            icon = "🚑" if (eid or "").startswith("ID_") else "📋"

            def _cell(text: str, bg: QColor, fg: QColor | None = None,
                      bold: bool = False, align=None) -> QTableWidgetItem:
                it = QTableWidgetItem(text)
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                it.setBackground(bg)
                if fg:
                    it.setForeground(fg)
                if bold:
                    f = QFont(); f.setBold(True); it.setFont(f)
                if align:
                    it.setTextAlignment(align)
                return it

            if kind == "header":
                self._table.setItem(r, 0, _cell(datum_fmt, self._COL_HEADER_BG, self._COL_HEADER_FG, bold=True))
                self._table.setItem(r, 1, _cell(f"{icon}  {einsatz_name}", self._COL_HEADER_BG, self._COL_HEADER_FG, bold=True))
                self._table.setItem(r, 3, _cell(b.get("von", ""), self._COL_HEADER_BG, self._COL_HEADER_FG, bold=True))
                self._table.setItem(r, 4, _cell("", self._COL_HEADER_BG, self._COL_HEADER_FG))
                self._table.setSpan(r, 1, 1, 2)  # Artikel + Menge-Spalte zusammenführen
                self._table.setRowHeight(r, 26)

            elif kind == "item":
                menge = abs(b.get("menge", 0))
                self._table.setItem(r, 0, _cell("", self._COL_ITEM_BG))
                self._table.setItem(r, 1, _cell(f"    ↳  {b.get('artikel_name', '')}", self._COL_ITEM_BG))
                self._table.setItem(r, 2, _cell(str(menge), self._COL_ITEM_BG,
                                                fg=QColor("#c0392b"),
                                                align=Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter))
                self._table.setItem(r, 3, _cell("", self._COL_ITEM_BG))
                _note_m = re.search(r"##\s*(.*)", bemerkung_raw)
                _item_note = _note_m.group(1).strip() if _note_m else ""
                self._table.setItem(r, 4, _cell(_item_note, self._COL_ITEM_BG))

            else:  # flat
                menge = abs(b.get("menge", 0))
                bemerkung = re.sub(r"\s*\(G?ID\s*\d+\)", "", bemerkung_raw).strip()
                self._table.setItem(r, 0, _cell(datum_fmt, self._COL_FLAT_BG))
                self._table.setItem(r, 1, _cell(b.get("artikel_name", ""), self._COL_FLAT_BG))
                self._table.setItem(r, 2, _cell(str(menge), self._COL_FLAT_BG,
                                                fg=QColor("#c0392b"),
                                                align=Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter))
                self._table.setItem(r, 3, _cell(b.get("von", ""), self._COL_FLAT_BG))
                self._table.setItem(r, 4, _cell(bemerkung, self._COL_FLAT_BG))

        self._table.blockSignals(False)

        pages = max(1, (self._total + PAGE_SIZE - 1) // PAGE_SIZE)
        self._lbl_info.setText(f"{self._total} Einträge  |  Seite {self._page + 1} / {pages}")
        self._btn_prev.setEnabled(self._page > 0)
        self._btn_next.setEnabled((self._page + 1) * PAGE_SIZE < self._total)

        # Checkboxen setzen (nach Befüllung der Zeilen)
        self._table.blockSignals(True)
        for r, (kind, eid, bid) in enumerate(self._display_meta):
            it = self._table.item(r, 0)
            if it is None:
                continue
            it.setFlags(it.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            if kind == "header":
                chk = Qt.CheckState.Unchecked if eid in self._excluded_eids else Qt.CheckState.Checked
            elif kind == "item":
                chk = Qt.CheckState.Unchecked if (
                    bid in self._excluded_ids or eid in self._excluded_eids
                ) else Qt.CheckState.Checked
            else:  # flat
                chk = Qt.CheckState.Unchecked if bid in self._excluded_ids else Qt.CheckState.Checked
            it.setCheckState(chk)
        self._table.blockSignals(False)

    # ── Slots ────────────────────────────────────────────────────────────────

    def _on_filter(self):
        self._page = 0
        self._excluded_ids.clear()
        self._excluded_eids.clear()
        self._laden()

    def _reset_filter(self):
        self._von.setDate(QDate(QDate.currentDate().year(), 1, 1))
        self._bis.setDate(QDate.currentDate())
        self._suche.clear()
        self._page = 0
        self._excluded_ids.clear()
        self._excluded_eids.clear()
        self._laden()

    def _prev_page(self):
        if self._page > 0:
            self._page -= 1
            self._laden()

    def _next_page(self):
        if (self._page + 1) * PAGE_SIZE < self._total:
            self._page += 1
            self._laden()

    def _export_csv(self):
        ausgeschlossen = len(self._excluded_ids) + len(self._excluded_eids)
        if ausgeschlossen > 0:
            antwort = QMessageBox.question(
                self, "Ausschlüsse vorhanden",
                f"{ausgeschlossen} Einträge / Einsätze sind vom Export ausgeschlossen.\n"
                f"Trotzdem exportieren (nur ausgewählte)?\n\n"
                f"'Ja' = nur ausgewählte  |  'Nein' = abbrechen",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if antwort != QMessageBox.StandardButton.Yes:
                return

        pfad, _ = QFileDialog.getSaveFileName(
            self, "Excel speichern", f"sanmat_verbrauch_{date.today()}.xlsx", "Excel (*.xlsx)"
        )
        if not pfad:
            return

        von_str = self._von.date().toString("yyyy-MM-dd")
        bis_str = self._bis.date().toString("yyyy-MM-dd")
        suche   = self._suche.text().strip() or None
        alle = self._db.get_buchungen(limit=99999, offset=0, typ="verbrauch",
                                      datum_von=von_str, datum_bis=bis_str, suche=suche)

        # ── Daten nach Datum + Gruppe strukturieren ────────────────────────
        nach_datum: dict[str, dict] = defaultdict(dict)
        for b in alle:
            bid = b.get("id")
            bem = b.get("bemerkung", "") or ""
            m = re.search(r"\((G?ID)\s*(\d+)\)", bem)
            group_key = f"{m.group(1)}_{m.group(2)}" if m else f"flat_{bid}"
            if bid in self._excluded_ids:
                continue
            if m and group_key in self._excluded_eids:
                continue
            datum_raw = b.get("datum", "")[:10]
            try:
                y, mo, d = datum_raw.split("-")
                datum_fmt = f"{d}.{mo}.{y}"
            except Exception:
                datum_fmt = datum_raw
            teile = bem.split("##")
            stichwort = re.sub(r"\s*\(G?ID\s*\d+\)", "", teile[0]).strip()
            notiz = teile[1].strip() if len(teile) > 1 else ""
            if datum_fmt not in nach_datum:
                nach_datum[datum_fmt] = {}
            if group_key not in nach_datum[datum_fmt]:
                nach_datum[datum_fmt][group_key] = {
                    "stichwort": stichwort,
                    "entnehmer": b.get("von", "") or "",
                    "artikel": [],
                }
            nach_datum[datum_fmt][group_key]["artikel"].append((
                b.get("artikel_name", ""),
                abs(b.get("menge", 0)),
                notiz,
            ))

        # ── Hilfsfunktionen ────────────────────────────────────────────────
        def _bg(color: str) -> PatternFill:
            return PatternFill("solid", fgColor=color)

        def _border() -> Border:
            s = Side(style="thin", color="00CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        CLR_FG = "00FFFFFF"

        # ── Workbook aufbauen ──────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verbrauch"

        # Titelzeile
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "Sanitätsmaterial-Verbrauchsprotokoll – DRK Erste-Hilfe-Station FKB"
        c.font = Font(bold=True, size=14, color=CLR_FG)
        c.fill = _bg("00334155")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Zeitraum-Zeile
        ws.merge_cells("A2:F2")
        c = ws["A2"]
        von_zeige = self._von.date().toString("dd.MM.yyyy")
        bis_zeige = self._bis.date().toString("dd.MM.yyyy")
        c.value = (f"Exportiert am: {date.today().strftime('%d.%m.%Y')}"
                   f"   |   Zeitraum: {von_zeige} – {bis_zeige}")
        c.font = Font(italic=True, size=10, color="00666666")
        c.fill = _bg("00F5F5F5")
        c.alignment = Alignment(horizontal="right")

        # Spalten-Header
        for col, h in enumerate(["Datum", "Einsatz / Grund", "Artikel", "Menge", "Entnehmer", "Notiz"], 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, color=CLR_FG)
            c.fill = _bg("001565A8")
            c.alignment = Alignment(horizontal="center")
            c.border = _border()
        ws.row_dimensions[3].height = 20

        # ── Datenzeilen ────────────────────────────────────────────────────
        artikel_gesamt: dict[str, int] = defaultdict(int)
        r = 4
        date_colors = ["00F5F5F5", "00FAFAFA"]
        exportiert = 0

        for di, datum in enumerate(sorted(nach_datum.keys(),
                                          key=lambda d: tuple(reversed(d.split("."))))):  # noqa: E501
            gruppen = nach_datum[datum]
            # Datums-Trennzeile
            ws.merge_cells(f"A{r}:F{r}")
            c = ws.cell(row=r, column=1, value=f"  {datum}")
            c.font = Font(bold=True, size=11, color=CLR_FG)
            c.fill = _bg("00546E7A")
            c.border = _border()
            ws.row_dimensions[r].height = 18
            r += 1

            item_bg = date_colors[di % len(date_colors)]
            for gruppe in gruppen.values():
                for i, (art, menge, notiz) in enumerate(gruppe["artikel"]):
                    ws.cell(row=r, column=1, value=datum if i == 0 else "").fill = _bg(item_bg)
                    ws.cell(row=r, column=2, value=gruppe["stichwort"] if i == 0 else "").fill = _bg(item_bg)
                    ws.cell(row=r, column=3, value=art).fill = _bg(item_bg)
                    c_m = ws.cell(row=r, column=4, value=menge)
                    c_m.font = Font(bold=True, color="00263238")
                    c_m.alignment = Alignment(horizontal="center")
                    c_m.fill = _bg(item_bg)
                    ws.cell(row=r, column=5, value=gruppe["entnehmer"] if i == 0 else "").fill = _bg(item_bg)
                    ws.cell(row=r, column=6, value=notiz).fill = _bg(item_bg)
                    for col in range(1, 7):
                        ws.cell(row=r, column=col).border = _border()
                    artikel_gesamt[art] += menge
                    exportiert += 1
                    r += 1
            r += 1  # Leerzeile

        # ── Gesamtübersicht ────────────────────────────────────────────────
        r += 1
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value="  Alle verbrauchten Artikel – Gesamtübersicht")
        c.font = Font(bold=True, size=12, color=CLR_FG)
        c.fill = _bg("00455A64")
        c.border = _border()
        ws.row_dimensions[r].height = 20
        r += 1

        for col, h in enumerate(["Artikel", "", "", "Verbraucht", "Einheit", ""], 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color=CLR_FG)
            c.fill = _bg("00607D8B")
            c.border = _border()
        r += 1

        for i, (art, menge) in enumerate(sorted(artikel_gesamt.items(), key=lambda x: x[0])):
            bg = "00F5F5F5" if i % 2 == 0 else "00FFFFFF"
            ws.cell(row=r, column=1, value=f"  {art}").fill = _bg(bg)
            ws.cell(row=r, column=1).border = _border()
            c_m = ws.cell(row=r, column=4, value=menge)
            c_m.font = Font(bold=True)
            c_m.alignment = Alignment(horizontal="center")
            c_m.fill = _bg(bg)
            c_m.border = _border()
            ws.cell(row=r, column=5, value="Stück").fill = _bg(bg)
            ws.cell(row=r, column=5).border = _border()
            for col in [2, 3, 6]:
                ws.cell(row=r, column=col).fill = _bg(bg)
                ws.cell(row=r, column=col).border = _border()
            r += 1

        # ── Spaltenbreiten / Freeze ────────────────────────────────────────
        for i, w in enumerate([12, 30, 42, 8, 15, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A4"
        ws.sheet_view.showGridLines = False

        try:
            wb.save(pfad)
            QMessageBox.information(self, "Export",
                                     f"Excel gespeichert ({exportiert} Zeilen):\n{pfad}")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", str(e))

    # ── Checkbox-Logik ───────────────────────────────────────────────────────

    def _on_checkbox_changed(self, item: QTableWidgetItem):
        """Checkbox in Spalte 0 – Einsatz-Header kaskadiert auf alle zugehörigen Items."""
        if item.column() != 0 or self._chk_lock:
            return
        row = item.row()
        if row >= len(self._display_meta):
            return
        kind, eid, bid = self._display_meta[row]
        checked = item.checkState() == Qt.CheckState.Checked

        if kind == "header":
            if checked:
                self._excluded_eids.discard(eid)
            else:
                self._excluded_eids.add(eid)
            # Items dieses Einsatzes kaskadieren
            self._chk_lock = True
            self._table.blockSignals(True)
            for r2, (k2, e2, b2) in enumerate(self._display_meta):
                if k2 == "item" and e2 == eid:
                    it2 = self._table.item(r2, 0)
                    if it2:
                        it2.setCheckState(item.checkState())
                    if checked:
                        self._excluded_ids.discard(b2)
                    else:
                        if b2 is not None:
                            self._excluded_ids.add(b2)
            self._table.blockSignals(False)
            self._chk_lock = False

        elif kind == "item":
            if checked:
                self._excluded_ids.discard(bid)
            else:
                if bid is not None:
                    self._excluded_ids.add(bid)

        else:  # flat
            if checked:
                self._excluded_ids.discard(bid)
            else:
                if bid is not None:
                    self._excluded_ids.add(bid)

    def _auswahl_alle(self):
        self._excluded_ids.clear()
        self._excluded_eids.clear()
        self._table.blockSignals(True)
        for r in range(self._table.rowCount()):
            it = self._table.item(r, 0)
            if it:
                it.setCheckState(Qt.CheckState.Checked)
        self._table.blockSignals(False)

    def _auswahl_keine(self):
        self._table.blockSignals(True)
        for r, (kind, eid, bid) in enumerate(self._display_meta):
            it = self._table.item(r, 0)
            if it:
                it.setCheckState(Qt.CheckState.Unchecked)
            if kind == "header" and eid is not None:
                self._excluded_eids.add(eid)
            elif bid is not None:
                self._excluded_ids.add(bid)
        self._table.blockSignals(False)

    # ── Kontextmenü ──────────────────────────────────────────────────────────

    def _kontextmenu(self, pos: QPoint):
        try:
            row = self._table.rowAt(pos.y())
            if row < 0 or row >= len(self._display_meta):
                return
            kind, eid, bid = self._display_meta[row]
        except Exception as e:
            QMessageBox.warning(self, "Kontextmenü Fehler", str(e))
            return

        menu = QMenu(self)

        if kind == "header":
            # Ganzen Einsatz löschen
            act_del_all = menu.addAction("🗑  Gesamten Einsatz löschen")
            act_undo    = menu.addAction("↩  Rückgängig") if self._undo_stack else None
            action = menu.exec(self._table.viewport().mapToGlobal(pos))
            if action == act_del_all:
                self._einsatz_loeschen(eid)
            elif act_undo and action == act_undo:
                self._undo()
        elif kind in ("item", "flat"):
            act_edit    = menu.addAction("✏  Bearbeiten")
            act_del     = menu.addAction("🗑  Position löschen")
            if kind == "item":
                act_del_all = menu.addAction("🗑  Gesamten Einsatz löschen")
            else:
                act_del_all = None
            menu.addSeparator()
            act_undo = menu.addAction("↩  Rückgängig") if self._undo_stack else None
            action = menu.exec(self._table.viewport().mapToGlobal(pos))
            if action == act_edit:
                self._position_bearbeiten(bid)
            elif action == act_del:
                self._position_loeschen(bid)
            elif act_del_all and action == act_del_all:
                self._einsatz_loeschen(eid)
            elif act_undo and action == act_undo:
                self._undo()

    def _position_loeschen(self, bid: int):
        if bid is None:
            return
        snap = self._db.get_buchung_by_id(bid)
        if not snap:
            return
        antwort = QMessageBox.question(
            self, "Position löschen",
            f"Buchung '{snap['artikel_name']}' (Menge {abs(snap['menge'])}) wirklich löschen?\n"
            "Der Bestand wird entsprechend wiederhergestellt.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort != QMessageBox.StandardButton.Yes:
            return
        ok, msg = self._db.delete_buchung(bid)
        if ok:
            self._undo_stack.append({"aktion": "loeschen", "snapshot": snap})
            self._btn_undo.setEnabled(True)
            self._laden()
        else:
            QMessageBox.warning(self, "Fehler", msg)

    def _einsatz_loeschen(self, eid: int):
        if eid is None:
            return
        # Alle Buchungs-IDs dieses Einsatzes sammeln
        bids = [bid for (kind, e, bid) in self._display_meta if e == eid and bid is not None]
        antwort = QMessageBox.question(
            self, "Einsatz löschen",
            f"Alle {len(bids)} Buchungen dieses Einsatzes löschen?\n"
            "Der Bestand wird entsprechend wiederhergestellt.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if antwort != QMessageBox.StandardButton.Yes:
            return
        snapshots = []
        fehler = []
        for bid in bids:
            snap = self._db.get_buchung_by_id(bid)
            if snap:
                snapshots.append(snap)
                ok, msg = self._db.delete_buchung(bid)
                if not ok:
                    fehler.append(msg)
        if snapshots:
            self._undo_stack.append({"aktion": "einsatz_loeschen", "snapshots": snapshots})
            self._btn_undo.setEnabled(True)
        if fehler:
            QMessageBox.warning(self, "Teilfehler", "\n".join(fehler))
        self._laden()

    def _position_bearbeiten(self, bid: int):
        if bid is None:
            return
        snap = self._db.get_buchung_by_id(bid)
        if not snap:
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Position bearbeiten")
        dlg.setMinimumWidth(400)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 18, 20, 18)

        lay.addWidget(QLabel(f"<b>{snap['artikel_name']}</b>"))

        form = QFormLayout()
        form.setSpacing(8)

        sb_menge = QSpinBox()
        sb_menge.setRange(1, 9999)
        sb_menge.setValue(abs(snap["menge"]))
        form.addRow("Menge:", sb_menge)

        le_von = QLineEdit(snap.get("von") or "")
        form.addRow("Entnehmer:", le_von)

        full_bem  = snap.get("bemerkung") or ""
        id_match  = re.search(r"\(G?ID\s*\d+\)", full_bem)
        if id_match:
            gruppe_name = re.sub(r"\s*\(G?ID\s*\d+\).*", "", full_bem).strip()
            note_m      = re.search(r"##\s*(.*)", full_bem)
            bem_note    = note_m.group(1).strip() if note_m else ""
            lbl_gruppe  = QLabel(f"<i>{gruppe_name}</i>")
            form.addRow("Einsatz / Gruppe:", lbl_gruppe)
            le_bem = QLineEdit(bem_note)
            form.addRow("Notiz:", le_bem)
        else:
            le_bem = QLineEdit(full_bem)
            form.addRow("Bemerkung:", le_bem)

        de_datum = QDateEdit()
        de_datum.setDisplayFormat("dd.MM.yyyy")
        de_datum.setCalendarPopup(True)
        try:
            parts = snap["datum"][:10].split("-")
            de_datum.setDate(QDate(int(parts[0]), int(parts[1]), int(parts[2])))
        except Exception:
            de_datum.setDate(QDate.currentDate())
        form.addRow("Datum:", de_datum)

        lay.addLayout(form)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ab = QPushButton("Abbrechen")
        btn_ab.clicked.connect(dlg.reject)
        btn_ok = QPushButton("💾 Speichern")
        btn_ok.setStyleSheet(
            "QPushButton{background:#1565a8;color:white;font-weight:bold;"
            "padding:6px 18px;border-radius:4px;}"
            "QPushButton:hover{background:#1976d2;}"
        )
        btn_row.addWidget(btn_ab)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        def _speichern():
            neue_menge = sb_menge.value()
            von = le_von.text().strip()
            if id_match:
                # Prefix bis inkl. ID-Tag beibehalten, Notiz mit ## anhängen
                prefix_m  = re.match(r"(.*?\(G?ID\s*\d+\))", full_bem)
                prefix    = prefix_m.group(1) if prefix_m else full_bem
                notiz     = le_bem.text().strip()
                bemerkung = f"{prefix}  ## {notiz}" if notiz else prefix
            else:
                bemerkung = le_bem.text().strip()
            datum_iso = de_datum.date().toString("yyyy-MM-dd")
            ok, msg = self._db.update_buchung(bid, neue_menge, von, bemerkung, datum_iso)
            if ok:
                self._undo_stack.append({"aktion": "bearbeiten", "snapshot": snap})
                self._btn_undo.setEnabled(True)
                dlg.accept()
                self._laden()
            else:
                QMessageBox.warning(dlg, "Fehler", msg)

        btn_ok.clicked.connect(_speichern)
        dlg.exec()

    def _undo(self):
        if not self._undo_stack:
            return
        eintrag = self._undo_stack.pop()
        aktion = eintrag["aktion"]

        if aktion == "loeschen":
            ok, msg = self._db.restore_buchung(eintrag["snapshot"])
            if not ok:
                QMessageBox.warning(self, "Rückgängig fehlgeschlagen", msg)
                self._undo_stack.append(eintrag)  # zurück auf Stack
                return

        elif aktion == "einsatz_loeschen":
            fehler = []
            for snap in eintrag["snapshots"]:
                ok, msg = self._db.restore_buchung(snap)
                if not ok:
                    fehler.append(msg)
            if fehler:
                QMessageBox.warning(self, "Rückgängig teilweise fehlgeschlagen", "\n".join(fehler))

        elif aktion == "bearbeiten":
            snap = eintrag["snapshot"]
            ok, msg = self._db.update_buchung(
                snap["id"], abs(snap["menge"]),
                snap.get("von") or "",
                snap.get("bemerkung") or "",
                snap["datum"]
            )
            if not ok:
                QMessageBox.warning(self, "Rückgängig fehlgeschlagen", msg)
                self._undo_stack.append(eintrag)
                return

        self._btn_undo.setEnabled(bool(self._undo_stack))
        self._laden()

    # ── Manueller Verbrauch ──────────────────────────────────────────────────

    def _verbrauch_anlegen(self):
        """Dialog: Verbrauchsgruppe mit mehreren Artikeln anlegen."""
        artikel_liste_db = self._db.get_artikel()
        if not artikel_liste_db:
            QMessageBox.information(self, "Kein Artikel", "Keine Artikel in der Datenbank.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Verbrauchsgruppe anlegen")
        dlg.setMinimumWidth(760)
        dlg.setMinimumHeight(540)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 18, 20, 18)

        lay.addWidget(QLabel("<b>📋  Verbrauchsgruppe anlegen</b>"))
        lay.addWidget(QLabel(
            "<span style='color:#666;font-size:11px;'>"
            "Für Verbräuche außerhalb von Einsätzen (Übung, Ablauf, Eigenbedarf …)"
            "</span>"
        ))

        # ── Kopffelder ──────────────────────────────────────────
        form_top = QFormLayout()
        form_top.setSpacing(8)
        le_stichwort = QLineEdit()
        le_stichwort.setPlaceholderText("z.B. Übung 02.04.2026, MHD-Abgang April …")
        form_top.addRow("Bezeichnung:", le_stichwort)
        cb_grund = QComboBox()
        cb_grund.addItems(self.GRUENDE)
        form_top.addRow("Grund:", cb_grund)
        le_entnehmer = QLineEdit()
        le_entnehmer.setPlaceholderText("Name …")
        form_top.addRow("Entnehmer:", le_entnehmer)
        de_datum = QDateEdit(QDate.currentDate())
        de_datum.setDisplayFormat("dd.MM.yyyy")
        de_datum.setCalendarPopup(True)
        form_top.addRow("Datum:", de_datum)
        lay.addLayout(form_top)

        # ── Splitter: Suche | Warenkorb ─────────────────────────
        split_lay = QHBoxLayout()
        split_lay.setSpacing(12)

        # Linke Seite: Artikelsuche
        left = QVBoxLayout()
        left.setSpacing(5)
        left.addWidget(QLabel("<b>Artikel hinzufügen:</b>"))
        le_art_suche = QLineEdit()
        le_art_suche.setPlaceholderText("Name oder Nummer …")
        left.addWidget(le_art_suche)
        tbl_art = QTableWidget(0, 3)
        tbl_art.setHorizontalHeaderLabels(["Bezeichnung", "Art.-Nr.", "Bestand"])
        tbl_art.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        tbl_art.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        tbl_art.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        tbl_art.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        tbl_art.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl_art.setAlternatingRowColors(True)
        tbl_art.verticalHeader().setVisible(False)
        left.addWidget(tbl_art, 1)
        menge_row = QHBoxLayout()
        menge_row.addWidget(QLabel("Menge:"))
        sb_menge = QSpinBox()
        sb_menge.setRange(1, 9999)
        sb_menge.setValue(1)
        menge_row.addWidget(sb_menge)
        btn_add_art = QPushButton("➕ Hinzufügen")
        menge_row.addWidget(btn_add_art)
        menge_row.addStretch()
        left.addLayout(menge_row)
        split_lay.addLayout(left, 3)

        # Rechte Seite: Warenkorb
        right = QVBoxLayout()
        right.setSpacing(5)
        right.addWidget(QLabel("<b>Warenkorb:</b>"))
        tbl_korb = QTableWidget(0, 3)
        tbl_korb.setHorizontalHeaderLabels(["Artikel", "Menge", ""])
        tbl_korb.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        tbl_korb.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        tbl_korb.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        tbl_korb.horizontalHeader().resizeSection(2, 32)
        tbl_korb.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl_korb.verticalHeader().setVisible(False)
        right.addWidget(tbl_korb, 1)
        split_lay.addLayout(right, 2)
        lay.addLayout(split_lay, 1)

        warenkorb: list[dict] = []

        def _fill_artikel(filter_text: str = ""):
            ft = filter_text.lower()
            tbl_art.setRowCount(0)
            for a in artikel_liste_db:
                if ft and ft not in a["bezeichnung"].lower() \
                        and ft not in (a.get("artikelnr") or "").lower():
                    continue
                r = tbl_art.rowCount()
                tbl_art.insertRow(r)
                tbl_art.setItem(r, 0, QTableWidgetItem(a["bezeichnung"]))
                tbl_art.setItem(r, 1, QTableWidgetItem(a.get("artikelnr") or ""))
                vorrat = a.get("menge", 0)
                it_v = QTableWidgetItem(str(vorrat))
                it_v.setForeground(QColor("#c0392b") if vorrat <= 0 else QColor("#1a7a1a"))
                tbl_art.setItem(r, 2, it_v)
                tbl_art.item(r, 0).setData(Qt.ItemDataRole.UserRole, a["id"])
                tbl_art.item(r, 0).setData(Qt.ItemDataRole.UserRole + 1, a["bezeichnung"])

        def _refresh_korb():
            tbl_korb.setRowCount(0)
            for i, pos in enumerate(warenkorb):
                r = tbl_korb.rowCount()
                tbl_korb.insertRow(r)
                tbl_korb.setItem(r, 0, QTableWidgetItem(pos["bezeichnung"]))
                tbl_korb.setItem(r, 1, QTableWidgetItem(str(pos["menge"])))
                btn_del = QPushButton("✕")
                btn_del.setFixedSize(28, 22)
                btn_del.clicked.connect(lambda _, idx=i: _remove(idx))
                tbl_korb.setCellWidget(r, 2, btn_del)

        def _remove(idx: int):
            if 0 <= idx < len(warenkorb):
                warenkorb.pop(idx)
                _refresh_korb()

        def _add_to_korb():
            row = tbl_art.currentRow()
            if row < 0:
                return
            art_id   = tbl_art.item(row, 0).data(Qt.ItemDataRole.UserRole)
            art_name = tbl_art.item(row, 0).data(Qt.ItemDataRole.UserRole + 1)
            menge = sb_menge.value()
            for pos in warenkorb:
                if pos["artikel_id"] == art_id:
                    pos["menge"] += menge
                    _refresh_korb()
                    return
            warenkorb.append({"artikel_id": art_id, "bezeichnung": art_name, "menge": menge})
            _refresh_korb()

        le_art_suche.textChanged.connect(_fill_artikel)
        btn_add_art.clicked.connect(_add_to_korb)
        tbl_art.doubleClicked.connect(lambda: _add_to_korb())
        _fill_artikel()

        # Buttons
        btn_row_lay = QHBoxLayout()
        btn_row_lay.addStretch()
        btn_ab = QPushButton("Abbrechen")
        btn_ab.clicked.connect(dlg.reject)
        btn_ok = QPushButton("✅ Buchen")
        btn_ok.setStyleSheet(
            "QPushButton{background:#1565a8;color:white;font-weight:bold;"
            "padding:6px 18px;border-radius:4px;}"
            "QPushButton:hover{background:#1976d2;}"
        )
        btn_row_lay.addWidget(btn_ab)
        btn_row_lay.addWidget(btn_ok)
        lay.addLayout(btn_row_lay)

        def _buchen():
            if not warenkorb:
                QMessageBox.warning(dlg, "Kein Artikel",
                                    "Bitte mindestens einen Artikel in den Warenkorb legen.")
                return
            bezeichnung = le_stichwort.text().strip()
            grund = cb_grund.currentText()
            stichwort = f"{grund}: {bezeichnung}" if bezeichnung else grund
            entnehmer = le_entnehmer.text().strip()
            datum_iso = de_datum.date().toString("yyyy-MM-dd")
            ok, msg = self._db.buche_verbrauch_gruppe(
                stichwort=stichwort,
                artikel_liste=warenkorb,
                datum=datum_iso,
                entnehmer=entnehmer,
                negativ_erlaubt=True,
            )
            if ok:
                QMessageBox.information(dlg, "Gebucht", msg)
                dlg.accept()
                self._laden()
            else:
                QMessageBox.warning(dlg, "Fehler", msg)

        btn_ok.clicked.connect(_buchen)
        dlg.exec()

    def showEvent(self, event):
        super().showEvent(event)
        self._laden()
