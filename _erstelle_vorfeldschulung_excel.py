# -*- coding: utf-8 -*-
"""
Erstellt eine Excel-Datei aller aktiven Mitarbeiter mit Vorfeldschulung-Daten.
Spalten: Nr. | Nachname | Vorname | Qualifikation | Datum absolviert | Gültig bis | Status | Bemerkung

Aufruf: python _erstelle_vorfeldschulung_excel.py
Ausgabe: Daten/Schulungen/Vorfeldschulung_JJJJMMTT.xlsx  (+ wird direkt geöffnet)
"""
import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from functions.schulungen_db import lade_mitarbeiter_mit_schulungen

_BASE = Path(os.path.dirname(os.path.abspath(__file__)))
ZIEL  = _BASE / "Daten" / "Schulungen"
ZIEL.mkdir(parents=True, exist_ok=True)

DATEINAME = ZIEL / f"Vorfeldschulung_{date.today().strftime('%Y%m%d')}.xlsx"

# ── Daten laden ───────────────────────────────────────────────────────────────
alle = lade_mitarbeiter_mit_schulungen()

zeilen = []
for ma in alle:
    eintrag = ma["schulungen"].get("Vorfeldschulung")
    zeilen.append({
        "nachname":         ma.get("nachname", ""),
        "vorname":          ma.get("vorname", ""),
        "qualifikation":    ma.get("qualifikation", ""),
        "datum_absolviert": eintrag.get("datum_absolviert", "") if eintrag else "",
        "gueltig_bis":      eintrag.get("gueltig_bis", "")      if eintrag else "",
        "status":           eintrag.get("status", "")            if eintrag else "",
        "bemerkung":        eintrag.get("bemerkung", "")         if eintrag else "",
    })

# Sortierung: Mitarbeiter ohne Vorfeldschulung ans Ende, dann alphabetisch
zeilen.sort(key=lambda z: (z["gueltig_bis"] == "", z["nachname"].lower(), z["vorname"].lower()))

# ── Excel erstellen ────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Vorfeldschulung"

# Farben
HEADER_BG   = "1565a8"   # DRK-Blau
HEADER_FG   = "FFFFFF"
ROW_ALT     = "EEF4FB"   # helles Blau alternierend
ROW_NORMAL  = "FFFFFF"
GRUEN       = "C8E6C9"   # gültig
ROT         = "FFCDD2"   # abgelaufen / fehlt
ORANGE_BG   = "FFE0B2"   # bald ablaufend
GRAU        = "F5F5F5"   # kein Eintrag

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

SPALTEN = [
    ("Nr.",              7),
    ("Nachname",        18),
    ("Vorname",         16),
    ("Qualifikation",   18),
    ("Datum absolviert", 18),
    ("Gültig bis",      16),
    ("Status",          14),
    ("Bemerkung",       30),
]

# Titel-Zeile
ws.merge_cells("A1:H1")
titel_cell = ws["A1"]
titel_cell.value = f"Vorfeldschulung – Übersicht (Stand: {date.today().strftime('%d.%m.%Y')})"
titel_cell.font      = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
titel_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
titel_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# Header-Zeile
for col, (name, breite) in enumerate(SPALTEN, start=1):
    c = ws.cell(row=2, column=col, value=name)
    c.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border
    ws.column_dimensions[get_column_letter(col)].width = breite
ws.row_dimensions[2].height = 20

# Daten
heute = date.today()
for idx, z in enumerate(zeilen, start=1):
    row = idx + 2

    # Zeilenhintergrund bestimmen
    gb_str = z["gueltig_bis"]
    hat_eintrag = bool(gb_str or z["datum_absolviert"])
    if not hat_eintrag:
        row_fill = PatternFill("solid", fgColor="F0F0F0")
    else:
        # Ablauf-Datum auswerten
        try:
            tag, mon, jahr = gb_str.split(".")
            gb_date = date(int(jahr), int(mon), int(tag))
            diff = (gb_date - heute).days
            if diff < 0:
                row_fill = PatternFill("solid", fgColor=ROT)
            elif diff <= 90:
                row_fill = PatternFill("solid", fgColor=ORANGE_BG)
            else:
                row_fill = PatternFill("solid", fgColor=GRUEN)
        except Exception:
            row_fill = PatternFill("solid", fgColor=ROW_ALT if idx % 2 == 0 else ROW_NORMAL)

    werte = [
        idx,
        z["nachname"],
        z["vorname"],
        z["qualifikation"],
        z["datum_absolviert"],
        z["gueltig_bis"],
        z["status"],
        z["bemerkung"],
    ]
    for col, wert in enumerate(werte, start=1):
        c = ws.cell(row=row, column=col, value=wert)
        c.font      = Font(name="Calibri", size=11)
        c.fill      = row_fill
        c.border    = border
        c.alignment = Alignment(vertical="center", wrap_text=(col == 8))
        if col == 1:
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 16

# Autofilter auf Header
ws.auto_filter.ref = f"A2:H{len(zeilen) + 2}"

# Zeile fixieren
ws.freeze_panes = "A3"

# ── Speichern & öffnen ─────────────────────────────────────────────────────────
wb.save(DATEINAME)
print(f"Gespeichert: {DATEINAME}")
os.startfile(str(DATEINAME))
