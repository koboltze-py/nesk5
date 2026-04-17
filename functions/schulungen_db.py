"""
Schulungen-Datenbank (schulungen.db)
Tabellen:
  mitarbeiter        – Stammdaten aller Mitarbeiter
  schulungseintraege – Schulungs-/Zertifizierungs-Einträge pro Mitarbeiter (mit Ablauf)
  schulungen_manuell – Ältere manuelle Einträge (Backward-Compat.)

Excel-Stammdaten werden über excel_importieren() eingelesen.
"""
import sqlite3
import re
from pathlib import Path
from datetime import datetime, date, timedelta
from config import BASE_DIR as _BASE_DIR

_DB_PFAD   = Path(_BASE_DIR) / "database SQL" / "schulungen.db"
_EXCEL_PFAD = (
    Path(_BASE_DIR) / "Daten" / "Stammdaten Schulungen"
    / "2026_03_18_Mitarbeiter.xlsx"
)

# ─── Schulungstyp-Konfiguration ───────────────────────────────────────────────
# ablauf:
#   "direkt"    → gueltig_bis = gespeicherter Excel-Wert
#   "intervall" → gueltig_bis = datum_absolviert + N Jahre
#   "einmalig"  → kein Ablauf, nur Ja/Nein
# laeuft_nicht_ab: kein Farbwarner im Kalender (z. B. ärztl. Untersuchung)
SCHULUNGSTYPEN_CFG = {
    "ZÜP":                    {"anzeige": "ZÜP",                   "ablauf": "direkt",    "intervall": None, "laeuft_nicht_ab": False},
    "EH":                     {"anzeige": "EH",                    "ablauf": "intervall", "intervall": 2,    "laeuft_nicht_ab": False},
    "Refresher":              {"anzeige": "Refresher",             "ablauf": "intervall", "intervall": 1,    "laeuft_nicht_ab": False},
    "Aerztl_Untersuchung":    {"anzeige": "Ärztl. Untersuchung",  "ablauf": "direkt",    "intervall": None, "laeuft_nicht_ab": True},
    "Fuehrerschein_Kont":     {"anzeige": "Führerschein Kontrolle","ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Einw_Zertifikate":       {"anzeige": "Einweisung Zertifikate","ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Fixierung":              {"anzeige": "Fixierung",             "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Einw_eMobby":            {"anzeige": "Einweisung e-Mobby",   "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Bulmor":                 {"anzeige": "Bulmor/Staplerschein", "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Arbeitsschutz":          {"anzeige": "Arbeitsschutz",         "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Einw_QM":                {"anzeige": "Einweisung QM",         "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Fragebogen_Schulung":    {"anzeige": "Fragebogen Schulung",  "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Personalausweis":        {"anzeige": "Personalausweis/Pass",  "ablauf": "einmalig",  "intervall": None, "laeuft_nicht_ab": True},
    "Sicherheitsschulung":    {"anzeige": "Sicherheitsschulung",   "ablauf": "intervall", "intervall": 5,    "laeuft_nicht_ab": False},
    "Vorfeldschulung":         {"anzeige": "Vorfeldschulung",        "ablauf": "direkt",    "intervall": None, "laeuft_nicht_ab": False},
    "PRM_Schulung":           {"anzeige": "PRM-Schulung",           "ablauf": "direkt",    "intervall": None, "laeuft_nicht_ab": False},
    "Sonstiges":              {"anzeige": "Sonstiges",              "ablauf": "direkt",    "intervall": None, "laeuft_nicht_ab": False},
}

# Listen für UI-Dropdowns (Backward-Compat.)
SCHULUNGSARTEN = [cfg["anzeige"] for cfg in SCHULUNGSTYPEN_CFG.values()]
STATUS_OPTIONEN = ["gültig", "abgelaufen", "ausstehend", "nicht erforderlich"]

# Anzeigenamen → Key
_ANZEIGE_ZU_KEY = {cfg["anzeige"]: key for key, cfg in SCHULUNGSTYPEN_CFG.items()}



# ─── DB-Verbindung ────────────────────────────────────────────────────────────
def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(_DB_PFAD, timeout=5)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous  = NORMAL")
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout  = 5000")
    return conn


def _init_db():
    _DB_PFAD.parent.mkdir(parents=True, exist_ok=True)
    with _connect() as conn:
        # ── Mitarbeiter-Stammdaten ──────────────────────────────────────────
        conn.execute("""
        CREATE TABLE IF NOT EXISTS mitarbeiter (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            nachname      TEXT NOT NULL,
            vorname       TEXT NOT NULL,
            geburtsdatum  TEXT,
            anstellung    TEXT,
            qualifikation TEXT,
            bemerkung     TEXT,
            aktiv         INTEGER DEFAULT 1,
            erstellt_am   TEXT NOT NULL
        )""")
        # ── Schulungseinträge pro Mitarbeiter ───────────────────────────────
        conn.execute("""
        CREATE TABLE IF NOT EXISTS schulungseintraege (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            mitarbeiter_id   INTEGER NOT NULL REFERENCES mitarbeiter(id) ON DELETE CASCADE,
            schulungstyp     TEXT NOT NULL,
            datum_absolviert TEXT,
            gueltig_bis      TEXT,
            laeuft_nicht_ab  INTEGER DEFAULT 0,
            status           TEXT DEFAULT 'gültig',
            bemerkung        TEXT,
            zuletzt_akt      TEXT
        )""")
        # ── Manuelle Einträge (Backward-Compat. + alte Tabelle) ─────────────
        conn.execute("""
        CREATE TABLE IF NOT EXISTS schulungen_manuell (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            erstellt_am      TEXT NOT NULL,
            mitarbeiter      TEXT NOT NULL,
            schulungsart     TEXT NOT NULL,
            datum            TEXT NOT NULL,
            gueltig_bis      TEXT,
            status           TEXT NOT NULL DEFAULT 'gültig',
            bemerkung        TEXT,
            aufgenommen_von  TEXT
        )""")
        # Migration: alte "schulungen"-Tabelle → schulungen_manuell
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]
        if "schulungen" in tables and "schulungen_manuell" in tables:
            count = conn.execute("SELECT COUNT(*) FROM schulungen").fetchone()[0]
            if count > 0:
                conn.execute("""
                    INSERT OR IGNORE INTO schulungen_manuell
                      (id, erstellt_am, mitarbeiter, schulungsart, datum, gueltig_bis, status, bemerkung, aufgenommen_von)
                    SELECT id, erstellt_am, mitarbeiter, schulungsart, datum, gueltig_bis, status, bemerkung, aufgenommen_von
                    FROM schulungen
                """)
        # Spalten-Migration: informiert / informiert_am
        for _col_def in [
            "ALTER TABLE schulungseintraege ADD COLUMN informiert INTEGER DEFAULT 0",
            "ALTER TABLE schulungseintraege ADD COLUMN informiert_am TEXT",
        ]:
            try:
                conn.execute(_col_def)
            except Exception:
                pass
        conn.commit()



# ─── Hilfsfunktionen ──────────────────────────────────────────────────────────
def _parse_datum(val) -> "date | None":
    """Wandelt Excel-Datetime, Datum-Objekt oder String in date-Objekt um."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val or val.upper() in ("X", "JA", "NEIN", "AB APRIL", "—", ""):
            return None
        m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", val)
        if m:
            try:
                return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                return None
        m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", val)
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None
    return None


def _datum_str(d: "date | None") -> str:
    return d.strftime("%d.%m.%Y") if d else ""


def _berechne_gueltig_bis(key: str, datum_absolviert: "date | None",
                           gueltig_bis_direkt: "date | None") -> "date | None":
    cfg = SCHULUNGSTYPEN_CFG.get(key, {})
    ablauf = cfg.get("ablauf", "einmalig")
    if ablauf == "direkt":
        return gueltig_bis_direkt or datum_absolviert
    elif ablauf == "intervall":
        if datum_absolviert is None:
            return None
        jahre = cfg.get("intervall") or 1
        try:
            return datum_absolviert.replace(year=datum_absolviert.year + jahre)
        except ValueError:
            return date(datum_absolviert.year + jahre, datum_absolviert.month, 28)
    return None


def _berechne_status(gueltig_bis: "date | None", laeuft_nicht_ab: bool) -> str:
    if laeuft_nicht_ab:
        return "gültig"
    if gueltig_bis is None:
        return "ausstehend"
    if gueltig_bis < date.today():
        return "abgelaufen"
    return "gültig"


def _dringlichkeit(gueltig_bis: "date | None", laeuft_nicht_ab: bool,
                   referenz: "date | None" = None) -> str:
    """Gibt 'rot', 'orange', 'gelb', 'ok', 'abgelaufen' oder '' zurück.
    referenz: Vergleichsdatum (Standard: heute). Im Kalender wird der
    erste Tag des angezeigten Monats übergeben, damit die Farbe zum
    dargestellten Zeitraum passt und nicht immer von heute aus gerechnet wird.
    """
    if laeuft_nicht_ab or gueltig_bis is None:
        return ""
    basis = referenz if referenz is not None else date.today()
    diff  = (gueltig_bis - basis).days
    if diff < 0:
        return "abgelaufen"
    if diff <= 31:
        return "rot"
    if diff <= 61:
        return "orange"
    if diff <= 92:
        return "gelb"
    return "ok"


# ─── Mitarbeiter-CRUD ─────────────────────────────────────────────────────────
def lade_alle_mitarbeiter(aktiv_only: bool = True) -> list[dict]:
    _init_db()
    sql = "SELECT * FROM mitarbeiter"
    if aktiv_only:
        sql += " WHERE aktiv=1"
    sql += " ORDER BY nachname, vorname"
    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql).fetchall()
    return [dict(r) for r in rows]


def speichere_mitarbeiter(daten: dict) -> int:
    _init_db()
    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        cur = conn.execute(
            """INSERT INTO mitarbeiter
               (nachname, vorname, geburtsdatum, anstellung, qualifikation, bemerkung, aktiv, erstellt_am)
               VALUES (?,?,?,?,?,?,1,?)""",
            (daten.get("nachname", ""),
             daten.get("vorname", ""),
             daten.get("geburtsdatum", ""),
             daten.get("anstellung", ""),
             daten.get("qualifikation", ""),
             daten.get("bemerkung", ""),
             now),
        )
        conn.commit()
        return cur.lastrowid


def aktualisiere_mitarbeiter(ma_id: int, daten: dict) -> None:
    _init_db()
    with _connect() as conn:
        conn.execute(
            """UPDATE mitarbeiter SET
               nachname=?, vorname=?, geburtsdatum=?, anstellung=?,
               qualifikation=?, bemerkung=?, aktiv=?
               WHERE id=?""",
            (daten.get("nachname", ""), daten.get("vorname", ""),
             daten.get("geburtsdatum", ""), daten.get("anstellung", ""),
             daten.get("qualifikation", ""), daten.get("bemerkung", ""),
             int(daten.get("aktiv", 1)), ma_id),
        )
        conn.commit()


def lade_mitarbeiter_namen() -> list[str]:
    """Vollständige Namen aller aktiven Mitarbeiter (für Dropdowns)."""
    alle = lade_alle_mitarbeiter(aktiv_only=True)
    return [f"{m['nachname']}, {m['vorname']}".strip(", ") for m in alle]


# ─── Schulungseinträge-CRUD ───────────────────────────────────────────────────
def speichere_schulungseintrag(daten: dict) -> int:
    _init_db()
    now = datetime.now().isoformat(timespec="seconds")
    key = daten.get("schulungstyp", "Sonstiges")
    cfg = SCHULUNGSTYPEN_CFG.get(key, SCHULUNGSTYPEN_CFG["Sonstiges"])
    with _connect() as conn:
        cur = conn.execute(
            """INSERT INTO schulungseintraege
               (mitarbeiter_id, schulungstyp, datum_absolviert, gueltig_bis,
                laeuft_nicht_ab, status, bemerkung, zuletzt_akt,
                informiert, informiert_am)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (daten.get("mitarbeiter_id"),
             key,
             daten.get("datum_absolviert", ""),
             daten.get("gueltig_bis", ""),
             int(cfg.get("laeuft_nicht_ab", False)),
             daten.get("status", "gültig"),
             daten.get("bemerkung", ""),
             now,
             int(daten.get("informiert", 0)),
             daten.get("informiert_am") or None),
        )
        conn.commit()
        return cur.lastrowid


def aktualisiere_schulungseintrag(eintrag_id: int, daten: dict) -> None:
    _init_db()
    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        conn.execute(
            """UPDATE schulungseintraege SET
               schulungstyp=?, datum_absolviert=?, gueltig_bis=?,
               laeuft_nicht_ab=?, status=?, bemerkung=?, zuletzt_akt=?,
               informiert=?, informiert_am=?
               WHERE id=?""",
            (daten.get("schulungstyp", ""),
             daten.get("datum_absolviert", ""),
             daten.get("gueltig_bis", ""),
             int(daten.get("laeuft_nicht_ab", 0)),
             daten.get("status", "gültig"),
             daten.get("bemerkung", ""),
             now,
             int(daten.get("informiert", 0)),
             daten.get("informiert_am") or None,
             eintrag_id),
        )
        conn.commit()


def loesche_schulungseintrag(eintrag_id: int) -> None:
    _init_db()
    with _connect() as conn:
        conn.execute("DELETE FROM schulungseintraege WHERE id=?", (eintrag_id,))
        conn.commit()


def loesche_mitarbeiter(ma_id: int) -> None:
    """Löscht einen Mitarbeiter und alle seine Schulungseinträge."""
    _init_db()
    with _connect() as conn:
        conn.execute("DELETE FROM schulungseintraege WHERE mitarbeiter_id=?", (ma_id,))
        conn.execute("DELETE FROM mitarbeiter WHERE id=?", (ma_id,))
        conn.commit()


def lade_schulungseintraege(ma_id: int) -> list[dict]:
    _init_db()
    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM schulungseintraege WHERE mitarbeiter_id=? ORDER BY schulungstyp",
            (ma_id,)
        ).fetchall()
    return [dict(r) for r in rows]


# ─── Kalender-Abfragen ────────────────────────────────────────────────────────
def lade_ablaufende(monate: int = 3) -> list[dict]:
    """
    Gibt Schulungseinträge zurück, die ablaufen oder abgelaufen sind.
    Pro Mitarbeiter + Schulungstyp wird nur der NEUESTE Eintrag berücksichtigt.
    Zeigt:
      - Einträge, die in den nächsten N Monaten ablaufen
      - Einträge, die in den letzten N Monaten abgelaufen sind
    Sehr alte abgelaufene Einträge (> N Monate her) werden ausgeblendet.
    """
    _init_db()
    heute    = date.today()
    fenster  = monate * 31  # Tage für Vor- und Rückblick
    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        # Nur den neuesten Eintrag pro MA + Typ (MAX gueltig_bis)
        rows = conn.execute(
            """SELECT se.*, m.nachname, m.vorname, m.qualifikation
               FROM schulungseintraege se
               JOIN mitarbeiter m ON m.id = se.mitarbeiter_id
               WHERE se.laeuft_nicht_ab = 0
                 AND se.gueltig_bis IS NOT NULL
                 AND se.gueltig_bis != ''
                 AND se.id IN (
                     SELECT id FROM schulungseintraege se2
                     WHERE se2.laeuft_nicht_ab = 0
                       AND se2.mitarbeiter_id = se.mitarbeiter_id
                       AND se2.schulungstyp   = se.schulungstyp
                       AND se2.gueltig_bis = (
                           SELECT MAX(se3.gueltig_bis)
                           FROM schulungseintraege se3
                           WHERE se3.mitarbeiter_id = se.mitarbeiter_id
                             AND se3.schulungstyp   = se.schulungstyp
                       )
                 )
               ORDER BY se.gueltig_bis""",
        ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        gb = _parse_datum(d.get("gueltig_bis"))
        if gb is None:
            continue
        diff = (gb - heute).days
        # Nur Einträge innerhalb des Fensters (weder zu weit in der Zukunft,
        # noch schon zu lange abgelaufen)
        if -fenster <= diff <= fenster:
            d["_datum_obj"]     = gb
            d["_tage_rest"]     = diff
            d["_dringlichkeit"] = _dringlichkeit(gb, False)
            d["_name"] = f"{d.get('nachname','')} {d.get('vorname','')}".strip()
            result.append(d)
    return result


def lade_eintraege_fuer_export(von_datum: date, bis_datum: date,
                                schulungstypen: list | None = None) -> list[dict]:
    """
    Gibt Schulungseinträge zurück, bei denen gueltig_bis im Bereich
    [von_datum, bis_datum] liegt. Pro MA+Schulungstyp nur den neuesten Eintrag.
    schulungstypen: Liste von Typ-Keys (z. B. ['EH', 'Refresher']).
                    None = alle Typen berücksichtigen.
    """
    _init_db()
    heute = date.today()
    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """SELECT se.*, m.nachname, m.vorname, m.qualifikation
               FROM schulungseintraege se
               JOIN mitarbeiter m ON m.id = se.mitarbeiter_id
               WHERE se.laeuft_nicht_ab = 0
                 AND se.gueltig_bis IS NOT NULL
                 AND se.gueltig_bis != ''
                 AND se.id IN (
                     SELECT id FROM schulungseintraege se2
                     WHERE se2.laeuft_nicht_ab = 0
                       AND se2.mitarbeiter_id = se.mitarbeiter_id
                       AND se2.schulungstyp   = se.schulungstyp
                       AND se2.gueltig_bis = (
                           SELECT MAX(se3.gueltig_bis)
                           FROM schulungseintraege se3
                           WHERE se3.mitarbeiter_id = se.mitarbeiter_id
                             AND se3.schulungstyp   = se.schulungstyp
                       )
                 )
               ORDER BY se.gueltig_bis""",
        ).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        if schulungstypen and d.get("schulungstyp") not in schulungstypen:
            continue
        gb = _parse_datum(d.get("gueltig_bis"))
        if gb is None:
            continue
        if not (von_datum <= gb <= bis_datum):
            continue
        diff = (gb - heute).days
        d["_datum_obj"]     = gb
        d["_tage_rest"]     = diff
        d["_dringlichkeit"] = _dringlichkeit(gb, False)
        d["_name"] = f"{d.get('nachname','')} {d.get('vorname','')}".strip()
        result.append(d)
    return result


def lade_kalender_daten(jahr: int, monat: int) -> dict:
    """Gibt dict {date: [eintrag_dict, ...]} für den angegebenen Monat zurück.
    Pro MA + Typ wird nur der neueste Eintrag (höchstes gueltig_bis) gezeigt.

    Vorwarnung: Einträge die in den nächsten 1-3 Kalendermonaten ablaufen
    werden ebenfalls geladen und auf dem letzten Tag des angezeigten Monats
    eingeblendet (mit _vorwarnung=True), damit rechtzeitig gewarnt wird.
    """
    import calendar as _cal

    _init_db()

    # Aktuelle Monat + nächste 3 Monate für Vorwarnungen berechnen
    vorwarn_monate = []   # [(jahr, monat), ...]
    m, j = monat, jahr
    for _ in range(3):
        m += 1
        if m > 12:
            m = 1
            j += 1
        vorwarn_monate.append((j, m))

    # LIKE-Muster für die zukünftigen Monate (nur gueltig_bis, keine absolviert)
    vorwarn_likes = [f"%.{m:02d}.{j}" for j, m in vorwarn_monate]

    # SQL: aktueller Monat (gueltig_bis ODER datum_absolviert) + Vorwarnungen (nur gueltig_bis)
    vorwarn_clause = " OR ".join([f"se.gueltig_bis LIKE ?" for _ in vorwarn_likes])
    sql = f"""SELECT se.*, m.nachname, m.vorname, m.qualifikation
               FROM schulungseintraege se
               JOIN mitarbeiter m ON m.id = se.mitarbeiter_id
               WHERE (
                   (se.gueltig_bis LIKE ? OR se.datum_absolviert LIKE ?)
                   OR ({vorwarn_clause})
               )
                 AND se.id IN (
                     SELECT id FROM schulungseintraege se2
                     WHERE se2.mitarbeiter_id = se.mitarbeiter_id
                       AND se2.schulungstyp   = se.schulungstyp
                       AND se2.gueltig_bis = (
                           SELECT MAX(se3.gueltig_bis)
                           FROM schulungseintraege se3
                           WHERE se3.mitarbeiter_id = se.mitarbeiter_id
                             AND se3.schulungstyp   = se.schulungstyp
                       )
                 )
               ORDER BY se.gueltig_bis"""
    params = [f"%.{monat:02d}.{jahr}", f"%.{monat:02d}.{jahr}"] + vorwarn_likes

    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql, params).fetchall()

    # Referenzdaten: erster und letzter Tag des angezeigten Monats
    letzter_tag_nr   = _cal.monthrange(jahr, monat)[1]
    referenz_erster  = date(jahr, monat, 1)
    referenz_letzter = date(jahr, monat, letzter_tag_nr)
    letzter_tag      = date(jahr, monat, letzter_tag_nr)

    aus = {}
    for r in rows:
        d = dict(r)
        gb              = _parse_datum(d.get("gueltig_bis"))
        laeuft_nicht_ab = bool(d.get("laeuft_nicht_ab"))
        d["_name"]      = f"{d.get('nachname','')} {d.get('vorname','')}".strip()

        if gb and gb.year == jahr and gb.month == monat:
            # Normaler Eintrag: läuft in diesem Monat ab
            # Farbe relativ zum Monatsersten (bereits korrigiert)
            d["_datum_obj"]     = gb
            d["_dringlichkeit"] = _dringlichkeit(gb, laeuft_nicht_ab, referenz=referenz_erster)
            d["_vorwarnung"]    = False
            aus.setdefault(gb, []).append(d)

        elif gb and not laeuft_nicht_ab:
            # Mögliche Vorwarnung: läuft in M+1, M+2 oder M+3 ab
            # Farbe relativ zum letzten Tag des Monats (zeigt wie dringend
            # es am Monatsende war – erzeugt 3 Monate Vorwarnzeit)
            dring = _dringlichkeit(gb, laeuft_nicht_ab, referenz=referenz_letzter)
            if dring in ("rot", "orange", "gelb"):
                d["_datum_obj"]     = gb
                d["_dringlichkeit"] = dring
                d["_vorwarnung"]    = True
                aus.setdefault(letzter_tag, []).append(d)

    return aus


def lade_mitarbeiter_mit_schulungen() -> list[dict]:
    """
    Alle Mitarbeiter mit ihren aktuellsten Schulungseinträgen pro Typ.
    Rückgabe: Liste von dicts:
      {id, nachname, vorname, qualifikation,
       schulungen: {typ: {datum_absolviert, gueltig_bis, status, _dringlichkeit}}}
    Mitarbeiter ohne jeglichen Eintrag haben schulungen = {}.
    """
    _init_db()
    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        mitarbeiter = conn.execute(
            "SELECT id, nachname, vorname, qualifikation FROM mitarbeiter ORDER BY nachname, vorname"
        ).fetchall()
        # Neuester Eintrag pro MA+Typ (MAX id als Tiebreaker)
        eintraege = conn.execute(
            """SELECT se.*
               FROM schulungseintraege se
               WHERE se.id IN (
                   SELECT MAX(id) FROM schulungseintraege
                   GROUP BY mitarbeiter_id, schulungstyp
               )"""
        ).fetchall()

    eintraege_by_ma: dict[int, dict] = {}
    for e in eintraege:
        d = dict(e)
        ma_id = d["mitarbeiter_id"]
        typ   = d["schulungstyp"]
        cfg   = SCHULUNGSTYPEN_CFG.get(typ, {})
        gb    = _parse_datum(d.get("gueltig_bis"))
        d["_dringlichkeit"] = _dringlichkeit(gb, cfg.get("laeuft_nicht_ab", True)) if gb else ""
        eintraege_by_ma.setdefault(ma_id, {})[typ] = d

    result = []
    for m in mitarbeiter:
        md = dict(m)
        md["schulungen"] = eintraege_by_ma.get(m["id"], {})
        result.append(md)
    return result


# ─── Excel-Import ─────────────────────────────────────────────────────────────
# Spalten-Mapping für Blatt "laufend"
_EXCEL_SPALTEN = [
    ("nachname",          None),                  # 0
    ("vorname",           None),                  # 1
    ("geburtsdatum",      None),                  # 2
    ("anstellung",        None),                  # 3
    ("qualifikation",     None),                  # 4
    ("Fuehrerschein_Kont","datum_absolviert"),     # 5
    ("ZÜP",               "gueltig_bis"),          # 6  – gespeichert als Ablauf
    ("EH",                "datum_absolviert"),     # 7  – letztes Datum → +2J
    ("Refresher",         "datum_absolviert"),     # 8  – letztes Datum → +1J
    ("Aerztl_Untersuchung","gueltig_bis"),          # 9  – direkt Ablauf
    ("Personalausweis",   "datum_absolviert"),     # 10
    ("Einw_Zertifikate",  "datum_absolviert"),     # 11
    ("Fixierung",         "datum_absolviert"),     # 12
    ("Einw_eMobby",       "datum_absolviert"),     # 13
    ("Bulmor",            "datum_absolviert"),     # 14
    ("Arbeitsschutz",     "datum_absolviert"),     # 15
    ("Einw_QM",           "datum_absolviert"),     # 16
    ("Fuehrerschein_Verl","datum_absolviert"),     # 17
    ("Fragebogen_Schulung","datum_absolviert"),     # 18
    ("bemerkung",         None),                  # 19
]


def excel_importieren(pfad: str | None = None) -> tuple[int, int]:
    """
    Liest das Blatt 'laufend' aus der Stammdaten-Excel ein.
    Gibt (importiert, uebersprungen) zurück.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl nicht installiert – bitte: pip install openpyxl")

    import shutil, tempfile
    pfad = pfad or str(_EXCEL_PFAD)

    # Datei in Temp kopieren (verhindert OneDrive-Sperren)
    tmp = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(pfad, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        try:
            import os; os.unlink(tmp)
        except Exception:
            pass

    # Blatt "laufend" bevorzugen, sonst erstes Blatt
    ws = wb["laufend"] if "laufend" in wb.sheetnames else wb.active

    _init_db()
    importiert = 0
    uebersprungen = 0
    now = datetime.now().isoformat(timespec="seconds")

    with _connect() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            nachname = str(row[0]).strip() if row[0] else ""
            vorname  = str(row[1]).strip() if row[1] else ""
            if not nachname or nachname.lower() in ("name", "none", ""):
                uebersprungen += 1
                continue
            # Vorname kann mit Leerzeichen beginnen
            vorname = vorname.lstrip()

            # Geburtsdatum
            geb = _parse_datum(row[2])
            geb_str = _datum_str(geb)

            anst  = str(row[3]).strip() if row[3] else ""
            quali = str(row[4]).strip() if row[4] else ""
            bem   = str(row[19]).strip() if len(row) > 19 and row[19] else ""

            # Mitarbeiter vorhanden? (eindeutig per Name)
            ex = conn.execute(
                "SELECT id FROM mitarbeiter WHERE nachname=? AND vorname=?",
                (nachname, vorname)
            ).fetchone()
            if ex:
                ma_id = ex[0]
                conn.execute(
                    "UPDATE mitarbeiter SET anstellung=?, qualifikation=?, geburtsdatum=?, bemerkung=? WHERE id=?",
                    (anst, quali, geb_str, bem, ma_id)
                )
            else:
                cur = conn.execute(
                    "INSERT INTO mitarbeiter (nachname, vorname, geburtsdatum, anstellung, qualifikation, bemerkung, aktiv, erstellt_am) VALUES (?,?,?,?,?,?,1,?)",
                    (nachname, vorname, geb_str, anst, quali, bem, now)
                )
                ma_id = cur.lastrowid

            # Schulungseinträge
            for col_idx in range(5, 19):
                if col_idx >= len(row):
                    break
                key_info = _EXCEL_SPALTEN[col_idx]
                typ_key  = key_info[0]
                art      = key_info[1]  # "gueltig_bis" oder "datum_absolviert"
                if typ_key not in SCHULUNGSTYPEN_CFG:
                    continue

                raw = row[col_idx]
                if raw is None:
                    continue
                # Ja/X ohne Datum → einmalig abgehakt
                if isinstance(raw, str) and raw.strip().upper() in ("JA", "X", "AB APRIL"):
                    # Nur vorhandene einmalige eintragen
                    cfg = SCHULUNGSTYPEN_CFG[typ_key]
                    if cfg["ablauf"] != "einmalig":
                        continue
                    datum_parsed  = None
                    gueltig_bis_d = None
                else:
                    datum_parsed = _parse_datum(raw)
                    if datum_parsed is None:
                        continue
                    cfg = SCHULUNGSTYPEN_CFG[typ_key]

                    if art == "gueltig_bis":
                        gueltig_bis_d  = datum_parsed
                        datum_absolviert_d = None
                    else:
                        datum_absolviert_d = datum_parsed
                        gueltig_bis_d = _berechne_gueltig_bis(typ_key, datum_absolviert_d, None)

                laeuft_nicht_ab = int(cfg.get("laeuft_nicht_ab", False))
                gb_str  = _datum_str(gueltig_bis_d)
                dat_str = _datum_str(datum_parsed if art == "datum_absolviert" else None)
                status  = _berechne_status(gueltig_bis_d, bool(laeuft_nicht_ab))

                # Vorhandenen Eintrag aktualisieren oder neu anlegen
                ex_e = conn.execute(
                    "SELECT id FROM schulungseintraege WHERE mitarbeiter_id=? AND schulungstyp=?",
                    (ma_id, typ_key)
                ).fetchone()
                if ex_e:
                    conn.execute(
                        """UPDATE schulungseintraege SET
                           datum_absolviert=?, gueltig_bis=?, laeuft_nicht_ab=?,
                           status=?, zuletzt_akt=? WHERE id=?""",
                        (dat_str, gb_str, laeuft_nicht_ab, status, now, ex_e[0])
                    )
                else:
                    conn.execute(
                        """INSERT INTO schulungseintraege
                           (mitarbeiter_id, schulungstyp, datum_absolviert, gueltig_bis,
                            laeuft_nicht_ab, status, bemerkung, zuletzt_akt)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (ma_id, typ_key, dat_str, gb_str, laeuft_nicht_ab, status, "", now)
                    )
            importiert += 1
        conn.commit()

    return importiert, uebersprungen


# ─── Backward-Compat. API (manuell) ──────────────────────────────────────────
def schulung_speichern(daten: dict) -> int:
    _init_db()
    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        cur = conn.execute(
            """INSERT INTO schulungen_manuell
               (erstellt_am, mitarbeiter, schulungsart, datum, gueltig_bis, status, bemerkung, aufgenommen_von)
               VALUES (?,?,?,?,?,?,?,?)""",
            (now,
             daten.get("mitarbeiter", ""),
             daten.get("schulungsart", ""),
             daten.get("datum", ""),
             daten.get("gueltig_bis", ""),
             daten.get("status", "gültig"),
             daten.get("bemerkung", ""),
             daten.get("aufgenommen_von", "")),
        )
        conn.commit()
        return cur.lastrowid


def schulung_aktualisieren(schulung_id: int, daten: dict) -> None:
    _init_db()
    with _connect() as conn:
        conn.execute(
            """UPDATE schulungen_manuell SET
               mitarbeiter=?, schulungsart=?, datum=?, gueltig_bis=?,
               status=?, bemerkung=?, aufgenommen_von=?
               WHERE id=?""",
            (daten.get("mitarbeiter", ""), daten.get("schulungsart", ""),
             daten.get("datum", ""), daten.get("gueltig_bis", ""),
             daten.get("status", "gültig"), daten.get("bemerkung", ""),
             daten.get("aufgenommen_von", ""), schulung_id),
        )
        conn.commit()


def schulung_loeschen(schulung_id: int) -> None:
    _init_db()
    with _connect() as conn:
        conn.execute("DELETE FROM schulungen_manuell WHERE id=?", (schulung_id,))
        conn.commit()


def lade_schulungen(jahr: int | None = None, mitarbeiter: str | None = None) -> list[dict]:
    _init_db()
    sql = "SELECT * FROM schulungen_manuell WHERE 1=1"
    params: list = []
    if jahr:
        sql += " AND substr(datum, 1, 4) = ?"
        params.append(str(jahr))
    if mitarbeiter:
        sql += " AND mitarbeiter LIKE ?"
        params.append(f"%{mitarbeiter}%")
    sql += " ORDER BY datum DESC"
    with _connect() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def lade_jahre() -> list[int]:
    _init_db()
    with _connect() as conn:
        rows = conn.execute(
            "SELECT DISTINCT substr(datum,1,4) AS j FROM schulungen_manuell ORDER BY j DESC"
        ).fetchall()
    return [int(r[0]) for r in rows if r[0] and r[0].isdigit()]


# ─── Einmaliger Erstimport ────────────────────────────────────────────────────

def _korrigiere_eh_intervall():
    """
    Einmalige DB-Korrektur: EH-Einträge, deren güeltig_bis auf
    datum_absolviert + 2 Jahre berechnet wurde, auf + 1 Jahr setzen.
    Wird beim _init_db aufgerufen (idempotent durch Flag in settings).
    """
    try:
        with _connect() as conn:
            rows = conn.execute(
                "SELECT id, datum_absolviert FROM schulungseintraege WHERE schulungstyp='EH' AND datum_absolviert IS NOT NULL AND datum_absolviert != ''"
            ).fetchall()
            now = datetime.now().isoformat(timespec="seconds")
            for row_id, dat_str in rows:
                d = _parse_datum(dat_str)
                if d is None:
                    continue
                korrekt = _datum_str(date(d.year + 2, d.month, d.day))
                status = _berechne_status(date(d.year + 2, d.month, d.day), False)
                conn.execute(
                    "UPDATE schulungseintraege SET gueltig_bis=?, status=?, zuletzt_akt=? WHERE id=?",
                    (korrekt, status, now, row_id)
                )
            conn.commit()
    except Exception:
        pass


def _dedup_schulungseintraege():
    """
    Bereinigt doppelte Schulungseinträge:
    Pro Mitarbeiter + Schulungstyp wird nur der Eintrag mit dem höchsten
    gueltig_bis behalten. Alle übrigen (älteren) Duplikate werden gelöscht.
    Idempotent – kann jederzeit aufgerufen werden.
    """
    try:
        with _connect() as conn:
            # Finde alle IDs, die behalten werden sollen (MAX gueltig_bis pro MA+Typ)
            keeper_ids = [r[0] for r in conn.execute(
                """SELECT MAX(id)
                   FROM schulungseintraege
                   GROUP BY mitarbeiter_id, schulungstyp"""
            ).fetchall()]
            if not keeper_ids:
                return
            platzhalter = ",".join("?" * len(keeper_ids))
            conn.execute(
                f"DELETE FROM schulungseintraege WHERE id NOT IN ({platzhalter})",
                keeper_ids,
            )
            conn.commit()
    except Exception:
        pass


def erstimport_wenn_leer() -> "tuple[int, int] | None":
    """
    Importiert die Stammdaten-Excel automatisch, wenn die Mitarbeiter-Tabelle
    noch leer ist (erster Programmstart).
    Pfad: zuerst Setting 'schulungen_excel_pfad', dann Standard _EXCEL_PFAD.
    Gibt (importiert, uebersprungen) zurück oder None wenn bereits Daten vorhanden.
    """
    _init_db()
    # EH-Intervall-Korrektur + Dedup immer ausführen (idempotent)
    _korrigiere_eh_intervall()
    _dedup_schulungseintraege()
    with _connect() as conn:
        anzahl = conn.execute("SELECT COUNT(*) FROM mitarbeiter").fetchone()[0]
    if anzahl > 0:
        return None  # DB bereits befüllt – kein Auto-Import

    # Pfad: Einstellung bevorzugen
    pfad = ""
    try:
        from functions.settings_functions import get_setting
        pfad = get_setting("schulungen_excel_pfad", "").strip()
    except Exception:
        pass
    if not pfad:
        pfad = str(_EXCEL_PFAD)

    if not Path(pfad).is_file():
        return None  # Datei nicht gefunden – still überspringen

    try:
        return excel_importieren(pfad)
    except Exception:
        return None


# ─── PRM-Schulung Import ──────────────────────────────────────────────────────

_PRM_EXCEL_PFAD = Path(_BASE_DIR) / "Daten" / "PRM Schulung" / "zertifikate_aktuell.xlsx"


def prm_schulung_importieren(pfad: str | None = None) -> dict:
    """
    Importiert PRM-Schulungen aus zertifikate_aktuell.xlsx.
    Spalten: Name (D/Index 3), Nachgewiesen am (F/Index 5), Gültig bis (G/Index 6).

    Namens-Abgleich (mehrstufig):
      1. Exakt (normalisiert)
      2. Nachname exakt + Vorname-Wortüberlappung
      3. Nachname-Teil in zusammengesetztem Excel-Nachnamen + Vorname-Overlap
      4. Fuzzy Nachname (Einschluss) + Vorname-Overlap
      → Kein Treffer: neuen Mitarbeiter anlegen

    Rückgabe-Dict:
      importiert   – Anzahl verarbeiteter Excel-Zeilen
      neu_angelegt – Anzahl neu erzeugter Mitarbeiter
      kein_match   – Liste der Excel-Namen ohne Treffer (neu angelegt)
      details      – Liste pro Zeile {excel_name, ma_id, grund, aktion, datum_absolviert, gueltig_bis}
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl nicht installiert – bitte: pip install openpyxl")

    import unicodedata as _ud
    import shutil as _shutil
    import tempfile as _tempfile

    pfad = pfad or str(_PRM_EXCEL_PFAD)

    # Datei temporär kopieren (verhindert OneDrive-Sperren)
    tmp = _tempfile.mktemp(suffix=".xlsx")
    _shutil.copy2(pfad, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        try:
            import os as _os; _os.unlink(tmp)
        except Exception:
            pass

    ws = wb.active

    # ── Hilfsfunktionen ───────────────────────────────────────────────────────
    def _norm(s: str) -> str:
        """Lowercase + Unicode-Normalisierung (Diakritika entfernen)."""
        s = _ud.normalize("NFKD", s.lower())
        return "".join(c for c in s if not _ud.combining(c))

    def _is_upper_word(w: str) -> bool:
        """True wenn das Wort komplett in Großbuchstaben steht (ß=SS)."""
        chk = w.replace("ß", "SS").replace("-", "").replace("'", "").replace(".", "")
        return len(chk) > 1 and chk == chk.upper()

    def _parse_name(raw: str):
        """Trennt Excel-Name in (vorname, nachname).
        Schema: Vorname(n) NACHNAME – letztes Uppercase-Wortblock = Nachname.
        Sonderfälle: alles uppercase → letztes Wort = Nachname.
        """
        parts = raw.strip().split()
        if not parts:
            return "", ""
        # Von rechts: solange uppercase → Nachname
        i = len(parts) - 1
        nn_start = len(parts)
        while i >= 0 and _is_upper_word(parts[i]):
            nn_start = i
            i -= 1
        if nn_start == 0:
            # Alle Wörter uppercase → nur letztes Wort ist Nachname
            return " ".join(parts[:-1]), parts[-1]
        if nn_start == len(parts):
            # Kein Wort komplett uppercase → letztes Wort als Nachname
            return " ".join(parts[:-1]), parts[-1]
        return " ".join(parts[:nn_start]), " ".join(parts[nn_start:])

    def _wort_menge(s: str) -> set:
        """Alle Einzelwörter inkl. Bindestrich-Teile (normalisiert)."""
        result = set()
        for w in s.split():
            n = _norm(w)
            result.add(n)
            for sub in n.split("-"):
                if sub:
                    result.add(sub)
        return result

    # ── Mitarbeiter aus DB laden ──────────────────────────────────────────────
    _init_db()
    with _connect() as _c:
        _rows = _c.execute(
            "SELECT id, nachname, vorname FROM mitarbeiter WHERE aktiv=1"
        ).fetchall()
    # Liste: (id, nn_norm, vn_norm, nn_orig, vn_orig)
    ma_liste: list[tuple] = [
        (r[0], _norm(r[1]), _norm(r[2]), r[1], r[2])
        for r in _rows
    ]

    def _match_ma(vn_ex: str, nn_ex: str):
        """Gibt (ma_id, grund) zurück oder (None, grund_str) wenn kein Treffer."""
        nn_n = _norm(nn_ex)
        nn_ex_teile = {_norm(p) for p in nn_ex.split()}
        ex_vn_wmenge = _wort_menge(vn_ex)

        # 1. Exakt
        for ma_id, ma_nn_n, ma_vn_n, _, _ in ma_liste:
            if ma_nn_n == nn_n and ma_vn_n == _norm(vn_ex):
                return ma_id, "exakt"

        # 2. Nachname exakt + Vorname-Wortüberlappung
        treffer = []
        for ma_id, ma_nn_n, ma_vn_n, _, _ in ma_liste:
            if ma_nn_n == nn_n and _wort_menge(ma_vn_n) & ex_vn_wmenge:
                treffer.append((ma_id, "vorname-überlappung"))
        if len(treffer) == 1:
            return treffer[0]
        if len(treffer) > 1:
            return None, f"mehrdeutig({len(treffer)})"

        # 3. DB-Nachname ist Teil des Excel-Nachnamens + Vorname-Overlap
        treffer = []
        for ma_id, ma_nn_n, ma_vn_n, _, _ in ma_liste:
            if ma_nn_n in nn_ex_teile and _wort_menge(ma_vn_n) & ex_vn_wmenge:
                treffer.append((ma_id, "nachname-teil"))
        if len(treffer) == 1:
            return treffer[0]

        # 4. Fuzzy Nachname (Einschluss) + Vorname-Overlap
        treffer = []
        for ma_id, ma_nn_n, ma_vn_n, _, _ in ma_liste:
            nn_fuzzy = (
                nn_n in ma_nn_n
                or ma_nn_n in nn_n
                or nn_n.replace("-", "") == ma_nn_n.replace("-", "")
            )
            if nn_fuzzy and _wort_menge(ma_vn_n) & ex_vn_wmenge:
                treffer.append((ma_id, "fuzzy"))
        if len(treffer) == 1:
            return treffer[0]
        if len(treffer) > 1:
            return None, f"fuzzy-mehrdeutig({len(treffer)})"

        return None, "kein_match"

    # ── Haupt-Importschleife ──────────────────────────────────────────────────
    now = datetime.now().isoformat(timespec="seconds")
    typ_key = "PRM_Schulung"
    cfg = SCHULUNGSTYPEN_CFG[typ_key]
    laeuft_nicht_ab = int(cfg.get("laeuft_nicht_ab", False))

    importiert = 0
    neu_angelegt = 0
    kein_match: list[str] = []
    details: list[dict] = []

    with _connect() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_raw = row[3]
            if not name_raw or not str(name_raw).strip():
                continue
            name_raw = str(name_raw).strip()

            datum_absolviert_d = _parse_datum(row[5])   # Nachgewiesen am
            gueltig_bis_d      = _parse_datum(row[6])   # Gültig bis

            dat_str = _datum_str(datum_absolviert_d)
            gb_str  = _datum_str(gueltig_bis_d)
            status  = _berechne_status(gueltig_bis_d, bool(laeuft_nicht_ab))

            vn_ex, nn_ex = _parse_name(name_raw)
            ma_id, grund = _match_ma(vn_ex, nn_ex)

            if ma_id is None:
                # Neuen Mitarbeiter anlegen
                nn_neu = nn_ex.title() if nn_ex else name_raw.strip().split()[-1].title()
                vn_neu = vn_ex.title() if vn_ex else " ".join(name_raw.strip().split()[:-1]).title()
                cur = conn.execute(
                    "INSERT INTO mitarbeiter (nachname, vorname, qualifikation, aktiv, erstellt_am)"
                    " VALUES (?,?,?,1,?)",
                    (nn_neu, vn_neu, "PRM", now),
                )
                ma_id = cur.lastrowid
                # Sofort in ma_liste aufnehmen (vermeidet Duplikate bei mehrfachem Vorkommen)
                ma_liste.append((ma_id, _norm(nn_neu), _norm(vn_neu), nn_neu, vn_neu))
                neu_angelegt += 1
                kein_match.append(name_raw)
                grund = "neu_angelegt"

            # Vorhandenen PRM-Eintrag aktualisieren oder neu anlegen
            ex_e = conn.execute(
                "SELECT id FROM schulungseintraege WHERE mitarbeiter_id=? AND schulungstyp=?",
                (ma_id, typ_key),
            ).fetchone()
            if ex_e:
                conn.execute(
                    """UPDATE schulungseintraege SET
                       datum_absolviert=?, gueltig_bis=?, laeuft_nicht_ab=?,
                       status=?, zuletzt_akt=? WHERE id=?""",
                    (dat_str, gb_str, laeuft_nicht_ab, status, now, ex_e[0]),
                )
                aktion = "aktualisiert"
            else:
                conn.execute(
                    """INSERT INTO schulungseintraege
                       (mitarbeiter_id, schulungstyp, datum_absolviert, gueltig_bis,
                        laeuft_nicht_ab, status, bemerkung, zuletzt_akt)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (ma_id, typ_key, dat_str, gb_str, laeuft_nicht_ab, status, "", now),
                )
                aktion = "neu"

            importiert += 1
            details.append({
                "excel_name":        name_raw,
                "ma_id":             ma_id,
                "grund":             grund,
                "aktion":            aktion,
                "datum_absolviert":  dat_str,
                "gueltig_bis":       gb_str,
            })
        conn.commit()

    return {
        "importiert":   importiert,
        "neu_angelegt": neu_angelegt,
        "kein_match":   kein_match,
        "details":      details,
    }
