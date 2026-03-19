"""
Telefonnummern-Datenbank
Liest Telefonnummern aus Excel-Dateien und speichert sie in SQLite.
"""
import sqlite3
import os
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager

from config import BASE_DIR

_DB_DIR   = Path(BASE_DIR) / "database SQL"
_DB_PFAD  = _DB_DIR / "telefonnummern.db"
_EXCEL_DIR = Path(BASE_DIR) / "Daten" / "Telefonnummern"

_CREATE_SQL = """
CREATE TABLE IF NOT EXISTS telefonnummern (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    erstellt_am TEXT NOT NULL,
    quelle      TEXT NOT NULL,
    sheet       TEXT NOT NULL,
    kategorie   TEXT DEFAULT '',
    bezeichnung TEXT NOT NULL,
    nummer      TEXT DEFAULT '',
    email       TEXT DEFAULT '',
    bemerkung   TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS tel_import_log (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    importiert_am TEXT NOT NULL,
    quelle        TEXT NOT NULL,
    anzahl        INTEGER DEFAULT 0
);
"""


def _ensure_db():
    _DB_DIR.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(_DB_PFAD, timeout=5)
    con.execute("PRAGMA journal_mode = WAL")
    con.execute("PRAGMA synchronous  = NORMAL")
    con.execute("PRAGMA busy_timeout = 5000")
    for stmt in _CREATE_SQL.strip().split(";"):
        stmt = stmt.strip()
        if stmt:
            con.execute(stmt)
    con.commit()
    con.close()


@contextmanager
def _db():
    _ensure_db()
    con = sqlite3.connect(_DB_PFAD, timeout=5)
    con.execute("PRAGMA journal_mode = WAL")
    con.execute("PRAGMA synchronous  = NORMAL")
    con.row_factory = sqlite3.Row
    try:
        yield con
        con.commit()
    finally:
        con.close()


def _push(table: str, row_id: int) -> None:
    try:
        from database.turso_sync import push_row
        conn = sqlite3.connect(_DB_PFAD, timeout=5)
        conn.row_factory = sqlite3.Row
        row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (row_id,)).fetchone()
        conn.close()
        if row:
            push_row(str(_DB_PFAD), table, dict(row))
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────────────────────
#  Excel-Parser
# ──────────────────────────────────────────────────────────────────────────────

def _parse_kontaktliste(quelle: str, sheet_display: str, rows: list) -> list[dict]:
    """
    Parst eine einfache Kontaktliste mit Spalten: Abt., Name, Tel., E-Mail.
    Verwendet in: wichtige Tel nummern.xlsx – Sheet "Telefonnummern".
    """
    now = datetime.now().isoformat(timespec="seconds")
    entries = []
    in_data = False

    for row in rows:
        if not any(v is not None for v in row):
            continue

        vals = [str(v or "").strip() for v in row]

        # Titelzeilen überspringen
        if any("Telefonnummer" in v for v in vals[:2]):
            continue
        first = vals[0] if vals else ""
        if first.lower().startswith("stand:"):
            continue

        # Header-Zeile erkennen
        if not in_data:
            lower_vals = [v.lower() for v in vals]
            if any(h in lower_vals for h in ("abt.", "abt", "name", "name ", "tel.", "e-mail", "e-mail ")):
                in_data = True
                continue
            continue

        # Spalten-Offset ermitteln (manche Sheets haben None in Spalte 0)
        off = 0
        if row[0] is None and len(row) > 1 and row[1] is not None:
            off = 1

        abt  = str(row[off]   or "").strip()
        name = str(row[off+1] or "").strip() if len(row) > off + 1 else ""
        tel  = str(row[off+2] or "").strip() if len(row) > off + 2 else ""
        mail = str(row[off+3] or "").strip() if len(row) > off + 3 else ""

        if not name and not tel:
            continue

        # Wenn "Name" wie eine Telefonnummer aussieht und "Tel" leer ist – tauschen
        if tel == "" and name and all(c in "0123456789 /-+()" for c in name):
            tel  = name
            name = abt or name

        if not name and not tel:
            continue

        entries.append({
            "erstellt_am": now,
            "quelle":      quelle,
            "sheet":       sheet_display,
            "kategorie":   abt,
            "bezeichnung": name or abt,
            "nummer":      tel,
            "email":       mail,
            "bemerkung":   "",
        })

    return entries


# Kategorie-Normierung: rohe Excel-Bezeichner → lesbare Kurzbezeichnungen
_CAT_NORMIERUNG: dict[str, str] = {
    "Check In Nummern (02203 40-)": "Check In B",
    "Checkin C":                    "Check In C",
    "Checkin D 401-420":            "Check In D (401-420)",
    "Checkin D 421 - 440":          "Check In D (421-440)",
    "FKB Nummern":                  "FKB intern",
    "FKB Abteilungsbezeichungen":   "FKB intern",
}


def _parse_grid_sheet(quelle: str, sheet_display: str, rows: list) -> list[dict]:
    """
    Parst mehrspaltige Gitter-Sheets wie CIC und int Gate aus Telefonnummern FKB.xlsx.
    Schema: 4 Spaltengruppen an (1,2), (4,5), (7,8), (10,11) = (Bezeichnung, Nummer).
    Abschnitte werden über Kategorie-Headerzeilen erkannt.
    """
    now = datetime.now().isoformat(timespec="seconds")
    entries = []
    col_groups = [(1, 2), (4, 5), (7, 8), (10, 11)]
    categories = ["", "", "", ""]
    data_phase  = False

    for row in rows:
        if not any(v is not None for v in row):
            continue

        row_s = [str(v or "").strip() for v in row[:13]]

        # Spalten-Header-Zeile erkennen: enthält "Nummer" und typische Bezeichner
        if "Nummer" in row_s and any(v in ("OPS", "Gate", "CIC B", "CIC B ", "CIC D") for v in row_s):
            data_phase = True
            continue

        if not data_phase:
            # Kategorie/Abschnitt-Headerzeile: Text in Label-Spalten, keine Zahlen
            for idx, (li, _) in enumerate(col_groups):
                if li < len(row):
                    cat = str(row[li] or "").strip()
                    if cat and cat not in (" ", ""):
                        categories[idx] = _CAT_NORMIERUNG.get(cat, cat)
            continue

        # Datenzeile verarbeiten
        for idx, (li, ni) in enumerate(col_groups):
            label  = str(row[li] or "").strip() if li < len(row) else ""
            number = str(row[ni] or "").strip() if ni < len(row) else ""
            if not label:
                continue
            # Zeile ist ein Abschnittshinweis (Nummer enthält "Telefon (...)")
            if "Telefon" in number:
                categories[idx] = _CAT_NORMIERUNG.get(label, label)
                continue
            entries.append({
                "erstellt_am": now,
                "quelle":      quelle,
                "sheet":       sheet_display,
                "kategorie":   categories[idx],
                "bezeichnung": label,
                "nummer":      number,
                "email":       "",
                "bemerkung":   "",
            })

    return entries


# ──────────────────────────────────────────────────────────────────────────────
#  Import
# ──────────────────────────────────────────────────────────────────────────────

def importiere_aus_excel(clear_first: bool = True) -> int:
    """
    Liest alle Telefonnummern-Excel-Dateien ein und speichert sie in der DB.
    Bei clear_first=True werden bestehende Einträge vorher gelöscht.
    Gibt die Anzahl importierter Einträge zurück.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl ist nicht installiert. Bitte 'pip install openpyxl' ausführen."
        ) from exc

    alle_eintraege: list[dict] = []
    now = datetime.now().isoformat(timespec="seconds")

    # ── Datei 1: wichtige Tel nummern.xlsx ───────────────────────────────────
    fn1 = _EXCEL_DIR / "wichtige Tel nummern.xlsx"
    if fn1.exists():
        wb = openpyxl.load_workbook(fn1, read_only=True, data_only=True)
        # Beide möglichen Sheet-Namen (mit/ohne Leerzeichen)
        for sn in ["Telefonnummern ", "Telefonnummern"]:
            if sn in wb.sheetnames:
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                alle_eintraege += _parse_kontaktliste(fn1.name, "Kontakte", rows)
                break
        wb.close()

    # ── Datei 2: Telefonnummern FKB.xlsx ─────────────────────────────────────
    fn2 = _EXCEL_DIR / "Telefonnummern FKB.xlsx"
    if fn2.exists():
        wb = openpyxl.load_workbook(fn2, read_only=True, data_only=True)
        if "CIC" in wb.sheetnames:
            ws   = wb["CIC"]
            rows = list(ws.iter_rows(values_only=True))
            alle_eintraege += _parse_grid_sheet(fn2.name, "Check-In (CIC)", rows)
        if "int Gate" in wb.sheetnames:
            ws   = wb["int Gate"]
            rows = list(ws.iter_rows(values_only=True))
            alle_eintraege += _parse_grid_sheet(fn2.name, "Interne & Gate", rows)
        wb.close()

    with _db() as con:
        if clear_first:
            con.execute("DELETE FROM telefonnummern")
        for e in alle_eintraege:
            con.execute(
                """
                INSERT INTO telefonnummern
                    (erstellt_am, quelle, sheet, kategorie, bezeichnung, nummer, email, bemerkung)
                VALUES
                    (:erstellt_am, :quelle, :sheet, :kategorie, :bezeichnung, :nummer, :email, :bemerkung)
                """,
                e,
            )
        con.execute(
            "INSERT INTO tel_import_log (importiert_am, quelle, anzahl) VALUES (?, ?, ?)",
            (now, "alle Dateien", len(alle_eintraege)),
        )
    try:
        from database.turso_sync import push_clear_table, push_table_batch
        if clear_first:
            push_clear_table(str(_DB_PFAD), "telefonnummern")
        push_table_batch(str(_DB_PFAD), "telefonnummern")
        push_table_batch(str(_DB_PFAD), "tel_import_log")
    except Exception:
        pass

    return len(alle_eintraege)


# ──────────────────────────────────────────────────────────────────────────────
#  Laden / Suchen
# ──────────────────────────────────────────────────────────────────────────────

def lade_telefonnummern(
    suchtext: str | None = None,
    kategorie: str | None = None,
    quelle: str | None = None,
    sheet: str | None = None,
) -> list[dict]:
    """Gibt Telefonnummern-Einträge aus der DB zurück, optional gefiltert."""
    where: list[str] = []
    params: list = []

    if suchtext:
        t = f"%{suchtext}%"
        where.append(
            "(bezeichnung LIKE ? OR nummer LIKE ? OR kategorie LIKE ? OR email LIKE ?)"
        )
        params.extend([t, t, t, t])
    if kategorie:
        where.append("kategorie = ?")
        params.append(kategorie)
    if quelle:
        where.append("quelle = ?")
        params.append(quelle)
    if sheet:
        where.append("sheet = ?")
        params.append(sheet)

    sql = "SELECT * FROM telefonnummern"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY quelle, sheet, kategorie, bezeichnung"

    with _db() as con:
        rows = con.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def lade_kategorien() -> list[str]:
    """Gibt alle vorhandenen Kategorien zurück."""
    with _db() as con:
        rows = con.execute(
            "SELECT DISTINCT kategorie FROM telefonnummern WHERE kategorie != '' ORDER BY kategorie"
        ).fetchall()
    return [r[0] for r in rows]


def lade_quellen() -> list[str]:
    """Gibt alle vorhandenen Quell-Dateinamen zurück."""
    with _db() as con:
        rows = con.execute(
            "SELECT DISTINCT quelle FROM telefonnummern ORDER BY quelle"
        ).fetchall()
    return [r[0] for r in rows]


def lade_sheets() -> list[str]:
    """Gibt alle vorhandenen Sheet-Namen zurück."""
    with _db() as con:
        rows = con.execute(
            "SELECT DISTINCT sheet FROM telefonnummern ORDER BY sheet"
        ).fetchall()
    return [r[0] for r in rows]


def letzter_import() -> str | None:
    """Gibt Zeitstempel des letzten Imports zurück oder None."""
    with _db() as con:
        row = con.execute(
            "SELECT importiert_am, anzahl FROM tel_import_log ORDER BY id DESC LIMIT 1"
        ).fetchone()
    if row:
        return f"{row['importiert_am']}  ({row['anzahl']} Einträge)"
    return None


def ist_db_leer() -> bool:
    with _db() as con:
        row = con.execute("SELECT COUNT(*) AS n FROM telefonnummern").fetchone()
    return row["n"] == 0


def hat_veraltete_daten() -> bool:
    """True, wenn die DB noch Einträge mit alten (unnormierten) Kategorienamen enthält."""
    try:
        with _db() as con:
            row = con.execute(
                "SELECT COUNT(*) AS n FROM telefonnummern "
                "WHERE kategorie IN ('Check In Nummern (02203 40-)','Checkin C',"
                "'Checkin D 401-420','Checkin D 421 - 440','FKB Nummern','FKB Abteilungsbezeichungen')"
            ).fetchone()
        return row["n"] > 0
    except Exception:
        return False


# ──────────────────────────────────────────────────────────────────────────────
#  CRUD
# ──────────────────────────────────────────────────────────────────────────────

def eintrag_speichern(daten: dict) -> int:
    now = datetime.now().isoformat(timespec="seconds")
    with _db() as con:
        cur = con.execute(
            """
            INSERT INTO telefonnummern
                (erstellt_am, quelle, sheet, kategorie, bezeichnung, nummer, email, bemerkung)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now,
                daten.get("quelle", "Manuell"),
                daten.get("sheet", "Manuell"),
                daten.get("kategorie", ""),
                daten.get("bezeichnung", ""),
                daten.get("nummer", ""),
                daten.get("email", ""),
                daten.get("bemerkung", ""),
            ),
        )
        new_id = cur.lastrowid
    _push("telefonnummern", new_id)
    return new_id


def eintrag_aktualisieren(entry_id: int, daten: dict) -> None:
    with _db() as con:
        con.execute(
            """
            UPDATE telefonnummern
               SET kategorie=?, bezeichnung=?, nummer=?, email=?, bemerkung=?, sheet=?
             WHERE id=?
            """,
            (
                daten.get("kategorie", ""),
                daten.get("bezeichnung", ""),
                daten.get("nummer", ""),
                daten.get("email", ""),
                daten.get("bemerkung", ""),
                daten.get("sheet", "Manuell"),
                entry_id,
            ),
        )
    _push("telefonnummern", entry_id)


def eintrag_loeschen(entry_id: int) -> None:
    with _db() as con:
        con.execute("DELETE FROM telefonnummern WHERE id = ?", (entry_id,))
    try:
        from database.turso_sync import push_delete
        push_delete(str(_DB_PFAD), "telefonnummern", entry_id)
    except Exception:
        pass
