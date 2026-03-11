"""
Dienstplan-Parser für Excel-Dateien
Adaptiert aus Nesk2/Function/dienstplan_parser.py für Nesk3
Erkennt Namen, Dienstzeiten, Betreuer/Dispo-Unterscheidung dynamisch
"""
import os
import sys
import shutil
import tempfile
import openpyxl
import re
from pathlib import Path
from datetime import datetime, time
from typing import Optional
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _runde_auf_volle_stunde(zeit_str: Optional[str]) -> Optional[str]:
    """
    Rundet eine Zeitangabe auf die volle Stunde ab (Minutenanteil wird 0).
    '07:15' -> '07:00'   '19:45' -> '19:00'   '06:00' -> '06:00'
    Gibt None oder den Original-String zurück wenn nicht parsbar.
    """
    if not zeit_str:
        return zeit_str
    try:
        teile = zeit_str.split(':')
        stunde = int(teile[0])
        return f'{stunde:02d}:00'
    except Exception:
        return zeit_str


def _betr_zu_dispo_kuerzel(kuerzel: str) -> str:
    """
    Wandelt ein Betreuer-Dienstkürzel in das entsprechende Dispo-Kürzel um.
    Wird verwendet wenn ein Mitarbeiter im Dispo-Abschnitt als krank gemeldet ist.

      T    → DT      T10  → DT      T8   → DT
      T(?) → DT(?)
      N    → DN      N10  → DN
      N(?) → DN(?)   S(?) → DN3(?)
    """
    mapping = {
        'T':    'DT',
        'T10':  'DT',
        'T8':   'DT',
        'N':    'DN',
        'N10':  'DN',
    }
    if kuerzel in mapping:
        return mapping[kuerzel]
    if kuerzel.startswith('T'):
        return kuerzel.replace('T', 'DT', 1)
    if kuerzel.startswith('N'):
        return kuerzel.replace('N', 'DN', 1)
    if kuerzel.startswith('S'):
        return 'DN3(?)'
    return kuerzel or 'D(?)'


class DienstplanParser:
    """Parser für Dienstplan-Excel-Dateien.
    Sucht die Header-Zeile (NAME, DIENST, BEGINN, ENDE) automatisch
    innerhalb der ersten 20 Zeilen.
    """

    BETREUER_KATEGORIEN = ['T', 'T10', 'N', 'N10', 'NF', 'FB1', 'FB2', 'FB']
    DISPO_KATEGORIEN    = ['DT', 'DT3', 'DN', 'DN3', 'D']

    # Dienste die still (ohne Warnung) komplett ignoriert werden
    STILLE_DIENSTE: frozenset = frozenset({'R', 'B1', 'B2'})

    # Vollständig ausgeschlossene Personen (Vorname Nachname, lowercase, fix)
    AUSGESCHLOSSENE_VOLLNAMEN: frozenset = frozenset({'lars peters'})

    def __init__(self, excel_path: str, alle_anzeigen: bool = False, round_dispo: bool = True):
        self.excel_path    = Path(excel_path)
        self.alle_anzeigen = alle_anzeigen  # True = keine Ausschlüsse, für Anzeige
        self.round_dispo   = round_dispo    # False = Zeiten nie runden (für Roh-Anzeige)
        self.workbook      = None
        self.sheet         = None
        self.column_map    = None
        self.unbekannte_dienste: set = set()

    # ------------------------------------------------------------------
    # Öffentliche Methode
    # ------------------------------------------------------------------

    def parse(self) -> dict:
        """
        Parst Excel-Datei und liefert Dienstplan-Daten.

        Returns:
            {
                'success': bool,
                'betreuer': list[dict],
                'dispo':    list[dict],
                'kranke':   list[dict],
                'error':    str | None,
                'unbekannte_dienste': list[str]
            }
        """
        try:
            # Datei in temp-Kopie lesen – umgeht Windows-Sperren (z.B. Excel offen / OneDrive-Sync)
            tmp_file = None
            try:
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
                    tmp_file = tf.name
                shutil.copy2(self.excel_path, tmp_file)
                self.workbook = openpyxl.load_workbook(tmp_file, data_only=True)
            except PermissionError as pe:
                raise PermissionError(
                    f"Datei ist gesperrt (vermutlich in Excel geöffnet oder OneDrive synchronisiert):\n"
                    f"{self.excel_path}"
                ) from pe
            self.sheet = self.workbook.active

            self.column_map = self._find_columns()
            if not self.column_map:
                return {
                    'success': False,
                    'betreuer': [], 'dispo': [], 'kranke': [],
                    'error': 'Header-Zeile nicht gefunden (benötigt: NAME, DIENST).',
                    'unbekannte_dienste': []
                }

            betreuer_liste = []
            dispo_liste    = []
            kranke_liste   = []
            alle_nachnamen = []

            # Abschnitt-Tracking: 'betreuer' oder 'dispo'
            aktueller_abschnitt = 'betreuer'

            for row in self.sheet.iter_rows(min_row=1, values_only=False):
                row_list = list(row)

                # Abschnitts-Header prüfen BEVOR _parse_row aufgerufen wird
                neuer_abschnitt = self._detect_abschnitt_header(row_list)
                if neuer_abschnitt is not None:
                    aktueller_abschnitt = neuer_abschnitt
                    continue   # Header-Zeile selbst nicht als Person parsen

                person = self._parse_row(row_list)
                if person:
                    # Abschnitts-Kontext auf Person übertragen
                    if aktueller_abschnitt == 'dispo':
                        person['ist_dispo']       = True
                        # Für Kranke: Dispo-Status + abgeleitetes Kürzel anpassen
                        if person.get('ist_krank'):
                            person['krank_ist_dispo'] = True
                            # Betreuer-Kürzel → Dispo-Kürzel umwandeln
                            d = person.get('krank_abgeleiteter_dienst') or ''
                            person['krank_abgeleiteter_dienst'] = _betr_zu_dispo_kuerzel(d)
                            # Anzeigezeiten auf volle Stunde abrunden
                            if self.round_dispo:
                                person['start_zeit'] = _runde_auf_volle_stunde(person.get('start_zeit'))
                                person['end_zeit']   = _runde_auf_volle_stunde(person.get('end_zeit'))

                    alle_nachnamen.append(person['nachname'])
                    if person['ist_krank']:
                        kranke_liste.append(person)
                    elif person['ist_dispo']:
                        dispo_liste.append(person)
                    else:
                        betreuer_liste.append(person)

            # Ausgeschlossene Personen herausfiltern (nur im Export-Modus)
            if not self.alle_anzeigen:
                try:
                    from functions.settings_functions import get_ausgeschlossene_namen
                    settings_ausgeschlossen = set(get_ausgeschlossene_namen())
                except Exception:
                    settings_ausgeschlossen = set()

                alle_ausgeschlossen = self.AUSGESCHLOSSENE_VOLLNAMEN | settings_ausgeschlossen

                def _filter_ausgeschlossen(lst):
                    return [p for p in lst
                            if p['vollname'].lower() not in alle_ausgeschlossen]

                betreuer_liste = _filter_ausgeschlossen(betreuer_liste)
                dispo_liste    = _filter_ausgeschlossen(dispo_liste)
                kranke_liste   = _filter_ausgeschlossen(kranke_liste)

            # Doppelte Nachnamen → Initial anhängen
            nachname_counts  = Counter(alle_nachnamen)
            doppelte         = {n for n, c in nachname_counts.items() if c > 1}
            for gruppe in (betreuer_liste, dispo_liste, kranke_liste):
                self._generate_display_names(gruppe, doppelte)

            return {
                'success': True,
                'betreuer': betreuer_liste,
                'dispo':    dispo_liste,
                'kranke':   kranke_liste,
                'error':    None,
                'unbekannte_dienste': list(self.unbekannte_dienste),
                'datum':    self._find_datum(),
                'column_map': dict(self.column_map) if self.column_map else {},
                'excel_path': str(self.excel_path),
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'betreuer': [], 'dispo': [], 'kranke': [],
                'error': str(e),
                'unbekannte_dienste': []
            }
        finally:
            if self.workbook:
                self.workbook.close()
            if tmp_file and os.path.exists(tmp_file):
                try:
                    os.remove(tmp_file)
                except OSError:
                    pass

    # ------------------------------------------------------------------
    # Interne Hilfsmethoden
    # ------------------------------------------------------------------

    def _find_datum(self) -> Optional[str]:
        """Sucht in den ersten Zeilen nach einem Datumswert und gibt ihn als DD.MM.YYYY zurück."""
        try:
            header_row = (self.column_map or {}).get('header_row', 20)
            for row_idx in range(1, header_row):
                for cell in self.sheet[row_idx]:
                    val = cell.value
                    if isinstance(val, datetime):
                        return val.strftime('%d.%m.%Y')
                    if isinstance(val, str):
                        # DD.MM.YYYY oder YYYY-MM-DD Muster
                        m = re.search(r'(\d{1,2}\.\d{1,2}\.\d{2,4})', val)
                        if m:
                            return m.group(1)
                        m = re.search(r'(\d{4}-\d{2}-\d{2})', val)
                        if m:
                            try:
                                d = datetime.strptime(m.group(1), '%Y-%m-%d')
                                return d.strftime('%d.%m.%Y')
                            except ValueError:
                                pass
        except Exception:
            pass
        return None

    def _find_columns(self) -> Optional[dict]:
        """Sucht Header-Zeile und gibt Spalten-Indizes zurück."""
        for row_idx in range(1, min(20, self.sheet.max_row + 1)):
            row        = list(self.sheet[row_idx])
            row_values = [cell.value for cell in row]

            name_idx = dienst_idx = beginn_idx = ende_idx = None

            for col_idx, value in enumerate(row_values):
                if value and isinstance(value, str):
                    v = value.strip().upper()
                    if v == 'NAME':
                        name_idx = col_idx
                    elif v == 'DIENST':
                        dienst_idx = col_idx
                    elif v == 'BEGINN':
                        beginn_idx = col_idx
                    elif v == 'ENDE':
                        ende_idx = col_idx

            if name_idx is not None and dienst_idx is not None:
                return {
                    'name':       name_idx,
                    'dienst':     dienst_idx,
                    'beginn':     beginn_idx,
                    'ende':       ende_idx,
                    'header_row': row_idx
                }
        return None

    def _detect_abschnitt_header(self, row_list: list) -> Optional[str]:
        """
        Erkennt Abschnitts-Trennzeilen in der Excel-Datei.
        Gibt 'dispo', 'betreuer' oder None (kein Header) zurück.

        Bekannte Header-Muster:
          '[Stamm FH]'  / 'Stamm ...'  → 'betreuer'
          'Dispo'                      → 'dispo'
        """
        cells = [cell.value if hasattr(cell, 'value') else cell for cell in row_list]

        # Nur nicht-leere String-Werte betrachten
        texte = [
            str(c).strip().lower()
            for c in cells
            if c is not None and isinstance(c, str) and str(c).strip()
        ]
        if not texte:
            return None

        # Wenn Namens-Spalte befüllt ist, ist es eine normale Datenzeile → kein Header
        if self.column_map and self.column_map.get('name') is not None:
            name_idx = self.column_map['name']
            if name_idx < len(cells) and cells[name_idx] and \
               isinstance(cells[name_idx], str) and cells[name_idx].strip():
                return None

        for t in texte:
            if t.startswith('dispo'):
                return 'dispo'
            if 'stamm' in t or 'betreuer' in t or '[stamm' in t:
                return 'betreuer'
        return None

    def _parse_row(self, row) -> Optional[dict]:
        """Parst eine Zeile und gibt Person-Dict oder None zurück."""
        row_list    = list(row)
        excel_row   = row_list[0].row if row_list and hasattr(row_list[0], 'row') else None
        cells       = [cell.value if hasattr(cell, 'value') else cell for cell in row_list]

        max_col = max(
            self.column_map['name'],
            self.column_map['dienst'],
            self.column_map.get('beginn') or 0,
            self.column_map.get('ende')   or 0,
        )
        if len(cells) <= max_col:
            return None

        # Name prüfen
        name_cell = cells[self.column_map['name']]
        if not name_cell or not isinstance(name_cell, str):
            return None

        parsed_name = self._parse_name(name_cell)
        if not parsed_name:
            return None

        full_name = f"{parsed_name['vorname']} {parsed_name['nachname']}"

        # Zellfarbe prüfen (Bulmorfahrer = gelb)
        name_cell_obj   = row_list[self.column_map['name']]
        dienst_cell_obj = row_list[self.column_map['dienst']]
        ist_bulmorfahrer, zeilen_farbe, dienst_farbe, dienst_farbe_hex = \
            self._check_cell_colors(name_cell_obj, dienst_cell_obj)

        # Dienst-Kürzel
        dienst_text     = None
        dienst_kategorie = None
        ist_krank        = False

        raw_dienst = cells[self.column_map['dienst']]
        if raw_dienst:
            dienst_text = str(raw_dienst).strip().upper()
            if not self.alle_anzeigen and dienst_text in self.STILLE_DIENSTE:
                return None   # still ignorieren, keine Warnung
            if dienst_text in ('KRANK', 'K'):
                ist_krank = True
            elif dienst_text in self.BETREUER_KATEGORIEN or dienst_text in self.DISPO_KATEGORIEN:
                dienst_kategorie = dienst_text
            elif dienst_text:
                dienst_kategorie = dienst_text
                self.unbekannte_dienste.add(dienst_text)

        # Zeiten
        round_times = (dienst_kategorie in self.DISPO_KATEGORIEN if dienst_kategorie else False) and self.round_dispo
        start_zeit  = None
        end_zeit    = None

        if self.column_map.get('beginn') is not None and len(cells) > self.column_map['beginn']:
            start_zeit = self._parse_time(cells[self.column_map['beginn']], round_to_hour=round_times)
        if self.column_map.get('ende') is not None and len(cells) > self.column_map['ende']:
            end_zeit = self._parse_time(cells[self.column_map['ende']], round_to_hour=round_times)

        schicht_typ = self._ermittle_schichttyp(start_zeit, end_zeit)

        # Für kranke Mitarbeiter: Schichttyp und Dispo-Status aus Zeiten ableiten
        krank_schicht_typ       = None
        krank_ist_dispo         = False
        krank_abgeleiteter_dienst = None
        if ist_krank:
            krank_schicht_typ, krank_ist_dispo, krank_abgeleiteter_dienst = \
                self._ermittle_krank_typ(start_zeit, end_zeit, full_name)

        return {
            'vorname':                  parsed_name['vorname'],
            'nachname':                 parsed_name['nachname'],
            'vollname':                 full_name,
            'anzeigename':              parsed_name['nachname'],
            'dienst_kategorie':         dienst_kategorie,
            'start_zeit':               start_zeit,
            'end_zeit':                 end_zeit,
            'schicht_typ':              schicht_typ,
            'ist_dispo':                dienst_kategorie in self.DISPO_KATEGORIEN if dienst_kategorie else False,
            'ist_krank':                ist_krank,
            'krank_schicht_typ':        krank_schicht_typ,
            'krank_ist_dispo':          krank_ist_dispo,
            'krank_abgeleiteter_dienst': krank_abgeleiteter_dienst,
            'ist_bulmorfahrer':         ist_bulmorfahrer,
            'zeilen_farbe':             zeilen_farbe,
            'dienst_farbe':             dienst_farbe,
            'dienst_farbe_hex':         dienst_farbe_hex,
            'excel_row':                excel_row,
        }

    def _check_cell_colors(self, name_cell_obj, dienst_cell_obj):
        """Liest Zellfarben aus und erkennt Bulmorfahrer (gelb) und Zebra-Zeilen (grau)."""
        ist_bulmorfahrer = False
        zeilen_farbe     = None
        dienst_farbe     = None
        dienst_farbe_hex = None

        for cell_obj in (name_cell_obj, dienst_cell_obj):
            if not hasattr(cell_obj, 'fill') or not cell_obj.fill:
                continue
            try:
                fill = cell_obj.fill
                if fill.patternType in ('solid', 'solidFill'):
                    fg = fill.fgColor
                    if hasattr(fg, 'rgb') and fg.rgb:
                        color_hex = str(fg.rgb).upper()

                        if cell_obj == dienst_cell_obj and color_hex not in ('00000000', 'FFFFFFFF'):
                            dienst_farbe_hex = color_hex

                        if color_hex == 'FFF5F5F5':
                            zeilen_farbe = 'gray'

                        # Gelb erkennen (AARRGGBB oder RRGGBB)
                        if len(color_hex) >= 8:
                            rr_gg = color_hex[2:6]
                            bb    = color_hex[6:8]
                        elif len(color_hex) == 6:
                            rr_gg = color_hex[0:4]
                            bb    = color_hex[4:6]
                        else:
                            continue

                        if rr_gg == 'FFFF' and int(bb, 16) <= 0x4F:
                            ist_bulmorfahrer = True
                            dienst_farbe     = 'yellow'
                            break
            except Exception:
                pass

        return ist_bulmorfahrer, zeilen_farbe, dienst_farbe, dienst_farbe_hex

    def _parse_name(self, name_text: str) -> Optional[dict]:
        """Parst 'Nachname, Vorname' oder 'Vorname Nachname'."""
        name_text = name_text.strip()
        name_text = re.sub(r'[^\w\säöüÄÖÜß\-,]', '', name_text)

        if ',' in name_text:
            parts = name_text.split(',')
            if len(parts) == 2:
                nachname      = parts[0].strip()
                vorname_teil  = parts[1].strip()
                if not vorname_teil:
                    return None
                vorname = vorname_teil.split()[0] if vorname_teil.split() else None
                if not vorname:
                    return None
                return {'vorname': vorname, 'nachname': nachname}

        parts = name_text.split()
        if len(parts) >= 2:
            return {'vorname': parts[0], 'nachname': ' '.join(parts[1:])}

        return None

    def _parse_time(self, value, round_to_hour=False) -> Optional[str]:
        """Parst Zeitwert und gibt 'HH:MM'-String oder None zurück."""
        try:
            if isinstance(value, datetime):
                t = value.time()
            elif isinstance(value, time):
                t = value
            elif isinstance(value, str):
                value = value.strip()
                m = re.match(r'(\d{1,2}):(\d{2})', value)
                if m:
                    t = time(int(m.group(1)), int(m.group(2)))
                else:
                    m = re.match(r'(\d{2})(\d{2})', value)
                    if m:
                        t = time(int(m.group(1)), int(m.group(2)))
                    else:
                        return None
            else:
                return None

            if round_to_hour:
                t = time(t.hour, 0)
            return f"{t.hour:02d}:{t.minute:02d}"
        except Exception:
            return None

    def _ermittle_schichttyp(self, start_zeit: Optional[str], end_zeit: Optional[str]) -> Optional[str]:
        """Kategorisiert Schicht anhand der Startzeit."""
        if not start_zeit:
            return None
        try:
            hour = int(start_zeit.split(':')[0])
        except Exception:
            return None

        if 5 <= hour < 9 or 9 <= hour < 12:
            return 'tagdienst_vormittag'
        elif 12 <= hour < 15 or 15 <= hour < 19:
            return 'tagdienst_nachmittag'
        elif 19 <= hour < 23 or 23 <= hour <= 23:
            return 'nachtschicht_frueh'
        elif 0 <= hour < 5:
            return 'nachtschicht_spaet'
        return None

    def _ermittle_krank_typ(
        self,
        start_zeit: Optional[str],
        end_zeit:   Optional[str],
        vollname:   str,
    ):
        """
        Leitet für einen kranken Mitarbeiter ab:
          - krank_schicht_typ  : 'tagdienst' | 'nachtdienst' | 'sonderdienst'
          - krank_ist_dispo    : bool
          - krank_abgeleiteter_dienst : str  (z.B. 'T', 'DT', 'N', 'DN', ...)

        Bekannte Dienstzeiten (Startzeit → Dienst):
          Tag    Betreuer : T    06:00–18:00 | T10  09:00–19:00 | T8   10:00–18:00
          Tag    Dispo    : DT   07:00–19:00
          Nacht  Betreuer : N    18:00–06:00 | N10  21:00–07:00
          Nacht  Dispo    : DN   19:00–07:00 | DN3  19:00–07:00
          Tag    Dispo*   : DT3  19:00–07:00  (* schichttechnisch Nacht, aber als Tag klassifiziert)
          Rufbereitschaft : R    Sonderdienst
        """
        # Sonderfall Bauschke: immer Dispo
        ist_dispo_by_name = 'bauschke' in vollname.lower()

        if not start_zeit:
            return 'sonderdienst', ist_dispo_by_name, None

        try:
            sh = int(start_zeit.split(':')[0])
            sm = int(start_zeit.split(':')[1]) if ':' in start_zeit else 0
        except Exception:
            return 'sonderdienst', ist_dispo_by_name, None

        eh = None
        em = None
        if end_zeit and ':' in end_zeit:
            try:
                eh = int(end_zeit.split(':')[0])
                em = int(end_zeit.split(':')[1])
            except Exception:
                pass

        # ── Exaktes Zeitmatching → Dienst-Kürzel ableiten ─────────────────────
        abgeleitet = None
        ist_dispo  = ist_dispo_by_name

        if sh == 6 and sm == 0 and eh == 18:
            abgeleitet = 'T';   ist_dispo = False
        elif sh == 9 and sm == 0 and eh == 19:
            abgeleitet = 'T10'; ist_dispo = False
        elif sh == 10 and sm == 0 and eh == 18:
            abgeleitet = 'T8';  ist_dispo = False
        elif sh == 7 and sm == 0 and eh == 19:
            abgeleitet = 'DT';  ist_dispo = True
        elif sh == 18 and sm == 0 and eh == 6:
            abgeleitet = 'N';   ist_dispo = False
        elif sh == 21 and sm == 0 and eh == 7:
            abgeleitet = 'N10'; ist_dispo = False
        elif sh == 19 and sm == 0 and eh == 7:
            abgeleitet = 'DN';  ist_dispo = True   # DN / DN3 / DT3 – Zeiten identisch
        elif ist_dispo_by_name:
            ist_dispo = True  # Bauschke: Dispo egal welche Zeit

        # ── Schichttyp bestimmen (Tag = Start 05:00–14:59) ────────────────────
        if 5 <= sh < 15:
            schicht_typ = 'tagdienst'
        elif 15 <= sh <= 23 or 0 <= sh < 5:
            schicht_typ = 'nachtdienst'
        else:
            schicht_typ = 'sonderdienst'

        # Kein exakter Treffer und keine Dispo-Erkennung → Sonderdienst
        if abgeleitet is None and not ist_dispo_by_name:
            # Grobe Fallback-Kennzeichnung
            if schicht_typ == 'tagdienst':
                abgeleitet = 'T(?)'
            elif schicht_typ == 'nachtdienst':
                abgeleitet = 'N(?)'
            else:
                abgeleitet = 'S(?)'  # Sonderdienst
                schicht_typ = 'sonderdienst'

        return schicht_typ, ist_dispo, abgeleitet

    def _generate_display_names(self, personen_liste: list, doppelte_nachnamen: set):
        """Hängt bei doppelten Nachnamen die ersten zwei Vorname-Buchstaben an."""
        for person in personen_liste:
            nachname = person['nachname']
            vorname  = person['vorname']
            if nachname in doppelte_nachnamen or '-' in nachname:
                if len(vorname) >= 2:
                    kurz = vorname[0].upper() + vorname[1].lower()
                elif len(vorname) == 1:
                    kurz = vorname[0].upper()
                else:
                    kurz = ''
                person['anzeigename'] = f"{nachname} {kurz}".strip()
            else:
                person['anzeigename'] = nachname
